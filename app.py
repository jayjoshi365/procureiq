# ── Standard library ──────────────────────────────────────────────────────
import asyncio
import concurrent.futures
import html
import math
from typing import Any, cast, Dict, List, Mapping, Optional, Tuple
import time

# ── Third-party — core ────────────────────────────────────────────────────
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
import streamlit.components.v1 as components

# ── Third-party — ML / fuzzy matching ─────────────────────────────────────
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    _SKLEARN_AVAILABLE = True
except ImportError:
    _SKLEARN_AVAILABLE = False
    TfidfVectorizer = None
    cosine_similarity = None

# ── Third-party — finance data ─────────────────────────────────────────────
try:
    import yfinance as yf
    _YFINANCE_AVAILABLE = True
except ImportError:
    yf = None
    _YFINANCE_AVAILABLE = False

# ── Third-party — API framework (for integration) ─────────────────────────
try:
    from fastapi import FastAPI, HTTPException, Request
    from pydantic import BaseModel
    _FASTAPI_AVAILABLE = True
except ImportError:
    FastAPI = None
    BaseModel = None
    _FASTAPI_AVAILABLE = False

# ── Third-party — Security ────────────────────────────────────────────────
try:
    import jwt
    from passlib.context import CryptContext
    from datetime import datetime, timedelta
    _SECURITY_AVAILABLE = True
except ImportError:
    jwt = None
    CryptContext = None
    datetime = None
    timedelta = None
    _SECURITY_AVAILABLE = False

# ── Local modules ─────────────────────────────────────────────────────────
from config import (DIMENSIONS, CURRENT_DIMS, FUTURE_DIMS, FINANCIAL_FIELDS, POSITION_COLORS,
                     KRALJIC_INFO, CATEGORY_RULES, DEFAULT_RFP_STAKEHOLDERS, RFP_TIMELINE,
                     PHASE_COLORS, AUCTION_TYPES, INTAKE_QUESTIONS, AI_TOOLS, AI_PROMPT_MODES,
                     DEFAULT_RFP_QUESTIONS, SCORING_RUBRICS, USE_CASE_TEMPLATES)
from taxonomy import CATEGORY_TAXONOMY, SUBCATEGORY_TAXONOMY
from evaluation import (get_subcategory_weights, recommend_auction_type,
                        compute_financial_health, financial_risk_label,
                        compute_edgar_financial_health)
from rfp import get_rfp_questions
from market_data import (get_market_data_for_supplier, _MARKET_DATA_AVAILABLE,
                         MARKET_LEADERS, DEFAULT_MARKET_LEADERS, get_market_leaders_extended)
from utils import (sx, hex_to_rgba, format_currency, format_percentage, safe_divide,
                   validate_email, clean_text, truncate_text, get_unique_values, filter_data,
                   sort_data, paginate_data, export_to_csv, calculate_stats, generate_id,
                   parse_date, format_date, get_file_extension, is_valid_file_type,
                   calculate_completion_percentage, format_duration)
from auth import require_authentication
from validation import validate_all_inputs, error_handler
from database import get_database

# ── AI Agents ─────────────────────────────────────────────────────────────
try:
    from agents.supplier_discovery  import run_supplier_discovery_agent
    from agents.supplier_enrichment import run_supplier_enrichment_agent
    from agents.sanctions_check     import run_sanctions_check
    from agents.contract_generation import run_contract_generation_agent
    from agents.erp_connector       import run_erp_connector
    from agents.spend_anomaly       import run_spend_anomaly_agent
    from agents.intake_agent        import run_intake_agent, get_intake_session_values
    from agents.tenant_provisioning import provision_organization, get_org_config, list_organizations
    from agents.llm_router          import call_llm as _router_call_llm, stream_llm as _router_stream_llm, PROVIDERS as _LLM_PROVIDERS, get_server_key as _llm_server_key
    from agents.supplier_risk_monitor import run_supplier_risk_monitor, get_xbrl_financial_ratios
    from agents.contract_tracker    import run_contract_tracker
    from agents.spend_forecast      import run_spend_forecast, parse_spend_csv as _sf_parse_csv
    _AGENTS_AVAILABLE = True
except ImportError as _agent_import_err:
    _AGENTS_AVAILABLE = False
    def run_supplier_discovery_agent(*a, **kw): return {"error": "agents package not found", "suppliers": []}
    def run_supplier_enrichment_agent(*a, **kw): return {"error": "agents package not found"}
    def run_sanctions_check(*a, **kw): return {"overall_status": "SKIP", "summary": "agents package not found"}
    def run_contract_generation_agent(*a, **kw): return {"error": "agents package not found", "html": ""}
    def run_erp_connector(*a, **kw): return {"error": "agents package not found"}
    def run_spend_anomaly_agent(*a, **kw): return {"error": "agents package not found", "anomalies": []}
    def run_intake_agent(*a, **kw): return {"reply": "Install agents package.", "session_updates": {}, "complete": False, "conversation": []}
    def get_intake_session_values(*a, **kw): return {}
    def provision_organization(*a, **kw): return {"success": False, "error": "agents package not found"}
    def get_org_config(*a, **kw): return {}
    def list_organizations(*a, **kw): return []
    def _router_call_llm(*a, **kw): raise RuntimeError("Agent services unavailable — agents package not installed")
    def _router_stream_llm(*a, **kw): raise RuntimeError("Agent services unavailable — agents package not installed"); yield  # noqa: unreachable
    def _llm_server_key(*a, **kw): return ""
    _LLM_PROVIDERS = {"claude": {"name": "Claude (Anthropic)", "models": {"claude-sonnet-4-6": "Sonnet 4.6"}, "default_model": "claude-sonnet-4-6", "key_placeholder": "sk-ant-...", "key_help": "console.anthropic.com"}}
    def run_supplier_risk_monitor(*a, **kw): return {"error": "agents package not found", "supplier_scores": [], "portfolio_summary": {}}
    def run_contract_tracker(*a, **kw): return {"error": "agents package not found", "obligations": [], "risk_flags": []}
    def run_spend_forecast(*a, **kw): return {"error": "agents package not found", "forecast": {}, "narrative": {}}
    def get_xbrl_financial_ratios(*a, **kw): return {}
    def _sf_parse_csv(s): return []

# ── Optional advanced modules ─────────────────────────────────────────────
try:
    from realtime_data import get_realtime_provider
    _REALTIME_AVAILABLE = True
except ImportError:
    get_realtime_provider = None
    _REALTIME_AVAILABLE = False

try:
    from security import get_security_manager, get_mfa_manager, get_audit_logger
    _ADVANCED_SECURITY_AVAILABLE = True
except ImportError:
    get_security_manager = None
    get_mfa_manager = None
    get_audit_logger = None
    _ADVANCED_SECURITY_AVAILABLE = False

# ── Environment variables — load .env if present (never committed to git) ─
import os
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv(override=False)  # Does not override vars already set in the shell
except ImportError:
    pass  # python-dotenv not installed — env vars must be set in shell or deployment config

# ── Anthropic Claude API ──────────────────────────────────────────────────
try:
    import anthropic as _anthropic_module
    _ANTHROPIC_AVAILABLE = True
except ImportError:
    _anthropic_module = None
    _ANTHROPIC_AVAILABLE = False


_CLAUDE_MODEL = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-6")


def _get_model() -> str:
    """Return the currently selected model (from session or provider default)."""
    provider = _get_provider()
    default = _LLM_PROVIDERS.get(provider, {}).get("default_model", _CLAUDE_MODEL)
    return st.session_state.get("_user_ai_model", default)


def call_claude_api(prompt: str, model: str = "") -> str:
    """
    Call the active AI provider and return the full response text.
    Routes to Claude, OpenAI, Deepseek, or Grok based on user's sidebar selection.
    """
    api_key = _get_api_key()
    provider = _get_provider()
    resolved_model = model or _get_model()
    try:
        return _router_call_llm(
            messages=[{"role": "user", "content": prompt}],
            provider=provider,
            api_key=api_key,
            model=resolved_model,
        )
    except RuntimeError:
        return ""


def stream_claude_api(prompt: str, model: str = ""):
    """
    Stream tokens from the active AI provider.
    Routes to Claude, OpenAI, Deepseek, or Grok based on user's sidebar selection.
    """
    api_key = _get_api_key()
    provider = _get_provider()
    resolved_model = model or _get_model()
    try:
        yield from _router_stream_llm(
            messages=[{"role": "user", "content": prompt}],
            provider=provider,
            api_key=api_key,
            model=resolved_model,
        )
    except RuntimeError:
        return


def _get_provider() -> str:
    """Return the currently selected AI provider (from session or env default)."""
    return st.session_state.get("_user_ai_provider",
                                os.getenv("DEFAULT_AI_PROVIDER", "claude"))


def _get_api_key() -> str:
    """
    Resolve the API key for the current provider.
    Priority: user-entered session key → server env var for that provider.
    Never reads from code or the database.
    """
    user_key = st.session_state.get("_user_api_key", "").strip()
    if user_key:
        return user_key
    # Fall back to server-configured env var for the active provider
    provider = _get_provider()
    return _llm_server_key(provider)


def _test_api_key(provider: str, api_key: str) -> dict:
    """
    Send a minimal prompt to the provider.
    Returns {"ok": bool, "warning": bool, "message": str}
      ok=True, warning=False → verified and working
      ok=True, warning=True  → key is valid but rate-limited (429) — save it
      ok=False               → genuinely invalid (401/403) or unreachable

    Note: llm_router catches all provider exceptions and returns them as
    strings ("gpt-4o-mini API error: ...") rather than re-raising, so we
    must inspect the returned string — not just the except block.
    """
    try:
        result = _router_call_llm(
            messages=[{"role": "user", "content": "Reply with the single word: connected"}],
            provider=provider,
            api_key=api_key,
            system="",
            max_tokens=10,
        )
    except Exception as exc:
        result = str(exc)

    r = str(result).strip()

    # llm_router returns errors as plain strings — classify them
    is_error_string = "API error" in r or "Error code" in r or "error" in r.lower()[:60]

    if not is_error_string:
        # Clean model reply — key works
        return {"ok": True, "warning": False, "message": r[:80]}

    # 429 = rate limited — key is real, just throttled
    if "429" in r or "rate limit" in r.lower() or "too many requests" in r.lower() or "quota" in r.lower():
        return {
            "ok": True,
            "warning": True,
            "message": (
                f"Key saved — provider is rate-limiting requests (429). "
                f"Your key is valid but you may be on a free-tier plan with very low RPM limits. "
                f"All features will work once the rate limit resets (usually within 60 seconds)."
            ),
        }

    # 401/403/invalid key
    return {"ok": False, "warning": False, "message": r[:300]}


def _render_api_key_sidebar():
    """
    Sidebar: read-only AI status display. All interactive config is in the Settings tab.
    """
    with st.sidebar:
        st.markdown("---")
        provider = _get_provider()
        info = _LLM_PROVIDERS.get(provider, {})
        key = _get_api_key()
        if key:
            st.markdown(
                f'<div style="font-size:0.78rem;color:#4ADE80;font-family:monospace">'
                f'AI ✓ {info.get("name", provider)}</div>',
                unsafe_allow_html=True,
            )
            model = st.session_state.get("_user_ai_model", info.get("default_model", ""))
            models = info.get("models", {})
            st.caption(models.get(model, model))
        else:
            st.markdown(
                '<div style="font-size:0.78rem;color:#F59E0B;font-family:monospace">'
                '⚠ AI not configured</div>',
                unsafe_allow_html=True,
            )
            st.caption("→ Use the ⚙ AI Settings panel above the tabs")


def _render_no_key_banner():
    """
    Full-width amber warning banner shown when no API key is configured.
    Disappears once a key is set. User can also dismiss it for the session.
    """
    if _get_api_key():
        return
    if st.session_state.get("_banner_dismissed"):
        return

    col_text, col_btn = st.columns([5, 1])
    with col_text:
        st.markdown(
            '<div style="'
            'background:rgba(245,158,11,0.1);'
            'border:1px solid rgba(245,158,11,0.4);'
            'border-left:4px solid #F59E0B;'
            'border-radius:8px;'
            'padding:0.65rem 1rem;'
            'display:flex;align-items:center;gap:0.75rem'
            '">'
            '<span style="font-size:1.1rem">🔑</span>'
            '<div>'
            '<div style="font-family:monospace;font-size:0.72rem;color:#F59E0B;'
            'text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.15rem">'
            'AI Features Disabled</div>'
            '<div style="font-size:0.83rem;color:#D0E0EF">'
            'No API key is configured. All AI agents (contract generation, supplier discovery, '
            'intake chat, spend analysis) are unavailable. '
            '<strong style="color:#FCD34D">→ Click the ⚙ AI Settings panel just above the tabs to add your key.</strong>'
            '</div>'
            '</div>'
            '</div>',
            unsafe_allow_html=True,
        )
    with col_btn:
        if st.button("Dismiss", key="_dismiss_banner", use_container_width=True):
            st.session_state["_banner_dismissed"] = True
            st.rerun()

    st.markdown("<div style='margin-bottom:0.5rem'></div>", unsafe_allow_html=True)


def _render_settings_tab():
    """
    Full Settings tab — the single source of truth for AI provider configuration.
    """
    st.markdown("### ⚙ Settings")
    st.markdown(
        '<p style="color:#94A3B8;font-size:0.88rem;margin-top:-0.5rem">'
        'Configure your AI provider. Your key is stored only in your browser session — '
        'it is never saved to the database or shared with anyone.'
        '</p>',
        unsafe_allow_html=True,
    )

    # ── Current status card ──────────────────────────────────
    key = _get_api_key()
    provider = _get_provider()
    info = _LLM_PROVIDERS.get(provider, {})
    server_key = _llm_server_key(provider)

    if server_key:
        st.success(f"Admin-configured key active for {info.get('name', provider)}. No action needed.")
    elif key:
        st.success(f"API key active — {info.get('name', provider)} is ready.")
    else:
        st.warning("No API key configured. AI features are disabled until you add a key below.")

    st.markdown("---")

    # ── Provider selector ────────────────────────────────────
    st.markdown("#### Step 1 — Choose your AI provider")

    provider_keys = list(_LLM_PROVIDERS.keys())
    current_provider = st.session_state.get("_user_ai_provider", "claude")
    if current_provider not in provider_keys:
        current_provider = "claude"

    _PROVIDER_DETAILS = {
        "claude": {
            "icon": "🟣", "cost": "~$3 / 1M tokens", "link": "console.anthropic.com",
            "badge": "RECOMMENDED", "badge_color": "#4ADE80",
            "desc": "Built for this tool. All AI agents (supplier discovery, CFO narrative, risk analysis) require Claude. Best quality for complex procurement reasoning.",
            "rate_limits": "Generous limits on all paid tiers. New accounts get $5 free credit.",
            "warning": None,
        },
        "openai": {
            "icon": "🟢", "cost": "~$0.15–5 / 1M tokens", "link": "platform.openai.com",
            "badge": "LIMITED — agents disabled", "badge_color": "#F87171",
            "desc": "Text generation only. Supplier discovery, CFO narrative, and all tool-use agents do not work with OpenAI — those features require Claude.",
            "rate_limits": "Free tier: 3 RPM / 200 req/day. Paid tier ($5+ credit): 500 RPM.",
            "warning": "⚠ Switching away from Claude disables AI agents. Supplier discovery and CFO Challenge generation will not work.",
        },
        "deepseek": {
            "icon": "🔵", "cost": "~$0.27 / 1M tokens", "link": "platform.deepseek.com",
            "badge": "LIMITED — agents disabled", "badge_color": "#F87171",
            "desc": "Text generation only. AI agents (supplier discovery, CFO narrative) require Claude and will not work with DeepSeek.",
            "rate_limits": "60 RPM on free tier.",
            "warning": "⚠ Switching away from Claude disables AI agents. Supplier discovery and CFO Challenge generation will not work.",
        },
        "grok": {
            "icon": "⚫", "cost": "~$3 / 1M tokens", "link": "console.x.ai",
            "badge": "LIMITED — agents disabled", "badge_color": "#F87171",
            "desc": "Text generation only. AI agents (supplier discovery, CFO narrative) require Claude and will not work with Grok.",
            "rate_limits": "Standard paid-tier limits. No free tier.",
            "warning": "⚠ Switching away from Claude disables AI agents. Supplier discovery and CFO Challenge generation will not work.",
        },
    }

    # ── Claude card (primary, always visible) ───────────────
    _claude_info  = _LLM_PROVIDERS["claude"]
    _claude_det   = _PROVIDER_DETAILS["claude"]
    _is_claude    = (current_provider == "claude")
    _border       = "#3B82F6" if _is_claude else "rgba(148,163,184,0.15)"
    _bg           = "rgba(59,130,246,0.08)" if _is_claude else "#0A1628"
    _c_card, _c_pick = st.columns([5, 1])
    with _c_card:
        st.markdown(
            f'<div style="background:{_bg};border:1px solid {_border};'
            f'border-radius:8px;padding:0.65rem 0.9rem;margin-bottom:0.3rem">'
            f'<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.25rem;flex-wrap:wrap">'
            f'<span style="font-size:1rem">{_claude_det["icon"]}</span>'
            f'<strong style="color:#E2E8F0;font-size:0.9rem">{_claude_info["name"]}</strong>'
            f'<span style="font-family:monospace;font-size:0.7rem;color:#60A5FA;'
            f'background:rgba(96,165,250,0.1);border-radius:3px;padding:0.05rem 0.3rem">'
            f'{_claude_det["cost"]}</span>'
            f'<span style="background:#4ADE8022;border:1px solid #4ADE8055;border-radius:4px;'
            f'padding:0.1rem 0.45rem;font-size:0.68rem;color:#4ADE80;font-weight:700;'
            f'letter-spacing:0.05em">RECOMMENDED — FULL FUNCTIONALITY</span>'
            f'</div>'
            f'<div style="font-size:0.82rem;color:#94A3B8;margin-bottom:0.2rem">{_claude_det["desc"]}</div>'
            f'<div style="font-size:0.75rem;color:#64748B;margin-bottom:0.15rem">'
            f'Rate limits: {_claude_det["rate_limits"]}</div>'
            f'<div style="font-size:0.78rem;color:#60A5FA">'
            f'Get your key → <strong>{_claude_det["link"]}</strong></div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    with _c_pick:
        _lbl = "✓ Selected" if _is_claude else "Select"
        if st.button(_lbl, key="_pick_provider_claude", use_container_width=True,
                     type="primary" if _is_claude else "secondary"):
            if not _is_claude:
                st.session_state["_user_ai_provider"] = "claude"
                st.session_state["_user_api_key"] = ""
                st.session_state["_banner_dismissed"] = False
                st.rerun()

    # ── Alternative providers (experimental, collapsed) ──────
    with st.expander("🔬 Alternative providers (experimental — limited functionality)", expanded=False):
        st.warning(
            "**These providers do not support AI agents.** "
            "Supplier discovery, CFO narrative generation, and all tool-use features require Claude. "
            "Use an alternative provider only for basic text generation experiments.",
            icon="⚠️",
        )
        for pk in [p for p in provider_keys if p != "claude"]:
            pinfo   = _LLM_PROVIDERS[pk]
            pd_info = _PROVIDER_DETAILS.get(pk, {})
            is_sel  = (pk == current_provider)
            border  = "#3B82F6" if is_sel else "rgba(148,163,184,0.15)"
            bg      = "rgba(59,130,246,0.08)" if is_sel else "#0A1628"

            c_card, c_pick = st.columns([5, 1])
            with c_card:
                st.markdown(
                    f'<div style="background:{bg};border:1px solid {border};'
                    f'border-radius:8px;padding:0.65rem 0.9rem;margin-bottom:0.3rem">'
                    f'<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.25rem;flex-wrap:wrap">'
                    f'<span style="font-size:1rem">{pd_info.get("icon","⚪")}</span>'
                    f'<strong style="color:#E2E8F0;font-size:0.9rem">{pinfo["name"]}</strong>'
                    f'<span style="font-family:monospace;font-size:0.7rem;color:#60A5FA;'
                    f'background:rgba(96,165,250,0.1);border-radius:3px;padding:0.05rem 0.3rem">'
                    f'{pd_info.get("cost","")}</span>'
                    f'<span style="background:#F8717122;border:1px solid #F8717155;border-radius:4px;'
                    f'padding:0.1rem 0.45rem;font-size:0.68rem;color:#F87171;font-weight:700;'
                    f'letter-spacing:0.05em">LIMITED — AGENTS DISABLED</span>'
                    f'</div>'
                    f'<div style="font-size:0.82rem;color:#94A3B8;margin-bottom:0.2rem">{pd_info.get("desc","")}</div>'
                    f'<div style="font-size:0.75rem;color:#64748B;margin-bottom:0.15rem">'
                    f'Rate limits: {pd_info.get("rate_limits","")}</div>'
                    f'<div style="font-size:0.78rem;color:#60A5FA">'
                    f'Get your key → <strong>{pd_info.get("link","")}</strong></div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            with c_pick:
                btn_label = "✓ Selected" if is_sel else "Select"
                if st.button(btn_label, key=f"_pick_provider_{pk}", use_container_width=True,
                             type="primary" if is_sel else "secondary"):
                    if pk != current_provider:
                        st.session_state["_user_ai_provider"] = pk
                        st.session_state["_user_api_key"] = ""
                        st.session_state["_banner_dismissed"] = False
                        st.rerun()

    st.markdown("---")

    # ── API key input ────────────────────────────────────────
    st.markdown("#### Step 2 — Enter your API key")
    pinfo = _LLM_PROVIDERS.get(current_provider, {})

    if server_key:
        st.info(f"A server-configured key is already active for {pinfo.get('name', current_provider)}. "
                "You can override it below if you want to use your own key instead.")

    st.markdown(
        f'<p style="font-size:0.83rem;color:#94A3B8">Your key for '
        f'<strong style="color:#E2E8F0">{pinfo.get("name", current_provider)}</strong>. '
        f'Get it at <strong style="color:#60A5FA">{_PROVIDER_DETAILS.get(current_provider, {}).get("link", "")}</strong>. '
        f'Keys start with <code style="background:rgba(96,165,250,0.1);padding:0.1rem 0.3rem;border-radius:3px">'
        f'{pinfo.get("key_placeholder", "...")}</code></p>',
        unsafe_allow_html=True,
    )

    entered_key = st.text_input(
        "API Key",
        value=st.session_state.get("_user_api_key", ""),
        type="password",
        placeholder=pinfo.get("key_placeholder", "Enter your API key here..."),
        key="_settings_api_key_input",
        label_visibility="collapsed",
    )

    col_save, col_clear = st.columns([1, 1])
    with col_save:
        if st.button("💾 Save & Verify Key", key="_settings_save_key", type="primary", use_container_width=True):
            _k = entered_key.strip()
            if not _k:
                st.session_state["_key_test_result"] = {"ok": False, "message": "No key entered — paste your API key above."}
            else:
                with st.spinner("Verifying key with provider…"):
                    _vr = _test_api_key(current_provider, _k)
                st.session_state["_key_test_result"] = _vr
                if _vr["ok"]:  # save on both verified and rate-limited (both mean key is real)
                    st.session_state["_user_api_key"] = _k
                    st.session_state["_banner_dismissed"] = True
    with col_clear:
        if st.button("🗑 Clear Key", key="_settings_clear_key", use_container_width=True):
            st.session_state["_user_api_key"] = ""
            st.session_state["_banner_dismissed"] = False
            st.session_state.pop("_key_test_result", None)
            st.rerun()

    # Show validation result inline — always visible, not inside expander collapse
    _test_result = st.session_state.get("_key_test_result")
    if _test_result:
        if _test_result.get("ok") and not _test_result.get("warning"):
            st.success(f"✓ Key verified and saved — {_test_result['message']}")
        elif _test_result.get("ok") and _test_result.get("warning"):
            st.warning(_test_result["message"])
            _cur_prov = st.session_state.get("_user_ai_provider", "claude")
            if _cur_prov == "openai":
                st.markdown(
                    '<div style="background:rgba(96,165,250,0.07);border:1px solid rgba(96,165,250,0.2);'
                    'border-radius:8px;padding:0.7rem 1rem;font-size:0.85rem;color:#93C5FD">'
                    '<strong>Two fixes for 429:</strong><br/>'
                    '1. <strong>Switch to Claude</strong> — select it above, get a free key at '
                    '<strong>console.anthropic.com</strong> (new accounts get $5 credit, no card required for signup).<br/>'
                    '2. <strong>Add $5 credit to your OpenAI account</strong> — '
                    'platform.openai.com/settings/billing — this upgrades you from 3 RPM to 500 RPM instantly.'
                    '</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.error(
                f"✗ Invalid key — {_test_result['message']}\n\n"
                "Check the key is copied correctly and matches the selected provider above."
            )

    st.markdown("---")

    # ── Model selector ───────────────────────────────────────
    st.markdown("#### Step 3 — Choose model")
    models = pinfo.get("models", {})
    model_keys = list(models.keys())
    if model_keys:
        current_model = st.session_state.get("_user_ai_model", pinfo.get("default_model", model_keys[0]))
        if current_model not in model_keys:
            current_model = model_keys[0]

        _MODEL_NOTES = {
            "claude-sonnet-4-6":         "Best for contracts, strategy, and complex analysis. Slower and pricier.",
            "claude-haiku-4-5-20251001": "Fast and cheap. Best for intake chat, summaries, and quick lookups.",
            "gpt-4o":                    "OpenAI's best model. Matches Claude Sonnet quality.",
            "gpt-4o-mini":               "10× cheaper than GPT-4o. Good for most procurement tasks.",
            "deepseek-chat":             "Deepseek V3 — best value across all providers.",
            "deepseek-reasoner":         "Deepseek R1 — slower but better at multi-step reasoning.",
            "grok-3-mini":               "Fast and cheap. Good for standard analysis.",
            "grok-3":                    "Grok's best quality. Use for complex contract and strategy work.",
        }

        for mk in model_keys:
            is_model_selected = (mk == current_model)
            m_border = "#3B82F6" if is_model_selected else "rgba(148,163,184,0.12)"
            m_bg = "rgba(59,130,246,0.07)" if is_model_selected else "#0A1628"
            mc1, mc2 = st.columns([5, 1])
            with mc1:
                st.markdown(
                    f'<div style="background:{m_bg};border:1px solid {m_border};'
                    f'border-radius:6px;padding:0.5rem 0.75rem;margin-bottom:0.25rem">'
                    f'<div style="font-size:0.85rem;color:#E2E8F0;font-weight:600">{models[mk]}</div>'
                    f'<div style="font-size:0.78rem;color:#94A3B8">{_MODEL_NOTES.get(mk, "")}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            with mc2:
                lbl = "✓ Active" if is_model_selected else "Use"
                if st.button(lbl, key=f"_pick_model_{mk}", use_container_width=True,
                             type="primary" if is_model_selected else "secondary"):
                    st.session_state["_user_ai_model"] = mk
                    st.rerun()

    st.markdown("---")

    # ── Security note ────────────────────────────────────────
    st.markdown(
        '<div style="background:rgba(148,163,184,0.05);border:1px solid rgba(148,163,184,0.12);'
        'border-radius:8px;padding:0.75rem 1rem">'
        '<div style="font-family:monospace;font-size:0.7rem;color:#60A5FA;text-transform:uppercase;'
        'letter-spacing:0.1em;margin-bottom:0.4rem">🔒 Security</div>'
        '<div style="font-size:0.82rem;color:#94A3B8;line-height:1.6">'
        'Your API key is stored only in your browser session and is cleared when you close the tab. '
        'It is never written to the database, sent to ProcureIQ servers, or logged anywhere. '
        'All AI calls go directly from this app to the provider using your key — '
        'usage appears on your provider account, not ours.'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )


def display_ai_governance_banner():
    """Display governance banner for AI-generated content to comply with EU AI Act and procurement policies."""
    st.markdown(
        '<div style="background:rgba(59,130,246,0.08);border-left:4px solid #3B82F6;border-radius:0 8px 8px 0;'
        'padding:0.75rem 1rem;margin-bottom:1rem;margin-top:0.5rem">'
        '<div style="font-family:monospace;font-size:0.80rem;color:#3B82F6;text-transform:uppercase;'
        'letter-spacing:0.12em;margin-bottom:0.35rem">AI-Assisted Analysis — Review Required</div>'
        '<div style="font-size:0.8rem;color:#D0E0EF;line-height:1.5">'
        f'This analysis was produced by {_LLM_PROVIDERS.get(_get_provider(), {}).get("name", "AI")} using the data and parameters you entered. '
        'It is decision-support only. All outputs must be reviewed by a qualified procurement professional before use in sourcing decisions, contract awards, or stakeholder communications.'
        '</div></div>',
        unsafe_allow_html=True,
    )


def _load_hris_sample_data():
    """Populate session state with a realistic HRIS evaluation example."""
    # Event settings
    st.session_state["ctrl_event"]       = "HRIS Platform Selection — 2025"
    st.session_state["ctrl_category"]    = "HRIS / HCM Platform"
    st.session_state["ctrl_kraljic"]     = "Strategic"
    st.session_state["ctrl_suppliers"]   = 3
    st.session_state["ctrl_stakeholders"] = 3
    st.session_state["selected_parent_cat"] = "human_resources"
    st.session_state["selected_sub_name"]   = "HRIS / HCM Platform"

    # Supplier 0 — Workday
    st.session_state["name_0"]          = "Workday"
    st.session_state["raw_price_0"]     = 480000.0
    st.session_state["sla_0"]           = "Strong"
    st.session_state["risk_0"]          = "Low"
    st.session_state["stake_0"]         = 5
    st.session_state["strategic_0"]     = 5
    st.session_state["innovation_0"]    = 5
    st.session_state["relationship_0"]  = 4
    st.session_state["flexibility_0"]   = 3
    st.session_state["esg_0"]           = "Strong"
    st.session_state["diversity_0"]     = "Moderate"
    st.session_state["notes_0"]         = "Strong enterprise HR suite. High implementation complexity. Premium pricing justified by automation depth."
    st.session_state["ticker_0"]        = "WDAY"

    # Supplier 1 — UKG (UltiPro)
    st.session_state["name_1"]          = "UKG Pro"
    st.session_state["raw_price_1"]     = 390000.0
    st.session_state["sla_1"]           = "Moderate"
    st.session_state["risk_1"]          = "Medium"
    st.session_state["stake_1"]         = 3
    st.session_state["strategic_1"]     = 4
    st.session_state["innovation_1"]    = 3
    st.session_state["relationship_1"]  = 4
    st.session_state["flexibility_1"]   = 4
    st.session_state["esg_1"]           = "Moderate"
    st.session_state["diversity_1"]     = "Moderate"
    st.session_state["notes_1"]         = "Strong mid-market HCM. Good payroll integration. Less enterprise-scale than Workday."
    st.session_state["ticker_1"]        = ""

    # Supplier 2 — SAP SuccessFactors
    st.session_state["name_2"]          = "SAP SuccessFactors"
    st.session_state["raw_price_2"]     = 520000.0
    st.session_state["sla_2"]           = "Strong"
    st.session_state["risk_2"]          = "High"
    st.session_state["stake_2"]         = 3
    st.session_state["strategic_2"]     = 4
    st.session_state["innovation_2"]    = 3
    st.session_state["relationship_2"]  = 3
    st.session_state["flexibility_2"]   = 2
    st.session_state["esg_2"]           = "Strong"
    st.session_state["diversity_2"]     = "Moderate"
    st.session_state["notes_2"]         = "Excellent if already on SAP ERP. Highest implementation risk and rigidity. Strong compliance module."
    st.session_state["ticker_2"]        = "SAP"

    # Stakeholder 0 — CHRO
    st.session_state["stake_name_0"]    = "Sarah Chen"
    st.session_state["stake_role_0"]    = "CPO"
    st.session_state["stake_power_0"]   = 9
    st.session_state["stake_interest_0"] = 9
    st.session_state["stake_position_0"] = "Champion"
    st.session_state["stake_priority_0"] = "Quality / SLA"

    st.session_state["_onboarding_shown"]      = True
    st.session_state["_sample_data_loaded"]    = True


def show_onboarding_modal():
    """Display onboarding guide for first-time users with sample data option."""
    if st.session_state.get("_onboarding_shown"):
        return

    st.markdown(
        '<div style="background:rgba(34,197,94,0.08);border:2px solid rgba(74,222,128,0.3);border-radius:12px;'
        'padding:1.5rem;margin-bottom:1.5rem">'
        '<div style="display:flex;align-items:center;gap:0.8rem;margin-bottom:1rem">'
        '<div style="font-size:1.5rem">👋</div>'
        '<div style="font-size:1.2rem;font-weight:700;color:#86EFAC">Welcome to ProcureIQ!</div>'
        '</div>'
        '<div style="font-size:0.9rem;color:#CBD5E1;line-height:1.6;margin-bottom:1rem">'
        'Guides you from category intake through supplier evaluation, market analysis, and executive briefing — '
        'powered by Kraljic methodology and AI-enhanced contract guidance.'
        '</div>'
        '<div style="background:rgba(0,0,0,0.2);border-radius:8px;padding:1rem;margin-bottom:1rem">'
        '<div style="font-weight:600;color:#4ADE80;margin-bottom:0.5rem">⏱️ Typical workflow (10-15 min):</div>'
        '<ul style="font-size:0.85rem;color:#CBD5E1;margin:0.5rem 0 0 1.5rem">'
        '<li><strong>Intake:</strong> Select category, Kraljic posture, number of suppliers</li>'
        '<li><strong>Suppliers:</strong> Enter names, pricing, SLA, risk, ESG scores</li>'
        '<li><strong>Strategy:</strong> Review Kraljic positioning and 90-day plan</li>'
        '<li><strong>Award Brief:</strong> Export executive memo and one-pager</li>'
        '</ul>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    _ob_col1, _ob_col2 = st.columns(2)
    with _ob_col1:
        if st.button("📋 Load HRIS Sample Evaluation", use_container_width=True, key="onboard_sample_btn",
                     help="Pre-fills a realistic HRIS platform evaluation: Workday vs UKG Pro vs SAP SuccessFactors"):
            _load_hris_sample_data()
            st.success("Sample data loaded — see Suppliers and Strategy tabs.")
            st.rerun()
    with _ob_col2:
        if st.button("▶ Start Fresh", use_container_width=True, key="onboard_fresh_btn"):
            st.session_state["_onboarding_shown"] = True
            st.rerun()


# =========================================================
# INTEGRATION API (Adjacent to Enterprise Systems)
# =========================================================
if _FASTAPI_AVAILABLE:
    db = get_database()

    def _create_api_app():
        return FastAPI(
            title="ProcureIQ API",
            description="Enterprise-grade procurement decision support API for sourcing event management, supplier evaluation, and market intelligence",
            version="1.0.0",
            docs_url="/api/docs",
            redoc_url="/api/redoc",
            openapi_url="/api/openapi.json",
            contact={
                "name": "ProcureIQ Support",
                "url": "https://procureiq.dev",
                "email": "support@procureiq.dev"
            },
            license_info={
                "name": "Proprietary",
                "url": "https://procureiq.dev/license"
            },
            tags=[
                {"name": "Spend", "description": "Spend data import and classification endpoints"},
                {"name": "Evaluation", "description": "Sourcing event and supplier evaluation endpoints"},
                {"name": "Market Data", "description": "Real-time market intelligence and supplier financial signals"},
                {"name": "Authentication", "description": "Session and token management endpoints"},
                {"name": "Audit", "description": "Audit log and compliance endpoints"},
            ]
        )
    app_api = _create_api_app()

    # In-memory login rate limiter: ip → {"count": int, "window_start": float}
    _login_attempts: dict = {}
    _LOGIN_MAX_ATTEMPTS = 5   # max failures per window
    _LOGIN_WINDOW_SECS  = 300  # 5-minute sliding window

    def _check_login_rate_limit(ip: str) -> bool:
        """Return True if the IP is allowed to attempt login, False if locked out."""
        now = time.time()
        entry = _login_attempts.get(ip)
        if entry is None:
            return True
        if now - entry["window_start"] > _LOGIN_WINDOW_SECS:
            # Window expired — reset
            del _login_attempts[ip]
            return True
        return entry["count"] < _LOGIN_MAX_ATTEMPTS

    def _record_login_failure(ip: str) -> None:
        now = time.time()
        entry = _login_attempts.get(ip)
        if entry is None or now - entry["window_start"] > _LOGIN_WINDOW_SECS:
            _login_attempts[ip] = {"count": 1, "window_start": now}
        else:
            entry["count"] += 1

    def _clear_login_failures(ip: str) -> None:
        _login_attempts.pop(ip, None)

    class SpendData(BaseModel):
        """Spend line-item data model for import."""
        category: str
        spend: float
        supplier: str
        date: str

    @app_api.post("/import-spend", tags=["Spend"], summary="Import spend line items")
    def import_spend(data: SpendData, token: str = None):
        """
        Import spend data for automatic category classification and spend analytics.
        
        **Authentication:** Optional JWT token via query parameter.
        
        **Input:**
        - `category`: UNSPSC or internal spend category
        - `spend`: Annual spend amount in USD
        - `supplier`: Supplier name or ID
        - `date`: Date of transaction (YYYY-MM-DD format)
        
        **Returns:** Imported spend record with auto-classification if applicable.
        
        **Audit:** All imports are logged with user and timestamp.
        """
        # Validate session if token provided
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        try:
            category = data.category
            if category not in CATEGORY_TAXONOMY:
                raise HTTPException(status_code=400, detail=f"Invalid category: {category}")
            
            # Store in database
            spend_data = {
                "category": category,
                "spend": data.spend,
                "supplier": data.supplier,
                "date": data.date
            }
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "import_spend", f"category:{category}", spend_data)
            
            return {"message": "Spend data imported successfully", "data": spend_data}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app_api.get("/export-decision/{event_id}")
    def export_decision(event_id: str, token: str = None):
        # Validate session
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        try:
            # Get evaluation from database
            evaluation = db.get_evaluation(event_id)
            if not evaluation:
                raise HTTPException(status_code=404, detail="Evaluation not found")
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "export_decision", event_id)
            
            return {
                "event_id": event_id,
                "recommendation": evaluation["recommendation"],
                "scores": evaluation["scores"],
                "created_at": evaluation["created_at"]
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app_api.get("/market-data/{ticker}")
    def get_market_data(ticker: str, token: str = None):
        # Validate session
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        if not _MARKET_DATA_AVAILABLE:
            raise HTTPException(status_code=503, detail="Market data service unavailable")
        
        try:
            # Check cache first
            cached_data = db.get_cached_market_data(ticker)
            if cached_data:
                if user_id:
                    db.log_audit_event(user_id, "view_cached_market_data", ticker)
                return {"ticker": ticker, "data": cached_data, "cached": True}
            
            # Fetch fresh data
            data = get_market_data_for_supplier(ticker.upper())
            if "error" in data:
                raise HTTPException(status_code=404, detail=data["error"])
            
            # Cache the data
            db.cache_market_data(ticker, data)
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "fetch_market_data", ticker)
            
            return {"ticker": ticker, "data": data, "cached": False}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app_api.post("/validate-rfp")
    def validate_rfp_data(rfp_data: dict, token: str = None):
        # Validate session
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        try:
            validation_result = validate_all_inputs(rfp_data)
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "validate_rfp", None, {"valid": validation_result["is_valid"]})
            
            return {"valid": validation_result["is_valid"], "errors": validation_result["errors"]}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    # Enhanced authentication endpoints
    class UserCredentials(BaseModel):
        username: str
        password: str

    @app_api.post("/login")
    def login(credentials: UserCredentials, request: Request):
        client_ip = request.client.host if request.client else "unknown"

        if not _check_login_rate_limit(client_ip):
            db.log_audit_event(credentials.username, "login_rate_limited", "api",
                               {"ip": client_ip})
            raise HTTPException(status_code=429, detail="Too many login attempts — try again in 5 minutes")

        _api_admin_hash = os.getenv("API_ADMIN_PASSWORD_HASH", "")
        if _api_admin_hash and credentials.username == "admin" and verify_password(credentials.password, _api_admin_hash):
            _clear_login_failures(client_ip)
            session_data = {"username": credentials.username, "login_time": time.time()}
            session_id = db.create_session(credentials.username, session_data)
            access_token = create_access_token(data={"sub": credentials.username, "session_id": session_id})
            db.log_audit_event(credentials.username, "login", "api")
            if _ADVANCED_SECURITY_AVAILABLE:
                try:
                    get_audit_logger().log_event("login", credentials.username,
                                                  {"session_id": session_id}, ip_address=client_ip)
                except Exception:
                    pass
            return {"access_token": access_token, "token_type": "bearer", "session_id": session_id}
        else:
            _record_login_failure(client_ip)
            db.log_audit_event(credentials.username, "login_failed", "api",
                               {"ip": client_ip})
            if _ADVANCED_SECURITY_AVAILABLE:
                try:
                    get_audit_logger().log_event("login_failed", credentials.username,
                                                  {"ip": client_ip}, severity="WARNING", ip_address=client_ip)
                except Exception:
                    pass
            raise HTTPException(status_code=401, detail="Invalid credentials")

    @app_api.post("/logout")
    def logout(token: str = None):
        if token:
            try:
                payload = verify_token(token)
                session_id = payload.get("session_id")
                if session_id:
                    db.delete_session(session_id)
                    user_id = payload.get("sub")
                    if user_id:
                        db.log_audit_event(user_id, "logout", "api")
                        if _ADVANCED_SECURITY_AVAILABLE:
                            try:
                                get_audit_logger().log_event("logout", user_id,
                                                              {"session_id": session_id})
                            except Exception:
                                pass
                    return {"message": "Logged out successfully"}
            except:
                pass
        raise HTTPException(status_code=400, detail="Invalid token")

    @app_api.get("/protected-endpoint")
    def protected_endpoint(token: str = None):
        if not token:
            raise HTTPException(status_code=401, detail="Token required")
        payload = verify_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user_id = payload["sub"]
        db.log_audit_event(user_id, "access_protected_endpoint", "api")
        
        return {"message": "Access granted", "user": user_id}

    @app_api.get("/audit-log")
    def get_audit_log(limit: int = 50, token: str = None):
        # Validate session
        if not token:
            raise HTTPException(status_code=401, detail="Token required")
        payload = verify_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user_id = payload["sub"]
        # Only allow users to see their own audit logs
        audit_log = db.get_audit_log(user_id=user_id, limit=limit)
        
        return {"audit_log": audit_log}

    # Real-time data endpoints
    @app_api.get("/realtime/quote/{symbol}")
    def get_realtime_quote(symbol: str, token: str = None):
        if not _REALTIME_AVAILABLE:
            raise HTTPException(status_code=503, detail="Real-time data services unavailable")
        
        # Validate session
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        try:
            provider = get_realtime_provider()
            quote = provider.get_stock_quote(symbol.upper())
            
            if "error" in quote:
                raise HTTPException(status_code=404, detail=quote["error"])
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "realtime_quote", symbol)
            
            return quote
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app_api.get("/realtime/company/{symbol}")
    def get_realtime_company(symbol: str, token: str = None):
        if not _REALTIME_AVAILABLE:
            raise HTTPException(status_code=503, detail="Real-time data services unavailable")
        
        # Validate session
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        try:
            provider = get_realtime_provider()
            overview = provider.get_company_overview(symbol.upper())
            
            if "error" in overview:
                raise HTTPException(status_code=404, detail=overview["error"])
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "realtime_company", symbol)
            
            return overview
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app_api.get("/realtime/news")
    def get_realtime_news(query: str = "procurement OR supply chain", token: str = None):
        if not _REALTIME_AVAILABLE:
            raise HTTPException(status_code=503, detail="Real-time data services unavailable")
        
        # Validate session
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        try:
            provider = get_realtime_provider()
            news = provider.get_market_news(query)
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "realtime_news", query)
            
            return {"news": news}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app_api.get("/realtime/trends")
    def get_realtime_trends(token: str = None):
        if not _REALTIME_AVAILABLE:
            raise HTTPException(status_code=503, detail="Real-time data services unavailable")
        
        # Validate session
        user_id = None
        if token:
            session = db.get_session(token)
            if session:
                user_id = session["user_id"]
        
        try:
            provider = get_realtime_provider()
            trends = provider.get_procurement_market_trends()
            
            # Log audit event
            if user_id:
                db.log_audit_event(user_id, "realtime_trends", None)
            
            return trends
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
else:
    app_api = None  # Fallback if FastAPI not available


# =========================================================
# SECURITY & COMPLIANCE BASE
# =========================================================
try:
    import jwt
    from passlib.context import CryptContext
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

    def verify_password(plain_password, hashed_password):
        return pwd_context.verify(plain_password, hashed_password)

    def get_password_hash(password):
        return pwd_context.hash(password)

    # JWT Configuration — must be set via SECRET_KEY environment variable.
    # Do NOT fall back to a known default string: a public-repo fallback allows
    # anyone to forge valid API tokens. If SECRET_KEY is absent, JWT is disabled.
    SECRET_KEY = os.getenv("SECRET_KEY", "")
    if not SECRET_KEY:
        import warnings as _sec_warnings
        _sec_warnings.warn(
            "SECRET_KEY environment variable is not set. "
            "FastAPI JWT endpoints are disabled. Set SECRET_KEY=<random-string> in .env before deployment.",
            stacklevel=2,
        )
    ALGORITHM = "HS256"
    ACCESS_TOKEN_EXPIRE_MINUTES = 30

    def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
        if not SECRET_KEY:
            raise RuntimeError("SECRET_KEY is not configured — JWT token creation is disabled.")
        to_encode = data.copy()
        if expires_delta:
            expire = datetime.utcnow() + expires_delta
        else:
            expire = datetime.utcnow() + timedelta(minutes=15)
        to_encode.update({"exp": expire})
        encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
        return encoded_jwt

    def verify_token(token: str):
        if not SECRET_KEY:
            return None
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            return payload
        except jwt.ExpiredSignatureError:
            return None
        except jwt.JWTError:
            return None

except ImportError:
    jwt = None
    pwd_context = None

    # Safe stubs — always defined so callers never hit NameError.
    # verify_password returns False (never grants access).
    # create_access_token raises so callers see a clear error, not a silent pass.
    def verify_password(plain_password, hashed_password):  # type: ignore[misc]
        return False

    def get_password_hash(password):  # type: ignore[misc]
        raise RuntimeError("passlib is required for password hashing — install it with: pip install passlib[bcrypt]")

    def create_access_token(data: dict, expires_delta=None):  # type: ignore[misc]
        raise RuntimeError("PyJWT and passlib are required for JWT tokens — install them with: pip install PyJWT passlib[bcrypt]")

    def verify_token(token: str):  # type: ignore[misc]
        return None

    SECRET_KEY = ""


# =========================================================
# PAGE CONFIG
# =========================================================
st.set_page_config(
    page_title="ProcureIQ",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# =========================================================
# SESSION STATE
# =========================================================
if "entered_dashboard" not in st.session_state:
    st.session_state.entered_dashboard = False
if "entered_express" not in st.session_state:
    st.session_state.entered_express = False
if "entered_scan" not in st.session_state:
    st.session_state.entered_scan = False

# ── PERSISTENT SESSION RESTORE ──────────────────────────────────────
# Applied here — before any widgets render — so keys are set before widgets initialise.
_SESSION_KEYS_PREFIX = ("ctrl_", "supplier_name_", "supplier_price_", "supplier_ticker_",
                        "supplier_contact_", "supplier_location_", "supplier_notes_",
                        "score_quality_", "score_delivery_", "score_financial_", "score_risk_",
                        "score_innovation_", "score_esg_", "score_tech_",
                        "auto_renews_", "unbid_cycles_",
                        "contract_start_", "contract_end_", "contract_value_",
                        "award_decision", "award_notes", "selected_parent_cat", "selected_subcategory")

def _restore_session_from_db(session_id: str) -> bool:
    """Load a saved session snapshot into st.session_state. Returns True on success."""
    try:
        _db = get_database()
        _data = _db.get_config(f"session_snap_{session_id}")
        if not _data:
            return False
        for _k, _v in _data.items():
            if not _k.startswith("FormSubmitter") and not _k.startswith("__"):
                st.session_state[_k] = _v
        return True
    except Exception:
        return False

if st.session_state.get("_piq_restore_session_id"):
    _rid = st.session_state.pop("_piq_restore_session_id")
    _restore_session_from_db(_rid)








# =========================================================
# EXCEL IMPORT / EXPORT
# =========================================================
import io as _io


def build_executive_onepager_html(
    event_name: str,
    category: str,
    subcategory: str,
    kraljic: str,
    leader: Dict,
    runner_up: Optional[Dict],
    ranked: List[Dict],
    risk_flags: List[str],
    action_plan: List[Dict],
    category_rule: Dict,
    leader_weakest_dim: str,
) -> str:
    """Generate a clean, print-ready HTML executive one-pager."""
    from datetime import date as _date
    today = _date.today().strftime("%B %d, %Y")

    kraljic_colors = {
        "Strategic": "#DC2626", "Leverage": "#16A34A",
        "Bottleneck": "#D97706", "Non-Critical": "#64748B",
    }
    kc = kraljic_colors.get(kraljic, "#64748B")

    # Supplier rows
    supplier_rows = ""
    for s in ranked[:6]:
        is_rec = s["Supplier"] == leader["Supplier"]
        bg = "#EFF6FF" if is_rec else "#FFFFFF"
        supplier_rows += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:6px 10px;font-weight:{"700" if is_rec else "400"};color:#B0C4DC">'
            f'{"★ " if is_rec else ""}{html.escape(s["Supplier"])}</td>'
            f'<td style="padding:6px 10px;text-align:right;color:#B0C4DC">${s["Raw Price"]:,.0f}</td>'
            f'<td style="padding:6px 10px;text-align:right;font-weight:700;color:{"#1D4ED8" if is_rec else "#374151"}">{s["Weighted Score"]}/100</td>'
            f'<td style="padding:6px 10px;color:#9EB8CE">{s.get("Financial Health", "—")}/100</td>'
            f'</tr>'
        )

    # Risk flags
    risk_html = ""
    for rf in risk_flags[:5]:
        risk_html += f'<li style="margin-bottom:4px;color:#9EB8CE">{html.escape(rf)}</li>'

    # Action plan
    phase_colors = {"Foundation": "#1D4ED8", "Execution": "#D97706", "Optimization": "#16A34A"}
    action_html = ""
    for phase in action_plan:
        pc = phase_colors.get(phase.get("label", ""), "#64748B")
        actions_li = "".join(
            f'<li style="margin-bottom:3px;color:#9EB8CE">{html.escape(a)}</li>'
            for a in phase.get("actions", [])[:4]
        )
        action_html += (
            f'<div style="border-left:3px solid {pc};padding:8px 12px;margin-bottom:10px">'
            f'<div style="font-size:11px;font-weight:700;color:{pc};text-transform:uppercase;letter-spacing:0.05em">'
            f'{phase.get("phase","")}: {phase.get("label","")}</div>'
            f'<ul style="margin:4px 0 0 0;padding-left:16px;font-size:11px">{actions_li}</ul>'
            f'</div>'
        )

    # RAQSCI musts
    raqsci_html = ""
    raqsci_keys = ["requirements", "quality", "service", "cost"]
    raqsci_labels = ["Requirements", "Quality", "Service", "Cost"]
    for key, label in zip(raqsci_keys, raqsci_labels):
        val = category_rule.get(key, "")
        if val:
            raqsci_html += (
                f'<div style="margin-bottom:6px">'
                f'<span style="font-size:10px;font-weight:700;color:#1D4ED8;text-transform:uppercase">{label}: </span>'
                f'<span style="font-size:10px;color:#9EB8CE">{html.escape(val[:120])}{"…" if len(val) > 120 else ""}</span>'
                f'</div>'
            )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>ProcureIQ — {html.escape(event_name)}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, 'Helvetica Neue', Arial, sans-serif; color: #1E293B; background: #fff; padding: 28px 36px; font-size: 12px; line-height: 1.5; }}
  h1 {{ font-size: 20px; color: #0F172A; margin-bottom: 2px; }}
  h2 {{ font-size: 13px; color: #1D4ED8; text-transform: uppercase; letter-spacing: 0.06em; margin: 16px 0 6px; border-bottom: 1px solid #E2E8F0; padding-bottom: 3px; }}
  .meta {{ font-size: 11px; color: #64748B; margin-bottom: 16px; }}
  .kpis {{ display: flex; gap: 12px; margin: 12px 0; }}
  .kpi {{ background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 6px; padding: 10px 14px; flex: 1; }}
  .kpi-label {{ font-size: 9px; color: #64748B; text-transform: uppercase; letter-spacing: 0.08em; }}
  .kpi-value {{ font-size: 18px; font-weight: 700; color: #0F172A; margin-top: 1px; }}
  .kpi-sub {{ font-size: 9px; color: #94A3B8; margin-top: 1px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
  th {{ background: #F1F5F9; color: #475569; font-weight: 600; text-transform: uppercase; font-size: 9px; letter-spacing: 0.06em; padding: 6px 10px; text-align: left; }}
  td {{ border-bottom: 1px solid #F1F5F9; }}
  .two-col {{ display: flex; gap: 20px; }}
  .col {{ flex: 1; }}
  .badge {{ display: inline-block; background: {kc}22; color: {kc}; border: 1px solid {kc}44; border-radius: 4px; padding: 2px 8px; font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em; }}
  .footer {{ margin-top: 24px; padding-top: 10px; border-top: 1px solid #E2E8F0; font-size: 9px; color: #94A3B8; display: flex; justify-content: space-between; }}
  @media print {{ body {{ padding: 12px 18px; }} }}
</style>
</head>
<body>
  <h1>{html.escape(event_name)}</h1>
  <div class="meta">
    {html.escape(category)} &nbsp;·&nbsp; {html.escape(subcategory)} &nbsp;·&nbsp;
    <span class="badge">{html.escape(kraljic)}</span>
    &nbsp;·&nbsp; Generated {today}
  </div>

  <div class="kpis">
    <div class="kpi">
      <div class="kpi-label">Recommended Supplier</div>
      <div class="kpi-value" style="font-size:14px">{html.escape(leader["Supplier"])}</div>
      <div class="kpi-sub">Composite score: {leader["Weighted Score"]}/100</div>
    </div>
    <div class="kpi">
      <div class="kpi-label">Score Gap vs Runner-Up</div>
      <div class="kpi-value" style="color:#1D4ED8">{f'+{round(leader["Weighted Score"] - runner_up["Weighted Score"], 1)}' if runner_up else "—"} pts</div>
      <div class="kpi-sub">{f'vs {html.escape(runner_up["Supplier"])}' if runner_up else "Only supplier evaluated"}</div>
    </div>
    <div class="kpi">
      <div class="kpi-label">Weakest Dimension</div>
      <div class="kpi-value" style="font-size:13px;color:#DC2626">{html.escape(leader_weakest_dim)}</div>
      <div class="kpi-sub">Primary contract focus area</div>
    </div>
    <div class="kpi">
      <div class="kpi-label">Financial Health</div>
      <div class="kpi-value">{leader.get("Financial Health", "—")}/100</div>
      <div class="kpi-sub">Composite signal score</div>
    </div>
  </div>

  <h2>Supplier Ranking</h2>
  <table>
    <thead><tr><th>Supplier</th><th style="text-align:right">Quoted Price</th><th style="text-align:right">Score</th><th>Fin. Health</th></tr></thead>
    <tbody>{supplier_rows}</tbody>
  </table>

  <div class="two-col" style="margin-top:16px">
    <div class="col">
      <h2>90-Day Action Plan</h2>
      {action_html}
    </div>
    <div class="col">
      <h2>Top Risk Flags</h2>
      <ul style="padding-left:16px;font-size:11px">{risk_html}</ul>
      <h2 style="margin-top:12px">Contract Musts (RAQSCI)</h2>
      {raqsci_html}
    </div>
  </div>

  <div class="footer">
    <span>ProcureIQ · Confidential · {html.escape(event_name)}</span>
    <span>Prepared {today}</span>
  </div>
</body>
</html>"""

def build_supplier_template_df():
    return pd.DataFrame(columns=["Supplier Name","Ticker","Quoted Price ($)","SLA Strength (Strong/Moderate/Weak)","Execution Risk (Low/Medium/High)","Stakeholder Confidence (1-5)","Strategic Alignment (1-5)","Innovation Capacity (1-5)","Relationship Depth (1-5)","Commercial Flexibility (1-5)","Years in Business","Ownership Structure","Revenue Trajectory","Recent M&A Activity","Payment Terms Offered","Workforce Changes (12mo)","Notes"])


def parse_supplier_excel(uploaded_file):
    try:
        df = pd.read_excel(uploaded_file)
        if "Supplier Name" not in df.columns or "Quoted Price ($)" not in df.columns:
            return None
        sla_map = {"strong":"Strong","moderate":"Moderate","weak":"Weak"}
        risk_map = {"low":"Low","medium":"Medium","high":"High"}
        def safe_int(val, d=3):
            try: return max(1,min(5,int(float(val))))
            except: return d
        fin_fields = ["Years in Business","Ownership Structure","Revenue Trajectory","Recent M&A Activity","Payment Terms Offered","Workforce Changes (12mo)"]
        out = []
        for _, row in df.iterrows():
            out.append({"Supplier": str(row.get("Supplier Name",f"Supplier {len(out)+1}")), "Ticker": str(row.get("Ticker","")), "Raw Price": float(row.get("Quoted Price ($)",1000000)), "Notes": str(row.get("Notes","")), "SLA Strength": sla_map.get(str(row.get("SLA Strength (Strong/Moderate/Weak)","Moderate")).strip().lower(),"Moderate"), "Execution Risk": risk_map.get(str(row.get("Execution Risk (Low/Medium/High)","Medium")).strip().lower(),"Medium"), "Stakeholder Confidence": safe_int(row.get("Stakeholder Confidence (1-5)",3)), "Strategic Alignment": safe_int(row.get("Strategic Alignment (1-5)",3)), "Innovation Capacity": safe_int(row.get("Innovation Capacity (1-5)",3)), "Relationship Depth": safe_int(row.get("Relationship Depth (1-5)",3)), "Commercial Flexibility": safe_int(row.get("Commercial Flexibility (1-5)",3)), "Financial Inputs": {f: str(row.get(f,"")) for f in fin_fields}, "SEC Context": None, "Alpha Context": None})
        return out
    except Exception:
        return None


# CSV column → FINANCIAL_FIELDS key mapping (used by both parse and template)
_CSV_FIN_COLUMNS = {
    "years_in_business":   "Years in Business",
    "ownership_structure": "Ownership Structure",
    "revenue_trajectory":  "Revenue Trajectory",
    "recent_ma_activity":  "Recent M&A Activity",
    "payment_terms":       "Payment Terms Offered",
    "workforce_changes":   "Workforce Changes (12mo)",
}

_CSV_IMPORT_TEMPLATE = (
    "supplier_name,ticker,quoted_price,"
    "sla_strength,execution_risk,"
    "stakeholder_confidence,strategic_alignment,innovation_capacity,"
    "relationship_depth,commercial_flexibility,"
    "esg_sustainability,supplier_diversity,"
    "years_in_business,ownership_structure,revenue_trajectory,"
    "recent_ma_activity,payment_terms,workforce_changes,notes\n"
    'Acme Corp,ACME,1200000,Strong,Low,4,4,3,4,4,Strong,Moderate,"10–25 years","Publicly traded","Growing 5–15%","None in 2 years","Net 60","Stable",\n'
    'BetaTech,,950000,Moderate,Medium,3,3,3,3,3,Moderate,Moderate,"3–10 years","Founder/private","Flat","None in 2 years","Net 30","Stable","Strong RFP response"\n'
    'GammaCo,GMC,1100000,Weak,High,2,4,4,2,3,Weak,Strong,"25+ years","Private equity-backed","Growing 15%+","Acquired a company","Net 60","Significant hiring",\n'
)


def parse_supplier_csv(uploaded_file):
    """Parse a supplier CSV upload.

    Required column: supplier_name.
    Optional: ticker, quoted_price, notes, and the six financial field columns.

    Returns (list_of_cleaned_rows, list_of_warning_strings).
    Returns (None, warnings) on unrecoverable parse failure.
    Caps at 20 rows (matches UI slot maximum).
    """
    import csv as _csv
    import io as _io
    from validation import validate_supplier_csv_row

    try:
        raw = uploaded_file.read()
        text = raw.decode("utf-8-sig")
    except Exception as e:
        return None, [f"Could not read file: {e}"]

    try:
        reader = _csv.DictReader(_io.StringIO(text))
        fieldnames = reader.fieldnames or []
        if "supplier_name" not in fieldnames:
            return None, ["CSV must include a 'supplier_name' column as the first required field."]

        out: list = []
        all_warnings: list = []

        for row_num, row in enumerate(reader, start=2):
            if len(out) >= 20:
                all_warnings.append("Only the first 20 rows were imported (maximum per evaluation).")
                break
            cleaned, row_warnings = validate_supplier_csv_row(dict(row), row_num)
            all_warnings.extend(row_warnings)
            if cleaned is not None:
                out.append(cleaned)

        return out, all_warnings
    except Exception as e:
        return None, [f"Failed to parse CSV: {e}"]


def build_export_excel(ranked, stake_df, *, event_name, category, kraljic, auction_type, exec_summary, risk_flags):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        df = pd.DataFrame([{"Supplier":s["Supplier"],"Overall Score":s["Weighted Score"],"Current Fit":s["Current Fit"],"Future Fit":s["Future Fit"],"Financial Health":s["Financial Health"]} for s in ranked])
        return df.to_csv(index=False).encode()
    wb = Workbook()
    ws1 = cast(Any, wb.active)
    assert ws1 is not None
    def hs(cell, bg="1F3A6E", fg="F0F4FF", bold=True):
        cell.font = Font(bold=bold, color=fg, name="Calibri", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    def ds(cell, bg="111E33", fg="B8CCEA"):
        cell.font = Font(color=fg, name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws1 = cast(Any, wb.active); ws1.title = "Executive Summary"; ws1.sheet_properties.tabColor = "3B7BF8"
    ws1.column_dimensions["A"].width = 26; ws1.column_dimensions["B"].width = 70
    for r,(k,v) in enumerate([("Event Name",event_name),("Category",category),("Kraljic",kraljic),("Recommended Supplier",ranked[0]["Supplier"]),("Overall Score",f"{ranked[0]['Weighted Score']} / 100"),("Auction Recommendation",auction_type),("",""),("Executive Summary",exec_summary)],2):
        ws1.row_dimensions[r].height = 80 if k=="Executive Summary" else 22
        hs(ws1.cell(r,1,k)); ds(ws1.cell(r,2,v))
    ws2 = wb.create_sheet("Supplier Ranking"); ws2.sheet_properties.tabColor = "22C55E"
    h2 = ["Rank","Supplier","Overall Score","Current Fit","Future Fit","Financial Health","Fin. Risk","Quoted Price ($)"] + list(ranked[0]["Scores"].keys())
    for i in range(1,len(h2)+1): ws2.column_dimensions[get_column_letter(i)].width = 18
    ws2.column_dimensions["B"].width = 24
    for c,h in enumerate(h2,1): hs(ws2.cell(1,c,h))
    for ri,s in enumerate(ranked,1):
        for c,v in enumerate([ri,s["Supplier"],s["Weighted Score"],s["Current Fit"],s["Future Fit"],s["Financial Health"],s["Financial Risk Label"],f"${s['Raw Price']:,.0f}"]+[s["Scores"][d] for d in s["Scores"]],1): ds(ws2.cell(ri+1,c,v))
    ws3 = wb.create_sheet("Stakeholder Map"); ws3.sheet_properties.tabColor = "F59E0B"
    h3 = ["Name","Role","Power","Interest","Position","Priority","Action","Talk Track"]
    for i,h in enumerate(h3,1): ws3.column_dimensions[get_column_letter(i)].width = 22 if i<7 else 55; hs(ws3.cell(1,i,h))
    for r,row in stake_df.iterrows():
        for c,v in enumerate([row["Name"],row["Role"],row["Power"],row["Interest"],row["Position"],row["Priority"],row.get("Action",""),row.get("Talk Track","")],1): ds(ws3.cell(r+2,c,v))
    ws4 = wb.create_sheet("Risk Flags"); ws4.sheet_properties.tabColor = "EF4444"
    for c,h in enumerate(["Tier","Title","Description"],1): ws4.column_dimensions[get_column_letter(c)].width = [10,35,75][c-1]; hs(ws4.cell(1,c,h))
    for r,f in enumerate(risk_flags,2):
        ds(ws4.cell(r,1,f["tier"])); ds(ws4.cell(r,2,f["title"])); ds(ws4.cell(r,3,f["body"]))
    ws5 = wb.create_sheet("Import Template"); ws5.sheet_properties.tabColor = "8B5CF6"
    tdf = build_supplier_template_df()
    for c,col in enumerate(tdf.columns,1): ws5.column_dimensions[get_column_letter(c)].width = 26; hs(ws5.cell(1,c,col)); ds(ws5.cell(2,c,"← fill in"))
    buf = _io.BytesIO(); wb.save(buf); return buf.getvalue()

# =========================================================
# EMAIL GENERATOR
# =========================================================
def generate_stakeholder_emails(leader, runner_up, stake_df, event_name, weakest_dim, kraljic, blocker_row):
    supplier = leader["Supplier"]; score = leader["Weighted Score"]; emails = {}
    champions = stake_df[stake_df["Position"]=="Champion"]
    if not champions.empty:
        ch = champions.iloc[0]
        emails["Champion Brief"] = {"subject": f"[Action Needed] {event_name} — Briefing Before Recommendation Meeting", "body": f"Hi {ch['Name']},\n\nI wanted to brief you before we present the {event_name} recommendation.\n\nOur recommendation is {supplier}, which scored {score}/100 — the strongest balance of current execution and strategic fit under our {kraljic} sourcing posture.\n\nAs a Champion of this decision, your support in the room will matter. The area most likely to draw questions is {weakest_dim}, and we have a mitigation plan ready. I'd welcome the chance to walk you through it before the meeting.\n\nCan we find 20 minutes this week?\n\nBest,"}
    skeptics = stake_df[stake_df["Position"].isin(["Skeptic","Blocker"])]
    if not skeptics.empty:
        sk = skeptics.iloc[0]
        emails["Skeptic Pre-Engagement"] = {"subject": f"[Pre-Meeting] {event_name} — Your Input Before We Finalize", "body": f"Hi {sk['Name']},\n\nBefore we present the {event_name} recommendation, I wanted to connect with you directly to make sure your perspective is reflected.\n\nOur current recommendation is {supplier}. I know {sk['Priority']} is a priority for you, and I want to walk you through exactly how we evaluated that dimension and what protections we're building into the contract.\n\nI'd rather address your concerns one-on-one than have them surface for the first time in the group meeting.\n\nAre you available for a brief call this week?\n\nBest,"}
    emails["Executive Sponsor Summary"] = {"subject": f"{event_name} — Recommendation: {supplier}", "body": f"Hi [Executive Sponsor],\n\nI'm writing to share the {event_name} sourcing outcome ahead of the formal presentation.\n\nRecommendation: {supplier}\nOverall Score: {score}/100\nPosture: {kraljic} category\n\nWhy {supplier} was selected:\n- Highest weighted score across {len(leader['Scores'])} evaluation dimensions\n- Strongest combination of current execution capability and long-term strategic fit\n- Financial health: {leader['Financial Health']}/100\n\nKey risk to monitor: {weakest_dim} — addressed in contract negotiation.\n\n"+(f"Runner-up was {runner_up['Supplier']} at {runner_up['Weighted Score']}/100.\n\n" if runner_up else "")+"Happy to discuss before the meeting.\n\nBest,"}
    emails["Award Notification"] = {"subject": f"{event_name} — Award Decision", "body": f"Dear [Supplier Contact],\n\nThank you for your participation in the {event_name} sourcing evaluation.\n\nAfter a thorough evaluation, we are pleased to confirm that {supplier} has been selected as our preferred supplier.\n\nOur team will be in contact within [X business days] to initiate contract finalization and transition planning.\n\nWe look forward to building a strong working relationship.\n\nBest regards,"}
    emails["Decline Notification"] = {"subject": f"{event_name} — Sourcing Decision Update", "body": f"Dear [Supplier Contact],\n\nThank you for the time and effort you invested in the {event_name} sourcing evaluation.\n\nAfter careful evaluation, we have decided to move forward with another supplier. This decision was based on a comprehensive multi-criteria evaluation and was not a reflection of any single dimension.\n\nWe value the relationship and hope to work together in the future.\n\nBest regards,"}
    return emails

# =========================================================
# AI PROMPT GENERATOR
# =========================================================
AI_TOOLS = [
    {"name": "ChatGPT",  "url": "https://chat.openai.com/",   "icon": "🤖", "color": "#10A37F"},
    {"name": "Gemini",   "url": "https://gemini.google.com/", "icon": "✨", "color": "#4285F4"},
    {"name": "Grok",     "url": "https://grok.x.ai/",         "icon": "⚡", "color": "#1DA1F2"},
    {"name": "DeepSeek", "url": "https://chat.deepseek.com/", "icon": "🔍", "color": "#8B5CF6"},
    {"name": "Claude",   "url": "https://claude.ai/",         "icon": "🧠", "color": "#D97706"},
]







@st.cache_data(show_spinner=False, ttl=3600)
def enrich_market_leaders_with_live_data(subcategory_name: str) -> List[Dict]:
    """
    Enrich market leaders data with live yfinance data where available.
    Falls back to static curated data if yfinance unavailable or ticker missing.
    """
    # Tier 1: original 8 curated entries in MARKET_LEADERS
    # Tier 2: extended 92 entries from market_data.py (with alias + fuzzy fallback)
    base_leaders = MARKET_LEADERS.get(subcategory_name)
    if base_leaders is None and _MARKET_DATA_AVAILABLE:
        base_leaders = get_market_leaders_extended(subcategory_name)
    if base_leaders is None:
        base_leaders = DEFAULT_MARKET_LEADERS
    if not _YFINANCE_AVAILABLE:
        return base_leaders

    enriched = []
    for ldr in base_leaders:
        ticker = ldr.get("ticker", "")
        if not ticker or ticker in ("Private", "—", ""):
            enriched.append({**ldr, "live": None})
            continue
        try:
            assert yf is not None
            t = yf.Ticker(ticker)
            info = t.info
            live = {
                "market_cap_fmt": _fmt_market_cap(info.get("marketCap")),
                "revenue_growth": _fmt_pct(info.get("revenueGrowth")),
                "gross_margin": _fmt_pct(info.get("grossMargins")),
                "employees": f"{info.get('fullTimeEmployees', 0):,}" if info.get("fullTimeEmployees") else "N/A",
                "52w_change": _fmt_pct(info.get("52WeekChange")),
                "analyst_rating": info.get("recommendationKey", "N/A").replace("_", " ").title(),
                "sector": info.get("sector", ""),
                "short_name": info.get("shortName", ldr["name"]),
            }
            enriched.append({**ldr, "live": live})
        except Exception:
            enriched.append({**ldr, "live": None})
    return enriched


def _fmt_market_cap(val) -> str:
    if val is None:
        return "N/A"
    try:
        val = float(val)
    except (TypeError, ValueError):
        return "N/A"
    if val >= 1e12:
        return f"${val/1e12:.2f}T"
    if val >= 1e9:
        return f"${val/1e9:.2f}B"
    if val >= 1e6:
        return f"${val/1e6:.2f}M"
    if val >= 1e3:
        return f"${val/1e3:.2f}K"
    return f"${val:,.2f}"


def _fmt_pct(val) -> str:
    if val is None:
        return "N/A"
    try:
        pct = float(val)
    except (TypeError, ValueError):
        return "N/A"
    if pct == 0:
        return "0.0%"
    return f"{pct * 100:+.1f}%"



# =========================================================
# ENHANCED AI PROMPT BUILDER
# =========================================================
AI_PROMPT_MODES = {
    "🎯 Decision Pressure Test": {
        "desc": "Critique the recommendation. Find what's wrong before the meeting does.",
        "guardrails": [
            "Do NOT validate the recommendation — your job is to challenge it.",
            "Do NOT assume the scoring is accurate — question the inputs.",
            "Do NOT ignore stakeholder dynamics — they kill more deals than bad suppliers.",
            "Avoid generic procurement advice — be specific to this subcategory.",
            "Do NOT hallucinate supplier capabilities — only use what's in the context.",
        ],
    },
    "📝 Contract Risk Analysis": {
        "desc": "Identify the 5 contract clauses most likely to create problems post-award.",
        "guardrails": [
            "Do NOT give legal advice — give procurement risk framing.",
            "Avoid boilerplate contract language — focus on this specific subcategory's risk profile.",
            "Do NOT assume standard terms are sufficient — identify gaps.",
            "Avoid abstract risk — every risk must have a specific business consequence.",
            "Do NOT hallucinate specific regulatory requirements — flag what needs legal review.",
        ],
    },
    "🗣️ Stakeholder Battle Plan": {
        "desc": "Tell me exactly how to handle each stakeholder before the presentation.",
        "guardrails": [
            "Do NOT suggest avoiding the blockers — address them directly.",
            "Avoid generic stakeholder management advice — use the specific roles and positions provided.",
            "Do NOT assume all champions will show up — give activation tactics.",
            "Avoid over-relying on data — some stakeholders respond to narrative, not numbers.",
            "Do NOT hallucinate stakeholder motivations — only use what's in the context.",
        ],
    },
    "💰 Negotiation Intelligence": {
        "desc": "Build my BATNA and identify the 5 highest-leverage negotiation points.",
        "guardrails": [
            "Do NOT suggest lowball tactics that damage the relationship.",
            "Avoid focusing only on price — total value is the frame.",
            "Do NOT ignore the weakest dimension — it's the primary negotiation target.",
            "Avoid generic negotiation tactics — tie every point to this specific subcategory.",
            "Do NOT assume the supplier will accept everything — build a concession strategy.",
        ],
    },
    "❓ RFP Question Generator": {
        "desc": "Generate 10 additional capability questions tailored to this exact evaluation.",
        "guardrails": [
            "Do NOT generate generic questions — every question must be specific to this supplier and subcategory.",
            "Avoid yes/no questions — require evidence and metrics.",
            "Do NOT ask about capabilities already proven — focus on the weakest dimension.",
            "Avoid leading questions — they produce useless answers.",
            "Do NOT hallucinate technical specifications — flag where you're inferring.",
        ],
    },
}


def build_ai_prompt_v2(event_name, category, subcategory_name, kraljic, leader, runner_up,
                       weakest_dim, blocker_row, stake_df, auction_type, intake_answers,
                       selected_sub, mode_name):
    """Enhanced AI prompt builder with mode-specific framing, guardrails, and subcategory context."""
    mode = AI_PROMPT_MODES.get(mode_name, AI_PROMPT_MODES["🎯 Decision Pressure Test"])
    guardrails = "\n".join(f"  ⛔ {g}" for g in mode["guardrails"])
    blocker_line = (
        f"- Most likely internal blocker: {blocker_row['Name']} ({blocker_row['Role']}) "
        f"— {blocker_row['Position']} who prioritizes {blocker_row['Priority']}"
        if blocker_row is not None else "- No critical internal blocker identified"
    )
    runner_line = (
        f"- Runner-up: {runner_up['Supplier']} at {runner_up['Weighted Score']}/100 "
        f"(gap: {round(leader['Weighted Score'] - runner_up['Weighted Score'], 1)} pts)"
        if runner_up else "- Only one supplier evaluated"
    )
    intake_lines = "\n".join(f"  - {q}: {a}" for q, a in intake_answers.items())
    scores_detail = "\n".join(f"  - {dim}: {score}/100" for dim, score in leader["Scores"].items())
    sub_context = (
        f"  - Key risks for this subcategory: {selected_sub.get('key_risks', 'N/A')}\n"
        f"  - Switching cost: {selected_sub.get('switching_cost', 'N/A')}\n"
        f"  - Expert note: {selected_sub.get('notes', 'N/A')}"
    )

    mode_instructions = {
        "🎯 Decision Pressure Test": f"""
TASK: Pressure-test this recommendation as if you are a skeptical CPO reviewing it before approval.
1. What is the single biggest flaw in this recommendation?
2. What 3 assumptions are being made that could be wrong?
3. What is the highest-probability failure mode 12 months post-award?
4. What would make you reject this recommendation?
5. What one question would expose the most risk if asked in the presentation?""",

        "📝 Contract Risk Analysis": f"""
TASK: Identify contract risks specific to {subcategory_name} at this supplier.
1. What 3 clauses are most likely to be missing or weak in a standard {selected_sub.get('contract_type', 'service')} agreement?
2. What is the most dangerous assumption about {leader['Supplier']}'s performance that isn't contractually protected?
3. What SLA language would a smart lawyer exploit to avoid penalties?
4. Given the weakest dimension is {weakest_dim} ({leader['Scores'][weakest_dim]}/100), what specific contract protection is needed?
5. What post-award governance mechanism is most critical for this subcategory?""",

        "🗣️ Stakeholder Battle Plan": f"""
TASK: Give me a stakeholder-by-stakeholder battle plan for defending {leader['Supplier']} in the room.
{blocker_line}
For each stakeholder group (Champions, Neutrals, Skeptics/Blockers):
1. What is the opening message?
2. What evidence do they need to see?
3. What objection are they most likely to raise?
4. How do I pre-empt it before they speak?
5. Who should speak first to set the tone?""",

        "💰 Negotiation Intelligence": f"""
TASK: Build my negotiation strategy for {leader['Supplier']} in {subcategory_name}.
1. What is my BATNA given the runner-up is {runner_up['Supplier'] if runner_up else 'not evaluated'}?
2. What are the 5 highest-leverage commercial terms in this subcategory?
3. The weakest dimension is {weakest_dim} ({leader['Scores'][weakest_dim]}/100) — what contract language closes this gap?
4. What concessions can I offer that cost me little but create perceived value for the supplier?
5. What is the one term I should never concede on, and why?""",

        "❓ RFP Question Generator": f"""
TASK: Generate 10 capability questions I have NOT already asked for {subcategory_name}.
Context: The weakest dimension is {weakest_dim}. Switching cost is {selected_sub.get('switching_cost', 'unknown')}.
Rules: Questions must be answerable with evidence, not opinion. Each must target a specific risk.
Format: Number each question. After each, add (Tests: [what risk it reveals]).
Focus areas:
1. The weakest dimension ({weakest_dim})
2. The key risks: {selected_sub.get('key_risks', 'N/A')}
3. Post-award performance accountability
4. Hidden costs and commercial traps
5. Exit / transition flexibility""",
    }.get(mode_name, "Analyze this evaluation and provide 5 specific, actionable insights.")

    return f"""You are a senior procurement strategist and CPO advisor. You have been given a live sourcing evaluation and need to provide specific, evidence-based guidance.

══ GUARDRAILS — READ BEFORE RESPONDING ══
{guardrails}
════════════════════════════════════════

== EVALUATION CONTEXT ==
Event: {event_name}
Category: {category}
Subcategory: {subcategory_name}
Kraljic Position: {kraljic}
Recommended Event Type: {auction_type}

== SUBCATEGORY-SPECIFIC CONTEXT ==
{sub_context}

== RECOMMENDED SUPPLIER ==
- Name: {leader['Supplier']}
- Overall Score: {leader['Weighted Score']} / 100
- Current Fit: {leader['Current Fit']} | Future Fit: {leader['Future Fit']}
- Financial Health: {leader['Financial Health']} / 100
- Weakest Dimension: {weakest_dim} (score: {leader['Scores'][weakest_dim]}/100)
{runner_line}

== ALL DIMENSION SCORES ==
{scores_detail}

== STAKEHOLDER CONTEXT ==
{blocker_line}
- Champions: {len(stake_df[stake_df['Position'] == 'Champion'])}
- Supporters: {len(stake_df[stake_df['Position'] == 'Supporter'])}
- Neutrals: {len(stake_df[stake_df['Position'] == 'Neutral'])}
- Skeptics/Blockers: {len(stake_df[stake_df['Position'].isin(['Skeptic', 'Blocker'])])}

== INTAKE ANSWERS ==
{intake_lines}

{mode_instructions}

Be direct. Be specific to this subcategory. Do not give generic procurement advice."""



# =========================================================
# CONSTANTS
# =========================================================




CATEGORY_RULES = {
    "technology": {
        "type": "Indirect",
        "tag": "Technology / SaaS",
        "requirements": "Define scope, implementation milestones, accountable owners, integration responsibilities, data migration obligations, security requirements, and acceptance criteria.",
        "assurance": "Protect continuity through uptime commitments, transition support, disaster recovery, incident response, and notice for major platform changes.",
        "quality": "Use measurable SLA language, incident severity definitions, root-cause expectations, service credits, and issue-resolution timelines.",
        "service": "Clarify support model, severity definitions, response times, escalation paths, reporting cadence, and governance structure.",
        "cost": "Specify user/module/usage pricing logic, implementation fees, renewal caps, pass-through restrictions, and billing transparency.",
        "innovation": "Tie roadmap visibility, product release transparency, and future capability support to periodic executive reviews.",
        "rfp_stakeholders": {
            "must": ["CIO / VP Technology", "Information Security", "Procurement Lead", "Legal Counsel"],
            "recommended": ["Finance / Budget Owner", "End-User Representative", "IT Architecture"],
            "nice": ["Change Management", "Data Privacy Officer"],
        },
    },
    "hr": {
        "type": "Indirect",
        "tag": "HR / People Services",
        "requirements": "Clarify deliverables, implementation milestones, employee-impact boundaries, data ownership, service scope, and business ownership.",
        "assurance": "Protect service continuity, support coverage, transition duties, disruption management, and employee-impact recovery plans.",
        "quality": "Focus on service reliability, response quality, issue resolution, employee experience impact, and performance reviews.",
        "service": "Define communication cadence, escalation structure, account ownership, stakeholder reviews, and service-level expectations.",
        "cost": "Clarify pricing structure, renewal controls, implementation costs, change-order logic, and scaling as usage grows.",
        "innovation": "Tie improvement commitments to workforce experience, automation, reporting, and operating-model evolution.",
        "rfp_stakeholders": {
            "must": ["CHRO / VP People", "HR Operations Lead", "Procurement Lead", "Legal Counsel"],
            "recommended": ["Finance / Budget Owner", "IT (data/integration)", "Employee Experience"],
            "nice": ["DEI Lead", "Communications"],
        },
    },
    "finance": {
        "type": "Indirect",
        "tag": "Finance Software / Services",
        "requirements": "Define scope, approval workflows, data ownership, implementation milestones, control requirements, reporting needs, and acceptance criteria.",
        "assurance": "Protect business continuity, reporting availability, audit support, transition support, and critical-process coverage.",
        "quality": "Use measurable controls, audit support, issue resolution timelines, reporting accuracy, and service accuracy expectations.",
        "service": "Clarify account governance, escalation paths, reporting cadence, finance stakeholder support, and service ownership.",
        "cost": "Clarify subscription structure, renewal caps, implementation fees, change orders, pass-through costs, and billing transparency.",
        "innovation": "Tie roadmap commitments to automation, controls improvement, analytics, and process efficiency.",
        "rfp_stakeholders": {
            "must": ["CFO / VP Finance", "Controller", "Procurement Lead", "Legal / Compliance"],
            "recommended": ["Internal Audit", "IT Architecture", "Finance Operations"],
            "nice": ["Tax", "Treasury"],
        },
    },
    "marketing": {
        "type": "Indirect",
        "tag": "Marketing Services",
        "requirements": "Define campaign scope, deliverables, timelines, usage rights, approval workflows, ownership, and acceptance criteria.",
        "assurance": "Protect continuity through staffing commitments, backup account coverage, transition support, and continuity planning.",
        "quality": "Use measurable output expectations, revision rights, performance reporting, and remediation steps.",
        "service": "Clarify account management, response expectations, reporting cadence, escalation process, and stakeholder communication.",
        "cost": "Define rate cards, media pass-through rules, markup transparency, scope-change logic, and renewal limits.",
        "innovation": "Tie improvement commitments to creative performance, channel testing, analytics, and campaign optimization.",
        "rfp_stakeholders": {
            "must": ["CMO / VP Marketing", "Brand Lead", "Procurement Lead", "Legal / IP Counsel"],
            "recommended": ["Finance", "Digital / Analytics", "Creative Director"],
            "nice": ["PR / Communications", "Product Marketing"],
        },
    },
    "services": {
        "type": "Indirect",
        "tag": "Professional Services",
        "requirements": "Define scope, staffing assumptions, deliverables, role ownership, timeline, and acceptance standards tightly.",
        "assurance": "Protect staffing continuity, substitution rules, knowledge transfer, transition support, and documentation ownership.",
        "quality": "Use milestone quality, output standards, remediation rights, and acceptance criteria instead of vague satisfaction language.",
        "service": "Clarify governance, reporting cadence, communication norms, escalation points, and stakeholder alignment.",
        "cost": "Define rate cards, out-of-scope work, change-request logic, rate-increase boundaries, and expense rules.",
        "innovation": "Require practical problem-solving, process improvement, and improvement contribution over the contract term.",
        "rfp_stakeholders": {
            "must": ["Executive Sponsor", "Project Owner", "Procurement Lead", "Legal Counsel"],
            "recommended": ["Finance", "IT (if technical)", "End Users"],
            "nice": ["PMO", "Risk / Compliance"],
        },
    },
    "packaging": {
        "type": "Direct",
        "tag": "Packaging / Direct Material",
        "requirements": "Specify tolerances, MOQ, artwork/specification control, tooling assumptions, approved materials, and change-control rules.",
        "assurance": "Protect continuity through capacity commitments, lead-time expectations, interruption notice, safety stock, and backup supply expectations.",
        "quality": "Use defect thresholds, traceability expectations, corrective action timing, audit rights, and incoming-quality requirements.",
        "service": "Focus on delivery performance, shortage communication, production coordination, and operational escalation.",
        "cost": "Clarify indexation, resin/commodity pass-through, freight treatment, volume tiers, tooling charges, and surcharge triggers.",
        "innovation": "Emphasize redesign, sustainability options, cost-down support, waste reduction, and material optimization.",
        "rfp_stakeholders": {
            "must": ["Supply Chain / Operations", "Engineering / R&D", "Procurement Lead", "Quality"],
            "recommended": ["Finance", "Sustainability", "Marketing (brand specs)"],
            "nice": ["Legal", "Logistics"],
        },
    },
    "manufacturing": {
        "type": "Direct",
        "tag": "Manufacturing / Direct Material",
        "requirements": "Define specs, tolerances, engineering change handling, qualification rules, approved materials, and production constraints.",
        "assurance": "Protect continuity through capacity commitments, geographic risk visibility, dual-source logic, and lead-time commitments.",
        "quality": "Use non-conformance logic, incoming-quality expectations, corrective-action timelines, traceability, and audit rights.",
        "service": "Focus on shortage response, production communication, operational escalation, and supplier responsiveness.",
        "cost": "Clarify commodity exposure, cost transparency, indexation, productivity expectations, and surcharge discipline.",
        "innovation": "Use process improvement, manufacturability support, cost-down, quality improvement, and resilience improvements.",
        "rfp_stakeholders": {
            "must": ["VP Operations / Plant Manager", "Engineering", "Procurement Lead", "Quality"],
            "recommended": ["Finance", "Supply Chain Planning", "Logistics"],
            "nice": ["Sustainability", "Legal"],
        },
    },
    "logistics": {
        "type": "Indirect",
        "tag": "Logistics / Transportation",
        "requirements": "Define lane scope, equipment needs, reporting requirements, shipment visibility, and carrier accountability.",
        "assurance": "Protect continuity through surge coverage, backup options, disruption response, and coverage commitments.",
        "quality": "Measure on-time performance, claims, exceptions, damage rates, service recovery, and issue resolution.",
        "service": "Define dispatch responsiveness, communication cadence, escalation behavior, and reporting expectations.",
        "cost": "Clarify fuel treatment, accessorials, lane assumptions, surcharge triggers, and invoice audit rights.",
        "innovation": "Focus on optimization, visibility, route efficiency, emissions reporting, and continuous improvement.",
        "rfp_stakeholders": {
            "must": ["VP Supply Chain", "Transportation Manager", "Procurement Lead", "Operations"],
            "recommended": ["Finance", "Customer Service", "IT (TMS/visibility)"],
            "nice": ["Sustainability", "Legal"],
        },
    },
}

DEFAULT_RFP_STAKEHOLDERS = {
    "must": ["Executive Sponsor", "Business Owner", "Procurement Lead", "Legal Counsel"],
    "recommended": ["Finance / Budget Owner", "IT (if applicable)", "End Users"],
    "nice": ["Risk / Compliance", "Communications"],
}

FINANCIAL_FIELDS = {
    "Years in Business": {
        "options": ["<3 years", "3–10 years", "10–25 years", "25+ years"],
        "scores": {"<3 years": 20, "3–10 years": 55, "10–25 years": 80, "25+ years": 95},
    },
    "Ownership Structure": {
        "options": ["Publicly traded", "Private equity-backed", "Founder/private", "Subsidiary"],
        "scores": {"Publicly traded": 85, "Private equity-backed": 55, "Founder/private": 70, "Subsidiary": 80},
    },
    "Revenue Trajectory": {
        "options": ["Growing 15%+", "Growing 5–15%", "Flat", "Declining", "Unknown"],
        "scores": {"Growing 15%+": 95, "Growing 5–15%": 78, "Flat": 55, "Declining": 25, "Unknown": 40},
    },
    "Recent M&A Activity": {
        "options": ["None in 2 years", "Acquired a company", "Being acquired", "Recently spun off"],
        "scores": {"None in 2 years": 85, "Acquired a company": 65, "Being acquired": 35, "Recently spun off": 45},
    },
    "Payment Terms Offered": {
        "options": ["Net 90+", "Net 60", "Net 30", "Net 15 or less"],
        "scores": {"Net 90+": 90, "Net 60": 75, "Net 30": 60, "Net 15 or less": 40},
    },
    "Workforce Changes (12mo)": {
        "options": ["Significant hiring", "Stable", "Minor layoffs <5%", "Major layoffs >10%"],
        "scores": {"Significant hiring": 90, "Stable": 80, "Minor layoffs <5%": 55, "Major layoffs >10%": 25},
    },
}

RFP_TIMELINE = {
    "Strategic": [
        {"week": "Week 1–2", "phase": "Align", "tasks": ["Define scope and business requirements", "Confirm stakeholder team and executive sponsor", "Establish evaluation criteria and weighting", "Brief legal on category risk profile"]},
        {"week": "Week 3–4", "phase": "Launch", "tasks": ["Issue RFP to shortlisted suppliers", "Host supplier briefing / Q&A session", "Set scoring committee and conflict-of-interest check", "Define clarification and scoring process"]},
        {"week": "Week 5–6", "phase": "Evaluate", "tasks": ["Score responses against weighted criteria", "Run supplier clarification sessions", "Complete financial and risk diligence", "Build recommendation narrative"]},
        {"week": "Week 7–8", "phase": "Negotiate", "tasks": ["Enter best-and-final with top 1–2 suppliers", "Close weakest-dimension gaps in contract", "Confirm stakeholder alignment before award", "Finalize commercial and legal terms"]},
        {"week": "Week 9", "phase": "Award", "tasks": ["Present recommendation to executive sponsor", "Communicate award and decline decisions", "Execute contract", "Kick off transition / implementation planning"]},
    ],
    "Leverage": [
        {"week": "Week 1", "phase": "Align", "tasks": ["Define scope and required outcomes", "Identify competitive supplier pool (3–5)", "Set price and performance benchmarks"]},
        {"week": "Week 2–3", "phase": "Launch", "tasks": ["Issue RFP with competitive intent visible", "Run structured Q&A", "Set scoring and scoring ownership"]},
        {"week": "Week 4", "phase": "Evaluate", "tasks": ["Score responses", "Identify leverage gaps and negotiation targets", "Short-list to 2 finalists"]},
        {"week": "Week 5", "phase": "Negotiate", "tasks": ["Run competitive negotiation", "Lock pricing, SLA, and renewal logic", "Close commercial terms"]},
        {"week": "Week 6", "phase": "Award", "tasks": ["Award and notify", "Execute contract", "Begin transition"]},
    ],
    "Bottleneck": [
        {"week": "Week 1", "phase": "Align", "tasks": ["Map current supply risk and single-source exposure", "Identify backup or secondary options", "Define continuity as primary criteria"]},
        {"week": "Week 2–3", "phase": "Launch", "tasks": ["Issue RFP with continuity requirements explicit", "Probe supplier capacity and backup coverage", "Assess financial stability closely"]},
        {"week": "Week 4", "phase": "Evaluate", "tasks": ["Score with continuity weighted highest", "Flag any capacity or financial risk signals", "Do not sacrifice continuity for price"]},
        {"week": "Week 5", "phase": "Negotiate", "tasks": ["Lock interruption handling and notice terms", "Establish safety stock or backup obligations", "Secure transition support language"]},
        {"week": "Week 6", "phase": "Award", "tasks": ["Award and confirm backup coverage plan", "Execute contract", "Build continuity review cadence"]},
    ],
    "Non-Critical": [
        {"week": "Week 1", "phase": "Align", "tasks": ["Define scope and confirm low complexity", "Identify 2–3 qualified suppliers", "Keep evaluation criteria simple"]},
        {"week": "Week 2", "phase": "Launch & Evaluate", "tasks": ["Issue short RFQ (not full RFP)", "Score on price, lead time, and ease of doing business"]},
        {"week": "Week 3", "phase": "Negotiate & Award", "tasks": ["Short negotiation on price and terms", "Execute standard agreement or PO", "Minimize admin overhead"]},
    ],
}

PHASE_COLORS = {
    "Align": "#3B82F6",
    "Launch": "#8B5CF6",
    "Evaluate": "#F59E0B",
    "Negotiate": "#EF4444",
    "Award": "#22C55E",
    "Launch & Evaluate": "#8B5CF6",
    "Negotiate & Award": "#22C55E",
}


# =========================================================
# DESIGN SYSTEM v3 — Maximum readability + Premium visual
# Reference: Linear, Stripe, Vercel, Apple, Netflix
# =========================================================
st.markdown(
    """
<style>
    /* ═══════════════════════════════════════════════════════
       PROCUREIQ — MISSION CONTROL DESIGN SYSTEM
       SpaceX ops center aesthetic · Clarity-first · No waste
    ═══════════════════════════════════════════════════════ */

    /* ── HIDE STREAMLIT CHROME ── */
    header[data-testid="stHeader"] { display: none !important; }
    [data-testid="stSidebar"]      { display: none !important; }
    #MainMenu                      { display: none !important; }
    footer                         { display: none !important; }
    [data-testid="stToolbar"]      { display: none !important; }

    /* ── BASE LAYOUT ── */
    .stApp {
        background: #03080F !important;
    }
    .block-container {
        padding: 0.8rem 1.6rem !important;
        max-width: 100% !important;
        margin-left: 0 !important;
    }

    /* ── FONTS ── */
    @import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=JetBrains+Mono:wght@300;400;500;700&family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ── DESIGN TOKENS ── */
    :root {
        /* Backgrounds — five-layer navy depth ramp */
        --bg-page:     #03080F;
        --bg-s1:       #060D1A;
        --bg-s2:       #0A1628;
        --bg-s3:       #0F1F38;
        --bg-panel:    #040B16;
        --bg-hover:    #0D1E35;

        /* Accent — single electric blue, three intensities */
        --blue:        #3B82F6;
        --blue-glow:   #60A5FA;
        --blue-dim:    #1D4ED8;
        --blue-bg:     rgba(29,78,216,0.06);
        --blue-border: rgba(96,165,250,0.12);
        --blue-active: rgba(96,165,250,0.35);

        /* Semantic traffic-light — never decorative */
        --red:         #F87171;
        --red-bg:      rgba(248,113,113,0.07);
        --amber:       #FCD34D;
        --amber-bg:    rgba(252,211,77,0.07);
        --green:       #4ADE80;
        --green-bg:    rgba(74,222,128,0.07);
        --neutral:     #334155;
        --purple:      #A78BFA;
        --purple-bg:   rgba(167,139,250,0.07);
        --cyan:        #22D3EE;

        /* Text hierarchy — WCAG AA+ compliant on #03080F–#0A1628 backgrounds */
        --t1:  #F1F5F9;   /* 16.4:1 — headings, labels, hero numbers */
        --t2:  #E2EBF7;   /* 13.8:1 — body paragraphs, primary content */
        --t3:  #C8D8EC;   /* 10.6:1 — secondary content, descriptions */
        --t4:  #B0C4DC;   /*  8.2:1 — metadata, supporting text */
        --t5:  #8BAAC4;   /*  5.0:1 — minimum: hints, placeholders only */


        /* Typography stacks */
        --serif:  'DM Serif Display', Georgia, serif;
        --mono:   'JetBrains Mono', 'Fira Code', monospace;
        --sans:   'Inter', system-ui, sans-serif;

        /* Borders */
        --b-struct:  rgba(96,165,250,0.08);
        --b-active:  rgba(96,165,250,0.18);
        --b-select:  rgba(96,165,250,0.35);

        /* Radius scale — strict */
        --r-sm:  6px;
        --r-md:  10px;
        --r-lg:  14px;
        --r-xl:  20px;

        /* Glass surface tokens (Microsoft Fluent physical material) */
        --glass-surface:          rgba(255,255,255,0.03);
        --glass-surface-elevated: rgba(255,255,255,0.06);
        --glass-border-top:       rgba(255,255,255,0.12);
        --glass-border-bottom:    rgba(0,0,0,0.30);
        --glass-border-side:      rgba(255,255,255,0.06);
        --glass-blur:             blur(20px);
        --glass-blur-elevated:    blur(40px);

        /* Shadows */
        --shadow-sm:   0 1px 4px rgba(0,0,0,0.50);
        --shadow-md:   0 4px 16px rgba(0,0,0,0.60);
        --shadow-lg:   0 12px 40px rgba(0,0,0,0.70);
        --shadow-glow: 0 0 0 3px rgba(96,165,250,0.18);
        --shadow-blue: 0 0 24px rgba(59,130,246,0.35);
        --shadow-lift: 0 20px 40px rgba(0,0,0,0.50), 0 0 30px rgba(96,165,250,0.05);

        /* Motion */
        --t-fast: 150ms ease;
        --t-med:  250ms ease;
        --t-slow: 400ms ease;

        /* Legacy aliases */
        --bg-base:          #060D1A;
        --bg-surface:       #0A1628;
        --bg-raised:        #0F1F38;
        --bg-overlay:       #162236;
        --border:           rgba(96,165,250,0.08);
        --border-bright:    rgba(96,165,250,0.18);
        --border-focus:     rgba(59,130,246,0.6);
        --text-primary:     #F1F5F9;
        --text-secondary:   #C8D8EC;
        --text-muted:       #B0C4DC;
        --text-inverse:     #03080F;
        --accent-blue:      #60A5FA;
        --accent-blue-bg:   rgba(96,165,250,0.08);
        --accent-green:     #4ADE80;
        --accent-green-bg:  rgba(74,222,128,0.08);
        --accent-amber:     #FCD34D;
        --accent-amber-bg:  rgba(252,211,77,0.08);
        --accent-red:       #F87171;
        --accent-red-bg:    rgba(248,113,113,0.08);
        --accent-purple:    #A78BFA;
        --accent-purple-bg: rgba(167,139,250,0.08);
        --font-display:     'DM Serif Display', Georgia, serif;
        --font-body:        'Inter', system-ui, sans-serif;
        --font-mono:        'JetBrains Mono', monospace;
    }

    /* ── GLOBAL BASE ── */
    html, body, .stApp {
        font-family: var(--sans);
        color: var(--t1);
        background: var(--bg-page);
    }

    /* ── NOISE GRAIN TEXTURE — physical material feel ── */
    html::before {
        content: "";
        position: fixed;
        inset: 0;
        z-index: 0;
        pointer-events: none;
        opacity: 0.035;
        background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='noise'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23noise)'/%3E%3C/svg%3E");
        background-repeat: repeat;
        background-size: 200px 200px;
    }

    /* ── ANIMATIONS ── */
    @keyframes fadeUp {
        from { opacity: 0; transform: translateY(14px); }
        to   { opacity: 1; transform: translateY(0); }
    }
    @keyframes fadeIn {
        from { opacity: 0; } to { opacity: 1; }
    }
    @keyframes pulseGreen {
        0%, 100% { opacity: 1; box-shadow: 0 0 0 0 rgba(74,222,128,0.4); }
        50%       { opacity: 0.7; box-shadow: 0 0 0 6px rgba(74,222,128,0); }
    }
    @keyframes pulseBlue {
        0%, 100% { box-shadow: 0 0 8px rgba(96,165,250,0.3); }
        50%       { box-shadow: 0 0 22px rgba(96,165,250,0.8); }
    }
    @keyframes drawBar {
        from { width: 0; opacity: 0.6; }
        to   { width: var(--bar-w, 100%); opacity: 1; }
    }
    @keyframes barRace {
        from { width: 0; }
        to   { width: var(--race-w, 80%); }
    }
    @keyframes countUp {
        from { opacity: 0; transform: translateY(6px); }
        to   { opacity: 1; transform: translateY(0); }
    }
    @keyframes terminalBlink {
        0%, 100% { opacity: 1; } 50% { opacity: 0; }
    }
    @keyframes assembleIn {
        from { opacity: 0; transform: scale(0.97) translateY(8px); }
        to   { opacity: 1; transform: scale(1) translateY(0); }
    }
    @keyframes assembleInSpring {
        from { opacity: 0; transform: scale(0.97) translateY(10px); }
        to   { opacity: 1; transform: scale(1) translateY(0); }
    }
    @keyframes shimmer {
        0%   { background-position: -400px 0; }
        100% { background-position: 400px 0; }
    }
    @keyframes tickerScroll {
        from { transform: translateX(0); }
        to   { transform: translateX(-50%); }
    }
    @keyframes tabSlideIn {
        from { opacity: 0; transform: translateX(24px); }
        to   { opacity: 1; transform: translateX(0); }
    }

    /* ── MISSION HEADER STRIP ── */
    .mission-header {
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 0.7rem 0 0.7rem;
        border-bottom: 1px solid var(--b-struct);
        margin-bottom: 0.8rem;
        animation: fadeIn 0.3s ease both;
    }
    .mission-breadcrumb {
        font-family: var(--mono);
        font-size: 0.62rem;
        letter-spacing: 0.2em;
        text-transform: uppercase;
        color: var(--t4);
    }
    .mission-breadcrumb span { color: var(--blue-glow); }
    .mission-event {
        font-family: var(--serif);
        font-size: 1.1rem;
        color: var(--t3);
        font-style: italic;
        letter-spacing: -0.01em;
    }
    .mission-status {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        font-family: var(--mono);
        font-size: 0.6rem;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        color: var(--t4);
    }
    .mission-dot {
        width: 6px; height: 6px;
        border-radius: 50%;
        background: var(--green);
        animation: pulseGreen 2s infinite;
        flex-shrink: 0;
    }

    /* ── TABS — precision instrument style ── */
    .stTabs [data-baseweb="tab-list"] {
        background: transparent !important;
        border-bottom: 1px solid var(--b-struct) !important;
        gap: 0 !important;
        padding: 0 !important;
    }
    .stTabs [data-baseweb="tab"] {
        font-family: var(--mono) !important;
        font-size: 0.62rem !important;
        letter-spacing: 0.12em !important;
        text-transform: uppercase !important;
        color: var(--t4) !important;
        background: transparent !important;
        border: none !important;
        border-bottom: 2px solid transparent !important;
        padding: 0.65rem 1.1rem !important;
        transition: all 150ms ease !important;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: var(--t3) !important;
        background: rgba(96,165,250,0.03) !important;
        border-bottom-color: rgba(96,165,250,0.12) !important;
    }
    .stTabs [aria-selected="true"] {
        color: var(--t1) !important;
        border-bottom: 2px solid var(--blue) !important;
        background: rgba(59,130,246,0.05) !important;
    }
    .stTabs [data-baseweb="tab-panel"] {
        animation: tabSlideIn 280ms cubic-bezier(0.4,0,0.2,1) both;
    }
    .stTabs [data-baseweb="tab-highlight"] { display: none !important; }
    .stTabs [data-baseweb="tab-border"]    { display: none !important; }

    /* ── RIGHT CONTROL PANEL — premium sidebar ── */
    .ctrl-panel {
        background: linear-gradient(180deg, #040B16 0%, #060D1A 100%);
        border: 1px solid var(--b-struct);
        border-radius: var(--r-lg);
        padding: 1rem 0.9rem;
        position: sticky;
        top: 0.5rem;
        box-shadow: var(--shadow-md);
    }
    .ctrl-header {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.22em;
        text-transform: uppercase;
        color: var(--blue);
        padding-bottom: 0.65rem;
        border-bottom: 1px solid var(--b-struct);
        margin-bottom: 0.85rem;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.5rem;
    }
    .ctrl-header-left {
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .ctrl-header::before {
        content: "";
        width: 4px; height: 4px;
        border-radius: 50%;
        background: var(--blue);
        animation: pulseBlue 2.5s infinite;
        flex-shrink: 0;
    }
    /* Icon-led section labels */
    .ctrl-section {
        font-family: var(--mono);
        font-size: 0.56rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        color: var(--t5);
        margin-top: 1rem;
        margin-bottom: 0.35rem;
        display: flex;
        align-items: center;
        gap: 0.4rem;
        padding-top: 0.65rem;
        border-top: 1px solid rgba(96,165,250,0.06);
    }
    .ctrl-section:first-of-type { border-top: none; padding-top: 0; }
    /* Category readout — terminal style */
    .ctrl-readout {
        background: var(--bg-page);
        border: 1px solid var(--b-struct);
        border-left: 2px solid var(--blue-dim);
        border-radius: 6px;
        padding: 0.4rem 0.6rem;
        margin-bottom: 0.6rem;
    }
    .ctrl-readout-label {
        font-family: var(--mono);
        font-size: 0.56rem;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        color: var(--t5);
        margin-bottom: 0.2rem;
    }
    .ctrl-readout-value {
        font-family: var(--mono);
        font-size: 0.82rem;
        color: #93C5FD;
        letter-spacing: 0.02em;
    }
    .ctrl-readout-sub {
        font-family: var(--mono);
        font-size: 0.6rem;
        color: var(--t5);
        margin-top: 0.15rem;
    }
    /* Sync badges */
    .sync-badge-green {
        font-family: var(--mono);
        font-size: 0.62rem;
        color: var(--green);
        margin-top: -0.3rem;
        margin-bottom: 0.4rem;
        letter-spacing: 0.04em;
    }
    .sync-badge-amber {
        font-family: var(--mono);
        font-size: 0.62rem;
        color: var(--amber);
        margin-top: -0.3rem;
        margin-bottom: 0.4rem;
        letter-spacing: 0.04em;
    }

    /* ── GLASS CARDS — Microsoft Fluent physical material ── */
    /* Base glass card: one light source top-left */
    .glass-card {
        position: relative;
        background: var(--glass-surface);
        border-radius: var(--r-md);
        padding: 1.1rem;
        transition: border-color 200ms ease, background 200ms ease;
        animation: fadeUp 0.4s ease both;
        /* Simulate edge lighting: lighter top, darker bottom */
        border-top:    1px solid var(--glass-border-top);
        border-left:   1px solid var(--glass-border-side);
        border-right:  1px solid var(--glass-border-side);
        border-bottom: 1px solid var(--glass-border-bottom);
    }
    .glass-card::before {
        content: "";
        position: absolute;
        inset: 0;
        border-radius: var(--r-md);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        z-index: -1;
    }
    .glass-card:hover {
        background: var(--glass-surface-elevated);
        border-top-color: rgba(255,255,255,0.18);
    }

    .glass-card-elevated {
        background: var(--glass-surface-elevated);
        border-top:    1px solid rgba(255,255,255,0.16);
        border-left:   1px solid rgba(255,255,255,0.08);
        border-right:  1px solid rgba(255,255,255,0.04);
        border-bottom: 1px solid rgba(0,0,0,0.40);
        box-shadow: var(--shadow-md);
    }

    /* KPI glass card for Overview strip */
    .kpi-glass {
        background: var(--glass-surface);
        border-top:    1px solid var(--glass-border-top);
        border-left:   1px solid var(--glass-border-side);
        border-right:  1px solid var(--glass-border-side);
        border-bottom: 1px solid var(--glass-border-bottom);
        border-radius: var(--r-md);
        padding: 1rem 1.1rem;
        text-align: center;
        transition: all 200ms ease;
        animation: fadeUp 0.4s ease both;
        box-shadow: var(--shadow-sm);
    }
    .kpi-glass:hover {
        background: var(--glass-surface-elevated);
        border-top-color: rgba(255,255,255,0.18);
        box-shadow: var(--shadow-md);
    }
    .kpi-glass-num {
        font-family: var(--serif);
        font-size: 2rem;
        color: var(--t1);
        letter-spacing: -0.02em;
        line-height: 1;
        animation: countUp 0.5s ease both;
    }
    .kpi-glass-label {
        font-family: var(--mono);
        font-size: 0.56rem;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        color: var(--t4);
        margin-top: 0.35rem;
    }

    /* Legacy .card — upgraded to glass */
    .card {
        background: var(--glass-surface);
        border-top:    1px solid var(--glass-border-top);
        border-left:   1px solid var(--glass-border-side);
        border-right:  1px solid var(--glass-border-side);
        border-bottom: 1px solid var(--glass-border-bottom);
        border-radius: var(--r-md);
        padding: 1.2rem 1.4rem;
        transition: all 200ms ease;
        animation: fadeUp 0.4s ease both;
    }
    .card:hover {
        background: var(--glass-surface-elevated);
        border-top-color: rgba(255,255,255,0.18);
        box-shadow: var(--shadow-lift);
    }
    .card-winner {
        border-top: 2px solid var(--blue);
        background: var(--blue-bg);
        box-shadow: 0 0 40px rgba(59,130,246,0.07);
    }
    .card-risk {
        border-left: 3px solid var(--red);
        background: var(--red-bg);
        border-radius: 0 var(--r-md) var(--r-md) 0;
    }
    .card-caution {
        border-left: 3px solid var(--amber);
        background: var(--amber-bg);
        border-radius: 0 var(--r-md) var(--r-md) 0;
    }
    .card-approved {
        border-left: 3px solid var(--green);
        background: var(--green-bg);
        border-radius: 0 var(--r-md) var(--r-md) 0;
    }

    /* ── SHIMMER LOADING STATE ── */
    .shimmer {
        background: linear-gradient(90deg,
            rgba(96,165,250,0.04) 0%,
            rgba(96,165,250,0.10) 50%,
            rgba(96,165,250,0.04) 100%);
        background-size: 400px 100%;
        animation: shimmer 1.5s linear infinite;
        border-radius: var(--r-sm);
        height: 1rem;
    }

    /* ── TICKER STRIP ── */
    .ticker-wrap {
        overflow: hidden;
        position: relative;
        height: 26px;
        background: rgba(4,9,16,0.85);
        border-bottom: 1px solid rgba(96,165,250,0.05);
        display: flex;
        align-items: center;
    }
    .ticker-wrap::before {
        content: "";
        position: absolute;
        left: 0; top: 0; bottom: 0;
        width: 40px;
        background: linear-gradient(90deg, rgba(4,9,16,1), transparent);
        z-index: 2;
        pointer-events: none;
    }
    .ticker-wrap::after {
        content: "";
        position: absolute;
        right: 0; top: 0; bottom: 0;
        width: 40px;
        background: linear-gradient(270deg, rgba(4,9,16,1), transparent);
        z-index: 2;
        pointer-events: none;
    }
    .ticker-inner {
        display: flex;
        gap: 2.5rem;
        animation: tickerScroll 40s linear infinite;
        white-space: nowrap;
        padding: 0 1rem;
    }
    .ticker-item {
        display: inline-flex;
        align-items: center;
        gap: 0.4rem;
        font-family: var(--mono);
        font-size: 0.58rem;
    }
    .ticker-sym  { color: var(--t4); letter-spacing: 0.06em; }
    .ticker-val  { color: var(--t1); }
    .ticker-up   { color: var(--green); }
    .ticker-down { color: var(--red); }

    /* ── SCORE BAR RACE ── */
    .race-row {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        padding: 0.55rem 1rem;
        border-radius: var(--r-md);
        transition: background 150ms ease;
        animation: fadeUp 0.4s ease both;
    }
    .race-row:hover { background: rgba(96,165,250,0.03); }
    .race-row-winner {
        background: rgba(74,222,128,0.04);
        border-left: 3px solid var(--green);
        border-radius: 0 var(--r-md) var(--r-md) 0;
        padding-left: 0.75rem;
    }
    .race-name {
        font-family: var(--sans);
        font-size: 0.85rem;
        color: var(--t1);
        width: 180px;
        flex-shrink: 0;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .race-track {
        flex: 1;
        height: 8px;
        background: rgba(96,165,250,0.07);
        border-radius: 999px;
        overflow: visible;
        position: relative;
    }
    .race-fill {
        height: 100%;
        border-radius: 999px;
        background: rgba(96,165,250,0.25);
        animation: barRace 800ms cubic-bezier(0.22,1,0.36,1) both;
    }
    .race-fill-leader {
        background: linear-gradient(90deg, #1D4ED8, #3B82F6, #60A5FA);
        box-shadow: 4px 0 16px rgba(96,165,250,0.45);
    }
    .race-score {
        font-family: var(--mono);
        font-size: 1rem;
        width: 52px;
        text-align: right;
        flex-shrink: 0;
    }
    .race-score-leader { color: var(--blue-glow); }
    .race-score-other  { color: var(--t4); }
    .race-badge-rec {
        font-family: var(--mono);
        font-size: 0.5rem;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        color: var(--green);
        border: 1px solid rgba(74,222,128,0.25);
        background: rgba(74,222,128,0.08);
        border-radius: 4px;
        padding: 0.1rem 0.35rem;
        flex-shrink: 0;
        animation: pulseGreen 3s infinite;
    }

    /* ── VERDICT HERO CARD (Award Brief) ── */
    .verdict-hero {
        position: relative;
        width: 100%;
        min-height: 200px;
        background: linear-gradient(135deg, #04090F 0%, #091428 55%, #0A1A35 100%);
        border-radius: var(--r-lg);
        overflow: hidden;
        display: flex;
        flex-direction: column;
        justify-content: center;
        padding: 2rem 2.5rem;
        margin-bottom: 1.5rem;
        animation: assembleInSpring 700ms cubic-bezier(0.34,1.56,0.64,1) both;
        /* Top border gradient simulating light from left */
        border-top: 1px solid transparent;
        background-clip: padding-box;
        box-shadow: var(--shadow-lg);
    }
    .verdict-hero::before {
        content: "";
        position: absolute;
        inset: 0;
        border-radius: var(--r-lg);
        padding: 1px;
        background: linear-gradient(90deg, #60A5FA 0%, rgba(96,165,250,0.15) 50%, transparent 100%);
        -webkit-mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
        mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
        -webkit-mask-composite: xor;
        mask-composite: exclude;
        pointer-events: none;
    }
    .verdict-hero::after {
        content: "";
        position: absolute;
        top: -60px; right: -60px;
        width: 320px; height: 280px;
        background: radial-gradient(ellipse, rgba(59,130,246,0.08), transparent 70%);
        pointer-events: none;
    }
    .verdict-eyebrow {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.24em;
        text-transform: uppercase;
        color: var(--green);
        margin-bottom: 0.7rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
        position: relative;
        z-index: 1;
    }
    .verdict-eyebrow::before {
        content: "";
        width: 6px; height: 6px;
        border-radius: 50%;
        background: var(--green);
        animation: pulseGreen 2s infinite;
        flex-shrink: 0;
    }
    .verdict-name {
        font-family: var(--serif);
        font-size: 3.8rem;
        color: var(--t1);
        letter-spacing: -0.02em;
        line-height: 1;
        margin-bottom: 0.5rem;
        position: relative;
        z-index: 1;
    }
    .verdict-score {
        font-family: var(--mono);
        font-size: 1.1rem;
        color: var(--blue-glow);
        letter-spacing: 0.06em;
        position: relative;
        z-index: 1;
    }
    .verdict-ci {
        position: absolute;
        bottom: 1.5rem;
        right: 2rem;
        font-family: var(--mono);
        font-size: 0.65rem;
        color: var(--t3);
        background: rgba(96,165,250,0.07);
        border: 1px solid rgba(96,165,250,0.18);
        border-radius: 999px;
        padding: 0.2rem 0.7rem;
        z-index: 1;
    }

    /* ── DIMENSION COLOR BARS ── */
    .dim-row {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        padding: 0.3rem 0;
        border-bottom: 1px solid rgba(96,165,250,0.04);
    }
    .dim-row:last-child { border-bottom: none; }
    .dim-label {
        font-family: var(--mono);
        font-size: 0.6rem;
        letter-spacing: 0.06em;
        text-transform: uppercase;
        color: var(--t3);
        width: 160px;
        flex-shrink: 0;
    }
    .dim-track {
        flex: 1;
        height: 5px;
        background: rgba(96,165,250,0.07);
        border-radius: 999px;
        overflow: hidden;
    }
    .dim-fill {
        height: 100%;
        border-radius: 999px;
        animation: drawBar 0.6s ease-out both;
    }
    .dim-val {
        font-family: var(--mono);
        font-size: 0.7rem;
        width: 36px;
        text-align: right;
        flex-shrink: 0;
    }
    /* Score-value color thresholds */
    .dim-high  { background: var(--green); }
    .dim-mid   { background: var(--amber); }
    .dim-low   { background: var(--red);   }
    .val-high  { color: var(--green); }
    .val-mid   { color: var(--amber); }
    .val-low   { color: var(--red);   }

    /* ── KRALJIC POSTURE BADGE — with glow ── */
    .kraljic-badge {
        width: 100%;
        height: 38px;
        display: flex;
        align-items: center;
        justify-content: center;
        border-radius: var(--r-sm);
        font-family: var(--serif);
        font-size: 1rem;
        margin-bottom: 0.6rem;
        transition: box-shadow 300ms ease;
    }
    .kraljic-strategic  { color:#DC2626; background:rgba(220,38,38,0.08); border:1px solid rgba(220,38,38,0.25); box-shadow:0 0 20px rgba(220,38,38,0.12); }
    .kraljic-leverage   { color:#16A34A; background:rgba(22,163,74,0.08);  border:1px solid rgba(22,163,74,0.25);  box-shadow:0 0 20px rgba(22,163,74,0.12);  }
    .kraljic-bottleneck { color:#D97706; background:rgba(217,119,6,0.08);  border:1px solid rgba(217,119,6,0.25);  box-shadow:0 0 20px rgba(217,119,6,0.12);  }
    .kraljic-noncrit    { color:#D0E0EF; background:rgba(100,116,139,0.08);border:1px solid rgba(100,116,139,0.25);box-shadow:0 0 20px rgba(100,116,139,0.10); }

    /* ── LIGHT/DARK TOGGLE PILL ── */
    .theme-toggle {
        display: inline-flex;
        align-items: center;
        gap: 0.3rem;
        background: rgba(96,165,250,0.08);
        border: 1px solid rgba(96,165,250,0.15);
        border-radius: 999px;
        padding: 0.15rem 0.55rem;
        font-family: var(--mono);
        font-size: 0.56rem;
        letter-spacing: 0.08em;
        color: var(--blue-glow);
        cursor: pointer;
        transition: all 200ms ease;
        text-transform: uppercase;
    }
    .theme-toggle:hover {
        background: rgba(96,165,250,0.15);
        border-color: rgba(96,165,250,0.30);
    }

    /* ── SCORE DISPLAY ── */
    .score-num {
        font-family: var(--serif);
        font-size: 3rem;
        color: var(--blue-glow);
        letter-spacing: -0.02em;
        line-height: 1;
        font-style: italic;
        animation: countUp 0.4s ease both;
    }
    .score-label {
        font-family: var(--mono);
        font-size: 0.6rem;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        color: var(--t4);
        margin-top: 0.2rem;
    }
    .score-bar-track {
        height: 5px;
        background: rgba(96,165,250,0.07);
        border-radius: 3px;
        overflow: hidden;
        margin-top: 0.3rem;
    }
    .score-bar-fill {
        height: 100%;
        background: linear-gradient(90deg, #1D4ED8, #3B82F6);
        border-radius: 3px;
        animation: drawBar 0.5s ease-out both;
    }
    .score-bar-fill-winner {
        height: 7px;
        margin-top: -1px;
        background: linear-gradient(90deg, #1D4ED8, #60A5FA);
    }

    /* ── RECOMMENDED BADGE ── */
    .badge-recommended {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        color: var(--green);
        border: 1px solid rgba(74,222,128,0.25);
        background: var(--green-bg);
        border-radius: 6px;
        padding: 0.12rem 0.45rem;
        animation: pulseGreen 3s infinite;
    }

    /* ── TERMINAL / INTELLIGENCE STRIP ── */
    .terminal-strip {
        background: var(--bg-page);
        border-left: 2px solid var(--blue-dim);
        border-radius: 0 6px 6px 0;
        padding: 0.7rem 0.9rem;
        margin-top: 0.5rem;
        font-family: var(--mono);
        font-size: 0.75rem;
    }
    .terminal-strip .t-key   { color: var(--t4); }
    .terminal-strip .t-val   { color: #93C5FD; }
    .terminal-strip .t-arrow { color: var(--blue-dim); }
    .terminal-strip .t-row {
        display: flex;
        gap: 1rem;
        padding: 0.12rem 0;
        border-bottom: 1px solid rgba(96,165,250,0.04);
    }
    .terminal-strip .t-row:last-child { border-bottom: none; }

    /* ── EYEBROW / SECTION LABEL ── */
    .eyebrow {
        font-family: var(--mono);
        font-size: 0.62rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        color: var(--t4);
        margin-bottom: 0.5rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .eyebrow::before {
        content: "";
        display: inline-block;
        width: 16px; height: 1px;
        background: var(--t5);
    }

    /* ── RISK FLAGS ── */
    .risk-flag {
        display: flex;
        align-items: flex-start;
        gap: 0.6rem;
        padding: 0.5rem 0.7rem;
        border-left: 2px solid var(--red);
        background: var(--red-bg);
        border-radius: 0 6px 6px 0;
        margin-bottom: 0.4rem;
        font-size: 0.82rem;
        color: var(--t2);
    }
    .caution-flag {
        border-left-color: var(--amber);
        background: var(--amber-bg);
    }
    .info-flag {
        border-left-color: var(--blue);
        background: var(--blue-bg);
    }

    /* ── SIGNAL CARDS (M&A news) ── */
    .signal-card {
        background: var(--bg-s1);
        border: 1px solid var(--b-struct);
        border-radius: 6px;
        padding: 0.6rem 0.8rem;
        margin-bottom: 0.3rem;
        font-size: 0.78rem;
        animation: fadeUp 0.3s ease both;
    }
    .signal-type {
        font-family: var(--mono);
        font-size: 0.6rem;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        font-weight: 700;
        margin-bottom: 0.15rem;
    }

    /* ── INTEL STRIP (supplier card) ── */
    .intel-strip {
        background: var(--bg-page);
        border: 1px solid var(--b-struct);
        border-left: 2px solid var(--blue-dim);
        border-radius: 0 6px 6px 0;
        padding: 0.65rem 0.9rem;
        margin-top: 0.5rem;
        font-size: 0.8rem;
        color: var(--t2);
        animation: fadeUp 0.3s ease both;
    }
    .intel-eyebrow {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        color: var(--t4);
        margin-bottom: 0.4rem;
    }

    /* ── STAKEHOLDER PILLS ── */
    .stakeholder-pill {
        display: inline-flex;
        align-items: center;
        gap: 5px;
        background: rgba(15,31,56,0.8);
        border: 1px solid var(--b-active);
        border-radius: 999px;
        padding: 0.25rem 0.65rem;
        font-family: var(--mono);
        font-size: 0.7rem;
        color: var(--blue-glow);
        margin: 0.2rem 0.2rem 0.2rem 0;
        letter-spacing: 0.04em;
    }

    /* ── COVER PAGE ── */
    .cover-outer {
        position: relative;
        width: 100%;
        min-height: 78vh;
        border-radius: var(--r-xl);
        overflow: hidden;
        border: 1px solid rgba(96,165,250,0.12);
        box-shadow: 0 80px 180px rgba(0,0,0,0.9);
        display: flex;
        animation: assembleIn 0.7s ease both;
    }
    .cover-left {
        position: relative;
        flex: 0 0 58%;
        padding: 3.5rem;
        display: flex;
        flex-direction: column;
        justify-content: center;
        background: linear-gradient(155deg, #030810 0%, #060F1E 50%, #091628 100%);
        border-right: 1px solid rgba(96,165,250,0.08);
        overflow: hidden;
    }
    #piq-canvas {
        position: absolute;
        inset: 0;
        width: 100%; height: 100%;
        opacity: 0.28;
        pointer-events: none;
    }
    .cover-content { position: relative; z-index: 2; }
    .cover-eyebrow {
        font-family: var(--mono);
        font-size: 0.65rem;
        letter-spacing: 0.28em;
        text-transform: uppercase;
        color: var(--blue);
        margin-bottom: 1.4rem;
        display: flex;
        align-items: center;
        gap: 0.7rem;
    }
    .cover-eyebrow::before {
        content: "";
        display: inline-block;
        width: 32px; height: 1px;
        background: var(--blue);
        opacity: 0.5;
    }
    .cover-title {
        font-family: var(--serif);
        font-size: clamp(4rem, 8vw, 7rem);
        line-height: 0.88;
        color: var(--t1);
        letter-spacing: -0.03em;
        font-weight: 400;
        margin-bottom: 1.6rem;
    }
    .cover-title .iq {
        -webkit-text-stroke: 2.5px #60A5FA;
        color: transparent;
        font-style: italic;
        letter-spacing: -0.04em;
    }
    .cover-tagline {
        font-family: var(--sans);
        font-size: 1.02rem;
        font-weight: 300;
        color: #94A3B8;
        line-height: 1.8;
        max-width: 460px;
        margin-bottom: 0;
    }
    .cover-tagline strong { color: #CBD5E1; font-weight: 500; }
    .cover-stats {
        display: flex;
        gap: 2.2rem;
        margin-top: 2.4rem;
        padding-top: 1.8rem;
        border-top: 1px solid rgba(148,163,184,0.07);
    }
    .cover-stat { display: flex; flex-direction: column; gap: 0.2rem; }
    .cover-stat-num {
        font-family: var(--mono);
        font-size: 1.5rem;
        font-weight: 700;
        color: var(--blue-glow);
        line-height: 1;
        letter-spacing: -0.02em;
    }
    .cover-stat-label {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        color: var(--t5);
    }
    .cover-cta {
        margin-top: 2.6rem;
        display: inline-block;
        font-family: var(--mono);
        font-size: 0.75rem;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        color: #93C5FD;
        background: #1E3A5F;
        border: 1px solid var(--blue);
        border-radius: 6px;
        padding: 0.75rem 2.2rem;
        cursor: pointer;
        transition: all 220ms ease;
        text-decoration: none;
    }
    .cover-cta:hover {
        background: var(--blue-dim);
        border-color: var(--blue-glow);
        box-shadow: 0 0 28px rgba(59,130,246,0.35);
        color: #fff;
    }
    .cover-footer {
        margin-top: auto;
        padding-top: 2rem;
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.14em;
        color: var(--t5);
    }

    /* Cover right panel — live preview */
    .cover-right {
        flex: 0 0 42%;
        background: #040B16;
        padding: 2.4rem 2rem;
        display: flex;
        flex-direction: column;
        gap: 0;
    }
    .preview-header {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        font-family: var(--mono);
        font-size: 0.6rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        color: var(--t4);
        margin-bottom: 1.4rem;
        padding-bottom: 0.8rem;
        border-bottom: 1px solid var(--b-struct);
    }
    .preview-dot {
        width: 5px; height: 5px;
        border-radius: 50%;
        background: var(--green);
        animation: pulseGreen 2s infinite;
    }
    .preview-module {
        padding: 1rem 0;
        border-bottom: 1px solid rgba(96,165,250,0.05);
    }
    .preview-module:last-child { border-bottom: none; }
    .preview-module-label {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        color: var(--t5);
        margin-bottom: 0.8rem;
    }

    /* Preview score bars */
    .preview-supplier {
        display: flex;
        align-items: center;
        gap: 0.6rem;
        margin-bottom: 0.5rem;
        animation: fadeUp 0.4s ease both;
    }
    .preview-sup-name {
        font-family: var(--mono);
        font-size: 0.68rem;
        color: var(--t3);
        min-width: 72px;
        letter-spacing: 0.02em;
    }
    .preview-bar-track {
        flex: 1;
        height: 4px;
        background: rgba(96,165,250,0.07);
        border-radius: 2px;
        overflow: hidden;
    }
    .preview-bar-fill {
        height: 100%;
        border-radius: 2px;
        animation: drawBar 0.8s ease-out both;
    }
    .preview-score {
        font-family: var(--mono);
        font-size: 0.72rem;
        min-width: 24px;
        text-align: right;
    }
    .preview-badge {
        font-family: var(--mono);
        font-size: 0.54rem;
        letter-spacing: 0.08em;
        color: var(--green);
        border: 1px solid rgba(74,222,128,0.2);
        background: var(--green-bg);
        border-radius: 3px;
        padding: 0.08rem 0.3rem;
        white-space: nowrap;
    }

    /* Preview terminal */
    .preview-terminal {
        font-family: var(--mono);
        font-size: 0.72rem;
        background: rgba(3,8,15,0.6);
        border-left: 2px solid var(--blue-dim);
        border-radius: 0 4px 4px 0;
        padding: 0.6rem 0.8rem;
    }
    .preview-terminal .pt-row {
        display: flex;
        gap: 0.8rem;
        padding: 0.1rem 0;
        color: #93C5FD;
    }
    .preview-terminal .pt-arrow { color: var(--blue-dim); }
    .preview-terminal .pt-key   { color: var(--t4); min-width: 80px; }

    .cover-bottom-note {
        font-family: var(--sans);
        font-size: 0.7rem;
        color: var(--t5);
        text-align: center;
        margin-top: auto;
        padding-top: 1.2rem;
    }

    /* ── PILLAR CARDS (below main cover on mobile) ── */
    .pillar {
        background: rgba(6,13,26,0.72);
        border: 1px solid var(--b-struct);
        border-radius: var(--r-lg);
        padding: 1.5rem;
        backdrop-filter: blur(8px);
        transition: all 250ms ease;
        position: relative;
        overflow: hidden;
        height: 100%;
        animation: fadeUp 0.5s ease both;
    }
    .pillar::after {
        content: "";
        position: absolute;
        bottom: 0; left: 0; right: 0;
        height: 2px;
        background: linear-gradient(90deg, transparent, rgba(96,165,250,0.35), transparent);
        opacity: 0;
        transition: opacity 250ms ease;
    }
    .pillar:hover {
        background: rgba(10,22,40,0.9);
        border-color: var(--b-active);
        transform: translateY(-3px);
        box-shadow: 0 20px 40px rgba(0,0,0,0.5), 0 0 30px rgba(96,165,250,0.06);
    }
    .pillar:hover::after { opacity: 1; }
    .pillar-icon   { font-size: 1.4rem; margin-bottom: 0.7rem; display: block; }
    .pillar-num    { font-family: var(--mono); font-size: 0.6rem; letter-spacing: 0.14em; color: var(--blue); margin-bottom: 0.45rem; text-transform: uppercase; }
    .pillar-title  { font-size: 1rem; font-weight: 700; color: var(--t1); margin-bottom: 0.45rem; letter-spacing: -0.01em; }
    .pillar-body   { font-size: 0.82rem; color: var(--t3); line-height: 1.62; }
    .pillar-tag    {
        display: inline-block; margin-top: 0.8rem;
        font-family: var(--mono); font-size: 0.6rem; letter-spacing: 0.08em;
        color: var(--blue-glow);
        background: rgba(96,165,250,0.06);
        border: 1px solid rgba(96,165,250,0.14);
        border-radius: 4px; padding: 0.14rem 0.5rem;
        text-transform: uppercase;
    }

    /* ── HERO BANNER (dashboard) ── */
    .hero {
        background: linear-gradient(135deg, #04090F 0%, #060F1A 60%, #091828 100%);
        border: 1px solid var(--b-struct);
        border-radius: var(--r-lg);
        padding: 2rem 2.2rem;
        margin-bottom: 1rem;
        position: relative;
        overflow: hidden;
        animation: fadeUp 0.4s ease both;
    }
    .hero::before {
        content: "";
        position: absolute;
        top: -80px; right: -80px;
        width: 320px; height: 320px;
        border-radius: 999px;
        background: radial-gradient(circle, rgba(59,130,246,0.08) 0%, transparent 65%);
        pointer-events: none;
    }
    .hero-eyebrow {
        font-family: var(--mono);
        font-size: 0.65rem;
        letter-spacing: 0.2em;
        text-transform: uppercase;
        color: var(--blue);
        margin-bottom: 0.6rem;
    }
    .hero-title {
        font-family: var(--serif);
        font-size: 2.6rem;
        color: var(--t1);
        line-height: 1;
        letter-spacing: -0.02em;
        margin-bottom: 0.6rem;
    }
    .hero-sub {
        font-size: 0.9rem;
        color: var(--t3);
        line-height: 1.6;
        max-width: 700px;
    }

    /* ── METRIC CARDS ── */
    .metric-card {
        background: var(--bg-s1);
        border: 1px solid var(--b-struct);
        border-radius: var(--r-md);
        padding: 1rem 1.1rem;
        text-align: center;
        transition: all 200ms ease;
        animation: fadeUp 0.35s ease both;
    }
    .metric-card:hover {
        border-color: var(--b-active);
        background: var(--bg-s2);
    }
    .metric-num {
        font-family: var(--mono);
        font-size: 1.9rem;
        font-weight: 700;
        color: var(--blue-glow);
        letter-spacing: -0.02em;
        line-height: 1.1;
    }
    .metric-label {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        color: var(--t4);
        margin-top: 0.3rem;
    }

    /* ── EXECUTIVE SUMMARY CARDS ── */
    .exec-card {
        background: var(--bg-s1);
        border: 1px solid var(--b-struct);
        border-radius: var(--r-md);
        padding: 1.2rem;
        animation: fadeUp 0.4s ease both;
        transition: border-color 200ms ease;
    }
    .exec-card:hover { border-color: var(--b-active); }
    .exec-card-title {
        font-family: var(--mono);
        font-size: 0.62rem;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        color: var(--blue);
        margin-bottom: 0.6rem;
    }
    .exec-card-body {
        font-size: 0.84rem;
        color: var(--t2);
        line-height: 1.6;
    }

    /* ── SUBCATEGORY INTELLIGENCE ── */
    .sub-intel {
        background: var(--bg-s1);
        border: 1px solid var(--b-struct);
        border-radius: var(--r-md);
        padding: 0.8rem 1rem;
        margin-top: 1rem;
        animation: fadeUp 0.35s ease both;
    }
    .sub-intel-eyebrow {
        font-family: var(--mono);
        font-size: 0.58rem;
        letter-spacing: 0.16em;
        text-transform: uppercase;
        color: var(--t5);
        margin-bottom: 0.5rem;
    }

    /* ── RFP / PLAYBOOK ── */
    .rfp-week-card {
        background: var(--bg-s1);
        border: 1px solid var(--b-struct);
        border-radius: var(--r-md);
        padding: 1rem 1.2rem;
        margin-bottom: 0.6rem;
        transition: border-color 200ms ease;
    }
    .rfp-week-card:hover { border-color: var(--b-active); }
    .rfp-week-label {
        font-family: var(--mono);
        font-size: 0.6rem;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        color: var(--blue);
        margin-bottom: 0.3rem;
    }
    .rfp-stakeholder-group-title {
        font-family: var(--mono);
        font-size: 0.62rem;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        font-weight: 700;
        margin-bottom: 0.6rem;
        color: var(--t3);
    }
    .rfp-stakeholder-pill {
        display: inline-flex;
        align-items: center;
        gap: 5px;
        background: var(--bg-s2);
        border: 1px solid var(--b-active);
        border-radius: 999px;
        padding: 0.25rem 0.6rem;
        font-size: 0.76rem;
        font-weight: 500;
        color: var(--t2);
        margin: 0.15rem 0.15rem 0.15rem 0;
    }

    /* ── WEIGHT BADGE ── */
    .weight-rec {
        display: inline-block;
        background: rgba(96,165,250,0.08);
        border: 1px solid rgba(96,165,250,0.22);
        color: var(--blue-glow);
        border-radius: 999px;
        padding: 0.1rem 0.45rem;
        font-size: 0.64rem;
        font-weight: 700;
        font-family: var(--mono);
        margin-left: 0.35rem;
    }

    /* ── STREAMLIT WIDGET OVERRIDES ── */
    /* Inputs */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input {
        background: #060F20 !important;
        border: 1px solid rgba(96,165,250,0.18) !important;
        border-radius: 6px !important;
        color: var(--t1) !important;
        font-family: var(--sans) !important;
        font-size: 0.85rem !important;
    }
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        border-color: var(--blue) !important;
        box-shadow: 0 0 0 3px rgba(59,130,246,0.12) !important;
    }
    /* Selectbox */
    .stSelectbox > div > div {
        background: #060F20 !important;
        border: 1px solid rgba(96,165,250,0.18) !important;
        border-radius: 6px !important;
        color: var(--t1) !important;
    }
    /* Sliders */
    .stSlider > div > div > div > div {
        background: var(--blue) !important;
    }
    [data-baseweb="slider"] [role="slider"] {
        background: #fff !important;
        border: 2px solid var(--blue) !important;
        box-shadow: 0 0 8px rgba(59,130,246,0.4) !important;
    }
    /* Labels */
    .stTextInput label, .stSelectbox label,
    .stSlider label, .stNumberInput label,
    .stTextArea label, .stToggle label {
        font-family: var(--mono) !important;
        font-size: 0.82rem !important;
        letter-spacing: 0.05em !important;
        color: var(--t3) !important;
        text-transform: uppercase !important;
    }
    /* ══════════════════════════════════════════════════════════════
       BUTTONS — Universal dark-theme fix.
       Targets BOTH class selectors (older Streamlit) and data-testid
       selectors (Streamlit 1.35+) with hardcoded colors so CSS vars
       never fail to inherit. Every interactive state is covered.
       ══════════════════════════════════════════════════════════════ */

    /* ── Secondary / default buttons ── */
    .stButton > button,
    button[data-testid="stBaseButton-secondary"],
    button[data-testid="stBaseButton-tertiary"],
    button[data-testid="stBaseButton-minimal"],
    [data-testid="stFormSubmitButton"] > button {
        font-family: 'JetBrains Mono', 'Fira Mono', monospace !important;
        font-size: 0.72rem !important;
        letter-spacing: 0.08em !important;
        text-transform: uppercase !important;
        border-radius: 6px !important;
        background: #0A1628 !important;
        color: #F1F5F9 !important;
        border: 1px solid rgba(96,165,250,0.22) !important;
        white-space: normal !important;
        word-break: break-word !important;
        min-height: 2.1rem !important;
        line-height: 1.3 !important;
        padding: 0.35rem 0.75rem !important;
        transition: background 120ms ease, border-color 120ms ease, box-shadow 120ms ease !important;
    }
    .stButton > button:hover,
    button[data-testid="stBaseButton-secondary"]:hover,
    button[data-testid="stBaseButton-tertiary"]:hover,
    button[data-testid="stBaseButton-minimal"]:hover,
    [data-testid="stFormSubmitButton"] > button:hover {
        background: #0F1F38 !important;
        color: #F1F5F9 !important;
        border-color: rgba(96,165,250,0.50) !important;
        box-shadow: 0 0 10px rgba(96,165,250,0.10) !important;
    }
    .stButton > button:active,
    .stButton > button:focus,
    .stButton > button:focus-visible,
    .stButton > button:focus:not(:active),
    button[data-testid="stBaseButton-secondary"]:active,
    button[data-testid="stBaseButton-secondary"]:focus,
    button[data-testid="stBaseButton-secondary"]:focus-visible,
    button[data-testid="stBaseButton-tertiary"]:active,
    button[data-testid="stBaseButton-tertiary"]:focus,
    button[data-testid="stBaseButton-tertiary"]:focus-visible,
    button[data-testid="stBaseButton-minimal"]:active,
    button[data-testid="stBaseButton-minimal"]:focus,
    button[data-testid="stBaseButton-minimal"]:focus-visible,
    [data-testid="stFormSubmitButton"] > button:active,
    [data-testid="stFormSubmitButton"] > button:focus,
    [data-testid="stFormSubmitButton"] > button:focus-visible {
        background: #0F1F38 !important;
        color: #F1F5F9 !important;
        border-color: rgba(96,165,250,0.50) !important;
        outline: none !important;
        box-shadow: 0 0 0 2px rgba(96,165,250,0.18) !important;
    }

    /* ── Primary buttons ── */
    .stButton > button[kind="primary"],
    button[data-testid="stBaseButton-primary"] {
        background: #1D4ED8 !important;
        color: #ffffff !important;
        border: 1px solid #3B82F6 !important;
        font-family: 'JetBrains Mono', 'Fira Mono', monospace !important;
        font-size: 0.72rem !important;
        letter-spacing: 0.08em !important;
        text-transform: uppercase !important;
        border-radius: 6px !important;
        min-height: 2.1rem !important;
        white-space: normal !important;
        word-break: break-word !important;
        transition: background 120ms ease, box-shadow 120ms ease !important;
    }
    .stButton > button[kind="primary"]:hover,
    button[data-testid="stBaseButton-primary"]:hover {
        background: #2563EB !important;
        color: #ffffff !important;
        border-color: #60A5FA !important;
        box-shadow: 0 0 20px rgba(59,130,246,0.30) !important;
    }
    .stButton > button[kind="primary"]:active,
    .stButton > button[kind="primary"]:focus,
    .stButton > button[kind="primary"]:focus-visible,
    button[data-testid="stBaseButton-primary"]:active,
    button[data-testid="stBaseButton-primary"]:focus,
    button[data-testid="stBaseButton-primary"]:focus-visible {
        background: #2563EB !important;
        color: #ffffff !important;
        border-color: #60A5FA !important;
        outline: none !important;
        box-shadow: 0 0 0 2px rgba(59,130,246,0.30) !important;
    }
    /* Expander */
    .streamlit-expanderHeader {
        font-family: var(--mono) !important;
        font-size: 0.86rem !important;
        letter-spacing: 0.04em !important;
        color: var(--t2) !important;
        background: var(--bg-s1) !important;
        border: 1px solid var(--b-struct) !important;
        border-radius: 6px !important;
    }
    /* Caption */
    .stCaption { color: var(--t3) !important; font-size: 0.82rem !important; }

    /* ── PERSISTENT SUPPLY CHAIN BACKGROUND ── */
    /* CSS-only approach: SVG pattern via background-image data URI.
       Works in Streamlit because it's in the stylesheet, not a canvas element.
       The pattern recreates the node-and-edge aesthetic without JavaScript. */
    .stApp::before {
        content: "";
        position: fixed;
        inset: 0;
        z-index: 0;
        pointer-events: none;
        background-image:
            /* Diagonal route lines — cargo paths */
            repeating-linear-gradient(
                -45deg,
                transparent,
                transparent 120px,
                rgba(96,165,250,0.022) 120px,
                rgba(96,165,250,0.022) 121px
            ),
            repeating-linear-gradient(
                45deg,
                transparent,
                transparent 180px,
                rgba(96,165,250,0.018) 180px,
                rgba(96,165,250,0.018) 181px
            ),
            /* Horizontal shipping lanes */
            repeating-linear-gradient(
                0deg,
                transparent,
                transparent 200px,
                rgba(96,165,250,0.012) 200px,
                rgba(96,165,250,0.012) 201px
            ),
            /* Radial node glows at key positions */
            radial-gradient(ellipse at 15% 30%, rgba(59,130,246,0.06) 0%, transparent 35%),
            radial-gradient(ellipse at 85% 20%, rgba(59,130,246,0.05) 0%, transparent 30%),
            radial-gradient(ellipse at 50% 70%, rgba(59,130,246,0.04) 0%, transparent 28%),
            radial-gradient(ellipse at 25% 75%, rgba(59,130,246,0.04) 0%, transparent 25%),
            radial-gradient(ellipse at 75% 60%, rgba(59,130,246,0.05) 0%, transparent 30%),
            radial-gradient(ellipse at 10% 60%, rgba(96,165,250,0.03) 0%, transparent 20%),
            radial-gradient(ellipse at 90% 75%, rgba(96,165,250,0.03) 0%, transparent 20%);
        background-size: auto;
        background-attachment: fixed;
    }

    /* Ensure all Streamlit content sits above the background */
    .stApp > div { position: relative; z-index: 1; }
    [data-testid="stAppViewContainer"] { background: transparent; }

    /* ── READABILITY OVERRIDES — Applied globally ── */
    /* Body text */
    .stMarkdown p {
        color: #CBD5E1 !important;
        font-size: 0.88rem !important;
        line-height: 1.7 !important;
    }
    /* Section headers */
    .stMarkdown h3 {
        color: #F1F5F9 !important;
        font-size: 1.35rem !important;
        letter-spacing: -0.02em !important;
        margin-bottom: 0.3rem !important;
    }
    .stMarkdown h4 {
        color: #E2E8F0 !important;
        font-size: 0.95rem !important;
        font-weight: 700 !important;
        margin-bottom: 0.3rem !important;
    }
    /* Mission header — larger, readable */
    .mission-header {
        padding: 0.85rem 0 0.85rem !important;
        margin-bottom: 1rem !important;
    }
    .mission-breadcrumb {
        font-size: 0.68rem !important;
        color: #60A5FA !important;
        letter-spacing: 0.14em !important;
    }
    .mission-event {
        font-size: 1.15rem !important;
        color: #E2E8F0 !important;
    }
    /* Expander headers — readable */
    details summary {
        font-size: 0.85rem !important;
        color: #E2E8F0 !important;
        padding: 0.6rem 0.8rem !important;
    }
    /* Right panel widget labels */
    .stTextInput label, .stSelectbox label,
    .stSlider label, .stNumberInput label,
    .stTextArea label, .stToggle label,
    .stCheckbox label, .stDateInput label, .stFileUploader label {
        font-size: 0.85rem !important;
        color: var(--t3) !important;
        text-transform: uppercase !important;
        letter-spacing: 0.05em !important;
    }
    /* Caption text */
    .stCaption { color: var(--t3) !important; font-size: 0.82rem !important; }
    /* Radio button labels */
    .stRadio label { color: var(--t2) !important; font-size: 0.92rem !important; }
    /* Muted utility */
    /* ── UTILITY ── */
    .muted     { color: var(--t2) !important; font-size: 0.95rem; line-height: 1.65; }
    .secondary { color: var(--t3) !important; }
    hr { border-color: var(--b-struct) !important; margin: 1.4rem 0 !important; }

    /* ── INTAKE RADIO CARDS — pill options that are easy to read and select ── */
    div[role="radiogroup"] {
        display: flex !important;
        flex-wrap: wrap !important;
        gap: 0.45rem !important;
        margin-top: 0.2rem !important;
    }
    div[role="radiogroup"] > label {
        background: #060D1A !important;
        border: 1px solid rgba(96,165,250,0.14) !important;
        border-radius: 8px !important;
        padding: 0.4rem 0.85rem !important;
        cursor: pointer !important;
        transition: background 120ms ease, border-color 120ms ease !important;
        display: flex !important;
        align-items: center !important;
        gap: 0.35rem !important;
    }
    div[role="radiogroup"] > label:hover {
        background: #0A1628 !important;
        border-color: rgba(96,165,250,0.35) !important;
    }
    div[role="radiogroup"] > label p,
    div[role="radiogroup"] > label span:not([data-testid]) {
        font-size: 0.83rem !important;
        color: var(--t3) !important;
        font-weight: 500 !important;
        margin: 0 !important;
        line-height: 1.4 !important;
    }
    div[role="radiogroup"] > label:has(input[type="radio"]:checked) {
        background: rgba(29,78,216,0.18) !important;
        border-color: #3B82F6 !important;
    }
    div[role="radiogroup"] > label:has(input[type="radio"]:checked) p,
    div[role="radiogroup"] > label:has(input[type="radio"]:checked) span:not([data-testid]) {
        color: #F1F5F9 !important;
        font-weight: 600 !important;
    }
    div[role="radiogroup"] > label input[type="radio"] {
        position: absolute !important;
        opacity: 0 !important;
        pointer-events: none !important;
        width: 0 !important;
        height: 0 !important;
    }
    /* Intake question label */
    [data-testid="stRadio"] > label {
        font-size: 0.82rem !important;
        color: #CBD5E1 !important;
        font-weight: 600 !important;
        letter-spacing: 0 !important;
        text-transform: none !important;
        margin-bottom: 0.3rem !important;
    }

    /* ── Download buttons — same dark treatment as secondary ── */
    .stDownloadButton > button,
    button[data-testid="stBaseButton-download"] {
        background: #0A1628 !important;
        color: #F1F5F9 !important;
        border: 1px solid rgba(96,165,250,0.22) !important;
        border-radius: 6px !important;
        font-family: 'JetBrains Mono', 'Fira Mono', monospace !important;
        font-size: 0.72rem !important;
        text-transform: uppercase !important;
        letter-spacing: 0.08em !important;
        white-space: normal !important;
        word-break: break-word !important;
        min-height: 2.1rem !important;
        transition: background 120ms ease !important;
    }
    .stDownloadButton > button:hover,
    .stDownloadButton > button:focus,
    .stDownloadButton > button:active,
    .stDownloadButton > button:focus-visible,
    button[data-testid="stBaseButton-download"]:hover,
    button[data-testid="stBaseButton-download"]:focus,
    button[data-testid="stBaseButton-download"]:active,
    button[data-testid="stBaseButton-download"]:focus-visible {
        background: #0F1F38 !important;
        color: #F1F5F9 !important;
        border-color: rgba(96,165,250,0.50) !important;
        outline: none !important;
        box-shadow: 0 0 0 2px rgba(96,165,250,0.15) !important;
    }
    /* Tabs — focus state (keyboard nav) */
    .stTabs [data-baseweb="tab"]:focus,
    .stTabs [data-baseweb="tab"]:focus-visible {
        color: var(--t2) !important;
        background: transparent !important;
        outline: none !important;
        border-bottom-color: var(--b-active) !important;
    }
    .stTabs [aria-selected="true"]:focus,
    .stTabs [aria-selected="true"]:focus-visible,
    .stTabs [aria-selected="true"]:active {
        color: var(--t1) !important;
        background: transparent !important;
        border-bottom: 2px solid var(--blue) !important;
        outline: none !important;
    }
    /* Selectbox — dropdown popover */
    [data-baseweb="popover"] [data-baseweb="menu"],
    [data-baseweb="menu"] {
        background: #060D1A !important;
        border: 1px solid var(--b-active) !important;
        border-radius: 8px !important;
    }
    [data-baseweb="option"] {
        background: #060D1A !important;
        color: var(--t1) !important;
    }
    [data-baseweb="option"]:hover,
    [data-baseweb="option"][aria-selected="true"] {
        background: var(--bg-s2) !important;
        color: var(--t1) !important;
    }
    /* Selectbox trigger when open/focused */
    [data-baseweb="select"] > div:focus,
    [data-baseweb="select"] > div[aria-expanded="true"],
    .stSelectbox [data-baseweb="select"]:focus-within > div {
        background: var(--bg-s1) !important;
        border-color: var(--blue) !important;
        color: var(--t1) !important;
    }
    /* Text/number inputs focus */
    .stTextInput input:focus,
    .stNumberInput input:focus,
    .stTextArea textarea:focus {
        background: var(--bg-s1) !important;
        color: var(--t1) !important;
        border-color: var(--blue) !important;
        box-shadow: none !important;
        outline: none !important;
    }
    /* Expander — active state */
    details[open] > summary,
    .streamlit-expanderHeader[aria-expanded="true"] {
        color: var(--t1) !important;
        background: var(--bg-s2) !important;
        border-color: var(--b-active) !important;
    }

    /* ── METRIC COMPONENTS ── */
    [data-testid="stMetricLabel"] {
        color: var(--t3) !important;
        font-size: 0.78rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.06em !important;
        text-transform: uppercase !important;
    }
    [data-testid="stMetricValue"] {
        color: var(--t1) !important;
        font-size: 1.6rem !important;
        font-weight: 700 !important;
    }
    [data-testid="stMetricDelta"] svg { display: none !important; }
    [data-testid="stMetricDelta"] > div {
        color: var(--t2) !important;
        font-size: 0.82rem !important;
    }

    /* ── ALERT / CALLOUT BOXES ── */
    [data-testid="stAlert"] {
        border-radius: 8px !important;
        border-left-width: 3px !important;
    }
    [data-testid="stAlert"] p,
    [data-testid="stAlert"] li,
    [data-testid="stAlert"] span {
        color: var(--t1) !important;
        font-size: 0.88rem !important;
        line-height: 1.65 !important;
    }

    /* ── DATAFRAME / TABLE ── */
    [data-testid="stDataFrame"] th {
        background: #060D1A !important;
        color: var(--t3) !important;
        font-size: 0.78rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.05em !important;
        text-transform: uppercase !important;
    }
    [data-testid="stDataFrame"] td {
        color: var(--t2) !important;
        font-size: 0.84rem !important;
        background: transparent !important;
    }

    /* ── SIDEBAR ── */
    [data-testid="stSidebar"] {
        background: #060D1A !important;
        border-right: 1px solid var(--b-struct) !important;
    }
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span {
        color: var(--t2) !important;
    }
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] h4 {
        color: var(--t1) !important;
    }

    /* ── TOGGLE / CHECKBOX ── */
    .stToggle p { color: var(--t2) !important; font-size: 0.88rem !important; }
    .stCheckbox span { color: var(--t2) !important; font-size: 0.88rem !important; }

    /* ── MULTISELECT tags ── */
    [data-baseweb="tag"] {
        background: rgba(59,130,246,0.2) !important;
        border: 1px solid rgba(59,130,246,0.4) !important;
    }
    [data-baseweb="tag"] span { color: var(--t1) !important; font-size: 0.82rem !important; }

    /* ── PROGRESS BAR ── */
    .stProgress > div > div > div {
        background: linear-gradient(90deg, #1D4ED8, #3B82F6) !important;
    }

</style>
""",
    unsafe_allow_html=True,
)



# =========================================================
# SUPPLIER INTEL STRIP — shared across tabs
# =========================================================
def render_supplier_intel_strip(supplier: dict) -> None:
    """Render a compact live-data bar for a supplier record throughout the flow."""
    ticker = (supplier.get("Ticker") or "").strip().upper()
    sec_ctx = supplier.get("SEC Context") or {}
    yf_ctx  = supplier.get("Alpha Context") or {}
    name    = supplier.get("Supplier", "")

    if not ticker or ticker in ("PRIVATE", "—", ""):
        return

    parts = []
    if sec_ctx.get("found"):
        filings = sec_ctx.get("recent_filings", [])
        parts.append(f'<span style="color:#60A5FA;font-weight:700">{ticker}</span>')
        parts.append(f'<span style="color:#D0E0EF">{sec_ctx.get("company_name", name)}</span>')
        if filings:
            parts.append(f'<span style="color:#9EB8CE">SEC: {len(filings)} recent filings</span>')
    else:
        parts.append(f'<span style="color:#60A5FA;font-weight:700">{ticker}</span>')

    if yf_ctx:
        mc = yf_ctx.get("market_cap_fmt") or yf_ctx.get("MarketCapitalization")
        pe = yf_ctx.get("pe_ratio") or yf_ctx.get("PERatio")
        rev = yf_ctx.get("revenue_fmt") or yf_ctx.get("RevenueTTM")
        if mc and mc not in ("N/A", "None", ""):
            parts.append(f'<span style="color:#4ADE80">MCap: {mc}</span>')
        if pe and pe not in ("N/A", "None", ""):
            parts.append(f'<span style="color:#C4D3E8">P/E: {pe}</span>')
        if rev and rev not in ("N/A", "None", ""):
            parts.append(f'<span style="color:#C4D3E8">Rev: {rev}</span>')

    news = get_supplier_news_signals(ticker)
    risk_signals  = [s for s in news.get("signals", []) if s.get("risk")]
    oppty_signals = [s for s in news.get("signals", []) if not s.get("risk")]
    if risk_signals:
        parts.append(f'<span style="color:#F87171">{risk_signals[0]["icon"]} {risk_signals[0]["type"]}</span>')
    if oppty_signals:
        parts.append(f'<span style="color:#4ADE80">{oppty_signals[0]["icon"]} {oppty_signals[0]["type"]}</span>')

    if not parts:
        return

    st.markdown(
        '<div style="background:rgba(6,13,26,0.9);border:1px solid rgba(96,165,250,0.15);'
        'border-left:3px solid #3B82F6;border-radius:8px;padding:0.45rem 0.9rem;'
        'margin-bottom:0.8rem;display:flex;flex-wrap:wrap;gap:1rem;align-items:center">'
        '<span style="font-size:0.6rem;color:#3B82F6;text-transform:uppercase;'
        'letter-spacing:0.16em;font-family:monospace;white-space:nowrap">Live Intel</span>'
        + "  ·  ".join(f'<span style="font-size:0.78rem;font-family:monospace">{p}</span>' for p in parts)
        + "</div>",
        unsafe_allow_html=True,
    )


# =========================================================
# SEC / ALPHA VANTAGE
# =========================================================
SEC_HEADERS = {
    "User-Agent": "ProcureIQ public demo contact: procurement-demo@example.com",
    "Accept-Encoding": "gzip, deflate",
}

@st.cache_data(show_spinner=False, ttl=86400)
def get_sec_ticker_map() -> pd.DataFrame:
    url = "https://www.sec.gov/files/company_tickers.json"
    response = requests.get(url, headers=SEC_HEADERS, timeout=15)
    response.raise_for_status()
    raw = response.json()
    records = []
    for _, item in raw.items():
        records.append({
            "ticker": str(item.get("ticker", "")).upper(),
            "title": item.get("title", ""),
            "cik": str(item.get("cik_str", "")).zfill(10),
        })
    return pd.DataFrame(records)


@st.cache_data(show_spinner=False, ttl=86400)
def get_sec_company_context(ticker: str) -> Optional[dict]:
    ticker = (ticker or "").upper().strip()
    if not ticker:
        return None
    ticker_map = get_sec_ticker_map()
    row = ticker_map[ticker_map["ticker"] == ticker]
    if row.empty:
        return {"found": False, "message": "Ticker not found in SEC public-company list."}
    cik = row.iloc[0]["cik"]
    title = row.iloc[0]["title"]
    submissions_url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    response = requests.get(submissions_url, headers=SEC_HEADERS, timeout=15)
    response.raise_for_status()
    data = response.json()
    filings = data.get("filings", {}).get("recent", {})
    forms = filings.get("form", [])
    filing_dates = filings.get("filingDate", [])
    recent_items = []
    for i in range(min(5, len(forms), len(filing_dates))):
        recent_items.append(f"{forms[i]} ({filing_dates[i]})")
    # Oldest filing year — useful as a lower bound for Years in Business
    oldest_year = None
    if filing_dates:
        try:
            oldest_year = int(min(filing_dates)[:4])
        except (ValueError, TypeError):
            pass
    return {
        "found": True,
        "ticker": ticker,
        "company_name": title,
        "cik": cik,
        "recent_filings": recent_items,
        "oldest_filing_year": oldest_year,
    }


@st.cache_data(show_spinner=False, ttl=7200)
def fetch_edgar_xbrl_financials(cik: str) -> dict:
    """
    Pull actual financial metrics from SEC EDGAR XBRL company facts API.
    Returns revenue, net income, long-term debt, cash, and operating cash flow
    from the most recent annual 10-K filing. All free, no API key required.
    """
    _empty = {"found": False}
    if not cik:
        return _empty
    cik_padded = str(cik).zfill(10)
    url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik_padded}.json"
    try:
        resp = requests.get(url, headers=SEC_HEADERS, timeout=15)
        resp.raise_for_status()
        facts = resp.json().get("facts", {}).get("us-gaap", {})
    except Exception:
        return _empty

    def _latest_annual(tag_names: list) -> Optional[dict]:
        for tag in tag_names:
            entries = facts.get(tag, {}).get("units", {}).get("USD", [])
            annual  = [e for e in entries if e.get("form") == "10-K" and e.get("fp") == "FY" and e.get("val") is not None]
            if annual:
                return sorted(annual, key=lambda x: x.get("end", ""))[-1]
        return None

    def _latest_balance(tag_names: list) -> Optional[dict]:
        for tag in tag_names:
            entries = facts.get(tag, {}).get("units", {}).get("USD", [])
            qs = [e for e in entries if e.get("form") in ("10-K", "10-Q") and e.get("val") is not None]
            if qs:
                return sorted(qs, key=lambda x: x.get("end", ""))[-1]
        return None

    revenue  = _latest_annual(["Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax", "SalesRevenueNet", "SalesRevenueGoodsNet"])
    net_inc  = _latest_annual(["NetIncomeLoss", "NetIncome", "ProfitLoss"])
    op_cf    = _latest_annual(["NetCashProvidedByUsedInOperatingActivities", "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations"])
    debt     = _latest_balance(["LongTermDebt", "LongTermDebtNoncurrent", "LongTermDebtAndCapitalLeaseObligations"])
    cash     = _latest_balance(["CashAndCashEquivalentsAtCarryingValue", "CashCashEquivalentsAndShortTermInvestments", "CashAndCashEquivalents"])

    def _fmt_b(entry: Optional[dict]) -> Optional[str]:
        if not entry:
            return None
        v = entry["val"]
        if abs(v) >= 1e9:
            return f"${v/1e9:.1f}B"
        if abs(v) >= 1e6:
            return f"${v/1e6:.0f}M"
        return f"${v:,.0f}"

    def _trend(entry: Optional[dict]) -> str:
        if not entry:
            return ""
        return f"(FY {entry.get('end','')[:4]})"

    result = {"found": bool(revenue or net_inc)}
    if revenue:
        result["revenue"]          = _fmt_b(revenue)
        result["revenue_year"]     = revenue.get("end", "")[:4]
        result["revenue_raw"]      = revenue["val"]
    if net_inc:
        result["net_income"]       = _fmt_b(net_inc)
        result["net_income_year"]  = net_inc.get("end", "")[:4]
        result["net_income_raw"]   = net_inc["val"]
        result["profitable"]       = net_inc["val"] > 0
    if op_cf:
        result["op_cash_flow"]     = _fmt_b(op_cf)
        result["op_cf_year"]       = op_cf.get("end", "")[:4]
    if debt:
        result["long_term_debt"]   = _fmt_b(debt)
    if cash:
        result["cash"]             = _fmt_b(cash)
    # Net margin
    if revenue and net_inc and revenue["val"] > 0:
        result["net_margin_pct"]   = round(net_inc["val"] / revenue["val"] * 100, 1)
    # Debt-to-revenue ratio — crude liquidity signal
    if revenue and debt and revenue["val"] > 0:
        result["debt_to_rev"]      = round(debt["val"] / revenue["val"], 2)

    return result


def get_alpha_vantage_key() -> Optional[str]:
    try:
        key = st.secrets.get("ALPHAVANTAGE_API_KEY", "")
        return key.strip() if key else None
    except Exception:
        return None


@st.cache_data(show_spinner=False, ttl=3600)
def get_alpha_vantage_overview(ticker: str, api_key: str) -> Optional[dict]:
    if not ticker or not api_key:
        return None
    url = "https://www.alphavantage.co/query"
    params = {"function": "OVERVIEW", "symbol": ticker.upper().strip(), "apikey": api_key}
    response = requests.get(url, params=params, timeout=20)
    response.raise_for_status()
    data = response.json()
    if not data or "Symbol" not in data:
        return None
    return data


# =========================================================
# FREE PUBLIC API INTEGRATIONS
# =========================================================

# ── USASpending.gov — Federal contract awards benchmark ──
@st.cache_data(show_spinner=False, ttl=86400)
def fetch_usaspending_awards(keyword: str, limit: int = 10) -> list:
    """Return recent federal contract awards matching a keyword."""
    try:
        url  = "https://api.usaspending.gov/api/v2/search/spending_by_award/"
        body = {
            "filters": {
                "keywords": [keyword],
                "award_type_codes": ["A", "B", "C", "D"],
                "time_period": [{"start_date": "2022-01-01", "end_date": "2025-12-31"}],
            },
            "fields": ["Award ID", "Recipient Name", "Award Amount", "Description",
                       "Action Date", "Awarding Agency Name", "Period of Performance Current End Date"],
            "sort": "Award Amount",
            "order": "desc",
            "limit": limit,
            "page": 1,
        }
        resp = requests.post(url, json=body, timeout=15)
        resp.raise_for_status()
        return resp.json().get("results", [])
    except Exception:
        return []


@st.cache_data(show_spinner=False, ttl=86400)
def fetch_usaspending_summary(keyword: str) -> dict:
    """Return aggregate spend stats for a keyword category."""
    try:
        url  = "https://api.usaspending.gov/api/v2/search/spending_by_award_count/"
        body = {
            "filters": {
                "keywords": [keyword],
                "award_type_codes": ["A", "B", "C", "D"],
                "time_period": [{"start_date": "2022-01-01", "end_date": "2025-12-31"}],
            },
        }
        resp = requests.post(url, json=body, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception:
        return {}


# ── BLS Producer Price Index — commodity inflation data ──
@st.cache_data(show_spinner=False, ttl=86400)
def fetch_bls_ppi(series_ids: list) -> dict:
    """Fetch Producer Price Index series from BLS public API (no key required for 1 yr)."""
    try:
        url  = "https://api.bls.gov/publicAPI/v2/timeseries/data/"
        body = {"seriesid": series_ids, "startyear": "2023", "endyear": "2025"}
        resp = requests.post(url, json=body, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if data.get("status") != "REQUEST_SUCCEEDED":
            return {}
        out = {}
        for series in data.get("Results", {}).get("series", []):
            sid  = series["seriesID"]
            vals = series.get("data", [])[:6]  # last 6 periods
            out[sid] = [{"period": v["periodName"] + " " + v["year"], "value": float(v["value"])} for v in vals]
        return out
    except Exception:
        return {}


# BLS PPI series IDs mapped to procurement categories
BLS_PPI_MAP = {
    "Operations & MRO":        "PCU3312--3312--",   # Steel mill products
    "Direct Materials":        "WPU012",             # Crude materials
    "Logistics & Transportation": "PCU484121484121", # Truckload
    "Information Technology":  "PCU334614334614",   # Computer storage
    "Facilities & Real Estate":"PCU2361002361001",   # Construction
    "Professional Services":   "PCU541610541610",   # Management consulting
    "Finance & Accounting":    "PCU5221105221101",  # Commercial banking
    "Human Resources":         "PCU561310561310",   # Employment placement
    "Marketing & Communications": "PCU541810541810", # Advertising
    "Legal & Compliance":      "PCU541110541110",   # Legal services
}


# ── Open Corporates — global company registration lookup ──
@st.cache_data(show_spinner=False, ttl=86400)
def fetch_open_corporates(company_name: str) -> list:
    """Search Open Corporates for company registration data (free tier, no key required)."""
    try:
        url    = "https://api.opencorporates.com/v0.4/companies/search"
        params = {"q": company_name, "per_page": 5, "order": "score"}
        resp   = requests.get(url, params=params, timeout=12)
        resp.raise_for_status()
        return resp.json().get("results", {}).get("companies", [])
    except Exception:
        return []


# ── World Bank Commodity Prices ────────────────────────────
@st.cache_data(show_spinner=False, ttl=86400)
def fetch_worldbank_commodity(indicator: str, years: int = 3) -> list:
    """Fetch World Bank commodity price index data."""
    try:
        url    = f"https://api.worldbank.org/v2/en/indicator/{indicator}"
        params = {"format": "json", "per_page": years * 12, "mrv": years * 12}
        resp   = requests.get(url, params=params, timeout=12)
        resp.raise_for_status()
        raw = resp.json()
        if len(raw) < 2:
            return []
        return [{"date": r["date"], "value": r["value"]} for r in raw[1] if r.get("value") is not None][:12]
    except Exception:
        return []


# World Bank indicators by category
WB_COMMODITY_MAP = {
    "Energy (Electricity / Natural Gas)": "PNGAS.USD",
    "Fuel / Petroleum Products":          "POILWTI.USD",
    "Raw Materials (Metals / Resins / Chemicals)": "PIORECR.USD",
    "Lumber / Building Products":         "PLOGSK.USD",
    "Agricultural / Food Ingredients":    "PFOOD.USD",
    "Water / Industrial Gases":           "PCOPPUSDM.USD",
}


# ── USASpending category keyword map ──────────────────────
USASPENDING_KEYWORD_MAP = {
    "Heat Exchangers & Pressure Vessels":          "heat exchanger pressure vessel",
    "Rotating Equipment (Pumps / Compressors / Turbines)": "pump compressor turbine",
    "Instrumentation & Controls (I&C)":            "instrumentation control systems",
    "Industrial Valves & Actuators":               "industrial valves actuators",
    "Field Engineering & Turnaround Services":     "turnaround maintenance services",
    "Inspection, Testing & NDT Services":          "nondestructive testing inspection",
    "MRO Distribution":                            "MRO industrial supply",
    "Cybersecurity (EDR / SIEM / SOC)":           "cybersecurity managed security",
    "Cloud Infrastructure (AWS / Azure / GCP)":   "cloud computing infrastructure",
    "Management Consulting":                       "management consulting advisory",
    "IT Consulting & Systems Integration":         "systems integration IT services",
    "Construction / Capital Projects":             "construction capital project",
    "Truckload (TL) / Full Truckload":            "truckload transportation freight",
    "International / Ocean Freight":              "ocean freight shipping logistics",
    "Raw Materials (Metals / Resins / Chemicals)": "industrial chemicals raw materials",
    "Contract Manufacturing / OEM":               "contract manufacturing OEM",
    "Outside Counsel (Law Firm)":                 "legal services law firm",
    "Audit Services (External)":                  "audit accounting assurance",
}


# =========================================================
# ASYNC DATA FETCHING — concurrent SEC + yfinance lookups
# =========================================================

_THREAD_POOL = concurrent.futures.ThreadPoolExecutor(max_workers=6)


def _fetch_sec_sync(ticker: str) -> Optional[dict]:
    """Synchronous SEC fetch — runs in thread pool."""
    try:
        return get_sec_company_context(ticker)
    except Exception as e:
        return {"found": False, "message": str(e)}


def _fetch_yfinance_sync(ticker: str) -> Optional[dict]:
    """Synchronous yfinance fetch — runs in thread pool."""
    if not _YFINANCE_AVAILABLE or not ticker:
        return None
    try:
        assert yf is not None
        t = yf.Ticker(ticker)
        info = t.info
        # Derive first-trade year for years-in-business estimate
        _first_ms = info.get("firstTradeDateMilliseconds") or info.get("firstTradeDateEpoch", 0)
        _first_year = None
        if _first_ms:
            import datetime
            try:
                _first_year = datetime.datetime.fromtimestamp(_first_ms / 1000).year
            except Exception:
                pass
        return {
            "_v": 2,  # version marker — used to detect stale cache
            "market_cap_fmt": _fmt_market_cap(info.get("marketCap")),
            "revenue_growth": _fmt_pct(info.get("revenueGrowth")),
            "gross_margin": _fmt_pct(info.get("grossMargins")),
            "employees": f"{info.get('fullTimeEmployees', 0):,}" if info.get("fullTimeEmployees") else "N/A",
            "52w_change": _fmt_pct(info.get("52WeekChange")),
            "analyst_rating": (info.get("recommendationKey") or "N/A").replace("_", " ").title(),
            "short_name": info.get("shortName", ticker),
            # Raw fields for financial health auto-fill
            "revenue_growth_raw": info.get("revenueGrowth"),
            "market_cap_raw": info.get("marketCap"),
            "employees_raw": info.get("fullTimeEmployees"),
            "quote_type": info.get("quoteType", ""),
            "first_trade_year": _first_year,
        }
    except Exception:
        return None


def suggest_fin_fields(yf_data: dict, news_signals: list, sec_ctx: Optional[dict] = None) -> dict:
    """Map yfinance + news + SEC data to FINANCIAL_FIELDS option values for auto-fill."""
    import datetime, re
    suggestions = {}
    sec_ctx = sec_ctx or {}

    # ── Ownership Structure ──────────────────────────────────────────────────
    # Most reliable signal: SEC found = company is publicly registered.
    # Fallback: yfinance market cap present.
    if sec_ctx.get("found") or bool(yf_data.get("market_cap_raw")) or (
        yf_data.get("market_cap_fmt") not in (None, "N/A", "")
    ):
        suggestions["Ownership Structure"] = "Publicly traded"

    # ── Revenue Trajectory ───────────────────────────────────────────────────
    # Prefer raw float (v2); fall back to parsing the formatted string (old cache).
    rev = yf_data.get("revenue_growth_raw")
    if rev is None:
        rev_str = yf_data.get("revenue_growth", "")
        if rev_str and rev_str not in ("N/A", ""):
            try:
                rev = float(re.sub(r"[^0-9.\-]", "", rev_str.lstrip("+").replace(",", ""))) / 100
            except (ValueError, TypeError):
                pass
    if rev is not None:
        if rev >= 0.15:
            suggestions["Revenue Trajectory"] = "Growing 15%+"
        elif rev >= 0.05:
            suggestions["Revenue Trajectory"] = "Growing 5–15%"
        elif rev >= -0.02:
            suggestions["Revenue Trajectory"] = "Flat"
        else:
            suggestions["Revenue Trajectory"] = "Declining"

    # ── Recent M&A Activity ──────────────────────────────────────────────────
    ma_sigs = [s for s in news_signals if s.get("type") == "M&A"]
    if ma_sigs:
        t = ma_sigs[0].get("text", "").lower()
        if any(k in t for k in ("acqui", "purchase", "buyout")):
            suggestions["Recent M&A Activity"] = "Acquired a company"
        elif any(k in t for k in ("divest", "spinoff", "spin-off", "spun")):
            suggestions["Recent M&A Activity"] = "Recently spun off"
        else:
            suggestions["Recent M&A Activity"] = "Being acquired"
    else:
        suggestions["Recent M&A Activity"] = "None in 2 years"

    # ── Workforce Changes ────────────────────────────────────────────────────
    risk_texts = " ".join(s.get("text", "").lower() for s in news_signals if s.get("risk"))
    pos_texts  = " ".join(s.get("text", "").lower() for s in news_signals if not s.get("risk"))
    if any(k in risk_texts for k in ("layoff", "job cut", "restructur", "reduc")):
        suggestions["Workforce Changes (12mo)"] = "Minor layoffs <5%"
    elif any(k in pos_texts for k in ("hiring", "expand", "headcount", "workforce grow")):
        suggestions["Workforce Changes (12mo)"] = "Significant hiring"
    else:
        suggestions["Workforce Changes (12mo)"] = "Stable"

    # ── Years in Business ────────────────────────────────────────────────────
    # Primary: yfinance first-trade year (v2). Fallback: oldest SEC filing year.
    anchor_year = yf_data.get("first_trade_year") or sec_ctx.get("oldest_filing_year")
    if anchor_year:
        years = datetime.date.today().year - anchor_year
        if years >= 25:
            suggestions["Years in Business"] = "25+ years"
        elif years >= 10:
            suggestions["Years in Business"] = "10–25 years"
        elif years >= 3:
            suggestions["Years in Business"] = "3–10 years"
        else:
            suggestions["Years in Business"] = "<3 years"

    return suggestions




# ── Ticker fuzzy match ────────────────────────────────────────────────────────
# Common company name → ticker mapping for the most-entered names
_TICKER_HINTS: Dict[str, str] = {
    # Tech
    "microsoft": "MSFT", "apple": "AAPL", "google": "GOOGL", "alphabet": "GOOGL",
    "amazon": "AMZN", "aws": "AMZN", "meta": "META", "facebook": "META",
    "salesforce": "CRM", "oracle": "ORCL", "sap": "SAP", "workday": "WDAY",
    "servicenow": "NOW", "crowdstrike": "CRWD", "palo alto": "PANW",
    "okta": "OKTA", "snowflake": "SNOW", "databricks": "DBX",
    # Logistics
    "ups": "UPS", "fedex": "FDX", "jb hunt": "JBHT", "j.b. hunt": "JBHT",
    "ch robinson": "CHRW", "c.h. robinson": "CHRW", "old dominion": "ODFL",
    "xpo": "XPO", "knight swift": "KNX", "werner": "WERN",
    # Finance / Insurance
    "jpmorgan": "JPM", "jp morgan": "JPM", "bank of america": "BAC",
    "bofa": "BAC", "citibank": "C", "citi": "C", "wells fargo": "WFC",
    "amex": "AXP", "american express": "AXP", "visa": "V", "mastercard": "MA",
    "marsh": "MMC", "aon": "AON", "wtw": "WTW",
    # Professional Services
    "accenture": "ACN", "infosys": "INFY", "cognizant": "CTSH",
    "korn ferry": "KFY", "heidrick": "HSII", "robert half": "RHI",
    "gartner": "IT", "forrester": "FORR",
    # Facilities / Industrial
    "grainger": "GWW", "fastenal": "FAST", "msc industrial": "MSM",
    "waste management": "WM", "republic services": "RSG",
    "johnson controls": "JCI", "carrier": "CARR", "trane": "TT",
    "constellation energy": "CEG", "nrg": "NRG",
    # Direct Materials
    "nucor": "NUE", "dow": "DOW", "lyondellbasell": "LYB", "basf": "BASFY",
    "tsmc": "TSM", "texas instruments": "TXN", "arrow": "ARW", "avnet": "AVT",
    "international paper": "IP", "sealed air": "SEE", "amcor": "AMCR",
    "flex": "FLEX", "jabil": "JBL", "linde": "LIN", "air products": "APD",
}


def suggest_ticker(company_name: str) -> Optional[str]:
    """
    Given a free-text company name, return a suggested ticker.
    Checks _TICKER_HINTS first, then tries TF-IDF fuzzy match on hint keys.
    Returns None if no confident match.
    """
    if not company_name or not company_name.strip():
        return None
    normalized = company_name.lower().strip()

    # Exact hint match
    if normalized in _TICKER_HINTS:
        return _TICKER_HINTS[normalized]

    # Partial match — if the entered name contains a known company name
    for known, ticker in _TICKER_HINTS.items():
        if known in normalized or normalized in known:
            return ticker

    # TF-IDF fuzzy on hint keys
    if _SKLEARN_AVAILABLE and TfidfVectorizer is not None:
        try:
            keys = list(_TICKER_HINTS.keys())
            corpus = keys + [normalized]
            vec = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4))
            tfidf = vec.fit_transform(corpus)
            query_vec = tfidf.getrow(tfidf.shape[0] - 1)
            candidates = cast(Any, tfidf.toarray()[: tfidf.shape[0] - 1])
            assert cosine_similarity is not None
            scores = cosine_similarity(query_vec, candidates).flatten()
            best_idx = int(np.argmax(scores))
            if float(scores[best_idx]) >= 0.40:
                return _TICKER_HINTS[keys[best_idx]]
        except Exception:
            pass
    return None


# ── M&A / news sentiment layer ────────────────────────────────────────────────
@st.cache_data(show_spinner=False, ttl=3600)
def get_supplier_news_signals(ticker: str) -> Dict:
    """
    Pull recent news from yfinance and classify into signals:
    - M&A activity (acquisitions, mergers, divestitures)
    - Leadership changes (CEO/CFO/executive departures/appointments)
    - Financial distress signals (layoffs, restructuring, debt concerns)
    - Positive signals (new contracts, partnerships, growth)
    Returns dict with classified signals list and raw headlines.
    """
    if not _YFINANCE_AVAILABLE or not ticker or ticker in ("Private", "", "—"):
        return {"signals": [], "headlines": [], "has_risks": False}
    try:
        assert yf is not None
        t = yf.Ticker(ticker)
        news_items = t.news or []
        news_items = news_items[:8]  # cap at 8 most recent
    except Exception:
        return {"signals": [], "headlines": [], "has_risks": False}

    if not news_items:
        return {"signals": [], "headlines": [], "has_risks": False}

    # Keyword classifiers — no LLM needed for this layer
    MA_KEYWORDS = ["acqui", "merger", "takeover", "divest", "spinoff", "spin-off",
                   "buyout", "purchase", "deal closed", "transaction"]
    LEADERSHIP_KEYWORDS = ["ceo", "cfo", "coo", "chief", "president", "depart",
                           "resign", "appoint", "named", "steps down", "leaves"]
    RISK_KEYWORDS = ["layoff", "restructur", "bankrupt", "debt", "loss", "miss",
                     "warning", "downgrad", "investigat", "lawsuit", "breach"]
    POSITIVE_KEYWORDS = ["contract", "partner", "expand", "growth", "record",
                         "upgrade", "win", "award", "launch", "revenue beat"]

    signals = []
    headlines = []
    has_risks = False

    for item in news_items:
        title = (item.get("title") or "").lower()
        raw_title = item.get("title") or ""
        if not title:
            continue
        headlines.append(raw_title)

        if any(kw in title for kw in MA_KEYWORDS):
            signals.append({"type": "M&A", "icon": "🔀", "color": "#A78BFA",
                             "text": raw_title, "risk": True})
            has_risks = True
        elif any(kw in title for kw in LEADERSHIP_KEYWORDS):
            signals.append({"type": "Leadership", "icon": "👤", "color": "#FCD34D",
                             "text": raw_title, "risk": True})
            has_risks = True
        elif any(kw in title for kw in RISK_KEYWORDS):
            signals.append({"type": "Risk Signal", "icon": "⚠", "color": "#F87171",
                             "text": raw_title, "risk": True})
            has_risks = True
        elif any(kw in title for kw in POSITIVE_KEYWORDS):
            signals.append({"type": "Positive", "icon": "✓", "color": "#4ADE80",
                             "text": raw_title, "risk": False})

    return {
        "signals": signals[:5],  # show top 5 classified signals
        "headlines": headlines,
        "has_risks": has_risks,
        "item_count": len(news_items),
    }



def fetch_supplier_intelligence_concurrent(tickers: List[str], alpha_key: Optional[str]) -> Dict[str, dict]:
    """
    Fetch SEC + yfinance data for all supplier tickers concurrently using a thread pool.
    Returns dict keyed by ticker with 'sec' and 'yf' sub-keys.
    Much faster than sequential fetching when evaluating 3-4 suppliers.
    """
    results: Dict[str, dict] = {}
    if not tickers:
        return results

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(tickers) * 2, 8)) as pool:
        sec_futures  = {ticker: pool.submit(_fetch_sec_sync, ticker) for ticker in tickers if ticker}
        yf_futures   = {ticker: pool.submit(_fetch_yfinance_sync, ticker) for ticker in tickers if ticker}

        for ticker in tickers:
            if not ticker:
                continue
            sec_result = None
            yf_result  = None
            try:
                sec_result = sec_futures[ticker].result(timeout=12)
            except Exception:
                sec_result = {"found": False, "message": "Timeout or network error"}
            try:
                yf_result = yf_futures[ticker].result(timeout=12)
            except Exception:
                yf_result = None
            results[ticker] = {"sec": sec_result, "yf": yf_result}

    return results


# =========================================================
# HELPERS
# =========================================================
def normalize_weights(weight_dict: Mapping[str, Optional[int]]) -> Dict[str, float]:
    clean_weights = {k: v for k, v in weight_dict.items() if v is not None}
    if not clean_weights:
        return {}
    total = sum(clean_weights.values())
    if total == 0:
        return {k: 1 / len(clean_weights) for k in clean_weights}
    return {k: v / total for k, v in clean_weights.items()}


def score_price(price: float, all_prices: List[float]) -> int:
    """
    Sigmoid-based price scoring. Replaces linear scale which breaks with outliers.
    A supplier 2x the mean is penalized hard; 10x doesn't collapse model resolution.
    Formula: Score = 100 * (1 / (1 + e^(k * (price - mu) / mu)))
    k=4 gives a steep-but-not-cliff curve. Mean-centered so group context matters.
    """
    if len(all_prices) <= 1:
        return 80
    valid = [p for p in all_prices if p and p > 0]
    if not valid:
        return 80
    mu = sum(valid) / len(valid)
    if math.isclose(mu, 0):
        return 80
    k = 4.0  # sensitivity constant — higher = steeper penalty curve
    try:
        raw = 1.0 / (1.0 + math.exp(k * (price - mu) / mu))
    except OverflowError:
        raw = 0.0 if price > mu else 1.0
    # Scale to 0-100, ensure integer
    return max(0, min(100, round(raw * 100)))


def score_sla(v: str) -> int:
    return {"Strong": 92, "Moderate": 60, "Weak": 28}.get(v, 50)


def score_risk(v: str) -> int:
    return {"Low": 92, "Medium": 58, "High": 24}.get(v, 50)


def score_num_1_to_5(v: int) -> int:
    return round((v / 5) * 100)


# compute_financial_health() and financial_risk_label() live in evaluation.py
# and are imported at the top of this file.


def compute_supplier_scores(supplier: Dict, all_prices: List[float], fin_score: int) -> Dict[str, int]:
    base_exec = score_risk(supplier["Execution Risk"])
    fin_adjustment = round(((fin_score - 50) / 50) * 20)
    exec_risk_adjusted = max(0, min(100, base_exec + fin_adjustment))
    return {
        "Price / TCO": score_price(float(supplier["Raw Price"]), all_prices),
        "SLA Strength": score_sla(supplier["SLA Strength"]),
        "Execution Risk": exec_risk_adjusted,
        "Stakeholder Confidence": score_num_1_to_5(supplier["Stakeholder Confidence"]),
        "Strategic Alignment": score_num_1_to_5(supplier["Strategic Alignment"]),
        "Innovation Capacity": score_num_1_to_5(supplier["Innovation Capacity"]),
        "Relationship Depth": score_num_1_to_5(supplier["Relationship Depth"]),
        "Commercial Flexibility": score_num_1_to_5(supplier["Commercial Flexibility"]),
        "ESG / Sustainability": score_sla(supplier.get("ESG / Sustainability", "Moderate")),
        "Supplier Diversity": score_sla(supplier.get("Supplier Diversity", "Moderate")),
    }


def weighted_score(scores: Dict[str, int], weights: Dict[str, float]) -> float:
    return round(sum(scores[d] * weights[d] for d in DIMENSIONS), 1)


def current_fit(scores: Dict[str, int]) -> float:
    return round(sum(scores[d] for d in CURRENT_DIMS) / len(CURRENT_DIMS), 1)


def future_fit(scores: Dict[str, int]) -> float:
    return round(sum(scores[d] for d in FUTURE_DIMS) / len(FUTURE_DIMS), 1)


def classify_category(category_text: str) -> Dict:
    """
    Multi-keyword category classifier. Checks all alias lists before falling back
    to the general default. Order matters: more specific aliases first.
    """
    c = (category_text or "").lower()

    # ── alias map: each key in CATEGORY_RULES maps to a list of trigger phrases ──
    ALIASES: Dict[str, List[str]] = {
        "technology": [
            "technology", "tech", "saas", "software", "it ", "cloud",
            "erp", "crm", "hris", "cybersecurity", "infrastructure",
            "platform", "application", "digital", "data", "analytics",
            "telecom", "network", "hardware", "license",
        ],
        "hr": [
            "hr ", "human resource", "people", "talent", "recruiting",
            "payroll", "benefits", "learning", "training", "workforce",
            "hcm", "employee", "staffing", "temp labor", "contingent",
            "background check", "wellness", "relocation",
        ],
        "finance": [
            "finance", "financial", "accounting", "audit", "treasury",
            "tax", "insurance", "banking", "payment", "invoice",
            "expense", "procurement card", "p-card",
        ],
        "marketing": [
            "marketing", "media", "advertising", "creative", "brand",
            "agency", "pr ", "events", "sponsorship", "digital marketing",
            "content", "social media", "print",
        ],
        "services": [
            "consulting", "professional service", "advisory", "outsourcing",
            "bpo", "managed service", "legal", "counsel", "facility management",
            "security service", "janitorial", "cleaning", "food service",
        ],
        "packaging": [
            "packaging", "carton", "label", "container", "bag", "box",
            "flexible packaging", "corrugated", "bottle", "cap", "closure",
        ],
        "manufacturing": [
            "manufacturing", "raw material", "component", "subassembly",
            "mro", "maintenance", "repair", "operations supply",
            "tooling", "castings", "forgings", "chemical", "resin",
            "metal", "steel", "aluminum", "commodity",
        ],
        "logistics": [
            "logistics", "transportation", "freight", "carrier", "shipping",
            "3pl", "warehouse", "distribution", "fleet", "trucking",
            "courier", "parcel", "last mile", "customs", "brokerage",
        ],
    }

    for rule_key, aliases in ALIASES.items():
        for alias in aliases:
            if alias in c:
                return CATEGORY_RULES[rule_key]

    # Facilities / real estate catch-all → services rule
    if any(kw in c for kw in ["facilities", "real estate", "office", "lease", "utilities", "energy"]):
        return CATEGORY_RULES["services"]

    return {
        "type": "Mixed / Unclassified",
        "tag": "General Procurement",
        "requirements": "Define scope, deliverables, owners, milestones, and acceptance logic clearly.",
        "assurance": "Document continuity, interruption handling, transition support, and escalation paths.",
        "quality": "Set measurable performance standards and issue-handling expectations.",
        "service": "Clarify support, communication cadence, and governance structure.",
        "cost": "Make pricing structure, caps, invoicing, and pass-through logic visible.",
        "innovation": "Address improvement and future-state support where relevant.",
        "rfp_stakeholders": DEFAULT_RFP_STAKEHOLDERS,
    }


def type_badge(category_type: str) -> str:
    if category_type == "Direct":
        return '<span class="badge-direct">Direct</span>'
    if category_type == "Indirect":
        return '<span class="badge-indirect">Indirect</span>'
    return '<span class="badge-mixed">Mixed</span>'


def stakeholder_action(power: int, interest: int, position: str, priority: str) -> str:
    if power >= 8 and position in ["Skeptic", "Blocker"]:
        return f"High-risk: prepare a targeted defense tied to {priority.lower()} before the meeting."
    if power >= 7 and interest >= 7:
        return f"Manage closely: secure visible support and align early using a {priority.lower()} narrative."
    if power >= 7 and interest < 7:
        return f"Keep satisfied: bring concise business-impact updates framed around {priority.lower()}."
    if power < 7 and interest >= 7:
        return f"Keep informed: use as an advocate and pressure-test the recommendation through a {priority.lower()} lens."
    return "Monitor lightly: maintain visibility but do not overinvest unless their stance changes."


def talk_track(role: str, position: str, priority: str, leader_name: str,
               category: str = "", subcategory: str = "", kraljic: str = "") -> str:
    """
    Returns an HTML bullet list of 3–5 position-specific, category-aware talk track points.
    Rendered inside unsafe_allow_html=True markdown so HTML tags are intentional.
    """
    cat = category.lower()
    sub = subcategory.lower()
    prio = priority.lower()

    # Category-specific proof dimensions
    _proof = {
        "it": ["security certification (ISO 27001 / SOC 2 Type II)", "uptime SLA with financial credits", "implementation case studies in your industry"],
        "information technology": ["security certification (ISO 27001 / SOC 2 Type II)", "uptime SLA with financial credits", "implementation case studies in your industry"],
        "hr": ["GDPR/CCPA compliance posture", "employee NPS from reference clients", "depth of integration with your existing HRIS stack"],
        "human resources": ["GDPR/CCPA compliance posture", "employee NPS from reference clients", "depth of integration with your existing HRIS stack"],
        "finance": ["SOX compliance controls and audit trail", "accuracy and reconciliation track record", "CFO reference from a comparable company"],
        "marketing": ["ROAS benchmarks on comparable campaigns", "brand safety controls and content exclusion policies", "speed-to-launch history"],
        "logistics": ["on-time delivery rate (target >98%)", "carrier backup capacity and alternate routing", "claims resolution cycle time"],
        "logistics & transportation": ["on-time delivery rate (target >98%)", "carrier backup capacity and alternate routing", "claims resolution cycle time"],
        "operations / mro": ["parts availability SLA and emergency response time", "certified technician ratios", "planned vs. reactive work order ratio"],
        "facilities": ["OSHA/EPA compliance record", "energy efficiency and sustainability certifications", "tenant satisfaction scores from comparable properties"],
        "legal": ["matter cost predictability vs. budget", "outside counsel rate benchmarks", "litigation win/settle ratio"],
        "professional services": ["delivery-on-time track record for comparable engagements", "knowledge-transfer and documentation quality", "consultant retention and continuity commitments"],
    }
    proof_points = next((v for k, v in _proof.items() if k in cat or k in sub),
                        ["weighted evaluation score vs. alternatives", "reference check outcomes", "contractual performance commitments"])

    # Category-specific risk phrase
    _risk_phrase = {
        "it": "a failed implementation, security breach, or integration failure",
        "information technology": "a failed implementation, security breach, or integration failure",
        "hr": "a data breach, compliance gap, or failed system migration",
        "human resources": "a data breach, compliance gap, or failed system migration",
        "logistics": "a supply-chain disruption or missed delivery window",
        "logistics & transportation": "a supply-chain disruption or missed delivery window",
        "finance": "an audit finding or reconciliation failure",
        "operations / mro": "an unplanned production shutdown",
        "facilities": "a safety incident or regulatory citation",
    }
    risk_phrase = next((v for k, v in _risk_phrase.items() if k in cat or k in sub),
                       "a failure to deliver on the core business need")

    # Escape user-controlled values before any HTML interpolation.
    # cat/sub/prio (lowercase derivatives) are used only for dict key lookups above —
    # they are NOT inserted into HTML, so they stay unescaped for matching.
    _e_leader   = html.escape(leader_name)
    _e_priority = html.escape(priority)
    _e_prio     = html.escape(prio)
    _e_category = html.escape(category)

    # Kraljic framing sentence (uses escaped leader_name)
    _kq_map = {
        "strategic": "This is a Strategic category — the question is not who is cheapest today, it is who can grow with us and won't become a single point of failure over the next 3–5 years.",
        "leverage": f"We have market leverage here. {_e_leader} knows we can switch — that should be visible in pricing, flexibility, and SLA commitments.",
        "bottleneck": f"Supply continuity is the primary risk in this category. {_e_leader} offers the most credible assurance against disruption and long lead-time exposure.",
        "non-critical": "This is a routine operational category. The goal is speed and simplicity, not the perfect vendor. Avoid over-engineering the decision.",
    }
    kq_sentence = _kq_map.get(kraljic.lower(), "")

    def _bullets(items):
        lis = "".join(f'<li style="margin-bottom:0.5rem;line-height:1.55">{b}</li>' for b in items)
        return f'<ul style="margin:0.35rem 0 0 0;padding-left:1.3rem">{lis}</ul>'

    pos = position

    if pos in ("Skeptic", "Blocker"):
        points = [
            f"<strong>Go private first.</strong> Before the group meeting, request a 1:1. Frame it as: <em>\"I want your concerns in the recommendation, not overridden by it.\"</em> Skeptics who feel heard convert; ones who feel steamrolled escalate publicly.",
            f"<strong>Anchor on their stated priority — {_e_priority}.</strong> Bring one specific, hard data point: {proof_points[0]} and {proof_points[1]}. Don't walk in with a deck — walk in with a single fact they can't easily dismiss.",
            f"<strong>Reframe the risk.</strong> The risk is not choosing {_e_leader} — the risk is {risk_phrase} caused by delay or by selecting a vendor without proven {_e_prio} controls. Make inaction visible.",
        ]
        if kq_sentence:
            points.append(f"<strong>Strategic frame:</strong> {kq_sentence}")
        points.append(
            f"<strong>Get a conditional commitment.</strong> End with a direct ask: <em>\"If the {_e_prio} data holds up, would you move from Skeptic to Neutral?\"</em> A conditional yes is far more valuable than a vague 'I'll consider it.'"
        )
        return _bullets(points)

    if pos == "Champion":
        points = [
            f"<strong>Extract their killer fact.</strong> Ask: <em>\"What is the single strongest evidence you've seen for {_e_leader} on {_e_prio}?\"</em> Use their answer — not yours — as the anchor in the group presentation. It's harder to attack.",
            f"<strong>Rehearse the hardest objection together.</strong> The most common in {_e_category or 'this category'}: cost, switching risk, or lack of internal references. Prepare a 30-second response and practice it with them before the room.",
            f"<strong>Assign a specific moment.</strong> Don't just ask them to 'be supportive.' Give them a cue: <em>\"When Finance raises cost, I'd like you to speak to {proof_points[0]} from your experience.\"</em> A targeted endorsement lands harder than a general one.",
        ]
        if kq_sentence:
            points.append(f"<strong>Frame for the room:</strong> {kq_sentence}")
        points.append(
            f"<strong>Keep them informed of changes.</strong> If the recommendation or terms shift before the meeting, brief them immediately. A Champion who is surprised in the room goes quiet at exactly the wrong moment."
        )
        return _bullets(points)

    if pos == "Supporter":
        points = [
            f"<strong>Leverage their operational credibility.</strong> They may not be the highest-power person in the room, but they represent ground-level reality — and skeptics often trust that more than executive endorsement.",
            f"<strong>Brief them fully before the meeting.</strong> Supporters who walk in without context inadvertently introduce doubt by hedging. Give them the final recommendation and the two or three facts that support it.",
            f"<strong>Give them something specific to confirm.</strong> Ask them to validate one concrete point: {proof_points[2] if len(proof_points) > 2 else proof_points[0]}. A focused, firsthand endorsement is worth more than broad support.",
            f"<strong>Don't over-rely on them.</strong> If {_e_priority} is the main battleground and your Champion or Supporter is low-power, ensure you've addressed the objection with data — not just with social proof.",
        ]
        return _bullets(points)

    # Neutral / Informed
    points = [
        f"<strong>Don't try to convert them — prevent drift.</strong> Neutrals who feel included stay neutral; ones who feel bypassed drift toward the loudest voice in the room, which is usually the Skeptic.",
        f"<strong>Send a pre-read.</strong> Share a one-page summary of the evaluation outcome 48 hours before the meeting. Neutrals who arrive informed are less likely to ask destabilizing questions in real time.",
        f"<strong>Create a micro-ownership moment.</strong> If they have subject-matter knowledge in {_e_category or 'this area'}, ask them to validate one specific finding — {proof_points[0]}. Ownership, even minor, builds alignment.",
        f"<strong>Watch for late movement.</strong> A Neutral with high Power who shifts to Skeptic in the final 24 hours is your highest risk scenario. Keep the communication line open all the way to the meeting.",
    ]
    return _bullets(points)


def likely_blocker(stake_df: pd.DataFrame):
    candidates = stake_df[
        ((stake_df["Position"] == "Blocker") | (stake_df["Position"] == "Skeptic"))
        & (stake_df["Power"] >= 7)
    ].copy()
    if candidates.empty:
        return None
    candidates["Rank"] = candidates["Power"] * 2 + candidates["Interest"]
    candidates = candidates.sort_values(by="Rank", ascending=False)
    return candidates.iloc[0]


def make_recommendation_text(leader: Dict, runner_up: Optional[Dict], weakest_dim: str, kraljic: str) -> str:
    if runner_up is not None:
        gap = round(leader["Weighted Score"] - runner_up["Weighted Score"], 1)
        gap_text = f"Lead vs runner-up: {gap} pts."
    else:
        gap_text = "Only one supplier is fully comparable."
    return (
        f"Pick: {leader['Supplier']}\n\n"
        f"Why it wins:\n"
        f"- Highest weighted score at {leader['Weighted Score']} / 100\n"
        f"- Best balance of current execution and future fit\n"
        f"- Strongest defense under a {kraljic} sourcing posture\n\n"
        f"What to watch:\n"
        f"- Weakest dimension is {weakest_dim}\n"
        f"- This must become a negotiation and mitigation topic before award\n\n"
        f"{gap_text}"
    )


def make_tradeoff_text(leader: Dict, runner_up: Optional[Dict]) -> str:
    if not runner_up:
        return "Not enough suppliers were entered to compare trade-offs."
    leader_price = leader["Raw Price"]
    runner_price = runner_up["Raw Price"]
    pct = 0 if runner_price == 0 else round(((leader_price - runner_price) / runner_price) * 100, 1)
    risk_delta = round(leader["Scores"]["Execution Risk"] - runner_up["Scores"]["Execution Risk"], 1)
    stake_delta = round(leader["Scores"]["Stakeholder Confidence"] - runner_up["Scores"]["Stakeholder Confidence"], 1)
    future_delta = round(leader["Future Fit"] - runner_up["Future Fit"], 1)
    direction = "more" if pct >= 0 else "less"
    return (
        f"{leader['Supplier']} is {abs(pct)}% {direction} expensive than {runner_up['Supplier']}, "
        f"but improves execution-risk by {risk_delta} pts, stakeholder-confidence by {stake_delta} pts, "
        f"and future-fit by {future_delta} pts."
    )


def block_risk_text(blocker_row, leader_name: str) -> str:
    if blocker_row is None:
        return f"No high-power blocker visible yet. {leader_name} looks defendable if the room stays on business criteria."
    return (
        f"Most likely block: {blocker_row['Name']} ({blocker_row['Role']}). "
        f"They are a {blocker_row['Position']} with high power and care most about {blocker_row['Priority']}. "
        f"Expect the room to challenge the decision there first."
    )


def alt_supplier_text(leader: Dict, runner_up: Optional[Dict]) -> str:
    if not runner_up:
        return "No runner-up comparison available."
    return (
        f"If you choose {runner_up['Supplier']} instead of {leader['Supplier']}, you may gain on one variable "
        f"such as price or flexibility, but weaken the overall defense narrative. "
        f"The room will need a stronger explanation for why the compromise is worth it."
    )


def default_negotiation_points(kraljic: str, category_rule: Dict, weakest_dim: str) -> List[str]:
    if kraljic == "Strategic":
        points = [
            "Protect governance, resilience, and executive escalation before chasing cosmetic commercial wins.",
            f"Use negotiation to directly close the weakest dimension: {weakest_dim}.",
            "Push for roadmap visibility and future capability commitments.",
        ]
    elif kraljic == "Leverage":
        points = [
            "Use competition to improve pricing, rebate, and benchmark language.",
            "Do not leave annual increase logic vague.",
            "Strengthen service and reporting while leverage is on your side.",
        ]
    elif kraljic == "Bottleneck":
        points = [
            "Do not sacrifice continuity protections to save cost.",
            "Prioritize interruption handling, notice periods, and escalation rights.",
            "Use negotiation to reduce operational fragility.",
        ]
    else:
        points = [
            "Keep terms simple and easy to administer.",
            "Avoid unnecessary customization.",
            "Focus on clear pricing and easy renewal / exit control.",
        ]
    points.append(
        f"Category lens: this is {category_rule['tag']}, so negotiation should reflect category-specific protections, not only generic sourcing logic."
    )
    return points


def category_raqsci(kraljic: str, category_rule: Dict) -> Dict:
    is_direct = category_rule["type"] == "Direct"
    return {
        "Requirements": {
            "must": category_rule["requirements"],
            "recommended": (
                "For direct categories, lock specifications, tolerances, and engineering change ownership tightly."
                if is_direct
                else f"Use the {kraljic.lower()} posture to tie requirements tightly to accountability and acceptance logic."
            ),
            "nice": "Add review checkpoints for evolving business needs.",
        },
        "Assurance of Supply": {
            "must": category_rule["assurance"],
            "recommended": (
                "For direct categories, define shortage, interruption, and recovery expectations in operational terms."
                if is_direct
                else f"In a {kraljic.lower()} category, define interruption handling and escalation in business terms."
            ),
            "nice": "Request recurring continuity reviews.",
        },
        "Quality": {
            "must": category_rule["quality"],
            "recommended": (
                "For direct categories, use measurable incoming-quality or defect-management reviews."
                if is_direct
                else "Use recurring scorecards or measurable review checkpoints."
            ),
            "nice": "Include improvement expectations over the term.",
        },
        "Service": {
            "must": category_rule["service"],
            "recommended": (
                "For direct categories, make shortage communication and delivery escalation explicit."
                if is_direct
                else "Make governance and escalation visible in the contract itself."
            ),
            "nice": "Request named contacts or support structure where useful.",
        },
        "Cost": {
            "must": category_rule["cost"],
            "recommended": (
                "For direct categories, tie commercial terms to indexation and pass-through discipline."
                if is_direct
                else f"Align cost language to the {kraljic.lower()} strategy instead of treating price as a standalone line item."
            ),
            "nice": "Add review or benchmark rights where practical.",
        },
        "Innovation": {
            "must": category_rule["innovation"],
            "recommended": (
                "For direct categories, connect innovation to redesign, cost-down, and supply resilience."
                if is_direct
                else "Tie future-state value to visible review or roadmap commitments."
            ),
            "nice": "Include annual improvement sessions.",
        },
    }


def generate_rfp_risk_flags(leader: Dict, runner_up: Optional[Dict], blocker_row, kraljic: str, category_rule: Dict) -> List[Dict]:
    flags = []
    weakest = min(DIMENSIONS, key=lambda d: leader["Scores"][d])
    weakest_score = leader["Scores"][weakest]

    # High risk flags
    if weakest_score < 35:
        flags.append({
            "tier": "HIGH",
            "icon": "🔴",
            "title": f"Critical gap in {weakest}",
            "body": f"{leader['Supplier']} scores only {weakest_score}/100 on {weakest}. This is not a negotiation nice-to-have — it is a structural risk that must be closed before award or it becomes a contract performance problem.",
        })

    if blocker_row is not None:
        flags.append({
            "tier": "HIGH",
            "icon": "🔴",
            "title": f"High-power blocker: {blocker_row['Name']}",
            "body": f"{blocker_row['Name']} ({blocker_row['Role']}) is a {blocker_row['Position']} who cares about {blocker_row['Priority']}. If not engaged before the presentation, they will challenge the recommendation in the room. Engage them one-on-one first.",
        })

    if leader["Financial Risk Label"] == "HIGH":
        flags.append({
            "tier": "HIGH",
            "icon": "🔴",
            "title": f"Supplier financial health risk",
            "body": f"{leader['Supplier']} has a financial health score that signals elevated risk. Consider requesting audited financials, bank references, or performance bonds before award.",
        })

    # Medium risk flags
    if leader["Scores"]["Stakeholder Confidence"] < 50:
        flags.append({
            "tier": "MEDIUM",
            "icon": "🟡",
            "title": "Stakeholder confidence is weak",
            "body": f"Low stakeholder confidence in {leader['Supplier']} means adoption risk is real even after award. This is how the best supplier loses after selection. Secure visible champion support before the meeting.",
        })

    if runner_up and abs(leader["Weighted Score"] - runner_up["Weighted Score"]) < 5:
        flags.append({
            "tier": "MEDIUM",
            "icon": "🟡",
            "title": "Thin lead over runner-up",
            "body": f"The gap between {leader['Supplier']} and {runner_up['Supplier']} is under 5 points. A single dimension shift in the room could flip the decision. Know exactly which dimension is being challenged before you walk in.",
        })

    if kraljic in ["Strategic", "Bottleneck"] and leader["Scores"]["Execution Risk"] < 60:
        flags.append({
            "tier": "MEDIUM",
            "icon": "🟡",
            "title": f"Execution risk elevated for {kraljic} category",
            "body": f"In a {kraljic} sourcing situation, execution risk matters more than price. {leader['Supplier']} has room to improve here — use contract terms and milestone obligations to reduce exposure.",
        })

    # Hidden risk flags
    if leader["Scores"]["Price / TCO"] > 80 and leader["Scores"]["SLA Strength"] < 55:
        flags.append({
            "tier": "HIDDEN",
            "icon": "🟣",
            "title": "Strong price, weak SLA — hidden operational exposure",
            "body": f"{leader['Supplier']} looks attractive on price but the SLA strength is below par. The cheapest option that underdelivers on service will cost more to manage than it saves on paper.",
        })

    if leader["Scores"]["Innovation Capacity"] < 40 and kraljic == "Strategic":
        flags.append({
            "tier": "HIDDEN",
            "icon": "🟣",
            "title": "Innovation capacity too low for a Strategic supplier",
            "body": f"You are placing a Strategic category with a supplier that scores low on Innovation Capacity. This is not a short-term problem — it becomes a strategic constraint 18–24 months after award when the business asks what is next.",
        })

    if not flags:
        flags.append({
            "tier": "MEDIUM",
            "icon": "🟡",
            "title": "No critical flags detected — stay disciplined",
            "body": "No major red flags are visible from current inputs. The risk is complacency. Keep stakeholder alignment active and do not let contract terms drift during final negotiation.",
        })

    return flags


def build_executive_summary(leader: Dict, runner_up: Optional[Dict], blocker_row, event_name: str, kraljic: str, category_rule: Dict, weakest_dim: str) -> str:
    tradeoff = make_tradeoff_text(leader, runner_up)
    blocker = block_risk_text(blocker_row, leader["Supplier"])

    summary = (
        f"Recommendation: {leader['Supplier']} for {event_name}. "
        f"Overall score: {leader['Weighted Score']}/100 — highest in the evaluation with a current-fit of {leader['Current Fit']} and future-fit of {leader['Future Fit']}. "
        f"{tradeoff} "
        f"The weakest dimension is {weakest_dim}, which must be addressed in contract negotiation before award. "
        f"{blocker} "
        f"Category posture: {kraljic} — {KRALJIC_INFO[kraljic]['desc']}"
    )
    return summary


def build_cfo_challenge(
    leader: Dict,
    runner_up: Optional[Dict],
    event_name: str,
    kraljic: str,
    category_rule: Dict,
    weakest_dim: str,
    intake_answers: Dict,
) -> List[Dict]:
    """Return a list of {question, answer, severity} CFO challenge Q&A pairs, all deterministic."""
    challenges = []
    score_gap = round(leader["Weighted Score"] - runner_up["Weighted Score"], 1) if runner_up else None
    price_diff = leader["Raw Price"] - runner_up["Raw Price"] if runner_up else None
    price_pct = round(price_diff / runner_up["Raw Price"] * 100, 1) if runner_up and runner_up["Raw Price"] else None

    # Q1 — Price premium justification
    if runner_up and price_diff is not None and price_diff > 0:
        challenges.append({
            "question": f"Why pay more for {leader['Supplier']} when {runner_up['Supplier']} is cheaper?",
            "answer": (
                f"{leader['Supplier']} costs ${price_diff:,.0f} more (+{price_pct}%) but scored "
                f"{score_gap} points higher ({leader['Weighted Score']} vs {runner_up['Weighted Score']}/100). "
                f"The premium reflects stronger {weakest_dim} performance and lower execution risk. "
                f"A failed implementation or SLA miss would cost far more than the price delta."
            ),
            "severity": "HIGH",
        })
    elif runner_up and price_diff is not None and price_diff < 0:
        challenges.append({
            "question": f"{leader['Supplier']} is already the cheaper option — why not just go with price?",
            "answer": (
                f"{leader['Supplier']} scores {leader['Weighted Score']}/100 and is ${abs(price_diff):,.0f} less "
                f"than {runner_up['Supplier']}. This is a dominant position — lower cost AND higher score. "
                f"The only risk to validate before award is {weakest_dim}, which should be protected contractually."
            ),
            "severity": "LOW",
        })

    # Q2 — Downside / failure exposure (provenance-aware)
    fin_risk   = leader.get("Financial Risk Label", "MEDIUM")
    fin_score  = leader.get("Financial Health", 70)
    fin_source = leader.get("Financial Health Source", "User Assessment")

    # Build a provenance clause that reflects the actual data source and freshness
    _edgar_period_end = leader.get("EDGAR Period End", "")
    _edgar_stale_clause = ""
    if _edgar_period_end and "SEC EDGAR" in fin_source:
        try:
            from datetime import date as _cfo_date
            _pe = _cfo_date.fromisoformat(_edgar_period_end)
            _today_cfo = _cfo_date.today()
            _age_months = (_today_cfo.year - _pe.year) * 12 + (_today_cfo.month - _pe.month)
            if _age_months > 18:  # >18 months: stale — more than one full annual filing cycle
                _edgar_stale_clause = (
                    f" Note: this SEC data is from {_edgar_period_end} "
                    f"({_age_months} months ago) — financial health should be refreshed before final award."
                )
            elif _age_months > 12:  # 13–18 months: amber — approaching second filing cycle
                _edgar_stale_clause = (
                    f" SEC filing data is from {_edgar_period_end} — verify recency before award."
                )
        except (ValueError, TypeError):
            pass

    if "SEC EDGAR" in fin_source:
        _fin_prov = (
            f"Financial health score of {fin_score}/100 is sourced from SEC EDGAR/XBRL annual filing data "
            f"(revenue growth, profit margin, and debt-to-assets ratio from the most recent 10-K).{_edgar_stale_clause}"
        )
    elif "Partial EDGAR" in fin_source:
        _fin_prov = (
            f"Financial health score of {fin_score}/100 is partially sourced from SEC EDGAR/XBRL filings "
            f"(some metrics were unavailable; remaining fields reflect internal assessment — validate before award).{_edgar_stale_clause}"
        )
    else:
        _fin_prov = (
            f"Financial health score of {fin_score}/100 is based on internal assessment. "
            f"This requires validation against audited financial statements before contract award."
        )

    if fin_risk in ("HIGH", "MEDIUM"):
        fin_mitigation = (
            "Require a performance bond, parent guarantee, or escrow arrangement before contract execution."
            if fin_risk == "HIGH"
            else "Request two years of audited financials and build milestone-based payment terms into the contract."
        )
        challenges.append({
            "question": f"What's our exposure if {leader['Supplier']} underperforms or fails?",
            "answer": (
                f"Financial health is rated {fin_risk}. {_fin_prov} "
                f"Weakest evaluation dimension is {weakest_dim} — this must be contractually closed before award. "
                f"{fin_mitigation} "
                f"The 90-day plan includes milestone checkpoints to catch delivery gaps early."
            ),
            "severity": fin_risk,
        })
    else:
        challenges.append({
            "question": f"What's our exposure if {leader['Supplier']} underperforms?",
            "answer": (
                f"{leader['Supplier']} has strong financial health (rated {fin_risk}, score: {fin_score}/100). "
                f"{_fin_prov} "
                f"Primary risk area is {weakest_dim}. Contractual protections should include: "
                f"SLA penalties, step-in rights, and 90-day cure periods."
            ),
            "severity": "LOW",
        })

    # Q3 — Evaluation rigour
    dim_count = sum(1 for d in DIMENSIONS if leader["Scores"].get(d, 50) != 50)
    completeness_pct = round(dim_count / len(DIMENSIONS) * 100)
    if completeness_pct >= 40:
        _completeness_clause = (
            f"{dim_count} of {len(DIMENSIONS)} dimensions ({completeness_pct}%) were scored above the default midpoint, "
            f"reflecting active assessor judgment."
        )
    else:
        _completeness_clause = (
            f"Only {dim_count} of {len(DIMENSIONS)} dimensions ({completeness_pct}%) were scored above the default midpoint — "
            f"the evaluation is primarily price and financial health driven. "
            f"Assessor should supplement dimension scores before presenting to senior stakeholders."
        )
    challenges.append({
        "question": "How rigorous was the evaluation? Is this just a gut-feel score?",
        "answer": (
            f"The evaluation scored {len(DIMENSIONS)} weighted dimensions using the Kraljic {kraljic} weighting model. "
            f"{_completeness_clause} "
            f"Weights are not arbitrary — they shift systematically based on {kraljic} posture "
            f"(e.g., {'SLA and execution risk carry highest weight' if kraljic == 'Strategic' else 'price and commercial flexibility carry highest weight' if kraljic == 'Leverage' else 'supply assurance carries highest weight'}). "
            f"Financial health ({fin_score}/100) is {'sourced from SEC EDGAR/XBRL annual filings' if 'SEC EDGAR' in fin_source else 'based on internal assessment — validate against audited financials before award'}."
        ),
        "severity": "MEDIUM" if completeness_pct >= 40 else "HIGH",
    })

    # Q4 — Runner-up alternative
    if runner_up:
        runner_risk = runner_up.get("Financial Risk Label", "MEDIUM")
        challenges.append({
            "question": f"What if we went with {runner_up['Supplier']} instead?",
            "answer": (
                f"{runner_up['Supplier']} scored {runner_up['Weighted Score']}/100 — "
                f"{'only ' if score_gap and score_gap < 8 else ''}{score_gap} pts below the recommendation. "
                + (f"At ${abs(price_diff):,.0f} {'less' if price_diff > 0 else 'more'}, " if price_diff else "")
                + f"the trade-off is lower score on {weakest_dim} with financial risk rated {runner_risk}. "
                f"{'The gap is thin enough that this is a legitimate alternative if {leader[\"Supplier\"]} fails to meet contract requirements.' if score_gap and score_gap < 5 else 'The score gap is meaningful — switching would require justifying the capability downgrade to the business.'}"
            ),
            "severity": "MEDIUM" if (score_gap and score_gap < 5) else "LOW",
        })

    # Q5 — Switching cost / exit
    switching_q = next((q for q in intake_answers if "switch" in q.lower() or "exit" in q.lower()), None)
    switching_ans = intake_answers.get(switching_q, "") if switching_q else ""
    if switching_ans:
        switching_summary = switching_ans.split("—")[0].strip()
        switching_source = f"Switching cost assessed as: {switching_summary} (from intake)."
    else:
        switching_summary = "Medium"
        switching_source = (
            "Switching cost not captured in intake — classification defaults to Medium and is unvalidated. "
            "Quantify against implementation cost, data portability terms, and re-onboarding timeline before award."
        )
    challenges.append({
        "question": "If this supplier fails mid-contract, what's the exit cost?",
        "answer": (
            f"{switching_source} "
            f"To de-risk: include a 90-day cure clause, data portability requirement, and step-in rights in the contract. "
            f"{'For a Strategic category, prioritize continuity over contract exit.' if kraljic == 'Strategic' else 'For a Leverage category, maintain at least one qualified alternate to keep exit credible.'}"
        ),
        "severity": "MEDIUM",
    })

    # Q6 — Market benchmark
    _live_market = bool(os.environ.get("ALPHA_VANTAGE_API_KEY") or os.environ.get("NEWS_API_KEY"))
    _market_data_qualifier = (
        "live BLS PPI data" if _live_market
        else "BLS PPI index data (static reference — configure Alpha Vantage API key for live pricing)"
    )
    challenges.append({
        "question": "Have we benchmarked pricing against the market?",
        "answer": (
            f"Market Intelligence tab shows {_market_data_qualifier} for {category_rule.get('tag', 'this category')} "
            f"alongside market leader comparables. {leader['Supplier']}'s quote should be compared against "
            f"the BLS inflation index and ISM price trend for this sub-category. "
            f"If pricing is above benchmark, use the Negotiation tab to generate category-specific leverage points before award."
        ),
        "severity": "LOW",
    })

    return challenges


def build_briefing_action_plan(kraljic: str, category_rule: Dict, weakest_dim: str, blocker_row, intake_answers: Dict[str, str]) -> List[str]:
    plan = [
        f"Focus the first 90 days on the highest-risk dimension: {weakest_dim}.",
        f"Obtain sign-off from the executive sponsor before contract execution — confirm they have reviewed the risk flags and financial health provenance.",
        "Lock the top negotiation priorities and contract protections before award.",
        "Confirm the recommendation narrative with the stakeholder team before the final meeting.",
    ]
    if blocker_row is not None:
        plan.insert(1, f"Engage the likely blocker {blocker_row['Name']} ({blocker_row['Role']}) one-on-one before the recommendation meeting.")
    if category_rule.get("type") == "Direct":
        plan.append("Ensure specifications, tolerances, and engineering change ownership are locked into the contract.")
    elif category_rule.get("type") == "Indirect":
        plan.append("Keep contract language focused on service performance, governance, and renewal discipline.")
    return plan


def build_90day_action_plan(
    leader: Dict,
    runner_up,
    blocker_row,
    kraljic: str,
    category_rule: Dict,
    weakest_dim: str,
    selected_sub: Dict,
    intake_answers: Dict[str, str],
) -> List[Dict]:
    """
    Build a structured 3-phase post-award 90-day action plan.
    Returns a list of phase dicts with actions, owners, and visual metadata.
    """
    supplier = leader["Supplier"]
    cat_type = category_rule.get("type", "Indirect")
    switching_cost = selected_sub.get("switching_cost", "")
    key_risks = selected_sub.get("key_risks", "performance continuity")
    fin_health = leader.get("Financial Health", 70)

    # ── Phase 1: Day 1–30 Foundation ─────────────────────────
    _weakest_score = leader["Scores"].get(weakest_dim, 50)
    p1 = [
        f"{weakest_dim} scored {_weakest_score}/100 — the lowest of all evaluated dimensions. "
        f"Execute contract with {supplier} and require a written remediation plan for this gap before Day 5.",
        "Issue internal award communication with one-page rationale memo to all stakeholders.",
        "Stand up governance model: monthly SLA review cadence, escalation path, and named supplier owner.",
    ]
    if blocker_row is not None:
        p1.insert(1, f"One-on-one with {blocker_row['Name']} ({blocker_row['Role']}) — address their {blocker_row['Priority']} concern with specific contract evidence.")
    if cat_type == "Direct":
        p1.append("Align specs, tolerances, and engineering change-control rules with supplier quality team before production handoff.")
    else:
        p1.append("Complete supplier onboarding checklist: system access, billing setup, reporting templates, and primary contact confirmation.")
    if fin_health < 60:
        p1.append(f"Flag {supplier}'s financial health score ({fin_health}/100) to Finance — request quarterly cash-flow monitoring and early-warning trigger.")

    # ── Phase 2: Day 31–60 Execution ─────────────────────────
    p2 = [
        f"Run first formal SLA review — compare actual vs. contracted {weakest_dim} performance with supporting data.",
        "Complete transition from incumbent supplier (if applicable) — confirm no operational gaps in coverage.",
        f"Quantify savings realization: compare actual invoiced price vs. scoring model assumption.",
        "Score stakeholder satisfaction: re-assess champion/skeptic positions after first 30 days of delivery.",
    ]
    if "High" in switching_cost:
        p2.insert(0, "Run parallel operations with incumbent for critical-path processes until {supplier} confirms full readiness — do not force cutover.")
    if cat_type == "Indirect":
        p2.append("Review first invoice cycle for billing accuracy, scope creep, and unauthorized pass-throughs.")
    else:
        p2.append("Conduct first quality audit against incoming inspection standards — document non-conformances with 30-day corrective action deadline.")

    # ── Phase 3: Day 61–90 Optimization ──────────────────────
    p3 = [
        f"Present 90-day performance report to executive sponsor — show {leader['Weighted Score']}/100 scoring assumptions vs. actual delivery.",
        f"Formalize Year 1 negotiation agenda: renewal cap, scope expansion options, innovation pipeline commitments.",
        f"Re-score {weakest_dim} dimension with 90 days of real data — update the evaluation record.",
        f"Document top 3 lessons learned from this sourcing event for the category playbook.",
    ]
    if runner_up is not None:
        p3.append(f"Maintain relationship with {runner_up['Supplier']} — keep BATNA viable for Year 1 renegotiation leverage.")
    p3.append(f"Set 12-month milestone: evaluate whether {kraljic} posture still applies or category has shifted.")

    return [
        {
            "phase": "Day 1–30",
            "label": "Foundation",
            "color": "#3B82F6",
            "icon": "⚡",
            "objective": f"Lock contract, activate governance, align stakeholders on {supplier} award.",
            "actions": p1,
            "owner": "Procurement Lead + Legal Counsel",
        },
        {
            "phase": "Day 31–60",
            "label": "Execution",
            "color": "#F59E0B",
            "icon": "⚙",
            "objective": "Measure first SLA cycle, validate savings, complete transition.",
            "actions": p2,
            "owner": "Category Manager + Business Owner",
        },
        {
            "phase": "Day 61–90",
            "label": "Optimization",
            "color": "#22C55E",
            "icon": "🎯",
            "objective": "Board-ready performance report, Year 1 agenda, BATNA maintenance.",
            "actions": p3,
            "owner": "Category Manager + Executive Sponsor",
        },
    ]


def build_briefing_memo(
    leader: Dict,
    runner_up: Optional[Dict],
    blocker_row,
    event_name: str,
    category: str,
    kraljic: str,
    category_rule: Dict,
    weakest_dim: str,
    stake_df: pd.DataFrame,
) -> str:
    memo_lines = [
        f"Recommendation: {leader['Supplier']} for {event_name} in {category}.",
        f"Overall Score: {leader['Weighted Score']} / 100 | Current Fit: {leader['Current Fit']} | Future Fit: {leader['Future Fit']}",
        "",
        "Why this supplier:",
        "- Highest defensible weighted score with the best mix of execution and strategic fit.",
        f"- Strongest defense under a {kraljic} sourcing posture.",
        "",
        "Primary risk to address:",
        f"- Weakest dimension: {weakest_dim}. Close this gap in negotiation and contract terms.",
    ]
    if blocker_row is not None:
        memo_lines.extend([
            "",
            "Stakeholder risk:",
            f"- Most likely blocker: {blocker_row['Name']} ({blocker_row['Role']}) — {blocker_row['Position']} focused on {blocker_row['Priority']}.",
        ])
    if runner_up is not None:
        memo_lines.extend([
            "",
            "Runner-up context:",
            f"- {runner_up['Supplier']} is the next-best supplier and narrows the gap by {round(leader['Weighted Score'] - runner_up['Weighted Score'], 1)} points.",
            "- Be prepared to explain why the chosen supplier is the stronger option despite any price delta.",
        ])

    memo_lines.extend([
        "",
        "Next steps:",
        "- Finalize the evaluation narrative and stakeholder alignment before the decision meeting.",
        "- Confirm top negotiation trade-offs and contract requirements before award.",
    ])

    return "\n".join(memo_lines)


def build_express_brief(
    event_name: str,
    category: str,
    kraljic: str,
    annual_spend: float,
    months_since_bid: int,
    score: int,
    event_type: str,
    savings_low: int,
    savings_high: int,
    reasons: List[str],
    timeline: str,
    next_steps: List[str],
) -> str:
    brief = [
        f"Recommendation: Launch a {event_type} for {category}.",
        f"Kraljic posture: {kraljic}.",
        f"Current annual spend: ${annual_spend:,.0f}.",
        f"Last competitive bid: {months_since_bid} months ago.",
        f"Decision score: {score}/90.",
        f"Estimated savings opportunity: ${savings_low:,} – ${savings_high:,}.",
        f"Timing: {timeline}.",
        "",
        "Why this matters:",
    ]
    for reason in reasons:
        brief.append(f"- {reason}")
    brief.extend([
        "",
        "Recommended board-level ask:",
        f"- Approve the proposed {event_type} and stakeholder alignment plan.",
        "- Allocate procurement and category resources to deliver the event on the next 90-day cadence.",
        "",
        "Immediate next actions:",
    ])
    for step in next_steps:
        brief.append(f"- {step}")
    brief.extend([
        "",
        "Executive note:",
        "This brief is designed to move the discussion from analysis to a ready-to-share decision package. It prioritizes clarity, risk defense, and a one-page narrative that speaks to CPO, CFO, and business partners.",
    ])
    return "\n".join(brief)


def build_quickscan_brief(scored_items: List[Dict]) -> str:
    if not scored_items:
        return "No scan results available. Enter at least one category and run the scan."

    total_spend = sum(item["spend"] for item in scored_items)
    total_savings = sum(item["savings_est"] for item in scored_items)
    p1 = [item for item in scored_items if item["priority"].startswith("🔴")]
    p2 = [item for item in scored_items if item["priority"].startswith("🟡")]
    p3 = [item for item in scored_items if item["priority"].startswith("🟢")]

    brief = [
        "90-Day Portfolio Scan Summary",
        f"Categories scanned: {len(scored_items)}.",
        f"Total portfolio spend covered: ${total_spend:,.0f}.",
        f"Total estimated savings opportunity: ${total_savings:,}.",
        "",
        "Top priorities:",
    ]
    for item in p1[:3]:
        brief.append(f"- {item['name']} — {item['priority']} | ${item['spend']:,.0f} | {item['time_flag']}.")
    if not p1:
        brief.append("- No P1 actions identified. Focus on P2 categories with medium urgency.")
    brief.extend([
        "",
        "Recommended executive ask:",
        "- Approve targeted competitive sourcing events for the P1 categories within 90 days.",
        "- Direct procurement to complete market research, supplier screening, and stakeholder alignment in the next 30 days.",
        "",
        "Recommended board-slide bullets:",
    ])
    brief.append(f"- {len(p1)} P1 category actions, {len(p2)} P2 opportunities, {len(p3)} P3 monitor items.")
    brief.append(f"- Largest near-term opportunity: {p1[0]['name']} with ${p1[0]['spend']:,.0f} spend and {p1[0]['time_flag']}." if p1 else "- No immediate P1 candidate identified.")
    brief.append("- Use the attached action plan to assign owners, dates, and expected savings for each priority category.")
    brief.extend([
        "",
        "Owner recommendation:",
        "- Procurement Category Manager: lead event execution for P1 categories.",
        "- Finance / CFO: approve savings targets and guardrail thresholds.",
        "- Legal: validate contract terms before award for any P1 events.",
    ])
    return "\n".join(brief)


def build_express_board_bullets(
    category: str,
    event_type: str,
    annual_spend: float,
    savings_low: int,
    savings_high: int,
    kraljic: str,
    timeline: str,
) -> List[str]:
    return [
        f"Approve a {event_type} for {category} to capture an estimated ${savings_low:,}–${savings_high:,} savings opportunity.",
        f"Use a {kraljic} posture to balance continuity risk with savings execution and lock contract protections before award.",
        f"Target a {timeline} launch and secure CPO/CFO sign-off on the scope, supplier list, and savings guardrails.",
        "Frame the decision as a value-risk trade-off: defend the recommended supplier on execution strength, not just price.",
    ]


def build_express_cfo_narrative(
    category: str,
    annual_spend: float,
    savings_low: int,
    savings_high: int,
    event_type: str,
    kraljic: str,
    timeline: str,
) -> str:
    return (
        f"This recommendation targets a {event_type} in {category} with ${annual_spend:,.0f} annual spend. "
        f"The expected savings range is ${savings_low:,} to ${savings_high:,}, representing a focused procurement opportunity under a {kraljic} posture. "
        f"The proposed timeline is {timeline}. "
        "Procurement requests support for the savings target, event governance, and executive escalation criteria if supplier performance or contract milestones drift."
    )


def build_express_action_matrix(
    event_type: str,
    category: str,
    kraljic: str,
    score: int,
) -> List[Dict[str, str]]:
    matrix = [
        {
            "Task": "Finalize decision memo and executive ask",
            "Owner": "Category Manager",
            "Due": "7 days",
            "Impact": "Board readiness",
        },
        {
            "Task": "Brief CFO on savings target and risk coverage",
            "Owner": "Procurement Lead",
            "Due": "7 days",
            "Impact": "Financial alignment",
        },
        {
            "Task": "Validate contract must-haves and SLA guardrails",
            "Owner": "Legal / Contract",
            "Due": "14 days",
            "Impact": "Award defense",
        },
    ]
    if score >= 60:
        matrix.append({
            "Task": "Build supplier long list and evaluation shortlist",
            "Owner": "Sourcing Team",
            "Due": "14 days",
            "Impact": "Supplier competitiveness",
        })
    else:
        matrix.append({
            "Task": "Confirm incumbent performance and renewal terms",
            "Owner": "Category Manager",
            "Due": "30 days",
            "Impact": "Contract posture",
        })
    matrix.append({
        "Task": "Publish board-slide bullets and action summary",
        "Owner": "Procurement Communications",
        "Due": "7 days",
        "Impact": "Stakeholder alignment",
    })
    return matrix


def build_quickscan_board_bullets(scored_items: List[Dict]) -> List[str]:
    top_p1 = [item for item in scored_items if item["priority"].startswith("🔴")]
    if not top_p1:
        return [
            "No immediate P1 actions identified. Focus on building urgency for P2 categories.",
            "Prepare a 90-day category roadmap for medium urgency categories with clear ownership.",
        ]

    bullets = [
        f"Approve near-term competitive sourcing events for the top {len(top_p1)} P1 categories.",
        f"Prioritize the highest spend P1 category ({top_p1[0]['name']}) for immediate market engagement.",
        "Establish a savings governance cadence with CFO and procurement to track delivery against category targets.",
    ]
    return bullets


def build_quickscan_action_matrix(scored_items: List[Dict]) -> List[Dict[str, str]]:
    matrix = []
    for item in scored_items[:3]:
        urgency = "Immediate" if item["priority"].startswith("🔴") else "Quarter"
        owner = "Category Manager" if item["priority"].startswith("🔴") else "Procurement Analyst"
        matrix.append({
            "Category": item["name"],
            "Task": f"Execute {item['priority'].split(' — ')[-1]} activity",
            "Owner": owner,
            "Due": "30 days" if urgency == "Immediate" else "60 days",
            "Impact": f"${item['savings_est']:,} savings potential",
        })
    if len(scored_items) > 3:
        matrix.append({
            "Category": "Portfolio",
            "Task": "Finalize 90-day category plan for remaining items",
            "Owner": "Procurement Operations",
            "Due": "30 days",
            "Impact": "Organizational focus",
        })
    return matrix


def build_quickscan_cfo_summary(scored_items: List[Dict]) -> str:
    total_spend = sum(item["spend"] for item in scored_items)
    total_savings = sum(item["savings_est"] for item in scored_items)
    p1_count = sum(1 for item in scored_items if item["priority"].startswith("🔴"))
    return (
        f"The 90-Day scan covers ${total_spend:,.0f} of spend and identifies ${total_savings:,} in near-term sourcing opportunity. "
        f"There are {p1_count} P1 actions requiring executive approval and resourcing. "
        "Procurement recommends a governance review with Finance to lock savings targets and operational accountability for the highest-risk categories."
    )


# =========================================================
# SENSITIVITY ANALYSIS — What-If engine
# =========================================================
def run_sensitivity_analysis(
    suppliers: List[Dict],
    base_weights: Dict[str, float],
    fin_scores: Dict[str, int],
    price_shocks: Dict[str, float],      # ticker/name → % change e.g. 0.10 = +10%
    weight_overrides: Dict[str, float],  # dimension → override weight (0-1 normalized)
    sla_overrides: Dict[str, str],       # name → new SLA value
    risk_overrides: Dict[str, str],      # name → new Risk value
) -> List[Dict]:
    """
    Re-score all suppliers with user-defined shocks applied.
    Returns ranked list with delta vs base score for each supplier.
    """
    # Apply price shocks
    shocked_suppliers = []
    for s in suppliers:
        s2 = dict(s)
        shock = price_shocks.get(s["Supplier"], 0.0)
        s2["Raw Price"] = float(s["Raw Price"]) * (1 + shock)
        sla_ov = sla_overrides.get(s["Supplier"])
        if sla_ov:
            s2["SLA Strength"] = sla_ov
        risk_ov = risk_overrides.get(s["Supplier"])
        if risk_ov:
            s2["Execution Risk"] = risk_ov
        shocked_suppliers.append(s2)

    # Merge weight overrides
    effective_weights = dict(base_weights)
    if weight_overrides:
        total_override = sum(weight_overrides.values())
        if total_override > 0:
            effective_weights = {k: weight_overrides.get(k, base_weights.get(k, 0)) for k in base_weights}
            s = sum(effective_weights.values())
            if s > 0:
                effective_weights = {k: v/s for k, v in effective_weights.items()}

    shocked_prices = [float(s["Raw Price"]) for s in shocked_suppliers]
    results = []
    for s in shocked_suppliers:
        fin = fin_scores.get(s["Supplier"], 65)
        scores = compute_supplier_scores(s, shocked_prices, fin)
        total = weighted_score(scores, effective_weights)
        c_fit = current_fit(scores)
        f_fit = future_fit(scores)
        results.append({
            "Supplier": s["Supplier"],
            "Shocked Price": s["Raw Price"],
            "Shocked Score": total,
            "Scores": scores,
            "Current Fit": c_fit,
            "Future Fit": f_fit,
        })

    return sorted(results, key=lambda x: x["Shocked Score"], reverse=True)


# =========================================================
# COVER PAGE
# =========================================================
def render_cover():
    """Cover page — split into left/right using st.columns for reliable rendering."""

    # ── Inject cover-specific animations ────────────────────────────────────
    st.markdown("""
<style>
@keyframes coverIn {
    from { opacity:0; transform:translateY(12px); }
    to   { opacity:1; transform:translateY(0); }
}
@keyframes barDraw {
    from { width:0 !important; }
}
.cover-hero-wrap {
    animation: coverIn 0.7s ease both;
    position: relative;
    z-index: 2;
}
.bar-87 { width:87% !important; animation: barDraw 1s 0.5s ease both; }
.bar-74 { width:74% !important; animation: barDraw 1s 0.65s ease both; }
.bar-61 { width:61% !important; animation: barDraw 1s 0.8s ease both; }
</style>
""", unsafe_allow_html=True)

    # ── Split into columns: 58% left, 42% right ─────────────────────────────
    left, right = st.columns([0.58, 0.42])

    with left:
        st.markdown("""
<div class="cover-hero-wrap" style="
    padding:3rem 2.5rem;
    min-height:58vh;
    display:flex;
    flex-direction:column;
    justify-content:center;
    background:linear-gradient(155deg,rgba(3,8,14,0.95) 0%,rgba(5,16,30,0.9) 60%,rgba(7,22,37,0.88) 100%);
    border:1px solid rgba(96,165,250,0.12);
    border-radius:16px 0 0 16px;
    border-right:none;
">
    <div style="font-family:'JetBrains Mono',monospace;font-size:0.6rem;letter-spacing:0.28em;text-transform:uppercase;color:#3B82F6;margin-bottom:1.4rem;display:flex;align-items:center;gap:0.6rem">
        <span style="width:28px;height:1px;background:#3B82F6;opacity:0.5;display:inline-block"></span>
        Procurement Decision Intelligence
    </div>
    <div style="font-family:'DM Serif Display',serif;font-size:5.5rem;line-height:0.88;color:#F1F5F9;letter-spacing:-0.03em;margin-bottom:1.4rem;font-weight:400">
        Procure<span style="-webkit-text-stroke:2px #60A5FA;color:transparent;font-style:italic">IQ</span>
    </div>
    <div style="font-size:1rem;font-weight:300;color:#C4D3E8;line-height:1.85;max-width:440px">
        Choose the supplier.<br>
        <strong style="color:#CBD5E1;font-weight:500">Defend the choice.</strong><br>
        Predict where it breaks —<br>
        before you're in the room.
    </div>
    <div style="display:flex;gap:2rem;margin-top:2rem;padding-top:1.5rem;border-top:1px solid rgba(148,163,184,0.07)">
        <div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:1.35rem;font-weight:700;color:#60A5FA;line-height:1">100</div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:0.5rem;letter-spacing:0.12em;text-transform:uppercase;color:#B0C4DC;margin-top:0.15rem">Subcategories</div>
        </div>
        <div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:1.35rem;font-weight:700;color:#60A5FA;line-height:1">8</div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:0.5rem;letter-spacing:0.12em;text-transform:uppercase;color:#B0C4DC;margin-top:0.15rem">Eval Dimensions</div>
        </div>
        <div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:1.35rem;font-weight:700;color:#60A5FA;line-height:1">10</div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:0.5rem;letter-spacing:0.12em;text-transform:uppercase;color:#B0C4DC;margin-top:0.15rem">Workflow Tabs</div>
        </div>
        <div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:1.35rem;font-weight:700;color:#60A5FA;line-height:1">SEC</div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:0.5rem;letter-spacing:0.12em;text-transform:uppercase;color:#B0C4DC;margin-top:0.15rem">EDGAR</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

    with right:
        st.markdown("""
<div style="
    padding:2rem 1.6rem;
    min-height:58vh;
    display:flex;
    flex-direction:column;
    background:rgba(4,11,22,0.95);
    border:1px solid rgba(96,165,250,0.12);
    border-radius:0 16px 16px 0;
    border-left:1px solid rgba(96,165,250,0.07);
">
    <div style="display:flex;align-items:center;gap:0.4rem;font-family:'JetBrains Mono',monospace;font-size:0.54rem;letter-spacing:0.18em;text-transform:uppercase;color:#9EB8CE;margin-bottom:1.1rem;padding-bottom:0.6rem;border-bottom:1px solid rgba(96,165,250,0.06)">
        <span style="width:5px;height:5px;border-radius:50%;background:#4ADE80;display:inline-block;animation:pulseGreen 2s infinite"></span>
        LIVE SYSTEM PREVIEW &nbsp;·&nbsp; ACTIVE
    </div>
    <div style="font-family:'JetBrains Mono',monospace;font-size:0.52rem;letter-spacing:0.16em;text-transform:uppercase;color:#B0C4DC;margin-bottom:0.6rem">SUPPLIER EVALUATION</div>
    <div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.4rem">
        <span style="font-family:'JetBrains Mono',monospace;font-size:0.6rem;color:#A8BEDC;width:68px">Supplier A</span>
        <div style="flex:1;height:4px;background:rgba(96,165,250,0.07);border-radius:2px;overflow:hidden">
            <div style="height:100%;background:linear-gradient(90deg,#1D4ED8,#60A5FA);border-radius:2px;width:87%"></div>
        </div>
        <span style="font-family:'JetBrains Mono',monospace;font-size:0.66rem;color:#60A5FA;width:22px;text-align:right">87</span>
        <span style="font-family:'JetBrains Mono',monospace;font-size:0.46rem;color:#4ADE80;border:1px solid rgba(74,222,128,0.22);background:rgba(74,222,128,0.07);border-radius:3px;padding:0.04rem 0.22rem">REC</span>
    </div>
    <div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.4rem">
        <span style="font-family:'JetBrains Mono',monospace;font-size:0.6rem;color:#A8BEDC;width:68px">Supplier B</span>
        <div style="flex:1;height:4px;background:rgba(96,165,250,0.07);border-radius:2px;overflow:hidden">
            <div style="height:100%;background:linear-gradient(90deg,#1E3A5F,#3B82F6);border-radius:2px;width:74%"></div>
        </div>
        <span style="font-family:'JetBrains Mono',monospace;font-size:0.66rem;color:#A8BEDC;width:22px;text-align:right">74</span>
    </div>
    <div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:1rem">
        <span style="font-family:'JetBrains Mono',monospace;font-size:0.6rem;color:#A8BEDC;width:68px">Supplier C</span>
        <div style="flex:1;height:4px;background:rgba(96,165,250,0.07);border-radius:2px;overflow:hidden">
            <div style="height:100%;background:linear-gradient(90deg,#172032,#1D4ED8);border-radius:2px;width:61%"></div>
        </div>
        <span style="font-family:'JetBrains Mono',monospace;font-size:0.66rem;color:#9EB8CE;width:22px;text-align:right">61</span>
    </div>
    <div style="font-family:'JetBrains Mono',monospace;font-size:0.52rem;letter-spacing:0.16em;text-transform:uppercase;color:#B0C4DC;margin-bottom:0.5rem">RISK SIGNALS</div>
    <div style="padding:0.28rem 0.45rem;border-radius:4px;background:rgba(248,113,113,0.06);border-left:2px solid #F87171;margin-bottom:0.22rem;font-size:0.82rem;color:#C4D3E8">⚠ Transfer pricing exposure flagged</div>
    <div style="padding:0.28rem 0.45rem;border-radius:4px;background:rgba(74,222,128,0.06);border-left:2px solid #4ADE80;margin-bottom:0.9rem;font-size:0.82rem;color:#C4D3E8">✓ Stakeholder alignment confirmed</div>
    <div style="font-family:'JetBrains Mono',monospace;font-size:0.52rem;letter-spacing:0.16em;text-transform:uppercase;color:#B0C4DC;margin-bottom:0.5rem">SYSTEM RECOMMENDATION</div>
    <div style="background:rgba(3,8,15,0.7);border-left:2px solid #1D4ED8;border-radius:0 4px 4px 0;padding:0.5rem 0.7rem;font-family:'JetBrains Mono',monospace;font-size:0.78rem">
        <div style="display:flex;gap:0.6rem;padding:0.07rem 0"><span style="color:#1D4ED8">&gt;</span><span style="color:#9EB8CE;width:74px">DECISION</span><span style="color:#93C5FD">Supplier A</span></div>
        <div style="display:flex;gap:0.6rem;padding:0.07rem 0"><span style="color:#1D4ED8">&gt;</span><span style="color:#9EB8CE;width:74px">SCORE</span><span style="color:#93C5FD">87 / 100</span></div>
        <div style="display:flex;gap:0.6rem;padding:0.07rem 0"><span style="color:#1D4ED8">&gt;</span><span style="color:#9EB8CE;width:74px">POSTURE</span><span style="color:#93C5FD">Strategic</span></div>
        <div style="display:flex;gap:0.6rem;padding:0.07rem 0"><span style="color:#1D4ED8">&gt;</span><span style="color:#9EB8CE;width:74px">CONFIDENCE</span><span style="color:#4ADE80">High</span></div>
        <div style="display:flex;gap:0.6rem;padding:0.07rem 0"><span style="color:#1D4ED8">&gt;</span><span style="color:#9EB8CE;width:74px">WEAKEST DIM</span><span style="color:#FCD34D">Price / TCO</span></div>
        <div style="display:flex;gap:0.6rem;padding:0.07rem 0"><span style="color:#1D4ED8">&gt;</span><span style="color:#9EB8CE;width:74px">ACTION</span><span style="color:#D0E0EF">Negotiate floor before award</span></div>
    </div>
    <div style="font-family:'JetBrains Mono',monospace;font-size:0.52rem;color:#C8D8EC;text-align:center;margin-top:auto;padding-top:0.8rem;letter-spacing:0.1em">REAL DATA POPULATES WHEN YOU ENTER THE DASHBOARD</div>
</div>
""", unsafe_allow_html=True)

    # ── Three pillar cards ───────────────────────────────────────────────────
    st.markdown("<div style='height:1.2rem'></div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    pillars = [
        ("⚖", "01 / DECIDE", "Recommendation Engine",
         "Translate supplier inputs into a weighted score — defensible in an executive review, not just printable.",
         "Sigmoid pricing · 8 dimensions"),
        ("🛡", "02 / DEFEND", "Trade-off + Block Risk",
         "Surface the compromises. Map exactly who will push back — before they do it to you in the room.",
         "Stakeholder mapping · AI pressure test"),
        ("🗺", "03 / RUN IT", "Category Playbook",
         "Auto-generate your stakeholder team, RFP timeline, contract must-haves, and risk flags by subcategory.",
         "100 subcategories · What-If analysis"),
    ]
    for col, (icon, num, title, body, tag) in zip([c1, c2, c3], pillars):
        with col:
            st.markdown(
                f'<div class="pillar">'
                f'<span style="font-size:1.3rem;display:block;margin-bottom:0.6rem">{icon}</span>'
                f'<div style="font-family:monospace;font-size:0.78rem;letter-spacing:0.14em;color:#3B82F6;margin-bottom:0.4rem;text-transform:uppercase">{num}</div>'
                f'<div style="font-size:0.95rem;font-weight:700;color:#E2E8F0;margin-bottom:0.4rem">{title}</div>'
                f'<div style="font-size:0.8rem;color:#D0E0EF;line-height:1.6">{body}</div>'
                f'<div style="display:inline-block;margin-top:0.7rem;font-family:monospace;font-size:0.78rem;letter-spacing:0.08em;color:#60A5FA;background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.14);border-radius:4px;padding:0.12rem 0.45rem;text-transform:uppercase">{tag}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)
    _, c_exp, c_dash, c_scan, _ = st.columns([0.6, 1, 1.2, 1, 0.6])
    with c_exp:
        if st.button("⚡ Express Mode", use_container_width=True, help="Quick 3-field assessment — should I run an event?"):
            st.session_state.entered_express = True
            st.rerun()
    with c_dash:
        if st.button("Enter Dashboard →", use_container_width=True, type="primary"):
            st.session_state.entered_dashboard = True
            st.rerun()
    with c_scan:
        if st.button("🔎 90-Day Scan", use_container_width=True, help="Multi-category opportunity scan"):
            st.session_state.entered_scan = True
            st.rerun()
    st.markdown(
        "<p style='text-align:center;font-size:0.85rem;color:#A8BEDC;margin-top:0.75rem;'>"
        "Express Mode and 90-Day Scan generate a ready-to-share executive brief and action plan without requiring the full dashboard."
        "</p>",
        unsafe_allow_html=True,
    )

    # ── About / Credentials ─────────────────────────────────────────────────
    st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)
    with st.expander("About ProcureIQ — Methodology & Credentials", expanded=False):
        ab1, ab2 = st.columns(2)
        with ab1:
            st.markdown("""
**Procurement Frameworks Applied**

| Framework | Application |
|---|---|
| **Kraljic Matrix** | Category posture classification; dimension weight adjustment |
| **RAQSCI** | Contract risk lens per category (Requirements, Assurance, Quality, Schedule, Cost, Innovation) |
| **Stakeholder Power/Interest Map** | Blocker identification and talk-track generation |
| **TCO Analysis** | Total cost of ownership multi-year projection |
| **CSCP / ASCM** | Supply chain risk and category strategy principles |
| **ISM / CPSM Standards** | Evaluation rubrics and supplier assessment criteria |

**Scoring Model**
- 8-dimension weighted evaluation with dynamic Kraljic weights
- Financial health computed from 6 qualitative signals
- Confidence interval ±8 pts based on ordinal input uncertainty
- Score completeness tracks how many dimensions exceed the default midpoint
            """)
        with ab2:
            st.markdown("""
**What ProcureIQ is designed for**
- Competitive sourcing events with 2–8 shortlisted suppliers
- Strategic and leverage categories where defensible recommendation is required
- Pre-meeting preparation: stakeholder mapping, risk flagging, CFO challenge prep
- Teams without access to Ariba, Coupa, or full-suite CLM

**What it is not**
- Not a contract management or P2P system
- Not a compliance certification tool (sanctions, ESG audits)
- Not a replacement for qualified procurement judgment or legal review
- AI-generated outputs are decision-support only

**AI Safety**
- All AI outputs are clearly labeled and disclaimed
- Sanctions screening explicitly not certified as compliance
- Contract drafts labeled as starting points requiring legal review
- LLM output is never rendered as raw HTML

**Technical Stack**
- Python / Streamlit frontend
- SQLite for session and evaluation persistence
- Pluggable LLM: Claude, GPT-4o, DeepSeek, Grok
- Live data: SEC EDGAR, BLS PPI, OFAC SDN, yfinance
            """)
        st.markdown(
            '<div style="font-size:0.75rem;color:#64748B;margin-top:0.5rem;padding-top:0.5rem;'
            'border-top:1px solid rgba(148,163,184,0.1)">'
            'ProcureIQ is a decision-support tool. All outputs must be reviewed by a qualified procurement '
            'professional before use in sourcing decisions, contract awards, or regulatory filings.'
            '</div>',
            unsafe_allow_html=True,
        )


def render_express():
    _back_col, _ = st.columns([1, 5])
    with _back_col:
        if st.button("← Back to Cover", key="express_back_top"):
            st.session_state.entered_express = False
            st.rerun()

    st.title("⚡ Express Mode")
    st.caption("Three inputs. Instant sourcing decision. No spreadsheet required.")
    st.divider()

    cat_keys = list(CATEGORY_TAXONOMY.keys())
    cat_options = [f"{CATEGORY_TAXONOMY[k]['icon']} {CATEGORY_TAXONOMY[k]['label']}" for k in cat_keys]

    c1, c2, c3 = st.columns(3)
    with c1:
        cat_choice = st.selectbox("Category", cat_options, key="express_category")
        cat_key = cat_keys[cat_options.index(cat_choice)]
    with c2:
        annual_spend = st.number_input("Annual Spend ($)", min_value=0, value=500000, step=50000, key="express_spend")
    with c3:
        months_since_bid = st.number_input("Months Since Last Bid", min_value=0, max_value=240, value=18, step=1, key="express_months")

    cat_data = CATEGORY_TAXONOMY[cat_key]
    first_sub = cat_data["subcategories"][0]
    kraljic = first_sub["kraljic_default"]

    score = 0
    reasons = []

    if months_since_bid >= 36:
        score += 40
        reasons.append(f"🔴 **{months_since_bid} months since last bid** — well past the 24–36 month refresh window. Incumbent has likely repriced upward.")
    elif months_since_bid >= 24:
        score += 25
        reasons.append(f"🟡 **{months_since_bid} months since last bid** — approaching the standard refresh threshold. Begin market soundings.")
    elif months_since_bid >= 12:
        score += 10
        reasons.append(f"🟢 **{months_since_bid} months since last bid** — within window, but review whether scope has drifted since award.")
    else:
        score += 0
        reasons.append(f"🟢 **{months_since_bid} months since last bid** — recently awarded. Focus on onboarding and SLA measurement.")

    if annual_spend >= 1_000_000:
        score += 30
        reasons.append(f"🔴 **${annual_spend:,.0f}/yr spend** — Tier 1 category. Even 5% savings = ${int(annual_spend*0.05):,}. Event ROI is clear.")
    elif annual_spend >= 250_000:
        score += 15
        reasons.append(f"🟡 **${annual_spend:,.0f}/yr spend** — Tier 2. Competitive event typically yields 8–15% savings at this level.")
    else:
        score += 5
        reasons.append(f"🟢 **${annual_spend:,.0f}/yr spend** — Lower spend tier. Run a 3-bid quote rather than a full RFP.")

    if kraljic == "Strategic":
        score += 20
        reasons.append(f"🔴 **Strategic posture** — This category carries significant risk. Run an event for continuity protection, not just savings.")
    elif kraljic == "Bottleneck":
        score += 15
        reasons.append(f"🟡 **Bottleneck posture** — Few qualified alternatives. Event focus: lock in SLAs and supply continuity terms.")
    elif kraljic == "Leverage":
        score += 10
        reasons.append(f"🟡 **Leverage posture** — Competitive market. An RFP or reverse auction will generate measurable savings.")
    else:
        score += 3
        reasons.append(f"🟢 **Non-Critical posture** — Lightweight benchmark or 3-bid quote is sufficient.")

    # Savings benchmarks by Kraljic
    _sav_low  = {"Strategic": 0.03, "Bottleneck": 0.01, "Leverage": 0.08, "Non-Critical": 0.10}.get(kraljic, 0.05)
    _sav_high = {"Strategic": 0.08, "Bottleneck": 0.04, "Leverage": 0.15, "Non-Critical": 0.20}.get(kraljic, 0.12)
    savings_low  = int(annual_spend * _sav_low)
    savings_high = int(annual_spend * _sav_high)

    if score >= 60:
        event_type = "Full RFP" if (annual_spend >= 500_000 or kraljic == "Strategic") else "3-Bid Quote"
        timeline   = "Launch in 30 days · Award in 60–90 days"
        st.error(f"### ✅ RUN AN EVENT — Score {score}/90")
        st.markdown("**Strong signals across all three dimensions.** A competitive sourcing event is justified and likely to deliver measurable results. Delay increases incumbent leverage.")
    elif score >= 35:
        event_type = "RFQ / Benchmark"
        timeline   = "Plan for next quarter"
        st.warning(f"### ⚡ CONSIDER AN EVENT — Score {score}/90")
        st.markdown("**Mixed signals.** Not urgent today, but put it on the Q2 sourcing calendar. Use the next 30 days for stakeholder alignment and market research.")
    else:
        event_type = "Performance Review"
        timeline   = "Revisit in 6–12 months"
        st.success(f"### ✋ HOLD — Score {score}/90")
        st.markdown("**Contract is likely still competitive.** Focus on SLA performance, relationship management, and innovation asks. Set a calendar reminder to reassess.")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Recommended Action", event_type)
    m2.metric("Savings Estimate", f"${savings_low:,} – ${savings_high:,}")
    m3.metric("Timeline", timeline.split(" · ")[0])
    m4.metric("Kraljic Posture", kraljic)

    st.divider()
    st.markdown("**Why this score:**")
    for r in reasons:
        st.markdown(r)

    with st.expander("Category Context & Procurement Notes"):
        cx1, cx2 = st.columns(2)
        cx1.metric("Category", cat_data["label"])
        cx1.metric("Typical Contract Type", first_sub["contract_type"])
        cx2.metric("Auction Format", first_sub.get("auction", "RFP"))
        cx2.metric("Switching Cost", first_sub.get("switching_cost", "—"))
        st.info(first_sub["notes"])
        st.markdown("**Key Risks:** " + first_sub.get("key_risks", "—"))
        st.markdown("**Key Stakeholders:** " + ", ".join(first_sub.get("stakeholders", [])))

    with st.expander("📋 Next Steps Checklist"):
        if score >= 60:
            steps = [
                "[ ] Get CPO / CFO sign-off to launch event",
                "[ ] Build long list of qualified suppliers (target 4–6)",
                "[ ] Draft RFP scope — use the ProcureIQ template in the dashboard",
                "[ ] Brief key stakeholders before issuing — use Stakeholder tab",
                "[ ] Set bid deadline 21–28 days from issue date",
                "[ ] Plan negotiation strategy — use Negotiate tab in dashboard",
            ]
        elif score >= 35:
            steps = [
                "[ ] Add to Q2 sourcing calendar with a firm launch date",
                "[ ] Request updated pricing from incumbent in the next 30 days",
                "[ ] Run a market scan — who are the top 3 alternatives?",
                "[ ] Align with CFO on savings target before launching",
                "[ ] Review current contract for auto-renewal dates and notice periods",
            ]
        else:
            steps = [
                "[ ] Schedule QBR with incumbent in the next 60 days",
                "[ ] Review SLA performance data from the last 6 months",
                "[ ] Confirm contract expiry date and set calendar reminder 90 days out",
                "[ ] Document any scope changes since last award for future RFP",
                "[ ] Identify one innovation or value-add ask for the next QBR",
            ]
        for step in steps:
            st.markdown(step)

    express_brief = build_express_brief(
        event_name="Express Procurement Decision",
        category=cat_data["label"],
        kraljic=kraljic,
        annual_spend=annual_spend,
        months_since_bid=months_since_bid,
        score=score,
        event_type=event_type,
        savings_low=savings_low,
        savings_high=savings_high,
        reasons=reasons,
        timeline=timeline,
        next_steps=steps,
    )
    express_bullets = build_express_board_bullets(
        category=cat_data["label"],
        event_type=event_type,
        annual_spend=annual_spend,
        savings_low=savings_low,
        savings_high=savings_high,
        kraljic=kraljic,
        timeline=timeline,
    )
    express_matrix = build_express_action_matrix(
        event_type=event_type,
        category=cat_data["label"],
        kraljic=kraljic,
        score=score,
    )
    cfo_narrative = build_express_cfo_narrative(
        category=cat_data["label"],
        annual_spend=annual_spend,
        savings_low=savings_low,
        savings_high=savings_high,
        event_type=event_type,
        kraljic=kraljic,
        timeline=timeline,
    )

    st.markdown("#### 📄 Executive Briefing Summary")
    st.text_area("Express brief — copy directly into a decision memo", express_brief, height=220)
    st.download_button(
        "⬇ Download Express Brief (TXT)",
        express_brief,
        "express_brief.txt",
        "text/plain",
        key="express_brief_export",
    )

    st.markdown("#### 💼 CFO Savings Narrative")
    st.text_area("CFO-ready savings narrative", cfo_narrative, height=180)
    st.download_button(
        "⬇ Download CFO Narrative (TXT)",
        cfo_narrative,
        "express_cfo_narrative.txt",
        "text/plain",
        key="express_cfo_export",
    )

    st.markdown("#### 📌 Board Slide Bullets")
    st.text_area("Board bullets — paste into deck notes", "\n".join(express_bullets), height=180)
    st.download_button(
        "⬇ Download Board Bullets (TXT)",
        "\n".join(express_bullets),
        "express_board_bullets.txt",
        "text/plain",
        key="express_bullets_export",
    )

    st.markdown("#### 🧭 Execution Matrix")
    st.table(pd.DataFrame(express_matrix))
    st.download_button(
        "⬇ Download Execution Matrix (CSV)",
        pd.DataFrame(express_matrix).to_csv(index=False),
        "express_execution_matrix.csv",
        "text/csv",
        key="express_matrix_export",
    )

    st.divider()
    _, center, _ = st.columns([2, 1, 2])
    with center:
        if st.button("Enter Full Dashboard →", use_container_width=True, type="primary", key="express_to_dashboard"):
            st.session_state.entered_express = False
            st.session_state.entered_dashboard = True
            st.rerun()


# =========================================================
# 90-DAY QUICK SCAN
# =========================================================
def _score_scan_item(item: dict) -> dict:
    score = 0
    if item["months"] >= 36:
        score += 40; time_flag = "🔴 Overdue"; time_action = "Launch RFP or reverse auction immediately"
    elif item["months"] >= 24:
        score += 25; time_flag = "🟡 Due Soon"; time_action = "Plan competitive event this quarter"
    elif item["months"] >= 12:
        score += 10; time_flag = "🟡 Approaching"; time_action = "Begin market research and long-list building"
    else:
        score += 2; time_flag = "🟢 Recent"; time_action = "Monitor SLA performance and relationship health"

    if item["spend"] >= 1_000_000:
        score += 30; spend_tier = "Tier 1 — Enterprise (>$1M)"
    elif item["spend"] >= 250_000:
        score += 15; spend_tier = "Tier 2 — Mid-Market ($250K–$1M)"
    elif item["spend"] >= 50_000:
        score += 8;  spend_tier = "Tier 3 — Tactical ($50K–$250K)"
    else:
        score += 2;  spend_tier = "Tier 4 — Spot (<$50K)"

    if item["score_boost"]:
        score += 10

    _sav_pct = 0.10 if score >= 50 else 0.06 if score >= 30 else 0.03
    savings_est = int(item["spend"] * _sav_pct)

    if score >= 55:
        priority = "🔴 P1 — Act Now"; p_delta = "High urgency"
    elif score >= 30:
        priority = "🟡 P2 — This Quarter"; p_delta = "Medium urgency"
    else:
        priority = "🟢 P3 — Monitor"; p_delta = "Low urgency"

    return {**item, "score": score, "time_flag": time_flag, "time_action": time_action,
            "spend_tier": spend_tier, "priority": priority, "savings_est": savings_est, "p_delta": p_delta}


def render_quickscan():
    _back_col, _ = st.columns([1, 5])
    with _back_col:
        if st.button("← Back to Cover", key="scan_back_top"):
            st.session_state.entered_scan = False
            st.session_state.pop("scan_results", None)
            st.rerun()

    st.title("🔎 90-Day Quick Scan")
    st.caption("Add your active categories. Get a prioritized action plan ranked by savings opportunity and urgency.")
    st.divider()

    num_cats = st.slider("Number of Categories to Scan", 1, 10, 3, key="scan_num_cats")

    scan_items = []
    for i in range(num_cats):
        with st.expander(f"Category {i + 1}", expanded=True):
            c1, c2, c3, c4 = st.columns([2, 1.5, 1.5, 1])
            with c1:
                cat_name = st.text_input("Category / Description", key=f"scan_cat_{i}",
                                         placeholder="e.g. Temp Labor, Cloud Infrastructure")
            with c2:
                spend = st.number_input("Annual Spend ($)", min_value=0, value=250000, step=25000, key=f"scan_spend_{i}")
            with c3:
                months = st.number_input("Months Since Last Bid", min_value=0, max_value=240, value=12, step=1, key=f"scan_months_{i}")
            with c4:
                score_boost = st.checkbox("No contract?", key=f"scan_boost_{i}", help="Check if there's no formal contract in place — adds urgency")
            scan_items.append({"name": cat_name, "spend": spend, "months": months, "score_boost": score_boost})

    col_run, col_clear = st.columns([1, 5])
    with col_run:
        run_clicked = st.button("▶ Run Scan", type="primary", key="run_scan")
    with col_clear:
        if st.button("✕ Clear", key="scan_clear"):
            st.session_state.pop("scan_results", None)
            st.rerun()

    # Store results in session state so they survive reruns
    if run_clicked:
        valid_items = [item for item in scan_items if item["name"].strip()]
        if not valid_items:
            st.warning("Enter at least one category name to scan.")
        else:
            scored = [_score_scan_item(item) for item in valid_items]
            scored.sort(key=lambda x: x["score"], reverse=True)
            st.session_state["scan_results"] = scored

    # Always render from session state — survives reruns
    scored = st.session_state.get("scan_results")
    if scored:
        st.divider()
        total_spend  = sum(s["spend"] for s in scored)
        total_savings = sum(s["savings_est"] for s in scored)
        p1_count = sum(1 for s in scored if s["priority"].startswith("🔴"))

        sm1, sm2, sm3 = st.columns(3)
        sm1.metric("Categories Scanned", len(scored))
        sm2.metric("Total Portfolio Spend", f"${total_spend:,.0f}")
        sm3.metric("Estimated Savings Opportunity", f"${total_savings:,}", delta=f"{p1_count} P1 actions")

        st.markdown("### 📋 90-Day Action Plan")
        for rank, item in enumerate(scored):
            priority = item["priority"]
            if priority.startswith("🔴"):
                st.error(f"**#{rank+1} — {item['name']}** · {priority}")
            elif priority.startswith("🟡"):
                st.warning(f"**#{rank+1} — {item['name']}** · {priority}")
            else:
                st.success(f"**#{rank+1} — {item['name']}** · {priority}")

            detail_cols = st.columns(4)
            detail_cols[0].metric("Annual Spend", f"${item['spend']:,.0f}")
            detail_cols[1].metric("Months Since Bid", str(item["months"]))
            detail_cols[2].metric("Est. Savings", f"${item['savings_est']:,}")
            detail_cols[3].metric("Spend Tier", item["spend_tier"].split(" —")[0])
            st.caption(f"{item['time_flag']} · Recommended action: **{item['time_action']}**")
            st.markdown("---")

        # Excel-ready formula for prioritization
        with st.expander("📐 Excel Formula — Replicate This Scoring in Your Spreadsheet"):
            st.markdown("Set up columns: **A** = Category, **B** = Annual Spend ($), **C** = Months Since Last Bid")
            st.markdown("Paste this in column **D** (Priority Score) starting at D2:")
            formula = '=IF(C2>=36,40,IF(C2>=24,25,IF(C2>=12,10,2)))+IF(B2>=1000000,30,IF(B2>=250000,15,IF(B2>=50000,8,2)))'
            st.code(formula, language="text")
            st.markdown("Sort column D descending. Scores ≥55 = P1, 30–54 = P2, <30 = P3.")
            st.code('=IF(D2>=55,"P1 - Act Now",IF(D2>=30,"P2 - This Quarter","P3 - Monitor"))', language="text")

        # Export
        scan_df = pd.DataFrame([{
            "Rank": i+1,
            "Category": s["name"],
            "Annual Spend ($)": s["spend"],
            "Months Since Bid": s["months"],
            "Priority Score": s["score"],
            "Priority": s["priority"].split(" — ")[-1] if " — " in s["priority"] else s["priority"],
            "Spend Tier": s["spend_tier"],
            "Recommended Action": s["time_action"],
            "Est. Savings ($)": s["savings_est"],
        } for i, s in enumerate(scored)])
        st.download_button(
            "⬇ Export Action Plan (CSV)", scan_df.to_csv(index=False),
            "90day_scan.csv", "text/csv", key="scan_export",
        )

        scan_brief = build_quickscan_brief(scored)
        scan_bullets = build_quickscan_board_bullets(scored)
        scan_matrix = build_quickscan_action_matrix(scored)
        scan_cfo_narrative = build_quickscan_cfo_summary(scored)

        st.markdown("#### 💼 CFO Savings Narrative")
        st.text_area("CFO-ready quick scan narrative", scan_cfo_narrative, height=180)
        st.download_button(
            "⬇ Download Quick Scan CFO Narrative (TXT)",
            scan_cfo_narrative,
            "90day_scan_cfo_narrative.txt",
            "text/plain",
            key="scan_cfo_export",
        )

        st.markdown("#### 📌 Board Slide Bullets")
        st.text_area("Quick scan board bullets", "\n".join(scan_bullets), height=180)
        st.download_button(
            "⬇ Download Quick Scan Bullets (TXT)",
            "\n".join(scan_bullets),
            "90day_scan_bullets.txt",
            "text/plain",
            key="scan_bullets_export",
        )

        st.markdown("#### 🧭 Action Matrix")
        st.table(pd.DataFrame(scan_matrix))
        st.download_button(
            "⬇ Download Quick Scan Action Matrix (CSV)",
            pd.DataFrame(scan_matrix).to_csv(index=False),
            "90day_scan_action_matrix.csv",
            "text/csv",
            key="scan_matrix_export",
        )

        st.markdown("#### 📄 Quick Scan Briefing Summary")
        st.text_area("Copy-ready quick scan summary", scan_brief, height=260)
        st.download_button(
            "⬇ Download Quick Scan Brief (TXT)",
            scan_brief,
            "90day_scan_brief.txt",
            "text/plain",
            key="scan_brief_export",
        )

    st.divider()
    _, center, _ = st.columns([2, 1, 2])
    with center:
        if st.button("Enter Full Dashboard →", use_container_width=True, type="primary", key="scan_to_dashboard"):
            st.session_state.entered_scan = False
            st.session_state.entered_dashboard = True
            st.rerun()


def _score_color(score) -> str:
    """Return hex color for a 0-100 score: red <60, amber 60-75, green >75."""
    try:
        s = float(score)
    except (TypeError, ValueError):
        return "#94A3B8"
    if s >= 75:
        return "#4ADE80"
    if s >= 60:
        return "#FCD34D"
    return "#F87171"


def _confidence_label(supplier: Dict) -> str:
    """Return HIGH/MEDIUM/LOW confidence based on score spread and data completeness."""
    score = supplier.get("Weighted Score", 0)
    scores = supplier.get("Scores", {})
    filled = sum(1 for v in scores.values() if v != 50)
    if score >= 70 and filled >= 6:
        return "HIGH"
    if score >= 55 and filled >= 3:
        return "MEDIUM"
    return "LOW"


def _confidence_color(label: str) -> str:
    return {"HIGH": "#4ADE80", "MEDIUM": "#FCD34D", "LOW": "#F87171"}.get(label, "#94A3B8")


def render_dashboard():
    # ══ RIGHT PANEL LAYOUT ═══════════════════════════════════
    # All controls live in a right-side column (always visible).
    # Main content (tabs, hero, dashboard) is in the wider left column.
    # The Streamlit sidebar is disabled — no hidden arrows needed.

    # ranked/leader/runner_up are computed inside the Suppliers tab block;
    # initialize here so Overview tab references don't raise UnboundLocalError
    ranked: list = []
    leader = None
    runner_up = None

    # ── Pre-read weight sub for use in right panel ──────────
    # Normalize parent key: intake agent writes long-form keys ("information_technology")
    # but CATEGORY_TAXONOMY uses short keys ("IT"). Fall back to first valid key.
    _CAT_ALIAS = {
        "information_technology": "IT", "it": "IT",
        "human_resources": "HR", "hr": "HR",
        "finance_professional_services": "Finance", "finance": "Finance",
        "marketing_communications": "Marketing", "marketing": "Marketing",
        "logistics_transportation": "Logistics", "logistics": "Logistics",
        "operations_mro": "Operations / MRO", "operations": "Operations / MRO",
        "legal": "Legal",
        "facilities_real_estate": "Facilities", "facilities": "Facilities",
        "professional_services": "Professional Services",
        "direct_materials": "Direct Materials",
        "corporate_services": "HR",
        # New categories
        "travel": "Travel & Meetings", "travel_meetings": "Travel & Meetings",
        "travel_expense": "Travel & Meetings", "t&e": "Travel & Meetings",
        "healthcare": "Healthcare & Benefits", "benefits": "Healthcare & Benefits",
        "healthcare_benefits": "Healthcare & Benefits",
        "energy": "Energy & Utilities", "utilities": "Energy & Utilities",
        "energy_utilities": "Energy & Utilities",
        "construction": "Construction & Capital Projects",
        "capital_projects": "Construction & Capital Projects",
        "print": "Print & Document Services", "document": "Print & Document Services",
        "print_document": "Print & Document Services",
    }
    _raw_parent = st.session_state.get("selected_parent_cat", list(CATEGORY_TAXONOMY.keys())[0])
    _wt_parent  = _CAT_ALIAS.get(str(_raw_parent).lower(), _raw_parent)
    if _wt_parent not in CATEGORY_TAXONOMY:
        _wt_parent = list(CATEGORY_TAXONOMY.keys())[0]
    _wt_sub    = st.session_state.get("selected_sub_name", CATEGORY_TAXONOMY[_wt_parent]["subcategories"][0]["name"])

    # ── Agent availability banner ────────────────────────────
    if not _AGENTS_AVAILABLE:
        st.warning(
            "**AI features are unavailable** — the `agents` package is not installed. "
            "Supplier discovery, sanctions screening, contract generation, spend analysis, "
            "and all LLM calls are disabled. "
            "Run `pip install -r requirements.txt` and restart the app to enable them.",
            icon="⚠️",
        )

    # ── Layout: main content | right control panel ──────────
    main_col, ctrl_col = st.columns([3.2, 1])

    # ════════════════════════════════════════
    # RIGHT CONTROL PANEL
    # ════════════════════════════════════════
    with ctrl_col:
        st.markdown('<div class="ctrl-panel">', unsafe_allow_html=True)
        st.markdown('<div class="ctrl-header">PROCUREIQ CONTROLS</div>', unsafe_allow_html=True)

        use_case = st.selectbox("Template", list(USE_CASE_TEMPLATES.keys()), index=0, key="ctrl_use_case")
        template = USE_CASE_TEMPLATES[use_case]

        event_name = st.text_input(
            "Event Name",
            f"{use_case} Evaluation" if use_case != "Neutral Template" else "Sourcing Evaluation",
            key="ctrl_event_name",
        )
        # Category — derived from Intake, shown read-only so there's one source of truth
        _auto_cat_key   = st.session_state.get("selected_parent_cat", list(CATEGORY_TAXONOMY.keys())[0])
        _auto_cat_label = CATEGORY_TAXONOMY.get(_auto_cat_key, {}).get("label", template["category"])
        category        = _auto_cat_label if _auto_cat_label else template["category"]
        _auto_sub_name  = st.session_state.get("selected_sub_name", "")
        st.markdown(
            f'<div class="ctrl-readout">'
            f'<div class="ctrl-readout-label">Category</div>'
            f'<div class="ctrl-readout-value">{html.escape(str(category))}</div>'
            f'<div class="ctrl-readout-sub">↑ Intake · {html.escape(_auto_sub_name[:30]) if _auto_sub_name else "—"}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # Kraljic — sync from subcategory default unless user has manually overridden
        _sub_kraljic_default = None
        _sub_key = st.session_state.get("selected_sub_name", "")
        if _sub_key:
            _parent_k = st.session_state.get("selected_parent_cat", list(CATEGORY_TAXONOMY.keys())[0])
            _sub_list = CATEGORY_TAXONOMY.get(_parent_k, {}).get("subcategories", [])
            _sub_obj  = next((s for s in _sub_list if s["name"] == _sub_key), None)
            if _sub_obj:
                _sub_kraljic_default = _sub_obj.get("kraljic_default")
        # Use subcategory default unless user has touched the Kraljic control
        _last_sub = st.session_state.get("_last_synced_sub", "")
        if _sub_key and _sub_key != _last_sub and _sub_kraljic_default:
            st.session_state["ctrl_kraljic"] = _sub_kraljic_default
            st.session_state["_last_synced_sub"] = _sub_key

        kraljic_options = ["Strategic", "Leverage", "Bottleneck", "Non-Critical"]
        _default_kq = st.session_state.get("ctrl_kraljic", template["kraljic"])
        _kq_idx = kraljic_options.index(_default_kq) if _default_kq in kraljic_options else 0
        kraljic = st.selectbox("Kraljic", kraljic_options, index=_kq_idx, key="ctrl_kraljic")

        # Show sync badge when Kraljic matches subcategory default
        if _sub_kraljic_default and _sub_kraljic_default == kraljic:
            st.markdown(
                f'<div class="sync-badge-green">✓ Matches {html.escape(_sub_key[:20]) if _sub_key else ""}… default</div>',
                unsafe_allow_html=True,
            )
        elif _sub_kraljic_default and _sub_kraljic_default != kraljic:
            st.markdown(
                f'<div class="sync-badge-amber">⚡ Override — default: {html.escape(str(_sub_kraljic_default))}</div>',
                unsafe_allow_html=True,
            )

        st.markdown('<div class="ctrl-section">Scope</div>', unsafe_allow_html=True)
        # Apply any pending bump from discovery add-to-slot (must happen before widget renders)
        _bump = st.session_state.pop("_ctrl_suppliers_bump", None)
        if _bump is not None:
            st.session_state["ctrl_suppliers"] = _bump
        num_suppliers    = st.slider("Suppliers", 2, 8, 3, key="ctrl_suppliers")
        num_stakeholders = st.slider("Stakeholders", 2, 8, 4, key="ctrl_stakeholders")

        st.markdown('<div class="ctrl-section">Weighting</div>', unsafe_allow_html=True)

        _wt_rec = get_subcategory_weights(_wt_sub, template["kraljic"])

        use_recommended = st.toggle(
            "Smart weights",
            value=st.session_state.get("use_rec_weights", False),
            key="use_rec_weights",
            help=f"Recommended for '{_wt_sub}' under {template['kraljic']}",
        )

        _tmpl_weights = template.get("weights", {})
        if use_recommended:
            st.caption(f"✓ {_wt_sub[:22]}...")
            weights_raw = {dim: _wt_rec.get(dim, _tmpl_weights.get(dim, 5)) for dim in DIMENSIONS}
            chip_html = "".join(
                f'<span style="display:inline-block;background:rgba(96,165,250,0.08);border:1px solid rgba(96,165,250,0.18);'
                f'border-radius:999px;padding:0.08rem 0.4rem;font-size:0.64rem;font-weight:600;color:#93C5FD;margin:0.1rem">'
                f'{d.split("/")[0].strip()[:6]}: {w}</span>'
                for d, w in weights_raw.items()
            )
            st.markdown(chip_html, unsafe_allow_html=True)
        else:
            weights_raw = {
                dim: st.slider(
                    dim.split("/")[0].strip()[:18],
                    1, 10,
                    _wt_rec.get(dim, _tmpl_weights.get(dim, 5)),
                    key=f"w_{dim}"
                )
                for dim in DIMENSIONS
            }

        st.markdown('<div style="margin-top:0.8rem;border-top:1px solid rgba(148,163,184,0.12);padding-top:0.6rem"></div>', unsafe_allow_html=True)

        kq_color = {"Strategic": "#F87171", "Leverage": "#4ADE80", "Bottleneck": "#FCD34D", "Non-Critical": "#94A3B8"}.get(kraljic, "#60A5FA")
        st.markdown(
            f'<div style="font-size:0.85rem;color:{kq_color};font-weight:700;text-align:center;'
            f'background:rgba(96,165,250,0.04);border-radius:6px;padding:0.35rem;margin-bottom:0.5rem">'
            f'{kraljic} posture</div>',
            unsafe_allow_html=True,
        )

        # ── Light / Dark Toggle ──────────────────────────────
        if st.button("← Back to Cover", use_container_width=True, key="ctrl_back"):
            st.session_state.entered_dashboard = False
            st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)

    # ── Computed variables (need ctrl_col values, used everywhere) ──
    weights       = normalize_weights(weights_raw)
    category_rule = classify_category(category)
    alpha_key     = get_alpha_vantage_key()
    safe_event    = sx(event_name)
    safe_category = sx(category)
    safe_use_case = sx(use_case)
    safe_kraljic  = sx(kraljic)
    kinfo         = KRALJIC_INFO[kraljic]

    # ════════════════════════════════════════
    # MAIN CONTENT COLUMN
    # ════════════════════════════════════════
    with main_col:
        # ── Mission Header Strip ─────────────────────────────
        _sub_display = st.session_state.get("selected_sub_name", "")
        _breadcrumb = " · ".join(filter(None, [
            use_case if use_case != "Neutral Template" else None,
            category or None,
            f"Kraljic: {kraljic}" if kraljic else None,
        ]))
        st.markdown(
            f'<div class="mission-header">'
            f'<div>'
            f'<div class="mission-breadcrumb">{_breadcrumb}</div>'
            f'<div class="mission-event">{html.escape(event_name)}</div>'
            f'<div style="font-size:0.58rem;color:rgba(148,163,184,0.5);letter-spacing:0.14em;text-transform:uppercase;margin-top:0.15rem">Sourcing Decision Framework</div>'
            f'</div>'
            f'<div class="mission-status"><span class="mission-dot"></span>EVALUATION ACTIVE</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── No-key warning banner ─────────────────────────────
        _render_no_key_banner()

        # ── Guided Workflow Progress Stepper ─────────────────
        _has_category  = bool(st.session_state.get("selected_parent_cat"))
        _has_suppliers = any(
            st.session_state.get(f"supplier_name_{k}") for k in range(10)
        )
        _has_scores = any(
            st.session_state.get(f"score_quality_{k}") for k in range(10)
        )
        _has_award = bool(st.session_state.get("award_decision"))

        _steps = [
            ("1", "Configure Intake",     _has_category,  "📋 Intake tab — set category, Kraljic, use case"),
            ("2", "Score Suppliers",      _has_suppliers, "👥 Suppliers tab — add and score each supplier"),
            ("3", "Build Strategy",       _has_scores,    "📄 Strategy tab — review playbook and timeline"),
            ("4", "Generate Award Brief", _has_scores,    "Decision Brief tab — generate executive recommendation"),
            ("5", "Export & Communicate", _has_award,     "📡 Comms & AI tab — export brief, send emails"),
        ]

        _step_html = '<div style="display:flex;gap:0;margin-bottom:1rem;border-radius:10px;overflow:hidden;border:1px solid rgba(96,165,250,0.12)">'
        for _si, (_snum, _slabel, _sdone, _stip) in enumerate(_steps):
            _s_bg  = "rgba(29,78,216,0.20)" if _sdone else "rgba(6,13,26,0.6)"
            _s_col = "#60A5FA" if _sdone else "#6B8BAE"
            _s_num_bg = "#2563EB" if _sdone else "rgba(96,165,250,0.12)"
            _s_num_col = "#fff" if _sdone else "#6B8BAE"
            _s_check = "✓" if _sdone else _snum
            _sep = "border-right:1px solid rgba(96,165,250,0.10);" if _si < len(_steps) - 1 else ""
            _step_html += (
                f'<div style="flex:1;padding:0.55rem 0.75rem;background:{_s_bg};{_sep}'
                f'display:flex;align-items:center;gap:0.5rem;min-width:0" title="{_stip}">'
                f'<div style="width:20px;height:20px;border-radius:50%;background:{_s_num_bg};'
                f'color:{_s_num_col};font-size:0.80rem;font-weight:700;display:flex;align-items:center;'
                f'justify-content:center;flex-shrink:0">{_s_check}</div>'
                f'<div style="font-size:0.72rem;font-weight:600;color:{_s_col};white-space:nowrap;'
                f'overflow:hidden;text-overflow:ellipsis">{_slabel}</div>'
                f'</div>'
            )
        _step_html += '</div>'
        st.markdown(_step_html, unsafe_allow_html=True)

        # ── AI Settings expander — fixed above tabs, not mixed into tab row ─
        _key_configured = bool(_get_api_key())
        _exp_label = (
            "⚙ AI Settings  ✓ configured"
            if _key_configured else
            "⚙ AI Settings  ⚠ No key set — click here to configure"
        )
        with st.expander(_exp_label, expanded=not _key_configured):
            _render_settings_tab()

        (tab_overview, tab_intake, tab_market, tab_suppliers, tab_stakeholders,
         tab_strategy, tab_negotiate, tab_briefing, tab_comms, tab_spend) = st.tabs(
            ["Overview", "Intake", "Market Intelligence", "Supplier Evaluation", "Stakeholders",
             "Category Strategy", "Negotiation", "Decision Brief", "Stakeholder Comms", "Spend & Risk"]
        )

    suppliers: List[Dict] = []

    # ── OVERVIEW TAB ───────────────────────────────────────
    with tab_overview:
        st.markdown("### 📊 Evaluation Overview")
        
        # Key metrics cards
        m_col1, m_col2, m_col3, m_col4 = st.columns(4)
        supplier_count = st.session_state.get("ctrl_suppliers", 3)
        for _mc, _mlabel, _mval, _mcolor in [
            (m_col1, "Category",       safe_category or "—",  "#60A5FA"),
            (m_col2, "Kraljic Posture", safe_kraljic or "—",   "#4ADE80"),
            (m_col3, "Suppliers",      str(supplier_count),   "#FB923C"),
            (m_col4, "Event",          safe_event,            "#C084FC"),
        ]:
            with _mc:
                st.markdown(
                    f'<div style="background:{_mcolor}18;border:1px solid {_mcolor}44;border-radius:10px;'
                    f'padding:1.2rem;text-align:center">'
                    f'<div style="font-size:0.75rem;color:{_mcolor};text-transform:uppercase;letter-spacing:0.08em;margin-bottom:0.3rem">{_mlabel}</div>'
                    f'<div style="font-size:1.2rem;font-weight:700;color:#F1F5F9">{_mval}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("---")

        # ── Start Here: 4-step guide ──────────────────────────────────────
        st.markdown(
            '<div style="background:rgba(29,78,216,0.06);border:1px solid rgba(96,165,250,0.18);'
            'border-radius:12px;padding:1rem 1.2rem;margin-bottom:1rem">'
            '<div style="font-size:0.72rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.12em;'
            'font-weight:700;margin-bottom:0.3rem">New here? Follow this path →</div>'
            '<div style="font-size:0.82rem;color:#94A3B8;margin-bottom:0.8rem">'
            'Intake → Market Intelligence → Supplier Evaluation → Decision Brief</div>'
            '<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:0.8rem">'
            '<div style="background:rgba(96,165,250,0.07);border-radius:8px;padding:0.75rem 0.9rem">'
            '<div style="font-size:0.68rem;color:#60A5FA;font-family:monospace;font-weight:700;margin-bottom:0.3rem">① INTAKE</div>'
            '<div style="font-size:0.8rem;color:#CBD5E1;line-height:1.5">Set category, Kraljic posture, spend, and timeline. '
            'Every recommendation downstream reads from this.</div>'
            '</div>'
            '<div style="background:rgba(251,191,36,0.07);border-radius:8px;padding:0.75rem 0.9rem">'
            '<div style="font-size:0.68rem;color:#FCD34D;font-family:monospace;font-weight:700;margin-bottom:0.3rem">② MARKET INTELLIGENCE</div>'
            '<div style="font-size:0.8rem;color:#CBD5E1;line-height:1.5">Discover suppliers for your subcategory. '
            'Run live scoring with an API key, or use illustrative market leaders.</div>'
            '</div>'
            '<div style="background:rgba(74,222,128,0.07);border-radius:8px;padding:0.75rem 0.9rem">'
            '<div style="font-size:0.68rem;color:#4ADE80;font-family:monospace;font-weight:700;margin-bottom:0.3rem">③ SUPPLIER EVALUATION</div>'
            '<div style="font-size:0.8rem;color:#CBD5E1;line-height:1.5">Score suppliers across 10 dimensions with pricing and financial health. '
            'Adjust weights to reflect your priorities.</div>'
            '</div>'
            '<div style="background:rgba(168,85,247,0.07);border-radius:8px;padding:0.75rem 0.9rem">'
            '<div style="font-size:0.68rem;color:#C084FC;font-family:monospace;font-weight:700;margin-bottom:0.3rem">④ DECISION BRIEF</div>'
            '<div style="font-size:0.8rem;color:#CBD5E1;line-height:1.5">Generate your executive memo, CFO challenge Q&amp;A, '
            'risk flags, and 90-day action plan. Export to Excel or HTML.</div>'
            '</div>'
            '</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        
        # ── Demo Use Case Loader ──────────────────────────────────────────
        if not st.session_state.get("_sample_data_loaded"):
            st.markdown(
                '<div style="background:rgba(74,222,128,0.05);border:1px solid rgba(74,222,128,0.18);'
                'border-radius:10px;padding:0.8rem 1rem;margin-bottom:0.8rem;display:flex;align-items:center;gap:1rem">'
                '<div style="flex:1">'
                '<div style="font-size:0.75rem;color:#4ADE80;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.2rem">Try a Live Demo</div>'
                '<div style="font-size:0.82rem;color:#CBD5E1">Load a pre-built HRIS platform selection example '
                '(Workday vs UKG Pro vs SAP SuccessFactors) to explore ProcureIQ with real data.</div>'
                '</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            if st.button("Load HRIS Demo Scenario", key="load_hris_demo_btn"):
                _load_hris_sample_data()
                st.rerun()

        # ── ACTIVE EVENT SNAPSHOT ────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Active Event Snapshot")

        _active_event = st.session_state.get("ctrl_event", event_name) or "Untitled Event"
        _active_cat   = st.session_state.get("ctrl_category", category) or "—"
        _active_kq    = st.session_state.get("ctrl_kraljic", kraljic) or "—"
        _active_sups  = st.session_state.get("ctrl_suppliers", num_suppliers)

        # Determine if we have a scored leader yet
        _overview_leader_name  = leader["Supplier"] if ranked else "—"
        _overview_leader_score = leader["Weighted Score"] if ranked else None
        _overview_runner_score = runner_up["Weighted Score"] if runner_up else None
        _score_gap = round(_overview_leader_score - _overview_runner_score, 1) if (_overview_leader_score and _overview_runner_score) else None
        _top_risk  = leader["Financial Risk Label"] if ranked else "—"

        _kq_colors = {"Strategic": "#F87171", "Leverage": "#4ADE80", "Bottleneck": "#FCD34D", "Non-Critical": "#94A3B8"}
        _kc = _kq_colors.get(_active_kq, "#60A5FA")

        # Five-metric executive strip
        ov1, ov2, ov3, ov4, ov5 = st.columns(5)
        for _col, _label, _val, _color in [
            (ov1, "Recommended",   _overview_leader_name,
             "#F1F5F9"),
            (ov2, "Score",         f"{_overview_leader_score}/100" if _overview_leader_score else "—",
             "#60A5FA"),
            (ov3, "Score Gap",     f"+{_score_gap}" if _score_gap else "—",
             "#4ADE80" if (_score_gap and _score_gap > 5) else "#F59E0B"),
            (ov4, "Financial Risk", _top_risk,
             {"LOW": "#4ADE80", "MEDIUM": "#F59E0B", "HIGH": "#F87171"}.get(_top_risk, "#94A3B8")),
            (ov5, "Posture",       _active_kq, _kc),
        ]:
            _col.markdown(
                f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);border-radius:10px;'
                f'padding:0.9rem 0.7rem;text-align:center">'
                f'<div style="font-size:0.78rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.3rem">{_label}</div>'
                f'<div style="font-size:1rem;font-weight:700;color:{_color}">{html.escape(str(_val))}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

        # ── Score Bar Race ───────────────────────────────────────────
        if ranked:
            # Hero stat strip — make the numbers speak
            _leader_r   = ranked[0]
            _gap_pts    = round(_leader_r["Weighted Score"] - ranked[1]["Weighted Score"], 1) if len(ranked) > 1 else None
            _gap_color  = "#4ADE80" if (_gap_pts or 0) >= 5 else "#FCD34D" if (_gap_pts or 0) >= 2 else "#F87171"
            _hero_html  = (
                '<div style="display:grid;grid-template-columns:2fr 1fr 1fr;gap:1rem;margin-bottom:1.2rem">'
                f'<div style="background:rgba(29,78,216,0.12);border:1.5px solid rgba(59,130,246,0.30);'
                f'border-radius:12px;padding:1rem 1.4rem">'
                f'<div style="font-size:0.78rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.14em;font-weight:700">Recommended Award</div>'
                f'<div style="font-size:1.9rem;font-weight:800;color:#F1F5F9;margin-top:0.2rem;line-height:1">'
                f'{html.escape(_leader_r["Supplier"])}</div>'
                f'<div style="font-size:0.85rem;color:#A8BEDC;margin-top:0.3rem">'
                f'{_leader_r["Weighted Score"]} / 100 composite score</div>'
                f'</div>'
                + (
                    f'<div style="background:rgba(74,222,128,0.06);border:1px solid rgba(74,222,128,0.20);'
                    f'border-radius:12px;padding:1rem;text-align:center">'
                    f'<div style="font-size:0.78rem;color:#4ADE80;text-transform:uppercase;letter-spacing:0.12em;font-weight:700">Lead Margin</div>'
                    f'<div style="font-size:2.4rem;font-weight:800;color:{_gap_color};font-family:monospace;line-height:1;margin-top:0.2rem">'
                    f'+{_gap_pts}</div>'
                    f'<div style="font-size:0.78rem;color:#A8BEDC">pts over #{2}</div>'
                    f'</div>'
                    if _gap_pts is not None else ""
                ) +
                f'<div style="background:rgba(96,165,250,0.05);border:1px solid rgba(96,165,250,0.15);'
                f'border-radius:12px;padding:1rem;text-align:center">'
                f'<div style="font-size:0.78rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.12em;font-weight:700">Suppliers Evaluated</div>'
                f'<div style="font-size:2.4rem;font-weight:800;color:#F1F5F9;font-family:monospace;line-height:1;margin-top:0.2rem">'
                f'{len(ranked)}</div>'
                f'<div style="font-size:0.78rem;color:#A8BEDC">{html.escape(selected_sub_name[:28])}</div>'
                f'</div>'
                '</div>'
            )
            st.markdown(_hero_html, unsafe_allow_html=True)

            st.markdown("<div style='height:0.4rem'></div>", unsafe_allow_html=True)
            _race_rows_html = ""
            _race_max = ranked[0]["Weighted Score"] if ranked else 100
            for _ri, _rs in enumerate(ranked):
                _is_leader = _ri == 0
                _delay_ms  = _ri * 80
                _fill_pct  = round(_rs["Weighted Score"] / max(_race_max, 1) * 100, 1)
                _row_cls   = "race-row race-row-winner" if _is_leader else "race-row"
                _fill_cls  = "race-fill race-fill-leader" if _is_leader else "race-fill"
                _score_cls = "race-score race-score-leader" if _is_leader else "race-score race-score-other"
                _badge     = '<span class="race-badge-rec">REC</span>' if _is_leader else ""
                _race_rows_html += (
                    f'<div class="{_row_cls}" style="animation-delay:{_delay_ms}ms">'
                    f'<span class="race-name">{html.escape(_rs["Supplier"])}</span>'
                    f'<div class="race-track">'
                    f'<div class="{_fill_cls}" style="width:{_fill_pct}%;animation-delay:{_delay_ms}ms"></div>'
                    f'</div>'
                    f'<span class="{_score_cls}">{_rs["Weighted Score"]}</span>'
                    f'{_badge}'
                    f'</div>'
                )
            st.markdown(
                f'<div style="background:rgba(4,9,15,0.55);border:1px solid rgba(96,165,250,0.08);'
                f'border-radius:14px;padding:1rem 1.2rem;margin-bottom:0.5rem">'
                f'<div style="font-family:var(--mono);font-size:0.78rem;letter-spacing:0.2em;text-transform:uppercase;'
                f'color:#A8BEDC;margin-bottom:0.8rem">Score Ranking — {html.escape(event_name)}</div>'
                f'{_race_rows_html}'
                f'</div>',
                unsafe_allow_html=True,
            )

        # ── Portfolio — Save & History ───────────────────────────────
        st.markdown("<div style='height:0.6rem'></div>", unsafe_allow_html=True)

        _pf_col1, _pf_col2 = st.columns([1, 2])
        with _pf_col1:
            if st.button("💾 Save Event to Portfolio", key="save_portfolio_btn",
                         help="Saves a snapshot of the current evaluation state to your portfolio history."):
                try:
                    _db = get_database()
                    _snap_id = generate_id("EVT-", 6)
                    _snap_scores = {s["Supplier"]: s["Weighted Score"] for s in ranked} if ranked else {}
                    _snap_rec = _overview_leader_name
                    # Store event metadata in config table for portfolio display
                    _db.store_config(f"portfolio_{_snap_id}", {
                        "event_id": _snap_id,
                        "event_name": _active_event,
                        "category": _active_cat,
                        "subcategory": selected_sub_name,
                        "kraljic": _active_kq,
                        "recommendation": _snap_rec,
                        "score": _overview_leader_score,
                        "score_gap": _score_gap,
                        "suppliers": _active_sups,
                        "saved_at": time.time(),
                        "fin_source": st.session_state.get("_piq_save_fin_source", ""),
                        "edgar_period_end": st.session_state.get("_piq_save_edgar_period", ""),
                        "fin_score": st.session_state.get("_piq_save_fin_score", None),
                        "fin_risk": st.session_state.get("_piq_save_fin_risk", ""),
                        "defensibility": st.session_state.get("_piq_save_defensibility", ""),
                        "high_risk_count": st.session_state.get("_piq_save_high_risk_count", 0),
                    })
                    st.success(f"Saved as {_snap_id}")
                except Exception as _pe:
                    st.error(f"Save failed: {_pe}")

        # Show portfolio history from DB
        with _pf_col2:
            try:
                _db = get_database()
                _all_config = {r["key"]: r["value"] for r in _db.get_portfolio_events(limit=8)}

                if _all_config:
                    st.markdown(
                        '<div style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.4rem">Recent Portfolio Events</div>',
                        unsafe_allow_html=True,
                    )
                    for _pk, _pv in list(_all_config.items())[:5]:
                        _pname = html.escape(str(_pv.get("event_name", "Untitled")))
                        _pcat  = html.escape(str(_pv.get("category", "—")))
                        _pkq   = html.escape(str(_pv.get("kraljic", "—")))
                        _prec  = html.escape(str(_pv.get("recommendation", "—")))
                        _psc   = _pv.get("score")
                        _pid   = html.escape(str(_pv.get("event_id", _pk)))
                        _pkc   = _kq_colors.get(_pv.get("kraljic", ""), "#94A3B8")
                        st.markdown(
                            f'<div style="display:flex;align-items:center;gap:0.6rem;background:#060D1A;border:1px solid rgba(148,163,184,0.1);'
                            f'border-radius:8px;padding:0.5rem 0.8rem;margin-bottom:0.3rem">'
                            f'<div style="width:8px;height:8px;border-radius:50%;background:{_pkc};flex-shrink:0"></div>'
                            f'<div style="flex:1;min-width:0">'
                            f'<span style="font-weight:600;font-size:0.82rem;color:#E2E8F0">{_pname}</span>'
                            f'<span style="font-size:0.85rem;color:#D0E0EF;margin-left:0.5rem">{_pcat} · {_pkq}</span><br/>'
                            f'<span style="font-size:0.85rem;color:#C4D3E8">Rec: {_prec}</span>'
                            f'</div>'
                            f'<div style="font-size:0.9rem;font-weight:700;color:#60A5FA">{f"{_psc}/100" if _psc else "—"}</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                else:
                    st.caption("No saved events yet. Use 'Save Event to Portfolio' to track multiple sourcing events.")
            except Exception:
                st.caption("Portfolio history unavailable.")

        # ── SESSION PERSISTENCE ───────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 💾 Session Persistence")
        st.markdown(
            '<p class="muted">Save your full evaluation state — suppliers, scores, intake answers — and reload it in any future browser session.</p>',
            unsafe_allow_html=True,
        )

        _sess_col1, _sess_col2 = st.columns([1, 2])
        with _sess_col1:
            _sess_label = st.text_input(
                "Session label (optional)",
                placeholder=f"{_active_event[:30] or 'My Session'}",
                key="sess_save_label",
            )
            if st.button("💾 Save Full Session", key="sess_save_btn", help="Saves all supplier data, scores, and intake answers to SQLite."):
                try:
                    import json as _sj, sqlite3 as _sq3, time as _t
                    _snap_id = generate_id("SES-", 6)
                    _state_to_save = {}
                    for _sk, _sv in st.session_state.items():
                        if any(_sk.startswith(_pfx) for _pfx in _SESSION_KEYS_PREFIX) or _sk in _SESSION_KEYS_PREFIX:
                            try:
                                _sj.dumps(_sv)  # Only save JSON-serialisable values
                                _state_to_save[_sk] = _sv
                            except (TypeError, ValueError):
                                pass
                    _meta = {
                        "session_id": _snap_id,
                        "label": _sess_label or _active_event or "Untitled",
                        "category": _active_cat,
                        "subcategory": selected_sub_name,
                        "suppliers": _active_sups,
                        "recommendation": _overview_leader_name,
                        "saved_at": _t.time(),
                        "state": _state_to_save,
                    }
                    get_database().store_config(f"session_snap_{_snap_id}", _meta)
                    st.success(f"Saved — ID: **{_snap_id}**. Use this ID to restore in any future session.")
                except Exception as _se:
                    st.error(f"Save failed: {_se}")

        with _sess_col2:
            # List recent saved sessions
            try:
                import datetime as _dt_s
                _sess_raw = get_database().get_session_snapshots(limit=6)
                _sess_rows = [(_r["key"], _r["value"]) for _r in _sess_raw]

                if _sess_rows:
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.5rem">Saved Sessions</div>',
                        unsafe_allow_html=True,
                    )
                    for _srow in _sess_rows:
                        _smeta = _srow[1] if isinstance(_srow[1], dict) else {}
                        _slabel   = html.escape(str(_smeta.get("label", "Untitled")))
                        _scat     = html.escape(str(_smeta.get("category", "—")))
                        _srec     = html.escape(str(_smeta.get("recommendation", "—")))
                        _ssid     = str(_smeta.get("session_id", _srow[0].replace("session_snap_", "")))
                        _sts = _smeta.get("saved_at", 0)
                        _sdate = _dt_s.datetime.fromtimestamp(_sts).strftime("%b %d %H:%M") if _sts else "—"

                        _r1, _r2 = st.columns([3, 1])
                        with _r1:
                            st.markdown(
                                f'<div style="background:#060D1A;border:1px solid rgba(148,163,184,0.1);'
                                f'border-radius:8px;padding:0.45rem 0.75rem;margin-bottom:0.25rem">'
                                f'<span style="font-weight:600;font-size:0.85rem;color:#E2E8F0">{_slabel}</span> '
                                f'<span style="font-size:0.78rem;color:#94A3B8">{_scat} · {_sdate}</span><br/>'
                                f'<span style="font-size:0.82rem;color:#60A5FA">Rec: {_srec}</span> '
                                f'<span style="font-size:0.78rem;color:#64748B">ID: {_ssid}</span>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )
                        with _r2:
                            if st.button("Restore", key=f"sess_restore_{_ssid}",
                                         help=f"Load session {_ssid}"):
                                st.session_state["_piq_restore_session_id"] = _ssid
                                st.rerun()
                else:
                    st.caption("No saved sessions yet. Click 'Save Full Session' to create your first checkpoint.")
            except Exception as _le:
                st.caption(f"Session history unavailable: {_le}")

        st.markdown("---")

        # ── Methodology & Limitations expanders ───────────────────────────
        _meth_col, _lim_col = st.columns(2)
        with _meth_col:
            with st.expander("Scoring Methodology", expanded=False):
                st.markdown("""
**How suppliers are scored**

ProcureIQ uses an 8-dimension weighted scoring model based on the Kraljic matrix posture you set in Intake.

| Dimension | What it measures |
|---|---|
| Price / TCO | Quoted price relative to other suppliers, normalized 0-100 |
| SLA Strength | Contract performance commitments (Strong/Moderate/Weak) |
| Execution Risk | Operational delivery risk (Low/Medium/High) |
| Stakeholder Confidence | Internal champion support (1-5 scale) |
| Strategic Alignment | Fit with company direction (1-5 scale) |
| Innovation Capacity | R&D and co-development ability (1-5 scale) |
| Relationship Depth | Executive and operational relationship quality (1-5 scale) |
| Commercial Flexibility | Contract and deal structure flexibility (1-5 scale) |

**Weights:** Dimension weights shift automatically based on Kraljic posture. Strategic categories weight Execution Risk and SLA higher; Leverage categories weight Price/TCO higher.

**Financial Health** is calculated separately from the 8 dimensions using business age, ownership structure, revenue trajectory, M&A activity, payment terms, and workforce changes.

**Confidence labels** (HIGH/MEDIUM/LOW) reflect how many dimensions have been scored above the default midpoint — not the absolute score.
                """)

        with _lim_col:
            with st.expander("Known Limitations", expanded=False):
                st.markdown("""
**What ProcureIQ does not do**

- **No independent supplier verification.** Scores are based entirely on data you enter. The tool cannot confirm whether a supplier's stated capabilities are accurate.
- **No live pricing data.** Quoted prices are manually entered. Market benchmark pricing is sourced from BLS PPI indices and ISM estimates — directional only.
- **Sanctions screening is not a compliance certification.** The OFAC name match uses the Treasury SDN list at the time of query. A no-match result does not replace a formal compliance review.
- **Financial health scores are estimates.** Calculated from qualitative fields you enter; not from audited financials or credit bureau data.
- **AI-generated text requires review.** All negotiation prompts, strategy text, and briefing content must be reviewed by a qualified procurement professional before use.
- **Not a contract management system.** Renewal alerts are based on contract dates you enter; the tool does not integrate with your ERP, CLM, or P2P system.
- **Supplier discovery is research-grade.** Discovery agent outputs identify candidate suppliers — they are not due-diligence reports and do not constitute vendor approval.
                """)

        st.markdown("---")
        st.markdown("#### Key Features")

        features_col1, features_col2 = st.columns(2)
        with features_col1:
            st.markdown("""
            - **Dynamic Kraljic Weighting** — Dimension weights adjust based on your posture
            - **RAQSCI Contract Guidance** — Industry-standard contract risk lens per category
            - **Real-time Market Data** — SEC filings, BLS pricing, supplier financial signals
            - **90-Day Action Plan** — Tailored post-award execution roadmap
            """)

        with features_col2:
            st.markdown("""
            - **AI Negotiation Prompts** — Claude-powered talking points with guardrails
            - **Savings Tracking** — Baseline selector and savings calculation
            - **Contract Renewal Calendar** — Expiration alerts and timelines
            - **Executive One-Pager** — Print-ready HTML export in 30 seconds
            """)

    # ── INTAKE TAB ─────────────────────────────────────────
    with tab_intake:
        # ── AGENT #25: Conversational Intake toggle ───────────────────────
        _intake_mode = st.radio(
            "Intake mode",
            ["📋 Form", "💬 Chat with AI"],
            horizontal=True,
            label_visibility="collapsed",
            key="intake_mode_toggle",
        )
        st.markdown("---")

        if _intake_mode == "💬 Chat with AI":
            st.markdown("### 💬 Quick Start — Tell Me What You're Sourcing")
            st.markdown(
                '<p class="muted">Describe your procurement need in plain language. '
                'The AI will ask a few smart questions and auto-populate the intake form. '
                'Switch to Form mode anytime to edit manually.</p>',
                unsafe_allow_html=True,
            )

            if "intake_conversation" not in st.session_state:
                st.session_state["intake_conversation"] = []

            # Display conversation history
            for _msg in st.session_state["intake_conversation"]:
                if _msg["role"] == "user":
                    with st.chat_message("user"):
                        st.markdown(_msg["content"])
                else:
                    # Strip session_update JSON tags from display
                    import re as _re
                    _display = _re.sub(r"<session_update>.*?</session_update>", "",
                                       _msg["content"], flags=_re.DOTALL).strip()
                    with st.chat_message("assistant", avatar="🤖"):
                        st.markdown(_display)

            # Starter prompt if empty
            if not st.session_state["intake_conversation"]:
                with st.chat_message("assistant", avatar="🤖"):
                    st.markdown(
                        "Hi! Tell me what you're looking to source. For example: "
                        "*'We need to source IT managed services for our Chicago office, "
                        "roughly $2M annually'* — and I'll take it from there."
                    )

            _user_input = st.chat_input("Describe your procurement need…", key="intake_chat_input")

            if _user_input:
                existing = get_intake_session_values(dict(st.session_state))
                result = run_intake_agent(
                    conversation=st.session_state["intake_conversation"],
                    user_message=_user_input,
                    api_key=_get_api_key(),
                    existing_session=existing,
                    provider=_get_provider(),
                )
                st.session_state["intake_conversation"] = result["conversation"]

                # Apply session updates
                for _k, _v in result.get("session_updates", {}).items():
                    if _v is not None:
                        st.session_state[_k] = _v

                if result.get("complete"):
                    st.success(
                        "✅ Intake complete — form fields populated. "
                        "Switch to **Form** mode to review and adjust, or continue to Suppliers."
                    )
                st.rerun()

            if st.session_state["intake_conversation"] and st.button(
                "🗑️ Clear Chat", key="clear_intake_chat"
            ):
                st.session_state["intake_conversation"] = []
                st.rerun()

            # Still render the form below (read-only summary) when chat is active
            st.markdown("---")
            st.markdown("#### Current Intake Values (auto-populated from chat)")

        else:
            st.markdown("### Category & Subcategory Selection")
            st.markdown('<p class="muted">Select your category and subcategory first. This drives the contract guidance, stakeholder requirements, auction recommendation, and risk flags throughout the tool.</p>', unsafe_allow_html=True)

        cat_col, sub_col = st.columns([1, 1.4])

        with cat_col:
            st.markdown("#### Category")
            selected_parent = st.selectbox(
                "Parent Category",
                list(CATEGORY_TAXONOMY.keys()),
                format_func=lambda k: f"{CATEGORY_TAXONOMY[k]['icon']} {CATEGORY_TAXONOMY[k]['label']}",
                key="selected_parent_cat",
            )

        parent_data = CATEGORY_TAXONOMY[selected_parent]
        sub_names = [s["name"] for s in parent_data["subcategories"]]

        with sub_col:
            st.markdown("#### Subcategory")
            selected_sub_name = st.selectbox(
                "Subcategory",
                sub_names,
                key="selected_sub_name",
            )

        selected_sub = next(s for s in parent_data["subcategories"] if s["name"] == selected_sub_name)

        # ── Kraljic sync: sidebar override vs subcategory default ──
        sub_default_kraljic = selected_sub["kraljic_default"]
        user_kraljic = kraljic
        kraljic_match = sub_default_kraljic == user_kraljic
        kraljic_color_map = {"Strategic": "#F87171", "Leverage": "#4ADE80", "Bottleneck": "#FCD34D", "Non-Critical": "#94A3B8"}
        sc_color  = kraljic_color_map.get(sub_default_kraljic, "#60A5FA")
        usr_color = kraljic_color_map.get(user_kraljic, "#60A5FA")

        # ── Card: use native Streamlit components to avoid HTML rendering issues ──
        st.markdown(
            f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.22);border-radius:14px;padding:1.2rem;margin-top:0.8rem">',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div style="font-size:0.82rem;text-transform:uppercase;letter-spacing:0.14em;color:#D0E0EF;font-weight:700;margin-bottom:0.6rem">Subcategory Intelligence — {sx(selected_sub["name"])}</div>',
            unsafe_allow_html=True,
        )

        if not kraljic_match:
            st.markdown(
                f'<div style="background:rgba(252,211,77,0.08);border-left:3px solid #FCD34D;border-radius:0 8px 8px 0;padding:0.6rem 0.85rem;margin-bottom:0.7rem;font-size:0.82rem;color:#E2E8F0">'
                f'<strong style="color:#FCD34D">⚡ Kraljic Override Active</strong> &nbsp;—&nbsp; '
                f'Subcategory default is <strong style="color:{sc_color}">{sx(sub_default_kraljic)}</strong> '
                f'but sidebar is set to <strong style="color:{usr_color}">{sx(user_kraljic)}</strong>. '
                f'All scoring uses <strong style="color:{usr_color}">{sx(user_kraljic)}</strong>.'
                f'</div>',
                unsafe_allow_html=True,
            )

        # Metadata row — simple columns avoids CSS var issues entirely
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown(f'<div style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.08em">Contract Type</div><div style="font-weight:700;color:#E2E8F0;font-size:0.9rem">{sx(selected_sub["contract_type"])}</div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.08em">Typical Event</div><div style="font-weight:700;color:#60A5FA;font-size:0.9rem">{sx(selected_sub["auction"])}</div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.08em">Sub Default</div><div style="font-weight:700;color:{sc_color};font-size:0.9rem">{sx(sub_default_kraljic)}</div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.08em">Your Setting</div><div style="font-weight:700;color:{usr_color};font-size:0.9rem">{sx(user_kraljic)}</div>', unsafe_allow_html=True)
        with c5:
            sw_color = {"Very High": "#F87171", "High": "#FCD34D", "Medium": "#60A5FA", "Low": "#4ADE80"}.get(selected_sub["switching_cost"], "#94A3B8")
            st.markdown(f'<div style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.08em">Switching Cost</div><div style="font-weight:700;color:{sw_color};font-size:0.9rem">{sx(selected_sub["switching_cost"])}</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:0.7rem'></div>", unsafe_allow_html=True)

        # Risk and notes — simple bordered boxes with hardcoded colors
        st.markdown(
            f'<div style="background:rgba(252,211,77,0.06);border-left:3px solid #FCD34D;border-radius:0 8px 8px 0;padding:0.7rem 1rem;margin-bottom:0.5rem;font-size:0.87rem;color:#E2E8F0">'
            f'<strong style="color:#FCD34D">Key Risks:</strong> {sx(selected_sub["key_risks"])}</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div style="background:rgba(96,165,250,0.06);border-left:3px solid #60A5FA;border-radius:0 8px 8px 0;padding:0.7rem 1rem;margin-bottom:0.5rem;font-size:0.87rem;color:#E2E8F0">'
            f'<strong style="color:#60A5FA">Expert Note:</strong> {sx(selected_sub["notes"])}</div>',
            unsafe_allow_html=True,
        )

        # Stakeholder chips
        chips_html = "".join(
            f'<span style="display:inline-block;background:#1A2A45;border:1px solid rgba(148,163,184,0.22);color:#C4D3E8;border-radius:999px;padding:0.22rem 0.62rem;margin:0.2rem 0.2rem 0.2rem 0;font-size:0.85rem;font-weight:600">👤 {sx(s)}</span>'
            for s in selected_sub["stakeholders"]
        )
        st.markdown(
            f'<div style="margin-top:0.3rem"><span style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.08em">Required Stakeholders</span><br>{chips_html}</div>',
            unsafe_allow_html=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

        # ── AGENT #1: Supplier Discovery — Procurement Decision Intelligence ───
        st.markdown("---")
        st.markdown("#### Supplier Discovery")
        st.markdown(
            '<p class="muted">Scores suppliers across 7 dimensions and classifies them into three tiers: '
            '<strong style="color:#4ADE80">Established</strong> (proven market leaders), '
            '<strong style="color:#60A5FA">Emerging</strong> (challengers worth qualifying), and '
            '<strong style="color:#A78BFA">Watchlist</strong> (disruptors to monitor). '
            'Works with or without an API key — known market leaders load instantly from the curated database.</p>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="background:rgba(96,165,250,0.05);border-left:3px solid rgba(96,165,250,0.3);'
            'border-radius:0 6px 6px 0;padding:0.4rem 0.7rem;font-size:0.78rem;color:#93C5FD;margin-bottom:0.6rem">'
            '<strong>Evidence tiers:</strong> Tier 1 = Gartner/Forrester/IDC/Everest recognition · '
            'Tier 2 = Crunchbase funding, LinkedIn growth, press · '
            'Tier 3 = Reviews, conference, job postings. '
            'Confidence label reflects source quality, not just score.'
            '</div>',
            unsafe_allow_html=True,
        )

        # Configuration row
        _disc_c1, _disc_c2, _disc_c3, _disc_c4 = st.columns([2, 1, 1, 1])
        with _disc_c1:
            _disc_geo = st.text_input(
                "Geography",
                value="United States",
                key="disc_geography",
                placeholder="e.g. North America, EU, APAC",
            )
        with _disc_c2:
            _disc_size = st.selectbox(
                "Buyer size",
                ["Enterprise", "Mid-market", "SMB"],
                key="disc_company_size",
            )
        with _disc_c3:
            _disc_risk = st.selectbox(
                "Risk tolerance",
                ["Low", "Medium", "High"],
                index=1,
                key="disc_risk_tolerance",
            )
        with _disc_c4:
            st.markdown("<div style='height:1.85rem'></div>", unsafe_allow_html=True)
            _run_discovery = st.button(
                "🚀 Run Discovery",
                key="run_discovery_btn",
                type="primary",
                use_container_width=True,
                help="Scores suppliers across 7 dimensions. 20–40 sec with API key; instant fallback without.",
            )

        _disc_goal = st.text_input(
            "Sourcing goal (optional — shapes scoring)",
            value="Cost optimization and risk reduction",
            key="disc_sourcing_goal",
            placeholder="e.g. Innovation partner, Cost reduction, Risk diversification, Dual-source strategy",
        )

        if _run_discovery:
            import hashlib as _disc_hash
            _disc_cat = safe_category or category
            _disc_cache_raw = f"{_disc_cat}|{selected_sub_name}|{_disc_geo}|{_disc_size}|{_disc_risk}|{_disc_goal}"
            _disc_cache_key = _disc_hash.md5(_disc_cache_raw.encode()).hexdigest()
            _disc_db = get_database()
            _cached = _disc_db.get_discovery_cache(_disc_cache_key)
            if _cached and not _cached.get("_fallback"):
                # Use cached live result — mark it so UI can show cache indicator
                _cached["_from_cache"] = True
                st.session_state["_discovery_result"] = _cached
                st.session_state["_discovery_cache_key"] = _disc_cache_key
            else:
                with st.spinner("Discovering and scoring suppliers · Analyzing leadership evidence · Building procurement brief …"):
                    _disc_result = run_supplier_discovery_agent(
                        category=_disc_cat,
                        subcategory=selected_sub_name,
                        geography=_disc_geo,
                        company_size=_disc_size,
                        risk_tolerance=_disc_risk,
                        sourcing_goal=_disc_goal,
                        api_key=_get_api_key(),
                    )
                # Cache only live (non-fallback) results to avoid persisting illustrative data
                if not _disc_result.get("_fallback") and not _disc_result.get("error"):
                    _disc_db.save_discovery_cache(_disc_cache_key, _disc_result)
                st.session_state["_discovery_result"] = _disc_result
                st.session_state["_discovery_cache_key"] = _disc_cache_key

        if st.session_state.get("_discovery_result"):
            _dr = st.session_state["_discovery_result"]

            if _dr.get("error") and not _dr.get("executive_shortlist"):
                st.error(f"Discovery error: {_dr['error']}")
            else:
                _shortlist  = _dr.get("executive_shortlist", [])
                _longlist   = _dr.get("expanded_longlist", [])
                _watchlist  = _dr.get("emerging_watchlist", [])
                _exec_sum   = _dr.get("executive_summary", {})
                _is_fallback  = _dr.get("_fallback", False)
                _from_cache   = _dr.get("_from_cache", False)
                _total        = len(_shortlist) + len(_longlist) + len(_watchlist)

                # Prominent illustrative data warning
                if _is_fallback:
                    st.warning(
                        "**Illustrative data only — these scores are NOT evidence-based.** "
                        "Rankings and scores shown here are static estimates derived from general industry "
                        "knowledge. They have NOT been validated against live data sources. "
                        "**Do not use these scores to support a sourcing decision.** "
                        "Add an API key in AI Settings and re-run Discovery to get live AI-scored intelligence.",
                        icon="⚠️",
                    )
                elif _from_cache:
                    st.info(
                        "Showing cached results from a previous live discovery run (up to 24 hours old). "
                        "Click **Run Discovery** again to refresh.",
                        icon="💾",
                    )

                # Header bar
                if _is_fallback:
                    _mode_badge = (
                        '<span style="background:rgba(251,146,60,0.2);border:1px solid rgba(251,146,60,0.5);'
                        'border-radius:6px;padding:0.2rem 0.6rem;font-size:0.72rem;color:#FB923C;font-weight:700">'
                        '⚠️ ILLUSTRATIVE — Static Knowledge Only</span>'
                    )
                elif _from_cache:
                    _mode_badge = (
                        '<span style="background:rgba(96,165,250,0.12);border:1px solid rgba(96,165,250,0.3);'
                        'border-radius:6px;padding:0.2rem 0.6rem;font-size:0.72rem;color:#60A5FA">'
                        '💾 Cached · Live AI Scored</span>'
                    )
                else:
                    _mode_badge = (
                        '<span style="background:rgba(74,222,128,0.12);border:1px solid rgba(74,222,128,0.3);'
                        'border-radius:6px;padding:0.2rem 0.6rem;font-size:0.72rem;color:#4ADE80">✓ Live AI Scored</span>'
                    )
                _header_border = "rgba(251,146,60,0.3)" if _is_fallback else "rgba(74,222,128,0.18)"
                _header_bg = "rgba(251,146,60,0.04)" if _is_fallback else "rgba(74,222,128,0.05)"
                st.markdown(
                    f'<div style="background:{_header_bg};border:1px solid {_header_border};'
                    f'border-radius:10px;padding:0.7rem 1rem;margin-bottom:1rem;display:flex;'
                    f'justify-content:space-between;align-items:center;flex-wrap:wrap;gap:0.5rem">'
                    f'<div><span style="color:#4ADE80;font-weight:700;font-size:0.95rem">'
                    f'{_total} suppliers loaded</span>'
                    f' · <span style="color:#C4D3E8;font-size:0.85rem">'
                    f'<span style="color:#4ADE80">{len(_shortlist)}</span> established'
                    f' · <span style="color:#60A5FA">{len(_longlist)}</span> emerging'
                    f' · <span style="color:#A78BFA">{len(_watchlist)}</span> watchlist</span></div>'
                    f'<div style="display:flex;gap:0.4rem;align-items:center">{_mode_badge}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                # ── Executive Summary ─────────────────────────────────
                if _exec_sum:
                    with st.expander("📊 Executive Briefing — VP/CPO Ready", expanded=True):
                        if _exec_sum.get("headline"):
                            st.markdown(
                                f'<div style="background:rgba(96,165,250,0.07);border:1px solid rgba(96,165,250,0.2);'
                                f'border-radius:10px;padding:0.9rem 1.1rem;margin-bottom:0.8rem;'
                                f'font-size:0.95rem;color:#E2E8F0;font-style:italic">'
                                f'{html.escape(_exec_sum["headline"])}</div>',
                                unsafe_allow_html=True,
                            )
                        _es_cols = st.columns(2)
                        _es_pairs = [
                            ("Top Recommendation", _exec_sum.get("top_recommendation", "—"), "#4ADE80"),
                            ("Key Tradeoffs", _exec_sum.get("key_tradeoffs", "—"), "#FCD34D"),
                            ("CFO Challenge", _exec_sum.get("cfo_challenge", "—"), "#F87171"),
                            ("CPO Challenge", _exec_sum.get("cpo_challenge", "—"), "#FB923C"),
                        ]
                        for _epi, (_lbl, _val, _col) in enumerate(_es_pairs):
                            with _es_cols[_epi % 2]:
                                st.markdown(
                                    f'<div style="background:#060D1A;border-left:3px solid {_col};'
                                    f'border-radius:0 8px 8px 0;padding:0.55rem 0.8rem;margin-bottom:0.5rem">'
                                    f'<div style="font-size:0.7rem;color:{_col};text-transform:uppercase;font-weight:700;letter-spacing:0.07em">{_lbl}</div>'
                                    f'<div style="font-size:0.85rem;color:#C4D3E8;margin-top:0.2rem;line-height:1.5">{html.escape(str(_val)[:200])}</div>'
                                    f'</div>',
                                    unsafe_allow_html=True,
                                )
                        if _exec_sum.get("recommended_next_step"):
                            st.markdown(
                                f'<div style="background:rgba(29,78,216,0.1);border:1px solid rgba(96,165,250,0.25);'
                                f'border-radius:8px;padding:0.6rem 1rem;font-size:0.88rem;color:#93C5FD">'
                                f'<strong>Next Step:</strong> {html.escape(_exec_sum["recommended_next_step"])}</div>',
                                unsafe_allow_html=True,
                            )

                # ── Helper: render one supplier card ─────────────────
                def _render_disc_supplier(
                    _si: int, _sup: Dict, _tier_offset: int, _show_add: bool = True
                ) -> None:
                    _sname   = str(_sup.get("name", ""))
                    _sesc    = html.escape(_sname)
                    _sown    = _sup.get("ownership", "Unknown")
                    _sloc    = html.escape(str(_sup.get("location", "—")))
                    _sconf   = _sup.get("confidence", "Medium")
                    _srec    = _sup.get("recommendation", "Monitor only")
                    _srflags = _sup.get("red_flags", [])
                    _sscores = _sup.get("scores", {})
                    _sover   = _sup.get("overall_score", 0)
                    _sfed    = _sup.get("federal_experience", False)
                    _samt_raw = _sup.get("federal_award_amount") or 0
                    try:
                        _samt = float(_samt_raw)
                    except (ValueError, TypeError):
                        import re as _re
                        _samt_m = _re.search(r"\d[\d,]*(?:\.\d+)?", str(_samt_raw))
                        _samt = float(_samt_m.group().replace(",", "")) if _samt_m else 0.0
                    _stier   = _sup.get("source_tier", "Tier 2")

                    _conf_col  = {"High": "#4ADE80", "Medium": "#60A5FA", "Low": "#94A3B8"}.get(_sconf, "#60A5FA")
                    _rec_col   = {
                        "Invite to RFP": "#4ADE80",
                        "Invite to RFI": "#60A5FA",
                        "Monitor only": "#FCD34D",
                        "Exclude for now": "#F87171",
                    }.get(_srec, "#94A3B8")
                    _own_color = {
                        "Public": "#60A5FA", "VC-backed": "#A78BFA", "PE-backed": "#FB923C",
                        "Private": "#94A3B8", "Bootstrapped": "#4ADE80",
                    }.get(_sown, "#94A3B8")
                    _score_color = "#4ADE80" if _sover >= 70 else ("#FCD34D" if _sover >= 50 else "#F87171")

                    _card_idx = _tier_offset + _si
                    _illus_suffix = "  ·  ⚠️ ILLUSTRATIVE" if _sup.get("_is_illustrative") else ""
                    with st.expander(
                        f"{_sesc}  ·  {_sown}  ·  Score {_sover}/100{_illus_suffix}",
                        expanded=False,
                    ):
                        _card_left, _card_right = st.columns([3, 1])
                        with _card_left:
                            # Header row
                            _illus_badge = (
                                '<span style="background:rgba(251,146,60,0.15);border:1px solid rgba(251,146,60,0.4);'
                                'border-radius:6px;padding:0.15rem 0.55rem;font-size:0.72rem;color:#FB923C;font-weight:700">'
                                '⚠️ ILLUSTRATIVE</span>'
                                if _sup.get("_is_illustrative") else ""
                            )
                            st.markdown(
                                f'<div style="display:flex;gap:0.4rem;flex-wrap:wrap;margin-bottom:0.7rem;align-items:center">'
                                f'<span style="background:{_own_color}22;border:1px solid {_own_color}55;'
                                f'border-radius:6px;padding:0.15rem 0.55rem;font-size:0.72rem;color:{_own_color}">{_sown}</span>'
                                f'<span style="background:{_conf_col}18;border:1px solid {_conf_col}44;'
                                f'border-radius:6px;padding:0.15rem 0.55rem;font-size:0.72rem;color:{_conf_col}">{_sconf} Confidence</span>'
                                f'<span style="background:{_rec_col}18;border:1px solid {_rec_col}44;'
                                f'border-radius:6px;padding:0.15rem 0.55rem;font-size:0.72rem;color:{_rec_col};font-weight:700">{_srec}</span>'
                                + (f'<span style="background:rgba(96,165,250,0.1);border:1px solid rgba(96,165,250,0.2);'
                                   f'border-radius:6px;padding:0.15rem 0.55rem;font-size:0.72rem;color:#60A5FA">{_stier}</span>')
                                + (f'<span style="font-size:0.75rem;color:#4ADE80">Fed: ${_samt/1e6:.0f}M</span>' if _sfed and _samt > 0 else "")
                                + _illus_badge
                                + '</div>',
                                unsafe_allow_html=True,
                            )

                            # Intelligence panels
                            _intel_pairs = [
                                ("Why Included", _sup.get("why_included", ""), "#60A5FA"),
                                ("Leadership Evidence", _sup.get("leadership_evidence", ""), "#4ADE80"),
                                ("Key Differentiator", _sup.get("key_differentiator", ""), "#A78BFA"),
                                ("Best-Fit Use Case", _sup.get("best_fit_use_case", ""), "#FCD34D"),
                                ("Major Risks", _sup.get("major_risks", ""), "#F87171"),
                            ]
                            for _lbl, _val, _col in _intel_pairs:
                                if _val:
                                    st.markdown(
                                        f'<div style="background:#060D1A;border-left:2px solid {_col};'
                                        f'border-radius:0 6px 6px 0;padding:0.35rem 0.7rem;margin-bottom:0.3rem">'
                                        f'<div style="font-size:0.68rem;color:{_col};text-transform:uppercase;font-weight:700;letter-spacing:0.07em">{_lbl}</div>'
                                        f'<div style="font-size:0.83rem;color:#C4D3E8;margin-top:0.1rem;line-height:1.5">{html.escape(str(_val)[:220])}</div>'
                                        f'</div>',
                                        unsafe_allow_html=True,
                                    )

                            # CFO/CPO challenge
                            _chal_pairs = [
                                ("CFO Would Ask", _sup.get("cfp_challenge") or _sup.get("cfo_challenge", ""), "#FB923C"),
                                ("CPO Would Challenge", _sup.get("cpo_challenge", ""), "#F472B6"),
                            ]
                            _chal_html = ""
                            for _clbl, _cval, _ccol in _chal_pairs:
                                if _cval:
                                    _chal_html += (
                                        f'<div style="flex:1;background:#060D1A;border:1px solid rgba(255,255,255,0.07);'
                                        f'border-radius:6px;padding:0.4rem 0.7rem">'
                                        f'<div style="font-size:0.68rem;color:{_ccol};text-transform:uppercase;font-weight:700">{_clbl}</div>'
                                        f'<div style="font-size:0.8rem;color:#C4D3E8;margin-top:0.15rem">{html.escape(str(_cval)[:180])}</div>'
                                        f'</div>'
                                    )
                            if _chal_html:
                                st.markdown(
                                    f'<div style="display:flex;gap:0.5rem;margin-top:0.4rem;flex-wrap:wrap">{_chal_html}</div>',
                                    unsafe_allow_html=True,
                                )

                            # Red flags
                            if _srflags:
                                _rf_html = " ".join(
                                    f'<span style="background:rgba(248,113,113,0.12);border:1px solid rgba(248,113,113,0.3);'
                                    f'border-radius:5px;padding:0.12rem 0.5rem;font-size:0.72rem;color:#F87171">⚠ {html.escape(str(rf)[:80])}</span>'
                                    for rf in _srflags[:5]
                                )
                                st.markdown(
                                    f'<div style="margin-top:0.4rem;display:flex;flex-wrap:wrap;gap:0.3rem">{_rf_html}</div>',
                                    unsafe_allow_html=True,
                                )

                        with _card_right:
                            # Score gauge
                            st.markdown(
                                f'<div style="background:{_score_color}18;border:2px solid {_score_color};'
                                f'border-radius:12px;padding:0.9rem 0.6rem;text-align:center;margin-bottom:0.7rem">'
                                f'<div style="font-size:2rem;font-weight:800;color:{_score_color}">{_sover}</div>'
                                f'<div style="font-size:0.68rem;color:#94A3B8;text-transform:uppercase;letter-spacing:0.07em">Overall Score</div>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )
                            # Dimension mini-bars
                            _dim_labels = {
                                "market_leadership": "Mkt Lead",
                                "category_fit": "Cat Fit",
                                "financial_strength": "Financial",
                                "execution_capability": "Execution",
                                "innovation_strength": "Innovation",
                                "risk_profile": "Risk Prof",
                                "strategic_procurement_fit": "Proc Fit",
                            }
                            _dims_html = ""
                            for _dk, _dlbl in _dim_labels.items():
                                _dv = _sscores.get(_dk, 0)
                                _dc = "#4ADE80" if _dv >= 70 else ("#FCD34D" if _dv >= 50 else "#F87171")
                                _dims_html += (
                                    f'<div style="margin-bottom:0.3rem">'
                                    f'<div style="display:flex;justify-content:space-between;font-size:0.68rem;margin-bottom:0.1rem">'
                                    f'<span style="color:#94A3B8">{_dlbl}</span>'
                                    f'<span style="color:{_dc};font-weight:600">{_dv}</span></div>'
                                    f'<div style="background:rgba(255,255,255,0.06);border-radius:3px;height:5px">'
                                    f'<div style="background:{_dc};width:{_dv}%;height:5px;border-radius:3px"></div>'
                                    f'</div></div>'
                                )
                            st.markdown(
                                f'<div style="padding:0.3rem 0">{_dims_html}</div>',
                                unsafe_allow_html=True,
                            )

                            # Add to slot button
                            if _show_add:
                                _slot_status, _slot_idx = _disc_slot_map_global.get(_card_idx, ("full", None))
                                if _slot_status == "loaded":
                                    st.markdown(
                                        '<div style="color:#4ADE80;font-size:0.78rem;text-align:center;margin-top:0.4rem">✓ In Suppliers</div>',
                                        unsafe_allow_html=True,
                                    )
                                elif _slot_status == "available" and _slot_idx is not None:
                                    _within_range = _slot_idx < st.session_state.get("ctrl_suppliers", num_suppliers)
                                    _btn_label = (
                                        f"+ Slot {_slot_idx + 1}"
                                        if _within_range
                                        else f"+ Add (expands to {_slot_idx + 1})"
                                    )
                                    _btn_help = (
                                        f"Fills Supplier {_slot_idx + 1} — within your current {st.session_state.get('ctrl_suppliers', num_suppliers)}-supplier count."
                                        if _within_range
                                        else f"All current slots are full. Adds as Supplier {_slot_idx + 1} and increases your count by 1."
                                    )
                                    if st.button(
                                        _btn_label,
                                        key=f"disc_add_{_card_idx}",
                                        help=_btn_help,
                                        use_container_width=True,
                                    ):
                                        st.session_state[f"name_{_slot_idx}"] = _sname
                                        if _sup.get("ticker"):
                                            st.session_state[f"ticker_{_slot_idx}"] = _sup["ticker"]
                                        if not _within_range:
                                            # Always bump by exactly 1 — never jump multiple slots
                                            _new_n = min(8, st.session_state.get("ctrl_suppliers", num_suppliers) + 1)
                                            st.session_state["_ctrl_suppliers_bump"] = _new_n
                                        st.rerun()
                                else:
                                    st.markdown(
                                        '<div style="color:#64748B;font-size:0.72rem;text-align:center;margin-top:0.4rem">Slots full</div>',
                                        unsafe_allow_html=True,
                                    )

                # ── Pre-assign slots for ALL suppliers across all tiers ──────
                # Slot search order: within-range slots FIRST, overflow after.
                # This guarantees discovery fills existing slots before expanding
                # the slider, keeping the user's chosen supplier count stable.
                _MAX_DISC_SLOTS = 10
                _cur_n = st.session_state.get("ctrl_suppliers", num_suppliers)
                _slot_search_order = list(range(_cur_n)) + list(range(_cur_n, _MAX_DISC_SLOTS))

                _all_disc_suppliers = (
                    [(s, 0) for s in _shortlist]
                    + [(s, len(_shortlist)) for s in _longlist]
                    + [(s, len(_shortlist) + len(_longlist)) for s in _watchlist]
                )
                _used_slots: set = set()
                _disc_slot_map_global: dict = {}
                for _gi, (_gsup, _goffset) in enumerate(_all_disc_suppliers):
                    _gsup_name = _gsup.get("name", "")
                    _gexisting = next(
                        (k for k in range(_MAX_DISC_SLOTS)
                         if st.session_state.get(f"name_{k}", "").strip() == _gsup_name.strip()
                         and _gsup_name.strip()),
                        None,
                    )
                    if _gexisting is not None:
                        _disc_slot_map_global[_gi] = ("loaded", _gexisting)
                        _used_slots.add(_gexisting)
                    else:
                        _gfree = next(
                            (k for k in _slot_search_order  # within-range first
                             if k not in _used_slots
                             and (not st.session_state.get(f"name_{k}", "").strip()
                                  or st.session_state.get(f"name_{k}", "") == f"Supplier {k+1}")),
                            None,
                        )
                        _disc_slot_map_global[_gi] = ("available", _gfree) if _gfree is not None else ("full", None)
                        if _gfree is not None:
                            _used_slots.add(_gfree)

                # ── Tier 1: Established Suppliers ────────────────────
                if _shortlist:
                    st.markdown(
                        '<div style="background:rgba(74,222,128,0.05);border:1px solid rgba(74,222,128,0.2);'
                        'border-radius:10px;padding:0.6rem 1rem;margin:0.8rem 0 0.4rem">'
                        '<div style="display:flex;align-items:baseline;gap:0.6rem;flex-wrap:wrap">'
                        '<span style="font-family:monospace;font-size:0.6rem;color:#4ADE80;letter-spacing:0.12em;text-transform:uppercase">TIER 1</span>'
                        '<span style="color:#4ADE80;font-weight:700;font-size:0.92rem">Established Suppliers</span>'
                        '<span style="color:#64748B;font-size:0.8rem">·</span>'
                        '<span style="color:#94A3B8;font-size:0.8rem">Proven market leaders with Gartner/Forrester/IDC recognition · RFI/RFP-ready · sorted by score</span>'
                        '</div>'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                    for _si, _sup in enumerate(sorted(_shortlist, key=lambda x: x.get("overall_score", 0), reverse=True)):
                        _render_disc_supplier(_si, _sup, 0)

                # ── Tier 2: Emerging Challengers ─────────────────────
                if _longlist:
                    st.markdown(
                        '<div style="background:rgba(96,165,250,0.05);border:1px solid rgba(96,165,250,0.18);'
                        'border-radius:10px;padding:0.6rem 1rem;margin:1rem 0 0.4rem">'
                        '<div style="display:flex;align-items:baseline;gap:0.6rem;flex-wrap:wrap">'
                        '<span style="font-family:monospace;font-size:0.6rem;color:#60A5FA;letter-spacing:0.12em;text-transform:uppercase">TIER 2</span>'
                        '<span style="color:#60A5FA;font-weight:700;font-size:0.92rem">Emerging Challengers</span>'
                        '<span style="color:#64748B;font-size:0.8rem">·</span>'
                        '<span style="color:#94A3B8;font-size:0.8rem">Growing market presence · worth qualifying · PE/VC-backed · screening candidates</span>'
                        '</div>'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                    for _si, _sup in enumerate(sorted(_longlist, key=lambda x: x.get("overall_score", 0), reverse=True)):
                        _render_disc_supplier(_si, _sup, len(_shortlist))

                # ── Tier 3: Innovation Watchlist ─────────────────────
                if _watchlist:
                    st.markdown(
                        '<div style="background:rgba(167,139,250,0.05);border:1px solid rgba(167,139,250,0.18);'
                        'border-radius:10px;padding:0.6rem 1rem;margin:1rem 0 0.4rem">'
                        '<div style="display:flex;align-items:baseline;gap:0.6rem;flex-wrap:wrap">'
                        '<span style="font-family:monospace;font-size:0.6rem;color:#A78BFA;letter-spacing:0.12em;text-transform:uppercase">TIER 3</span>'
                        '<span style="color:#A78BFA;font-weight:700;font-size:0.92rem">Innovation Watchlist</span>'
                        '<span style="color:#64748B;font-size:0.8rem">·</span>'
                        '<span style="color:#94A3B8;font-size:0.8rem">VC-backed disruptors · niche specialists · adjacent players · monitor, do not shortlist yet</span>'
                        '</div>'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                    for _si, _sup in enumerate(sorted(_watchlist, key=lambda x: x.get("overall_score", 0), reverse=True)):
                        _render_disc_supplier(_si, _sup, len(_shortlist) + len(_longlist))

        st.markdown("---")
        st.markdown("#### Requirements Intake — 6 Questions That Shape Everything")
        st.markdown('<p class="muted">These answers reshape the Kraljic recommendation, risk flags, stakeholder priorities, and contract must-haves throughout the tool.</p>', unsafe_allow_html=True)

        intake_answers = {}
        for i, q in enumerate(INTAKE_QUESTIONS):
            intake_answers[q["question"]] = st.radio(
                q["question"],
                q["options"],
                horizontal=True,
                key=f"intake_{q['id']}",
                help=f"Impact: {q['impact']}",
            )

        # Store only non-widget keys in session state
        # Widget keys (selected_parent_cat, selected_sub_name) are owned by Streamlit
        # — do NOT write to them manually or Streamlit throws StreamlitAPIException
        st.session_state["intake_answers"] = intake_answers
        st.session_state["_selected_sub_obj"] = selected_sub

        st.markdown(
            """
            <div class="soft-green" style="margin-top:1rem; font-size:0.86rem">
                ✓ Intake complete. Your answers are now applied across the Overview, Sourcing Playbook, Risk Flags, and Contract Guidance tabs.
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Pull intake state for all other tabs ──────────────
    _raw_pkey   = st.session_state.get("selected_parent_cat", list(CATEGORY_TAXONOMY.keys())[0])
    _parent_key = _CAT_ALIAS.get(str(_raw_pkey).lower(), _raw_pkey)
    if _parent_key not in CATEGORY_TAXONOMY:
        _parent_key = list(CATEGORY_TAXONOMY.keys())[0]
    _sub_name   = st.session_state.get("selected_sub_name", CATEGORY_TAXONOMY[_parent_key]["subcategories"][0]["name"])
    _parent_subs = CATEGORY_TAXONOMY.get(_parent_key, CATEGORY_TAXONOMY["HR"])["subcategories"]
    selected_sub = next((s for s in _parent_subs if s["name"] == _sub_name), _parent_subs[0])
    selected_sub_name = selected_sub["name"]
    intake_answers = st.session_state.get("intake_answers", {q["question"]: q["options"][0] for q in INTAKE_QUESTIONS})
    switching_cost_answer = intake_answers.get(INTAKE_QUESTIONS[1]["question"], "Medium — some disruption expected")

    # ── Pre-fetch all supplier intelligence concurrently ──────
    # Runs once when tickers change; results cached in session state
    _current_tickers = [
        st.session_state.get(f"ticker_{i}", "").strip().upper()
        for i in range(num_suppliers)
        if st.session_state.get(f"ticker_{i}", "").strip()
    ]
    _last_tickers = st.session_state.get("_last_prefetch_tickers", [])
    if _current_tickers and _current_tickers != _last_tickers:
        with st.spinner("Fetching supplier intelligence..."):
            _prefetch_result = fetch_supplier_intelligence_concurrent(_current_tickers, alpha_key)
            st.session_state["_intel_prefetch"] = _prefetch_result
            st.session_state["_last_prefetch_tickers"] = _current_tickers

    # ── SUPPLIERS TAB ──────────────────────────────────────
    with tab_suppliers:
        st.markdown("### Supplier Inputs")
        st.markdown(
            '<p class="muted">Add up to 6 suppliers. Enter name, pricing, and scores — the tool pulls SEC financials, '
            'news signals, and company registration data automatically. '
            'Alpha Vantage enrichment is optional.</p>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="background:rgba(96,165,250,0.05);border-left:3px solid rgba(96,165,250,0.3);'
            'border-radius:0 6px 6px 0;padding:0.4rem 0.8rem;font-size:0.80rem;color:#93C5FD;margin-bottom:0.8rem">'
            '💡 <strong>Workflow tip:</strong> Check the <strong>Market Intelligence</strong> tab (to the left) '
            'for live price benchmarks, PPI trends, and SEC filing signals before calibrating '
            '<strong>Price / TCO</strong> scores. '
            'Run <strong>Supplier Discovery</strong> in the Intake tab to auto-populate supplier names.'
            '</div>',
            unsafe_allow_html=True,
        )

        # ── Empty state — cold start CTA ──────────────────────
        _any_named = any(
            st.session_state.get(f"name_{k}", f"Supplier {k+1}") != f"Supplier {k+1}"
            for k in range(num_suppliers)
        )
        if not _any_named:
            st.markdown(
                '<div style="background:rgba(29,78,216,0.06);border:1.5px dashed rgba(96,165,250,0.25);'
                'border-radius:14px;padding:2rem;text-align:center;margin-bottom:1.5rem">'
                '<div style="font-size:1.6rem;margin-bottom:0.5rem">👥</div>'
                '<div style="font-size:1.05rem;font-weight:700;color:#E2E8F0;margin-bottom:0.3rem">'
                'Add Your First Supplier</div>'
                '<div style="font-size:0.90rem;color:#A8BEDC;max-width:420px;margin:0 auto 1rem">'
                'Open Supplier 1 below. Enter the company name — the tool will suggest the stock ticker, '
                'pull SEC filings, and auto-score financial health dimensions.</div>'
                '<div style="font-size:0.80rem;color:#8BAAC4">'
                'Minimum 2 suppliers required for comparison · Maximum 6</div>'
                '</div>',
                unsafe_allow_html=True,
            )

        # ── CSV Import ────────────────────────────────────────────────────
        with st.expander("📥 Import Suppliers from CSV", expanded=False):
            st.markdown(
                '<p class="muted">Upload a CSV to pre-populate supplier slots. '
                'Download the template, fill it in, then upload here. '
                'Financial fields left blank will use the dropdown defaults.</p>',
                unsafe_allow_html=True,
            )
            st.download_button(
                "⬇ Download CSV Template",
                data=_CSV_IMPORT_TEMPLATE,
                file_name="procureiq_supplier_template.csv",
                mime="text/csv",
                key="csv_template_dl",
            )
            _csv_file = st.file_uploader("Upload supplier CSV", type=["csv"], key="csv_import_upload")
            if _csv_file is not None and _csv_file.size > 5 * 1024 * 1024:
                st.error("CSV file is too large (max 5 MB). The supplier template should be well under 1 MB.")
                _csv_file = None
            if _csv_file is not None:
                _parsed, _csv_warnings = parse_supplier_csv(_csv_file)
                if _parsed is None:
                    for _w in _csv_warnings:
                        st.error(_w)
                else:
                    # Preview
                    _preview_rows = [
                        {"Supplier": r["supplier_name"], "Ticker": r["ticker"],
                         "Price ($)": f'{r["raw_price"]:,.0f}',
                         "Scores set": sum(1 for v in r["scores"].values() if v is not None),
                         "Fin Fields": sum(1 for v in r["fin_inputs"].values() if v)}
                        for r in _parsed
                    ]
                    st.dataframe(_preview_rows, use_container_width=True, hide_index=True)
                    for _w in _csv_warnings:
                        st.warning(_w)
                    if st.button(
                        f"✅ Import {len(_parsed)} supplier{'s' if len(_parsed) != 1 else ''} into slots",
                        key="csv_import_confirm", type="primary",
                    ):
                        for _idx, _row in enumerate(_parsed):
                            st.session_state[f"name_{_idx}"] = _row["supplier_name"]
                            st.session_state[f"ticker_{_idx}"] = _row["ticker"]
                            st.session_state[f"raw_price_{_idx}"] = _row["raw_price"]
                            st.session_state[f"notes_{_idx}"] = _row["notes"]
                            for _fname, _fval in _row["fin_inputs"].items():
                                st.session_state[f"fin_{_fname}_{_idx}"] = _fval
                            # Dimension scores
                            _sc = _row.get("scores", {})
                            for _sk in ("sla", "risk", "esg", "diversity"):
                                if _sc.get(_sk) is not None:
                                    _key_map = {"sla": f"sla_{_idx}", "risk": f"risk_{_idx}",
                                                "esg": f"esg_{_idx}", "diversity": f"diversity_{_idx}"}
                                    st.session_state[_key_map[_sk]] = _sc[_sk]
                            for _sk in ("stake", "strategic", "innovation", "relationship", "flexibility"):
                                if _sc.get(_sk) is not None:
                                    st.session_state[f"{_sk}_{_idx}"] = _sc[_sk]
                        # Adjust supplier count to match import
                        _new_n = max(2, min(8, len(_parsed)))
                        st.session_state["ctrl_suppliers"] = _new_n
                        st.success(
                            f"Imported {len(_parsed)} suppliers. "
                            "Scroll down to review and adjust scores before running the evaluation."
                        )
                        st.rerun()

        for i in range(num_suppliers):
            with st.expander(f"Supplier {i + 1}", expanded=(i == 0)):
                left, right = st.columns([1.08, 0.92])

                with left:
                    name = st.text_input("Name", f"Supplier {i + 1}", key=f"name_{i}")
                    # Ticker input with inline auto-suggest
                    _suggested_ticker = suggest_ticker(name) if (name and name not in (f"Supplier {i+1}", "")) else None
                    _ticker_placeholder = f"e.g. {_suggested_ticker}" if _suggested_ticker else "e.g. MSFT, ORCL, AMZN"
                    # Pre-populate from "Use ticker" button click before widget renders
                    _fill = st.session_state.pop(f"ticker_fill_{i}", None)
                    if _fill:
                        st.session_state[f"ticker_{i}"] = _fill
                    ticker = st.text_input(
                        "Public Ticker (optional)",
                        "",
                        key=f"ticker_{i}",
                        placeholder=_ticker_placeholder,
                        help="Enter a stock ticker to pull SEC filings, news signals, and financial data. Type the company name first — we'll suggest the ticker.",
                    )
                    # If ticker is blank but we have a suggestion, show a one-click populate button
                    if not ticker.strip() and _suggested_ticker:
                        _use_col, _hint_col = st.columns([0.45, 0.55])
                        with _use_col:
                            if st.button(f"Use {_suggested_ticker}", key=f"use_ticker_{i}", help="Click to use this ticker"):
                                st.session_state[f"ticker_fill_{i}"] = _suggested_ticker
                                st.rerun()
                        with _hint_col:
                            st.markdown(
                                f'<div style="font-size:0.82rem;color:#60A5FA;padding-top:0.5rem">'
                                f'💡 Suggested for &quot;{html.escape(name[:20])}&quot;</div>',
                                unsafe_allow_html=True,
                            )
                    raw_price = st.number_input("Annual Contract Value $", min_value=0.0,
                                                value=float(1000000 + i * 150000),
                                                step=1000.0, key=f"raw_price_{i}",
                                                help="Base annual cost from the supplier's proposal.")
                    notes = st.text_area("Notes", "Concerns, differentiators, negotiation observations.",
                                         key=f"notes_{i}", height=80)

                    # ── Open Corporates Company Registration Check ─────────
                    if name.strip():
                        with st.expander("🌍 Company Registration Lookup (Open Corporates)", expanded=False):
                            st.markdown(
                                '<p style="font-size:0.85rem;color:#C4D3E8">Verify legal registration status, '
                                'jurisdiction of incorporation, and company number across 130+ jurisdictions. '
                                'Useful for due diligence on unfamiliar suppliers or international vendors.</p>',
                                unsafe_allow_html=True,
                            )
                            if st.button("Search Open Corporates", key=f"oc_search_{i}"):
                                with st.spinner("Querying Open Corporates..."):
                                    _oc_results = fetch_open_corporates(name.strip())
                                if _oc_results:
                                    for _oc in _oc_results[:3]:
                                        _co = _oc.get("company", {})
                                        _co_name  = _co.get("name", "Unknown")
                                        _co_num   = _co.get("company_number", "N/A")
                                        _co_juri  = _co.get("jurisdiction_code", "").upper()
                                        _co_inc   = _co.get("incorporation_date", "N/A")
                                        _co_type  = _co.get("company_type", "N/A")
                                        _co_stat  = _co.get("current_status", "Unknown")
                                        _stat_col = "#4ADE80" if "active" in str(_co_stat).lower() else "#FCD34D"
                                        st.markdown(
                                            f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(96,165,250,0.12);'
                                            f'border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.4rem">'
                                            f'<div style="font-weight:600;color:#E2E8F0;font-size:0.88rem">{html.escape(_co_name)}</div>'
                                            f'<div style="font-size:0.80rem;color:#A8BEDC;margin-top:0.2rem">'
                                            f'#{html.escape(_co_num)} · {html.escape(_co_juri)} · {html.escape(_co_type)}</div>'
                                            f'<div style="font-size:0.78rem;margin-top:0.2rem">'
                                            f'<span style="color:{_stat_col}">{html.escape(_co_stat)}</span> · '
                                            f'<span style="color:#8BAAC4">Inc: {html.escape(_co_inc)}</span></div>'
                                            f'</div>',
                                            unsafe_allow_html=True,
                                        )
                                else:
                                    st.info("No matches found. Try the legal entity name exactly as it appears on contracts.")

                    # ── AGENT #3: Supplier Enrichment ─────────────────────
                    if name.strip() and name.strip() != f"Supplier {i+1}":
                        with st.expander("🧠 AI Supplier Dossier (Enrichment Agent)", expanded=False):
                            st.markdown(
                                '<p style="font-size:0.85rem;color:#C4D3E8">Pulls SEC financials, '
                                'Open Corporates registration, and federal contract history — '
                                'synthesized into a procurement-grade risk brief.</p>',
                                unsafe_allow_html=True,
                            )
                            if st.button("Enrich Supplier Profile", key=f"enrich_btn_{i}"):
                                with st.spinner(f"Enriching {name} …"):
                                    _enrich = run_supplier_enrichment_agent(
                                        company_name=name.strip(),
                                        api_key=_get_api_key(),
                                        provider=_get_provider(),
                                    )
                                st.session_state[f"_enrich_{i}"] = _enrich

                            _enrich_data = st.session_state.get(f"_enrich_{i}")
                            if _enrich_data and _enrich_data.get("error"):
                                _err_msg = str(_enrich_data["error"])
                                _is_no_key = "api_key" in _err_msg.lower() or "no key" in _err_msg.lower() or "agents package" in _err_msg.lower()
                                if _is_no_key:
                                    st.info(
                                        "Enrichment requires an AI API key. Configure one in the AI Settings panel above the tabs, "
                                        "then click Enrich again."
                                    )
                                else:
                                    st.markdown(
                                        f'<div style="background:rgba(248,113,113,0.08);border-left:3px solid #F87171;'
                                        f'border-radius:0 6px 6px 0;padding:0.5rem 0.75rem;font-size:0.82rem;color:#FCA5A5">'
                                        f'<strong>Enrichment failed.</strong> This can happen when the supplier name returns no public '
                                        f'SEC filings or Open Corporates records, or when the AI provider is unreachable. '
                                        f'Try a more specific legal entity name, or enter financial data manually in the Financial Health Signals section below.'
                                        f'<br/><span style="color:#64748B;font-size:0.75rem">{html.escape(_err_msg[:200])}</span>'
                                        f'</div>',
                                        unsafe_allow_html=True,
                                    )
                            elif _enrich_data and not _enrich_data.get("error"):
                                _syn  = _enrich_data.get("synthesis", {})
                                _fed  = _enrich_data.get("federal_contract_history", {})
                                _sec  = _enrich_data.get("sec_financial_data", {})
                                _oc   = _enrich_data.get("registration_data", {})
                                _fh   = _syn.get("financial_health", "UNKNOWN")
                                _rec  = _syn.get("procurement_recommendation", "—")
                                _fh_color = {"STRONG":"#4ADE80","STABLE":"#60A5FA","WATCH":"#FCD34D","DISTRESSED":"#F87171"}.get(_fh, "#94A3B8")
                                _rec_color = {"SHORTLIST":"#4ADE80","QUALIFY_FIRST":"#FCD34D","MONITOR":"#FB923C","AVOID":"#F87171"}.get(_rec, "#94A3B8")
                                st.markdown(
                                    f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:0.6rem;margin-bottom:0.6rem">'
                                    f'<div style="background:#060D1A;border:1px solid rgba(148,163,184,0.12);border-radius:8px;padding:0.6rem 0.8rem">'
                                    f'<div style="font-size:0.75rem;color:#94A3B8;text-transform:uppercase">Financial Health</div>'
                                    f'<div style="font-size:1rem;font-weight:700;color:{_fh_color}">{_fh}</div>'
                                    f'<div style="font-size:0.82rem;color:#C4D3E8;margin-top:0.2rem">{html.escape(_syn.get("financial_summary","")[:120])}</div>'
                                    f'</div>'
                                    f'<div style="background:#060D1A;border:1px solid rgba(148,163,184,0.12);border-radius:8px;padding:0.6rem 0.8rem">'
                                    f'<div style="font-size:0.75rem;color:#94A3B8;text-transform:uppercase">Recommendation</div>'
                                    f'<div style="font-size:1rem;font-weight:700;color:{_rec_color}">{_rec}</div>'
                                    f'<div style="font-size:0.82rem;color:#C4D3E8;margin-top:0.2rem">{html.escape(_syn.get("recommendation_reason","")[:120])}</div>'
                                    f'</div>'
                                    f'</div>',
                                    unsafe_allow_html=True,
                                )
                                if _syn.get("risk_flags"):
                                    for _rf in _syn["risk_flags"]:
                                        st.markdown(f'<div style="font-size:0.84rem;color:#FCD34D;margin-bottom:0.2rem">⚠ {html.escape(_rf)}</div>', unsafe_allow_html=True)
                                if _fed.get("total_federal_awards_usd", 0) > 0:
                                    st.markdown(
                                        f'<div style="font-size:0.84rem;color:#4ADE80">✓ Federal contracts: '
                                        f'${_fed["total_federal_awards_usd"]:,.0f} across {_fed.get("federal_contract_count",0)} awards</div>',
                                        unsafe_allow_html=True,
                                    )

                    # ── Conflict of Interest Disclosure ────────────────────
                    st.markdown(
                        '<div style="background:rgba(239,68,68,0.08);border-left:3px solid #EF4444;border-radius:0 4px 4px 0;'
                        'padding:0.6rem 0.8rem;margin:0.8rem 0 0.8rem 0">'
                        '<div style="font-size:0.82rem;color:#EF4444;text-transform:uppercase;letter-spacing:0.08em;'
                        'font-weight:600;margin-bottom:0.3rem">⚠️ Procurement Governance</div>'
                        '<div style="font-size:0.75rem;color:#C4D3E8">Disclose any financial, personal, or professional relationship with this supplier that could influence your scoring or recommendation.</div>'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                    coi_disclosure = st.checkbox(
                        "I have a conflict of interest or relationship with this supplier that should be disclosed",
                        value=False,
                        key=f"coi_flag_{i}",
                        help="Check if you have any financial stake, family connection, prior employment, or other relationship that could bias scoring."
                    )
                    if coi_disclosure:
                        coi_description = st.text_area(
                            "Describe the relationship:",
                            placeholder="E.g., 'Former colleague now at supplier', 'Personal investment', 'Vendor's CEO is on my board'",
                            key=f"_coi_desc_text_{i}",
                            height=60,
                            help="This will be logged in the audit trail and flagged for transparency in the final recommendation."
                        )
                        st.session_state[f"coi_description_{i}"] = coi_description
                    else:
                        st.session_state[f"coi_description_{i}"] = None

                    # ── TCO Model ────────────────────────────────────────────
                    with st.expander("📐 TCO Model — Total Cost of Ownership", expanded=False):
                        st.markdown(
                            '<div style="font-size:0.82rem;color:#60A5FA;text-transform:uppercase;'
                            'letter-spacing:0.12em;margin-bottom:0.6rem;font-family:monospace">'
                            'True cost = Annual Value × Term + One-Time Costs</div>',
                            unsafe_allow_html=True,
                        )
                        # Benchmark defaults derived from annual contract value.
                        # Industry norms (Gartner / Hackett Group): impl 10%, integration 6%,
                        # training 3%, admin 4% p.a., switching 15%. Zero-out if no price entered.
                        _tco_base = max(raw_price, 0.0)
                        _def_impl  = round(_tco_base * 0.10 / 5000) * 5000
                        _def_intg  = round(_tco_base * 0.06 / 5000) * 5000
                        _def_train = round(_tco_base * 0.03 / 1000) * 1000
                        _def_admin = round(_tco_base * 0.04 / 1000) * 1000
                        _def_exit  = round(_tco_base * 0.15 / 5000) * 5000
                        if _tco_base > 0:
                            st.markdown(
                                '<div style="background:rgba(96,165,250,0.06);border-left:2px solid '
                                'rgba(96,165,250,0.3);padding:0.4rem 0.8rem;border-radius:0 4px 4px 0;'
                                'font-size:0.76rem;color:#94A3B8;margin-bottom:0.6rem">'
                                'Defaults are industry benchmarks (Gartner / Hackett Group) scaled to '
                                'the annual contract value above. Adjust to reflect actual vendor quotes '
                                'and internal capacity costs.</div>',
                                unsafe_allow_html=True,
                            )
                        _tco_term = st.slider("Contract Term (years)", 1, 5, 3, key=f"tco_term_{i}")
                        from datetime import date as _date
                        _contract_start = st.date_input(
                            "Contract Start Date",
                            value=_date.today(),
                            key=f"contract_start_{i}",
                            help="Actual signed contract start date — drives the renewal calendar.",
                        )
                        _notice_period = st.selectbox(
                            "Notice Period Required to Exit",
                            ["30 days", "60 days", "90 days", "120 days", "180 days", "1 year"],
                            index=1,
                            key=f"notice_period_{i}",
                            help="Days of advance notice required to terminate or not renew. This is your action deadline, not the expiry date.",
                        )
                        _auto_renews = st.checkbox(
                            "Auto-renews unless notice given",
                            value=False,
                            key=f"auto_renews_{i}",
                            help="If checked, this contract will roll automatically if the action deadline is missed — the most dangerous contract condition.",
                        )
                        _unbid_cycles = st.number_input(
                            "Times renewed without competitive bid",
                            min_value=0, max_value=20, value=0, step=1,
                            key=f"unbid_cycles_{i}",
                            help="How many consecutive renewal cycles this contract has rolled without an RFP or reverse auction.",
                        )
                        _tco_impl = st.number_input("Implementation / Onboarding ($)", min_value=0.0,
                                                     value=_def_impl, step=5000.0, key=f"tco_impl_{i}",
                                                     help="One-time cost: scoping, migration, setup, project management. Benchmark: ~10% of annual contract value.")
                        _tco_intg = st.number_input("Integration / IT Cost ($)", min_value=0.0,
                                                     value=_def_intg, step=5000.0, key=f"tco_intg_{i}",
                                                     help="APIs, middleware, data pipelines, SSO, security review. Benchmark: ~6% of annual contract value.")
                        _tco_train = st.number_input("Training & Change Management ($)", min_value=0.0,
                                                      value=_def_train, step=1000.0, key=f"tco_train_{i}",
                                                      help="End-user training, documentation, internal change effort. Benchmark: ~3% of annual contract value.")
                        _tco_admin = st.number_input("Annual Internal Admin Cost ($)", min_value=0.0,
                                                      value=_def_admin, step=1000.0, key=f"tco_admin_{i}",
                                                      help="Ongoing FTE time to manage this supplier relationship annually. Benchmark: ~4% of annual contract value.")
                        _tco_exit = st.number_input("Estimated Exit / Switching Cost ($)", min_value=0.0,
                                                     value=_def_exit, step=5000.0, key=f"tco_exit_{i}",
                                                     help="Data migration, transition support, and re-onboarding cost if you leave. Benchmark: ~15% of annual contract value.")
                        _tco_total = (raw_price * _tco_term) + _tco_impl + _tco_intg + _tco_train + (_tco_admin * _tco_term) + _tco_exit
                        _tco_per_year = _tco_total / _tco_term if _tco_term > 0 else 0
                        st.session_state[f"tco_total_{i}"] = _tco_total
                        st.session_state[f"tco_per_year_{i}"] = _tco_per_year
                        c_tco1, c_tco2 = st.columns(2)
                        with c_tco1:
                            st.markdown(
                                f'<div style="background:#060D1A;border:1px solid rgba(96,165,250,0.2);border-radius:8px;padding:0.7rem 1rem">'
                                f'<div style="font-size:0.6rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.1em">Total {_tco_term}yr TCO</div>'
                                f'<div style="font-size:1.4rem;font-weight:700;color:#F1F5F9;font-family:monospace">${_tco_total:,.0f}</div>'
                                f'</div>', unsafe_allow_html=True,
                            )
                        with c_tco2:
                            st.markdown(
                                f'<div style="background:#060D1A;border:1px solid rgba(96,165,250,0.2);border-radius:8px;padding:0.7rem 1rem">'
                                f'<div style="font-size:0.6rem;color:#4ADE80;text-transform:uppercase;letter-spacing:0.1em">Effective Annual Cost</div>'
                                f'<div style="font-size:1.4rem;font-weight:700;color:#F1F5F9;font-family:monospace">${_tco_per_year:,.0f}</div>'
                                f'</div>', unsafe_allow_html=True,
                            )

                # ── Resolve ticker intel BEFORE financial selects render ────────
                sec_context = None
                alpha_context = None
                _edgar_fin_result = None
                if ticker.strip():
                    _prefetch = st.session_state.get("_intel_prefetch", {})
                    _intel = _prefetch.get(ticker.strip().upper(), {})
                    sec_context = _intel.get("sec")
                    if not sec_context:
                        try:
                            sec_context = get_sec_company_context(ticker.strip())
                        except Exception as e:
                            sec_context = {"found": False, "message": str(e)}
                    alpha_context = _intel.get("yf") if _intel else None

                    # Auto-fill financial health fields (never overwrite user edits)
                    _auto_key = f"_fin_auto_{ticker.strip().upper()}_{i}"
                    _yf = alpha_context or {}
                    _is_stale_cache = bool(_yf) and _yf.get("_v", 1) < 2
                    if _is_stale_cache:
                        st.session_state.pop("_last_prefetch_tickers", None)
                        st.session_state.pop("_intel_prefetch", None)
                        st.session_state.pop(_auto_key, None)
                    _have_v2 = _yf.get("_v", 1) >= 2
                    if not st.session_state.get(_auto_key):
                        _news_for_fill = get_supplier_news_signals(ticker.strip().upper())
                        _suggestions = suggest_fin_fields(
                            _yf, _news_for_fill.get("signals", []), sec_ctx=sec_context
                        )
                        for _fname, _fval in _suggestions.items():
                            if st.session_state.get(f"fin_{_fname}_{i}", "") == "":
                                st.session_state[f"fin_{_fname}_{i}"] = _fval
                        # Only lock if we have full v2 data; otherwise retry next run
                        if _have_v2 or sec_context:
                            st.session_state[_auto_key] = True

                    # ── Edit A: Fetch EDGAR XBRL ratios for financial health scoring ──
                    # Cached per ticker+slot; fetched once per session.
                    # Non-blocking: any failure leaves _edgar_fin_result as None.
                    _edgar_cache_key = f"_edgar_fin_{ticker.strip().upper()}_{i}"
                    if _edgar_cache_key not in st.session_state:
                        _xbrl_raw = {}
                        try:
                            _cik_for_xbrl = str(sec_context.get("cik", "")) if sec_context and sec_context.get("found") else ""
                            if _cik_for_xbrl:
                                _xbrl_raw = get_xbrl_financial_ratios(_cik_for_xbrl)
                        except Exception:
                            _xbrl_raw = {}
                        st.session_state[_edgar_cache_key] = compute_edgar_financial_health(_xbrl_raw)
                    _edgar_fin_result = st.session_state.get(_edgar_cache_key)

                with right:
                    # Pull subcategory rubrics for contextual help text
                    _sub_rubrics = SCORING_RUBRICS.get(selected_sub_name, {})

                    def _rubric_help(dim: str, keys: list) -> str:
                        """Build help text from rubric definitions for a dimension."""
                        r = _sub_rubrics.get(dim, {})
                        if not r:
                            return ""
                        lines = [f"{k}: {r[k]}" for k in keys if k in r]
                        return "\n\n".join(lines)

                    sla = st.selectbox(
                        "SLA Strength", ["Strong", "Moderate", "Weak"], index=1, key=f"sla_{i}",
                        help=_rubric_help("SLA Strength", ["Strong", "Moderate", "Weak"]) or
                             "Strong: Contractually committed uptime with credits and named contact.\n"
                             "Moderate: Stated targets; remedies limited or manual.\n"
                             "Weak: Best-effort only; no credits defined.",
                    )
                    risk = st.selectbox(
                        "Execution Risk", ["Low", "Medium", "High"], index=1, key=f"risk_{i}",
                        help=_rubric_help("Execution Risk", ["Low", "Medium", "High"]) or
                             "Low: Proven comparable implementations; dedicated PM; rollback documented.\n"
                             "Medium: Some comparable history; methodology documented.\n"
                             "High: First of this scale; no dedicated team; untested plan.",
                    )
                    def _scored_slider(label: str, key: str, help_text: str) -> int:
                        val = st.slider(label, 1, 5, 3, key=key, help=help_text)
                        if val in (1, 2, 4, 5):
                            _just_key = f"just_{key}"
                            _direction = "high" if val >= 4 else "low"
                            _prompt    = (
                                f"Score is {'strong (' + str(val) + '/5)' if _direction == 'high' else 'weak (' + str(val) + '/5)'}. "
                                f"What specific evidence supports this? (Required for audit trail)"
                            )
                            _existing = st.session_state.get(_just_key, "")
                            _justification = st.text_input(
                                _prompt, value=_existing, key=_just_key,
                                placeholder="e.g. 'VP Finance confirmed support in 12/3 meeting' or 'No implementation team named as of RFP response'",
                            )
                            if not _justification.strip():
                                st.markdown(
                                    '<div style="font-size:0.82rem;color:#F59E0B;margin-top:-0.3rem;margin-bottom:0.4rem">'
                                    '⚠ Justification required for scores of 1, 2, 4, or 5 — this appears in the audit log.</div>',
                                    unsafe_allow_html=True,
                                )
                        return val

                    stake = _scored_slider(
                        "Stakeholder Confidence", f"stake_{i}",
                        "1 = Active opposition. 3 = Neutral / informed. 5 = Champion actively advocating for this supplier.",
                    )
                    strategic = _scored_slider(
                        "Strategic Alignment", f"strategic_{i}",
                        "1 = No overlap with company direction. 3 = Partial alignment. 5 = Roadmap directly supports our multi-year strategy.",
                    )
                    innovation = _scored_slider(
                        "Innovation Capacity", f"innovation_{i}",
                        "1 = Reactive, no R&D investment visible. 3 = Annual product updates. 5 = Dedicated innovation team, co-development possible.",
                    )
                    relationship = _scored_slider(
                        "Relationship Depth", f"relationship_{i}",
                        "1 = Transactional only. 3 = Regular QBRs and named account team. 5 = Executive-to-executive relationship, joint planning.",
                    )
                    flexibility = _scored_slider(
                        "Commercial Flexibility", f"flexibility_{i}",
                        "1 = Rigid standard contract; no negotiation. 3 = Some flexibility on terms and SLAs. 5 = Willing to structure custom deal with risk-sharing.",
                    )

                    # ── ESG & Supplier Diversity ─────────────────────────────
                    st.markdown(
                        '<div style="border-top:1px solid rgba(148,163,184,0.12);padding-top:0.6rem;'
                        'margin-top:0.4rem;font-size:0.82rem;color:#60A5FA;text-transform:uppercase;'
                        'letter-spacing:0.1em;font-weight:600;font-family:monospace">ESG/Sustainability Assessment (Internal)</div>',
                        unsafe_allow_html=True,
                    )
                    esg = st.selectbox(
                        "ESG / Sustainability", ["Strong", "Moderate", "Weak"], index=1, key=f"esg_{i}",
                        help=_rubric_help("ESG / Sustainability", ["Strong", "Moderate", "Weak"]) or
                             "Strong: Published report (GRI/SASB); Science-Based Target; ESG audit rights.\n"
                             "Moderate: Internal policy; some public reporting; no third-party verification.\n"
                             "Weak: No published ESG commitments; no audit rights.",
                    )
                    diversity = st.selectbox(
                        "Supplier Diversity", ["Strong", "Moderate", "Weak"], index=1, key=f"diversity_{i}",
                        help=_rubric_help("Supplier Diversity", ["Strong", "Moderate", "Weak"]) or
                             "Strong: Certified diverse-owned (WBENC/NMSDC/VOSB) OR documented Tier 2 diverse spend >10%.\n"
                             "Moderate: Diversity policy exists; spend tracked but Tier 2 not reported.\n"
                             "Weak: No supplier diversity program or tracking.",
                    )

                st.markdown("#### Financial Health Signals")
                if ticker.strip() and st.session_state.get(f"_fin_auto_{ticker.strip().upper()}_{i}"):
                    st.markdown(
                        '<div style="font-size:0.85rem;color:#4ADE80;margin-bottom:0.5rem;'
                        'font-family:monospace">✓ Auto-filled from ticker — edit any field to override</div>',
                        unsafe_allow_html=True,
                    )
                fin_col1, fin_col2 = st.columns(2)
                fin_data = {}
                field_names = list(FINANCIAL_FIELDS.keys())
                split_point = math.ceil(len(field_names) / 2)

                for idx, field_name in enumerate(field_names):
                    target_col = fin_col1 if idx < split_point else fin_col2
                    with target_col:
                        fin_data[field_name] = st.selectbox(
                            field_name,
                            [""] + FINANCIAL_FIELDS[field_name]["options"],
                            key=f"fin_{field_name}_{i}",
                        )

                # ── Edit B: EDGAR Financial Health Score Card ─────────────────
                if _edgar_fin_result:
                    _efr = _edgar_fin_result
                    _e_score = _efr["score"]
                    _e_conf  = _efr["confidence"]
                    _e_period = _efr.get("period_end", "")
                    _conf_color = "#4ADE80" if _e_conf == "high" else "#FCD34D"
                    _conf_label = "High confidence" if _e_conf == "high" else "Partial data"
                    _risk_lbl, _risk_hex, _ = financial_risk_label(_e_score)

                    # ── Freshness badge logic ─────────────────────────────────
                    # Thresholds based on SEC annual filing cycle:
                    _EDGAR_FRESH_LIMIT = 12   # ≤12 months: within one annual filing cycle
                    _EDGAR_AMBER_LIMIT = 18   # 13–18 months: approaching second cycle
                    # >18 months: stale — more than one full filing cycle behind

                    _freshness = "unknown"
                    _months_old = None
                    if _e_period:
                        try:
                            from datetime import date as _dt_date
                            _pe_date = _dt_date.fromisoformat(_e_period)
                            _today_dt = _dt_date.today()
                            _months_old = (
                                (_today_dt.year - _pe_date.year) * 12
                                + (_today_dt.month - _pe_date.month)
                            )
                            if _months_old <= _EDGAR_FRESH_LIMIT:
                                _freshness = "fresh"
                            elif _months_old <= _EDGAR_AMBER_LIMIT:
                                _freshness = "amber"
                            else:
                                _freshness = "stale"
                        except (ValueError, TypeError):
                            _freshness = "unknown"

                    _fresh_badge_map = {
                        "fresh":   ("#4ADE80", f"SEC data: {_e_period}"),
                        "amber":   ("#FCD34D", f"SEC data: {_e_period} — verify recency"),
                        "stale":   ("#F87171", f"SEC data: {_e_period} — stale, verify before use"),
                        "unknown": ("#94A3B8", "SEC data date unavailable — verify manually"),
                    }
                    _fresh_color, _fresh_text = _fresh_badge_map[_freshness]

                    # Source badge + score + freshness badge (all on main card, not hidden in expander)
                    st.markdown(
                        f'<div style="background:rgba(74,222,128,0.06);border:1px solid rgba(74,222,128,0.22);'
                        f'border-radius:10px;padding:0.8rem 1.1rem;margin:0.6rem 0">'
                        f'<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:0.4rem">'
                        f'<div style="display:flex;align-items:center;gap:0.5rem;flex-wrap:wrap">'
                        f'<span style="font-size:0.6rem;font-family:monospace;text-transform:uppercase;'
                        f'letter-spacing:0.14em;color:#4ADE80;font-weight:700">SEC EDGAR/XBRL</span>'
                        f'<span style="font-size:0.6rem;color:{_conf_color};font-family:monospace;'
                        f'text-transform:uppercase">· {_conf_label}</span>'
                        f'<span style="background:{_fresh_color}22;border:1px solid {_fresh_color}55;'
                        f'border-radius:4px;padding:0.1rem 0.45rem;font-size:0.65rem;color:{_fresh_color};'
                        f'font-family:monospace;font-weight:600">{html.escape(_fresh_text)}</span>'
                        f'</div>'
                        f'<span style="font-size:0.6rem;color:#7A9BC0;font-family:monospace">'
                        f'Review and adjust before use.</span>'
                        f'</div>'
                        f'<div style="display:flex;align-items:baseline;gap:0.6rem;margin-top:0.4rem">'
                        f'<span style="font-size:2.2rem;font-weight:800;font-family:monospace;'
                        f'color:{_risk_hex};letter-spacing:-0.03em">{_e_score}</span>'
                        f'<span style="font-size:1rem;color:#A8BEDC">/100</span>'
                        f'<span style="font-size:0.75rem;font-weight:700;color:{_risk_hex};'
                        f'background:rgba(255,255,255,0.06);border-radius:4px;padding:0.1rem 0.4rem">'
                        f'{_risk_lbl} RISK</span>'
                        f'</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                    # "Why this score?" expander
                    with st.expander("Why this score? (SEC EDGAR sources)", expanded=False):
                        _inputs = _efr.get("inputs", {})
                        _flags  = _efr.get("flags", [])
                        _rows = ""
                        _metric_colors = {
                            "Revenue Growth": "#60A5FA",
                            "Profit Margin": "#4ADE80",
                            "Debt-to-Assets": "#FCD34D",
                        }
                        for _mname, _mval in _inputs.items():
                            _mc = _metric_colors.get(_mname, "#E2E8F0")
                            _rows += (
                                f'<tr>'
                                f'<td style="padding:0.3rem 0.8rem 0.3rem 0;color:#A8BEDC;'
                                f'font-size:0.82rem;white-space:nowrap">{html.escape(_mname)}</td>'
                                f'<td style="padding:0.3rem 0;color:{_mc};font-size:0.82rem">'
                                f'{html.escape(_mval)}</td>'
                                f'</tr>'
                            )
                        st.markdown(
                            f'<table style="width:100%;border-collapse:collapse;margin-bottom:0.5rem">'
                            f'{_rows}'
                            f'</table>',
                            unsafe_allow_html=True,
                        )
                        if _flags:
                            for _fl in _flags:
                                st.markdown(
                                    f'<div style="background:rgba(248,113,113,0.08);border-left:3px solid #F87171;'
                                    f'border-radius:0 6px 6px 0;padding:0.3rem 0.7rem;'
                                    f'font-size:0.82rem;color:#FCA5A5;margin-bottom:0.3rem">'
                                    f'⚠ {html.escape(_fl)}</div>',
                                    unsafe_allow_html=True,
                                )
                        if _e_conf == "partial":
                            _missing = [k for k, v in _inputs.items() if "Not reported" in v]
                            if _missing:
                                st.markdown(
                                    f'<div style="font-size:0.75rem;color:#FCD34D;margin-top:0.3rem">'
                                    f'⚠ Partial data: {html.escape(", ".join(_missing))} not available in EDGAR — '
                                    f'score reflects {_efr["n_metrics"]} of 3 metrics.</div>',
                                    unsafe_allow_html=True,
                                )
                        st.markdown(
                            '<div style="font-size:0.72rem;color:#64748B;margin-top:0.5rem;border-top:1px solid '
                            'rgba(148,163,184,0.1);padding-top:0.4rem">'
                            'Source: SEC EDGAR XBRL (10-K annual filings). Not a substitute for audited financial '
                            'statements, credit ratings, or D&B reports. Verify before contract award.</div>',
                            unsafe_allow_html=True,
                        )

                    # Source selector: use EDGAR score or keep manual selects
                    _src_key = f"fin_source_{i}"
                    _src_options = ["Use SEC EDGAR Score", "Use My Manual Assessment"]
                    _default_src = 0  # default to EDGAR when available
                    if st.session_state.get(_src_key) not in _src_options:
                        st.session_state[_src_key] = _src_options[_default_src]
                    _chosen_src = st.radio(
                        "Financial health source for this evaluation:",
                        _src_options,
                        key=_src_key,
                        horizontal=True,
                        help="SEC EDGAR uses live annual filing data. My Manual Assessment uses the dropdown selections above.",
                    )
                elif ticker.strip() and sec_context and not sec_context.get("found"):
                    st.markdown(
                        '<div style="background:rgba(248,113,113,0.06);border:1px solid rgba(248,113,113,0.18);'
                        'border-radius:8px;padding:0.6rem 0.9rem;font-size:0.82rem;color:#FCA5A5;margin:0.4rem 0">'
                        '⚠ No SEC EDGAR financial data available for this ticker — '
                        'manual assessment required.</div>',
                        unsafe_allow_html=True,
                    )
                elif ticker.strip() and sec_context and sec_context.get("found") and _edgar_fin_result is None:
                    st.markdown(
                        '<div style="background:rgba(252,211,77,0.06);border:1px solid rgba(252,211,77,0.18);'
                        'border-radius:8px;padding:0.6rem 0.9rem;font-size:0.82rem;color:#FCD34D;margin:0.4rem 0">'
                        '⚠ EDGAR lookup attempted — XBRL financial data unavailable for this ticker. '
                        'Score reflects manual inputs. Validate against audited financials before award.</div>',
                        unsafe_allow_html=True,
                    )
                elif not ticker.strip():
                    st.markdown(
                        '<div style="font-size:0.78rem;color:#7A9BC0;margin:0.4rem 0">'
                        'Enter a public ticker above to auto-populate financial health from SEC EDGAR/XBRL. '
                        'Private company suppliers require manual assessment.</div>',
                        unsafe_allow_html=True,
                    )

                if ticker.strip():
                    if sec_context and sec_context.get("found"):
                        # ── Fetch real XBRL financials ───────────────────────
                        _xbrl = fetch_edgar_xbrl_financials(str(sec_context.get("cik", "")))

                        def _metric_cell(label: str, value: str, color: str = "#F1F5F9") -> str:
                            return (
                                f'<div style="background:rgba(96,165,250,0.05);border:1px solid rgba(96,165,250,0.1);'
                                f'border-radius:6px;padding:0.4rem 0.7rem;min-width:90px">'
                                f'<div style="font-size:0.55rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em">{label}</div>'
                                f'<div style="font-size:0.9rem;font-weight:700;color:{color};font-family:var(--mono)">{value}</div>'
                                f'</div>'
                            )

                        _cells = ""
                        if _xbrl.get("revenue"):
                            _cells += _metric_cell(f"Revenue ({_xbrl.get('revenue_year','')})", _xbrl["revenue"])
                        if _xbrl.get("net_income"):
                            _ni_color = "#4ADE80" if _xbrl.get("profitable") else "#F87171"
                            _cells += _metric_cell(f"Net Income ({_xbrl.get('net_income_year','')})", _xbrl["net_income"], _ni_color)
                        if _xbrl.get("net_margin_pct") is not None:
                            _m = _xbrl["net_margin_pct"]
                            _mc = "#4ADE80" if _m >= 10 else "#FCD34D" if _m >= 0 else "#F87171"
                            _cells += _metric_cell("Net Margin", f"{_m}%", _mc)
                        if _xbrl.get("op_cash_flow"):
                            _cells += _metric_cell(f"Op. Cash Flow ({_xbrl.get('op_cf_year','')})", _xbrl["op_cash_flow"])
                        if _xbrl.get("long_term_debt"):
                            _cells += _metric_cell("LT Debt", _xbrl["long_term_debt"], "#FCD34D")
                        if _xbrl.get("cash"):
                            _cells += _metric_cell("Cash", _xbrl["cash"], "#60A5FA")
                        if _xbrl.get("debt_to_rev") is not None:
                            _dr = _xbrl["debt_to_rev"]
                            _drc = "#4ADE80" if _dr < 0.5 else "#FCD34D" if _dr < 1.5 else "#F87171"
                            _cells += _metric_cell("Debt / Rev", f"{_dr}x", _drc)

                        _filings_text = " · ".join(sec_context.get("recent_filings", [])[:3]) or "—"
                        _xbrl_section = (
                            f'<div style="display:flex;flex-wrap:wrap;gap:0.5rem;margin:0.6rem 0">{_cells}</div>'
                            if _cells else
                            '<div style="font-size:0.75rem;color:#D0E0EF;margin:0.4rem 0">XBRL financial data not available for this company.</div>'
                        )
                        _debt_flag = ""
                        if _xbrl.get("debt_to_rev") is not None and _xbrl["debt_to_rev"] > 2:
                            _debt_flag = '<div style="background:rgba(248,113,113,0.08);border-left:3px solid #F87171;border-radius:0 4px 4px 0;padding:0.3rem 0.6rem;font-size:0.75rem;color:#FCA5A5;margin-top:0.4rem">⚠ Debt-to-revenue ratio exceeds 2x — elevated financial risk. Review most recent 10-K for covenant details.</div>'
                        if _xbrl.get("profitable") is False:
                            _debt_flag += '<div style="background:rgba(248,113,113,0.08);border-left:3px solid #F87171;border-radius:0 4px 4px 0;padding:0.3rem 0.6rem;font-size:0.75rem;color:#FCA5A5;margin-top:0.3rem">⚠ Supplier reported a net loss in the most recent fiscal year. Assess business continuity risk.</div>'

                        _sec_html = (
                            f'<div class="intel-strip">'
                            f'<div class="intel-eyebrow">Supplier Intelligence · SEC EDGAR XBRL (Live)</div>'
                            f'<strong style="color:#F1F5F9">{html.escape(sec_context["company_name"])}</strong> '
                            f'<span style="color:#D0E0EF">({html.escape(sec_context["ticker"])}) · CIK {html.escape(str(sec_context["cik"]))}</span><br>'
                            f'<span style="font-size:0.85rem;color:#9EB8CE">{html.escape(_filings_text)}</span>'
                            f'{_xbrl_section}'
                            f'{_debt_flag}'
                        )
                        if alpha_context:
                            market_cap = alpha_context.get("MarketCapitalization", "N/A")
                            pe_ratio = alpha_context.get("PERatio", "N/A")
                            _sec_html += f'<div style="font-size:0.85rem;color:#60A5FA;margin-top:0.3rem">Alpha Vantage · Market Cap: {market_cap} · P/E: {pe_ratio}</div>'
                        _sec_html += "</div>"
                        st.markdown(_sec_html, unsafe_allow_html=True)
                    else:
                        msg = sec_context.get("message", "No SEC context found.") if sec_context else "No SEC context found."
                        st.markdown(
                            f'<div class="intel-strip"><div class="intel-eyebrow">Supplier Intelligence</div>'
                            f'<span style="color:#D0E0EF">{msg}</span></div>',
                            unsafe_allow_html=True,
                        )

                    # ── M&A / News Signal Layer ──────────────────────────
                    _news = get_supplier_news_signals(ticker.strip().upper())
                    if _news.get("signals"):
                        _news_html = (
                            '<div style="background:#0D1526;border:1px solid rgba(167,139,250,0.2);'                            'border-radius:8px;padding:0.7rem 1rem;margin-top:0.5rem">'                            '<div style="font-size:0.82rem;color:#D0E0EF;text-transform:uppercase;'                            'letter-spacing:0.1em;margin-bottom:0.5rem">Recent Signals · yfinance News</div>'
                        )
                        for _sig in _news["signals"]:
                            _risk_border = "rgba(248,113,113,0.3)" if _sig["risk"] else "rgba(74,222,128,0.2)"
                            _news_html += (
                                f'<div style="display:flex;align-items:flex-start;gap:0.5rem;'                                f'margin-bottom:0.35rem;padding:0.3rem 0.5rem;'                                f'border-left:2px solid {_risk_border}">'                                f'<span style="font-size:0.8rem;min-width:1rem">{_sig["icon"]}</span>'                                f'<div>'                                f'<span style="font-size:0.80rem;color:{_sig["color"]};font-weight:700;'                                f'text-transform:uppercase;letter-spacing:0.08em">{_sig["type"]}</span>'                                f'<div style="font-size:0.78rem;color:#CBD5E1;margin-top:0.1rem">{_sig["text"][:90]}{"…" if len(_sig["text"]) > 90 else ""}</div>'                                f'</div></div>'
                            )
                        _news_html += '</div>'
                        st.markdown(_news_html, unsafe_allow_html=True)
                    elif ticker.strip() and ticker.strip() not in ("Private", "—", ""):
                        st.markdown(
                            '<div style="font-size:0.75rem;color:#A8BEDC;margin-top:0.4rem">'                            '📰 No recent news signals detected for this ticker.</div>',
                            unsafe_allow_html=True,
                        )

                # Edit C: determine EDGAR score and source before appending
                _use_edgar = (
                    _edgar_fin_result is not None
                    and st.session_state.get(f"fin_source_{i}", "Use SEC EDGAR Score") == "Use SEC EDGAR Score"
                )
                _edgar_score_final = _edgar_fin_result["score"] if _use_edgar else None
                if _edgar_score_final is not None:
                    _fin_health_source = (
                        "SEC EDGAR/XBRL"
                        if _edgar_fin_result["confidence"] == "high"
                        else "Partial EDGAR + User"
                    )
                else:
                    _fin_health_source = "User Assessment"

                suppliers.append({
                    "Supplier": name, "Ticker": ticker.strip().upper(),
                    "Raw Price": raw_price, "Notes": notes,
                    "SLA Strength": sla, "Execution Risk": risk,
                    "Stakeholder Confidence": stake, "Strategic Alignment": strategic,
                    "Innovation Capacity": innovation, "Relationship Depth": relationship,
                    "Commercial Flexibility": flexibility,
                    "ESG / Sustainability": esg, "Supplier Diversity": diversity,
                    "Financial Inputs": fin_data,
                    "SEC Context": sec_context, "Alpha Context": alpha_context,
                    "EDGAR Financial Score": _edgar_score_final,
                    "Financial Health Source": _fin_health_source,
                    "EDGAR Period End": _edgar_fin_result.get("period_end", "") if _edgar_fin_result else "",
                })

    all_prices = [float(s["Raw Price"]) for s in suppliers]

    scored_suppliers = []
    for s in suppliers:
        # Edit D: prefer EDGAR-sourced score when available and selected by user
        _edgar_s = s.get("EDGAR Financial Score")
        fin_score = _edgar_s if _edgar_s is not None else compute_financial_health(s["Financial Inputs"])
        scores = compute_supplier_scores(s, all_prices, fin_score)
        total = weighted_score(scores, weights)
        c_fit = current_fit(scores)
        f_fit = future_fit(scores)
        risk_label, risk_color, risk_bg = financial_risk_label(fin_score)
        scored_suppliers.append({
            "Supplier": s["Supplier"], "Ticker": s["Ticker"],
            "Notes": s["Notes"], "Raw Price": s["Raw Price"],
            "SEC Context": s["SEC Context"], "Alpha Context": s["Alpha Context"],
            "Financial Health": fin_score, "Financial Risk Label": risk_label,
            "Financial Risk Color": risk_color, "Financial Risk BG": risk_bg,
            "Financial Health Source": s.get("Financial Health Source", "User Assessment"),
            "EDGAR Period End": s.get("EDGAR Period End", ""),
            "Scores": scores, "Weighted Score": total,
            "Current Fit": c_fit, "Future Fit": f_fit,
        })

    ranked = sorted(scored_suppliers, key=lambda x: x["Weighted Score"], reverse=True)
    leader = ranked[0]
    runner_up = ranked[1] if len(ranked) > 1 else None
    leader_weakest_dim = min(DIMENSIONS, key=lambda d: leader["Scores"][d])

    # Persist leader financial intelligence so the save button (rendered earlier in the
    # execution order) can include it when the user clicks Save on the next render cycle.
    st.session_state["_piq_save_fin_source"] = leader.get("Financial Health Source", "User Assessment")
    st.session_state["_piq_save_edgar_period"] = leader.get("EDGAR Period End", "")
    st.session_state["_piq_save_fin_score"] = round(float(leader.get("Financial Health", 0) or 0))
    st.session_state["_piq_save_fin_risk"] = leader.get("Financial Risk Label", "")
    st.session_state["_piq_save_defensibility"] = _confidence_label(leader)
    st.session_state["_piq_save_high_risk_count"] = sum(
        1 for s in ranked if s.get("Financial Risk Label") == "HIGH"
    )

    # Warn if any supplier still has all subjective dimension sliders at the default midpoint.
    _dim_score_keys = ["Stakeholder Confidence", "Strategic Alignment", "Innovation Capacity",
                       "Relationship Depth", "Commercial Flexibility"]
    _all_default_sups = [
        s["Supplier"] for s in ranked
        if all(s["Scores"].get(k, 50) == 50 for k in _dim_score_keys)
    ]
    if len(_all_default_sups) > 0:
        _default_names = ", ".join(_all_default_sups)
        _all_default = len(_all_default_sups) == len(ranked)
        with tab_suppliers:
            st.warning(
                f"{'All suppliers have' if _all_default else f'{len(_all_default_sups)} supplier(s) ({_default_names}) have'} "
                "all subjective dimension scores at the default midpoint (50). "
                "Adjust Stakeholder Confidence, Strategic Alignment, Innovation Capacity, "
                "Relationship Depth, and Commercial Flexibility sliders to reflect real supplier "
                "differentiation — otherwise scores are driven by price and financial health alone."
            )

    stakeholder_rows = []

    # ── STAKEHOLDERS TAB ───────────────────────────────────
    with tab_stakeholders:
        st.markdown("### Who's in the Room — and How to Work Them")
        st.markdown('<p class="muted">This is not just a power/interest map. It is your preparation layer for how to defend the recommendation before anyone challenges it.</p>', unsafe_allow_html=True)

        for i in range(num_stakeholders):
            with st.expander(f"Stakeholder {i + 1}", expanded=(i == 0)):
                c1, c2, c3 = st.columns(3)
                with c1:
                    s_name = st.text_input("Name", f"Stakeholder {i + 1}", key=f"stake_name_{i}")
                    role = st.selectbox("Role / Title", [
                        "CPO", "CFO", "COO", "CIO", "VP Procurement", "VP Finance",
                        "Business Partner", "Legal Counsel", "Procurement Analyst",
                        "Category Manager", "Director", "Manager", "Other",
                    ], key=f"stake_role_{i}")
                with c2:
                    power = st.slider("Power", 1, 10, 6, key=f"stake_power_{i}")
                    interest = st.slider("Interest", 1, 10, 6, key=f"stake_interest_{i}")
                with c3:
                    position = st.selectbox("Position", [
                        "Champion", "Supporter", "Neutral", "Skeptic", "Blocker"
                    ], key=f"stake_position_{i}")
                    priority = st.selectbox("Primary Priority", [
                        "Cost / Savings", "Risk Reduction", "Quality / SLA", "Speed",
                        "Innovation", "Compliance / Legal", "Supplier Relationship",
                    ], key=f"stake_priority_{i}")
                st.markdown(
                    '<div style="border-top:1px solid rgba(148,163,184,0.12);margin:0.6rem 0 0.5rem 0"></div>',
                    unsafe_allow_html=True,
                )
                rc1, rc2, rc3 = st.columns(3)
                with rc1:
                    last_contact = st.date_input("Last Contact Date", value=None, key=f"stake_last_contact_{i}")
                with rc2:
                    rel_health = st.selectbox(
                        "Relationship Health",
                        ["— Select —", "🟢 Green — Strong", "🟡 Amber — Needs Attention", "🔴 Red — At Risk"],
                        key=f"stake_health_{i}",
                    )
                with rc3:
                    next_action = st.text_input("Next Action", placeholder="e.g. Schedule follow-up call", key=f"stake_next_action_{i}")
                stakeholder_rows.append({
                    "Name": s_name, "Role": role, "Power": power,
                    "Interest": interest, "Position": position, "Priority": priority,
                    "Last Contact": str(last_contact) if last_contact else "",
                    "Relationship Health": rel_health if rel_health != "— Select —" else "",
                    "Next Action": next_action,
                })

    stake_df = pd.DataFrame(stakeholder_rows)
    stake_df["Action"] = stake_df.apply(
        lambda row: stakeholder_action(row["Power"], row["Interest"], row["Position"], row["Priority"]), axis=1)
    stake_df["Talk Track"] = stake_df.apply(
        lambda row: talk_track(
            row["Role"], row["Position"], row["Priority"], leader["Supplier"],
            category=category, subcategory=selected_sub_name, kraljic=kraljic,
        ), axis=1)
    blocker = likely_blocker(stake_df)

    with tab_briefing:
        # ── Decision Brief variables ─────────────────────────
        _br_score     = leader["Weighted Score"]
        _br_ci_lo     = max(0, round(_br_score - 8))
        _br_ci_hi     = min(100, round(_br_score + 8))
        _br_gap       = round(_br_score - runner_up["Weighted Score"], 1) if runner_up else None
        _br_fin       = leader["Financial Health"]
        _completeness = sum(
            1 for d in DIMENSIONS if leader["Scores"].get(d, 50) != 50
        ) / len(DIMENSIONS) * 100
        _conf_color   = "#4ADE80" if _completeness >= 70 else "#F59E0B" if _completeness >= 40 else "#F87171"
        _conf_label   = "High" if _completeness >= 70 else "Medium" if _completeness >= 40 else "Low"
        _br_sc        = _score_color(_br_score)
        _fin_sc       = _score_color(_br_fin)
        _gap_color    = "#4ADE80" if (_br_gap and _br_gap >= 5) else "#F59E0B"
        _dec_conf_lbl = _confidence_label(leader)
        _dec_conf_col = _confidence_color(_dec_conf_lbl)

        # ── EDGAR staleness gate ─────────────────────────────────────────
        _gate_period     = leader.get("EDGAR Period End", "")
        _gate_source     = leader.get("Financial Health Source", "")
        _gate_needed     = False
        _gate_tier       = ""
        _gate_age_months = 0.0
        if "EDGAR" in _gate_source and _gate_period:
            try:
                from datetime import date as _gdate
                _gd = _gdate.fromisoformat(_gate_period[:10])
                _gate_age_months = (_gdate.today() - _gd).days / 30.44
                if _gate_age_months > 18:
                    _gate_needed = True
                    _gate_tier   = "STALE"
                elif _gate_age_months > 12:
                    _gate_needed = True
                    _gate_tier   = "AMBER"
            except Exception:
                pass
        _gate_ack_key = f"_piq_edgar_ack_{_gate_period}"
        _gate_acked   = st.session_state.get(_gate_ack_key, False)
        if _gate_needed and not _gate_acked:
            _gc = "#F87171" if _gate_tier == "STALE" else "#FCD34D"
            _gbg = "rgba(248,113,113,0.08)" if _gate_tier == "STALE" else "rgba(252,211,77,0.08)"
            _g_age_str = f"{int(_gate_age_months)} months"
            st.markdown(
                f'<div style="background:{_gbg};border:1px solid {_gc};border-radius:12px;'                f'padding:1.1rem 1.4rem;margin-bottom:1rem">'                f'<div style="font-size:0.68rem;color:{_gc};text-transform:uppercase;'                f'letter-spacing:0.15em;font-weight:700;margin-bottom:0.4rem">'                f'{"🔴 Stale" if _gate_tier == "STALE" else "🟡 Verify"} — EDGAR Data Age Notice</div>'                f'<div style="font-size:0.88rem;color:#E2E8F0;line-height:1.55">'                f'The recommended supplier\u2019s financial health score is sourced from '                f'<strong>SEC EDGAR/XBRL</strong> data filed <strong>{_g_age_str} ago</strong> '                f'({_gate_period[:10]}). '                f'{"This is more than 18 months old and should be treated as stale." if _gate_tier == "STALE" else "This is 12\u201318 months old — verify recency before final award."}'                f'</div>'                f'<div style="font-size:0.78rem;color:#94A3B8;margin-top:0.5rem">'                f'Obtain a current credit check, D&B report, or confirm the most recent 10-K '                f'before relying on the Financial Health score in this brief.</div>'                f'</div>',
                unsafe_allow_html=True,
            )
            if st.checkbox(
                f"I understand — financial data is {_g_age_str} old. Show the Decision Brief.",
                key=f"_piq_edgar_ack_cb_{_gate_period}",
            ):
                st.session_state[_gate_ack_key] = True
                st.rerun()
        _show_brief = not _gate_needed or _gate_acked
        if _show_brief:
            # ── Hero strip — full-width decision summary ──────────
            _hero_gap_block = (
                f'<div style="text-align:center">'
                f'<div style="font-size:0.65rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.2rem">Score Gap</div>'
                f'<div style="font-size:1.5rem;font-weight:800;color:{_gap_color};font-family:monospace">+{_br_gap}</div>'
                f'<div style="font-size:0.72rem;color:#64748B">pts vs {html.escape(runner_up["Supplier"][:14]) if runner_up else "—"}</div>'
                f'</div>'
            ) if runner_up else ""

            st.markdown(
                f'<div style="background:linear-gradient(135deg,rgba(29,78,216,0.12),rgba(6,13,26,0.9));'
                f'border:1px solid rgba(96,165,250,0.22);border-radius:16px;padding:1.4rem 1.8rem;'
                f'margin-bottom:1rem;border-top:3px solid {_br_sc}">'
                # Eyebrow
                f'<div style="font-size:0.6rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.2em;'
                f'font-family:monospace;margin-bottom:0.6rem">Decision Brief · {html.escape(event_name)} · {html.escape(category or "—")}</div>'
                # Name + score row
                f'<div style="display:flex;align-items:baseline;gap:1rem;flex-wrap:wrap;margin-bottom:0.8rem">'
                f'<div style="font-size:2.2rem;font-weight:800;color:#F1F5F9;letter-spacing:-0.02em;line-height:1">'
                f'{html.escape(leader["Supplier"])}</div>'
                f'<div style="font-size:1.1rem;font-weight:700;color:{_br_sc};font-family:monospace">'
                f'{_br_score}/100</div>'
                f'<div style="font-size:0.72rem;color:#64748B;font-family:monospace">CI {_br_ci_lo}–{_br_ci_hi}</div>'
                f'</div>'
                # KPI strip
                f'<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:0.8rem">'
                f'<div style="text-align:center">'
                f'<div style="font-size:0.65rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.2rem">Posture</div>'
                f'<div style="font-size:0.88rem;font-weight:700;color:{"#F87171" if kraljic=="Strategic" else "#4ADE80" if kraljic=="Leverage" else "#FCD34D"}">{html.escape(kraljic)}</div>'
                f'</div>'
                f'<div style="text-align:center">'
                f'<div style="font-size:0.65rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.2rem">Fin. Health</div>'
                f'<div style="font-size:0.88rem;font-weight:700;color:{_fin_sc}">{_br_fin}/100</div>'
                f'<div style="font-size:0.65rem;color:{leader["Financial Risk Color"]}">{leader["Financial Risk Label"]}</div>'
                f'</div>'
                f'<div style="text-align:center">'
                f'<div style="font-size:0.65rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.2rem">Decision Conf.</div>'
                f'<div style="font-size:0.88rem;font-weight:700;color:{_dec_conf_col}">{_dec_conf_lbl}</div>'
                f'<div style="font-size:0.65rem;color:#64748B">{_completeness:.0f}% dims scored</div>'
                f'</div>'
                + _hero_gap_block +
                f'<div style="text-align:center">'
                f'<div style="font-size:0.65rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.2rem">Watch Dim.</div>'
                f'<div style="font-size:0.75rem;font-weight:700;color:#FCD34D">{html.escape(leader_weakest_dim.split("/")[0].strip())}</div>'
                f'</div>'
                f'</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            display_ai_governance_banner()

            # ── Verdict Hero Card ────────────────────────────────
            _gap_color    = "#4ADE80" if (_br_gap and _br_gap >= 5) else "#F59E0B"
            _gap_html     = (
                f'<span style="color:{_gap_color};font-size:1rem;font-family:var(--mono)">+{_br_gap} pts</span>'
                f'<span style="color:#A8BEDC;font-size:0.82rem;margin-left:0.4rem">vs {html.escape(runner_up["Supplier"])}</span>'
                if runner_up else ""
            )
            _dim_bars_html = ""
            for _dim in DIMENSIONS:
                _dv = leader["Scores"].get(_dim, 50)
                _dc = "dim-high" if _dv >= 70 else "dim-mid" if _dv >= 45 else "dim-low"
                _vc = "val-high" if _dv >= 70 else "val-mid" if _dv >= 45 else "val-low"
                _dim_bars_html += (
                    f'<div class="dim-row">'
                    f'<span class="dim-label">{html.escape(_dim)}</span>'
                    f'<div class="dim-track"><div class="dim-fill {_dc}" style="width:{_dv}%"></div></div>'
                    f'<span class="dim-val {_vc}">{_dv}</span>'
                    f'</div>'
                )

            st.markdown(
                f'<div class="verdict-hero">'
                f'<div class="verdict-eyebrow">SYSTEM RECOMMENDATION &nbsp;·&nbsp; DECISION BRIEF</div>'
                f'<div class="verdict-name">{html.escape(leader["Supplier"])}</div>'
                f'<div class="verdict-score">{_br_score} / 100 &nbsp;·&nbsp; {_conf_label} Confidence'
                f'<span title="Score built from ordinal inputs (1–5 sliders). ±8 pts reflects inherent rating uncertainty."'
                f' style="cursor:help;color:#A8BEDC;font-size:0.75rem;margin-left:0.4rem">ⓘ</span></div>'
                f'<div style="margin-top:0.6rem;display:flex;align-items:center;gap:1.2rem;flex-wrap:wrap">'
                f'{_gap_html}'
                f'<span style="color:{leader["Financial Risk Color"]};font-size:0.82rem;font-family:var(--mono)">'
                f'FIN {_br_fin}/100 · {html.escape(leader["Financial Risk Label"])}</span>'
                f'<span style="color:{_conf_color};font-size:0.82rem;font-family:var(--mono)">'
                f'{_completeness:.0f}% dims scored</span>'
                f'</div>'
                f'<div class="verdict-ci">CI {_br_ci_lo}–{_br_ci_hi}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # ── Dimension score bars ─────────────────────────────
            st.markdown(
                f'<div style="background:rgba(4,9,15,0.6);border:1px solid rgba(96,165,250,0.08);'
                f'border-radius:12px;padding:1rem 1.2rem;margin-bottom:1.2rem">'
                f'<div style="font-family:var(--mono);font-size:0.78rem;letter-spacing:0.2em;text-transform:uppercase;'
                f'color:#A8BEDC;margin-bottom:0.7rem">Dimension Breakdown — {html.escape(leader["Supplier"])}</div>'
                f'{_dim_bars_html}'
                f'</div>',
                unsafe_allow_html=True,
            )

            # ── Thin-lead warning ───────────────────────────────
            if runner_up and _br_gap is not None and _br_gap < 5:
                st.markdown(
                    f'<div style="background:rgba(245,158,11,0.08);border-left:3px solid #F59E0B;'
                    f'border-radius:0 6px 6px 0;padding:0.4rem 0.7rem;font-size:0.77rem;color:#FCD34D;margin-bottom:1rem">'
                    f'⚠ Gap under 5 pts — a single dimension shift in the room could flip this. '
                    f'Know which dimension will be challenged before you present.</div>',
                    unsafe_allow_html=True,
                )

            # ── Executive Brief ─────────────────────────────────
            st.markdown(
                '<div style="display:flex;align-items:center;gap:0.7rem;margin-bottom:0.3rem">'
                '<span style="font-size:1.1rem;font-weight:700;color:#F1F5F9">Executive Decision Memo</span>'
                '<span style="font-family:monospace;font-size:0.6rem;color:#60A5FA;border:1px solid rgba(96,165,250,0.3);'
                'border-radius:4px;padding:0.08rem 0.4rem;letter-spacing:0.1em">COPY-READY</span>'
                '</div>',
                unsafe_allow_html=True,
            )
            briefing_memo = build_briefing_memo(
                leader, runner_up, blocker, event_name, category, kraljic,
                category_rule, leader_weakest_dim, stake_df,
            )

            # Format memo lines into a styled card
            _memo_lines = [l for l in briefing_memo.split("\n") if l.strip()]
            _memo_html  = ""
            for _ml in _memo_lines:
                if _ml.startswith("##"):
                    _memo_html += f'<div style="font-size:0.68rem;color:#60A5FA;font-family:monospace;text-transform:uppercase;letter-spacing:0.14em;margin-top:0.9rem;margin-bottom:0.2rem">{html.escape(_ml.lstrip("#").strip())}</div>'
                elif _ml.startswith("**") and _ml.endswith("**"):
                    _memo_html += f'<div style="font-size:0.88rem;font-weight:700;color:#F1F5F9;margin-top:0.4rem;margin-bottom:0.2rem">{html.escape(_ml.strip("*"))}</div>'
                elif _ml.startswith("•") or _ml.startswith("-"):
                    _memo_html += f'<div style="font-size:0.82rem;color:#CBD5E1;padding:0.1rem 0 0.1rem 0.8rem;line-height:1.5">→ {html.escape(_ml.lstrip("•- "))}</div>'
                else:
                    _memo_html += f'<div style="font-size:0.85rem;color:#C4D3E8;line-height:1.6;margin-bottom:0.3rem">{html.escape(_ml)}</div>'

            st.markdown(
                f'<div style="background:#060D1A;border:1px solid rgba(96,165,250,0.15);border-radius:12px;'
                f'padding:1.2rem 1.4rem;margin-bottom:0.6rem">'
                f'{_memo_html}'
                f'</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                "Download Decision Memo (TXT)",
                briefing_memo,
                file_name=f"decision_memo_{event_name[:30].replace(' ','_')}.txt",
                mime="text/plain",
                key="memo_dl_btn",
            )

            # ── 90-Day Execution Plan ────────────────────────────
            st.markdown("---")
            st.markdown(
                '<div style="display:flex;align-items:center;gap:0.7rem;margin-bottom:0.3rem">'
                '<span style="font-size:1.1rem;font-weight:700;color:#F1F5F9">90-Day Execution Plan</span>'
                '<span style="font-family:monospace;font-size:0.6rem;color:#22C55E;border:1px solid rgba(74,222,128,0.3);'
                'border-radius:4px;padding:0.08rem 0.4rem;letter-spacing:0.1em">POST-AWARD</span>'
                '</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                '<p class="muted" style="font-size:0.92rem">Three phases, specific to this supplier, Kraljic posture, '
                'and weakest dimension. Not a generic checklist.</p>',
                unsafe_allow_html=True,
            )

            plan_phases = build_90day_action_plan(
                leader, runner_up, blocker, kraljic, category_rule,
                leader_weakest_dim, selected_sub, intake_answers,
            )

            for phase in plan_phases:
                actions_html = "".join(
                    f'<div style="font-size:0.82rem;color:#CBD5E1;padding:0.22rem 0;'
                    f'border-bottom:1px solid rgba(96,165,250,0.05)">'
                    f'<span style="color:{phase["color"]};margin-right:0.4rem">→</span>{a}</div>'
                    for a in phase["actions"]
                )
                st.markdown(
                    f'<div style="border-left:4px solid {phase["color"]};background:rgba(6,13,26,0.8);'
                    f'border-radius:0 12px 12px 0;padding:1rem 1.3rem;margin-bottom:0.75rem;'
                    f'border-top:1px solid rgba(96,165,250,0.07);border-right:1px solid rgba(96,165,250,0.07);'
                    f'border-bottom:1px solid rgba(96,165,250,0.07)">'
                    f'<div style="display:flex;align-items:center;gap:0.7rem;margin-bottom:0.45rem">'
                    f'<span style="font-family:monospace;font-size:0.78rem;color:{phase["color"]};'
                    f'letter-spacing:0.16em;text-transform:uppercase">{phase["phase"]}</span>'
                    f'<span style="font-size:0.95rem;font-weight:700;color:#F1F5F9">{phase["label"]}</span>'
                    f'</div>'
                    f'<div style="font-size:0.8rem;color:#D0E0EF;margin-bottom:0.55rem;font-style:italic">{phase["objective"]}</div>'
                    f'{actions_html}'
                    f'<div style="font-size:0.82rem;color:#A8BEDC;margin-top:0.5rem;font-family:monospace;'
                    f'letter-spacing:0.06em">OWNER — {phase["owner"]}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # ── Contract Health Snapshot ─────────────────────────
            st.markdown("---")
            st.markdown("#### Contract & Risk Health")
            raqsci = category_raqsci(kraljic, category_rule)
            col_r1, col_r2 = st.columns(2)
            sections_left = ["Requirements", "Assurance of Supply"]
            sections_right = ["Cost", "Innovation"]
            for col, sections in [(col_r1, sections_left), (col_r2, sections_right)]:
                with col:
                    for sec in sections:
                        item = raqsci.get(sec)
                        if item:
                            st.markdown(
                                f'<div style="background:#060D1A;border-left:3px solid #3B82F6;'
                                f'border-radius:0 8px 8px 0;padding:0.6rem 0.9rem;margin-bottom:0.5rem">'
                                f'<div style="font-family:monospace;font-size:0.6rem;color:#3B82F6;'
                                f'text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.3rem">{sec}</div>'
                                f'<div style="font-size:0.82rem;color:#E2E8F0">★ {item["must"]}</div>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )

            # ── Risk Flags ──────────────────────────────────────
            st.markdown("---")
            st.markdown("#### Live Risk Flags")
            risk_flags = generate_rfp_risk_flags(leader, runner_up, blocker, kraljic, category_rule)
            for flag in risk_flags:
                tier_colors = {"HIGH": "#F87171", "MEDIUM": "#FCD34D", "HIDDEN": "#A78BFA"}
                color = tier_colors.get(flag["tier"], "#60A5FA")
                st.markdown(
                    f'<div style="border-left:4px solid {color};background:rgba(6,13,26,0.7);'
                    f'padding:0.6rem 0.9rem;margin-bottom:0.4rem;border-radius:0 8px 8px 0">'
                    f'<div style="font-family:monospace;font-size:0.78rem;color:{color};'
                    f'text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.2rem">{flag["tier"]} RISK</div>'
                    f'<strong style="color:#F1F5F9;font-size:0.88rem">{flag["title"]}</strong>'
                    f'<div style="font-size:0.8rem;color:#C4D3E8;margin-top:0.2rem">{flag["body"]}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # ── Stakeholder Snapshot ─────────────────────────────
            st.markdown("---")
            st.markdown("#### Stakeholder Snapshot")
            snap_col1, snap_col2 = st.columns(2)
            with snap_col1:
                if blocker is not None:
                    st.markdown(
                        f'<div style="background:rgba(248,113,113,0.06);border-left:3px solid #F87171;'
                        f'border-radius:0 8px 8px 0;padding:0.6rem 0.9rem">'
                        f'<div style="font-size:0.6rem;color:#F87171;font-family:monospace;text-transform:uppercase;'
                        f'letter-spacing:0.1em;margin-bottom:0.2rem">Most Likely Blocker</div>'
                        f'<strong style="color:#F1F5F9">{blocker["Name"]}</strong> — {blocker["Role"]}<br>'
                        f'<span style="font-size:0.78rem;color:#C4D3E8">{blocker["Position"]} · Focuses on {blocker["Priority"]}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="background:rgba(74,222,128,0.06);border-left:3px solid #22C55E;'
                        'border-radius:0 8px 8px 0;padding:0.6rem 0.9rem;font-size:0.92rem;color:#C4D3E8">'
                        'No high-power blocker detected — keep narrative crisp.</div>',
                        unsafe_allow_html=True,
                    )
            with snap_col2:
                top_champion = stake_df[stake_df["Position"] == "Champion"]
                if not top_champion.empty:
                    names = ", ".join(top_champion["Name"].tolist())
                    st.markdown(
                        f'<div style="background:rgba(74,222,128,0.06);border-left:3px solid #22C55E;'
                        f'border-radius:0 8px 8px 0;padding:0.6rem 0.9rem">'
                        f'<div style="font-size:0.6rem;color:#22C55E;font-family:monospace;text-transform:uppercase;'
                        f'letter-spacing:0.1em;margin-bottom:0.2rem">Champions — Activate Before Meeting</div>'
                        f'<strong style="color:#F1F5F9">{names}</strong>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="background:rgba(252,211,77,0.06);border-left:3px solid #FCD34D;'
                        'border-radius:0 8px 8px 0;padding:0.6rem 0.9rem;font-size:0.92rem;color:#C4D3E8">'
                        'No champion assigned yet — designate one before the presentation.</div>',
                        unsafe_allow_html=True,
                    )

            # ── CFO Challenge Preparation ────────────────────────
            st.markdown("---")
            st.markdown(
                '<div style="display:flex;align-items:center;gap:0.7rem;margin-bottom:0.3rem">'
                '<span style="font-size:1.1rem;font-weight:700;color:#F1F5F9">CFO Challenge Preparation</span>'
                '<span style="font-family:monospace;font-size:0.6rem;color:#FCD34D;border:1px solid rgba(252,211,77,0.3);'
                'border-radius:4px;padding:0.08rem 0.4rem;letter-spacing:0.1em">PREPARE THE ROOM</span>'
                '</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                '<p class="muted">Every procurement recommendation gets challenged in the room. '
                'These are the questions a CFO or COO will ask — with pre-drafted answers built from your evaluation data. '
                'Review before your presentation and adapt the language to your audience.</p>',
                unsafe_allow_html=True,
            )

            _cfo_challenges = build_cfo_challenge(
                leader, runner_up, event_name, kraljic, category_rule,
                leader_weakest_dim, intake_answers,
            )
            _sev_colors = {"HIGH": "#F87171", "MEDIUM": "#FCD34D", "LOW": "#4ADE80"}
            for _ci, _ch in enumerate(_cfo_challenges):
                _sev = _ch.get("severity", "MEDIUM")
                _scol = _sev_colors.get(_sev, "#94A3B8")
                with st.expander(
                    f"Q{_ci+1}: {_ch['question']}",
                    expanded=(_sev == "HIGH" and _ci == 0),
                ):
                    st.markdown(
                        f'<div style="border-left:3px solid {_scol};padding:0.6rem 0.9rem;'
                        f'background:rgba(6,13,26,0.6);border-radius:0 8px 8px 0;margin-bottom:0.3rem">'
                        f'<div style="font-size:0.62rem;color:{_scol};font-family:monospace;text-transform:uppercase;'
                        f'letter-spacing:0.12em;margin-bottom:0.4rem">{_sev} PRIORITY</div>'
                        f'<div style="font-size:0.88rem;color:#E2E8F0;line-height:1.6">{html.escape(_ch["answer"])}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            # ── AI CFO Narrative (requires key) ──────────────────
            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            if _get_api_key():
                with st.expander("Generate AI-Assisted CFO Narrative", expanded=False):
                    st.markdown(
                        '<p style="font-size:0.85rem;color:#C4D3E8">Uses your AI key to generate a polished '
                        'CFO-ready narrative paragraph that synthesizes the evaluation data into a compelling '
                        'financial case. Review and edit before presenting.</p>',
                        unsafe_allow_html=True,
                    )
                    if st.button("Generate CFO Narrative", key="cfo_narrative_btn", type="primary"):
                        _cfo_prompt = (
                            f"You are a senior procurement advisor preparing a CFO briefing.\n\n"
                            f"Event: {event_name}\nCategory: {category} ({kraljic})\n"
                            f"Recommended Supplier: {leader['Supplier']} — Score {leader['Weighted Score']}/100\n"
                            f"Financial Health: {leader['Financial Health']}/100 ({leader['Financial Risk Label']})\n"
                            + (f"Runner-up: {runner_up['Supplier']} — Score {runner_up['Weighted Score']}/100, "
                               f"Price ${runner_up['Raw Price']:,.0f}\n" if runner_up else "")
                            + f"Price: ${leader['Raw Price']:,.0f}\n"
                            f"Weakest dimension: {leader_weakest_dim}\n\n"
                            f"Write a single 3-4 sentence CFO-ready paragraph that: (1) states the recommendation "
                            f"with confidence, (2) quantifies the value case vs the alternative, (3) names the primary "
                            f"risk and how it is mitigated. No bullet points. No hedging language. "
                            f"Write as if you are the procurement lead speaking to the CFO in a 10-minute briefing."
                        )
                        with st.spinner("Generating CFO narrative …"):
                            try:
                                _cfo_result = _router_call_llm(
                                    messages=[
                                        {"role": "system", "content": "You are a senior procurement advisor. Be direct, precise, and confident."},
                                        {"role": "user", "content": _cfo_prompt},
                                    ],
                                    provider=_get_provider(),
                                    api_key=_get_api_key(),
                                    model=_get_model(),
                                )
                                st.session_state["_cfo_narrative"] = _cfo_result
                            except Exception as _ce:
                                st.session_state["_cfo_narrative"] = f"Error: {_ce}"

                    _cfo_text = st.session_state.get("_cfo_narrative", "")
                    if _cfo_text:
                        if not _cfo_text.startswith("Error:"):
                            st.markdown(
                                f'<div style="background:rgba(29,78,216,0.08);border:1px solid rgba(96,165,250,0.2);'
                                f'border-radius:10px;padding:1rem 1.2rem;font-size:0.9rem;color:#E2E8F0;line-height:1.7">'
                                f'{html.escape(_cfo_text)}</div>',
                                unsafe_allow_html=True,
                            )
                            st.download_button(
                                "Download CFO Narrative (TXT)",
                                _cfo_text,
                                file_name="cfo_narrative.txt",
                                mime="text/plain",
                                key="cfo_narrative_dl",
                            )
                        else:
                            st.error(_cfo_text)
            else:
                st.markdown(
                    '<div style="background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.15);'
                    'border-radius:8px;padding:0.6rem 0.9rem;font-size:0.82rem;color:#93C5FD">'
                    'Set an AI key in the AI Settings panel above to also generate a polished AI-written CFO narrative paragraph.'
                    '</div>',
                    unsafe_allow_html=True,
                )

            # ── Evidence & Assumptions ───────────────────────────
            st.markdown("---")
            with st.expander("📋 Evidence & Assumptions — What This Brief Is Built On", expanded=False):
                st.markdown(
                    '<p style="color:#C4D3E8;font-size:0.88rem;margin-bottom:0.9rem">'
                    'A procurement professional should be able to defend every line of this brief. '
                    'This section surfaces the data sources, scoring assumptions, and gaps that '
                    'a CFO, auditor, or legal reviewer is likely to probe.</p>',
                    unsafe_allow_html=True,
                )
                _ea_c1, _ea_c2, _ea_c3 = st.columns(3)

                # ── Column 1: Data Sources ──────────────────────
                with _ea_c1:
                    st.markdown(
                        '<div style="font-size:0.68rem;color:#60A5FA;text-transform:uppercase;'
                        'letter-spacing:0.12em;font-weight:700;margin-bottom:0.6rem">Data Sources</div>',
                        unsafe_allow_html=True,
                    )
                    _ea_fin_src   = leader.get("Financial Health Source", "User Assessment")
                    _ea_period    = leader.get("EDGAR Period End", "")
                    _ea_period_lbl = f" ({_ea_period[:10]})" if _ea_period else ""
                    _ea_fin_line  = f"SEC EDGAR/XBRL{_ea_period_lbl}" if "EDGAR" in _ea_fin_src else "User assessment (qualitative)"
                    _ea_fin_color = "#60A5FA" if "EDGAR" in _ea_fin_src else "#94A3B8"

                    _ea_dims_filled = sum(1 for d in DIMENSIONS if leader["Scores"].get(d, 50) != 50)
                    _ea_n_sup       = len(ranked)

                    _ea_sources = [
                        ("Financial health",     _ea_fin_line,                              _ea_fin_color),
                        ("Price inputs",         f"User-entered ({_ea_n_sup} supplier{'s' if _ea_n_sup != 1 else ''})", "#94A3B8"),
                        ("Dimensions scored",    f"{_ea_dims_filled}/{len(DIMENSIONS)} beyond default midpoint",
                         "#4ADE80" if _ea_dims_filled >= 7 else "#FCD34D" if _ea_dims_filled >= 4 else "#F87171"),
                        ("Market intelligence",  "Live (Alpha Vantage / SEC)" if st.session_state.get("alpha_key") or st.session_state.get("anthropic_key") else "Static fallback data", "#94A3B8"),
                    ]
                    for _src_lbl, _src_val, _src_col in _ea_sources:
                        st.markdown(
                            f'<div style="margin-bottom:0.45rem">'
                            f'<div style="font-size:0.67rem;color:#64748B;text-transform:uppercase;letter-spacing:0.08em">{_src_lbl}</div>'
                            f'<div style="font-size:0.78rem;color:{_src_col};font-weight:500">{html.escape(_src_val)}</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                # ── Column 2: Scoring Assumptions ──────────────
                with _ea_c2:
                    st.markdown(
                        '<div style="font-size:0.68rem;color:#A78BFA;text-transform:uppercase;'
                        'letter-spacing:0.12em;font-weight:700;margin-bottom:0.6rem">Scoring Assumptions</div>',
                        unsafe_allow_html=True,
                    )
                    _ea_gap_label = (
                        f"+{round(_br_gap, 1)} pts — {'strong' if _br_gap >= 5 else 'narrow'} lead"
                        if runner_up and _br_gap else "No runner-up to compare"
                    )
                    _ea_weight_src = "Subcategory-tuned" if weights != {d: 1/len(DIMENSIONS) for d in DIMENSIONS} else "Default equal weights"
                    _ea_assumptions = [
                        ("Price normalisation",  "Sigmoid (k=4) — outlier-resistant",          "#94A3B8"),
                        ("Weight profile",       _ea_weight_src,                                "#94A3B8"),
                        ("Fin. health method",   "EDGAR ratios (revenue, margin, D/A)" if "EDGAR" in _ea_fin_src else "Qualitative signals", "#94A3B8"),
                        ("Runner-up gap",        _ea_gap_label,                                 "#4ADE80" if (_br_gap and _br_gap >= 5) else "#FCD34D"),
                        ("Score confidence",     _dec_conf_lbl,                                 _dec_conf_col),
                    ]
                    for _as_lbl, _as_val, _as_col in _ea_assumptions:
                        st.markdown(
                            f'<div style="margin-bottom:0.45rem">'
                            f'<div style="font-size:0.67rem;color:#64748B;text-transform:uppercase;letter-spacing:0.08em">{_as_lbl}</div>'
                            f'<div style="font-size:0.78rem;color:{_as_col};font-weight:500">{html.escape(str(_as_val))}</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                # ── Column 3: What to Validate Before Award ────
                with _ea_c3:
                    st.markdown(
                        '<div style="font-size:0.68rem;color:#FCD34D;text-transform:uppercase;'
                        'letter-spacing:0.12em;font-weight:700;margin-bottom:0.6rem">Validate Before Award</div>',
                        unsafe_allow_html=True,
                    )
                    _ea_flags = []

                    # EDGAR staleness
                    if _gate_needed and _gate_tier == "STALE":
                        _ea_flags.append(("🔴", f"Financial data {int(_gate_age_months)}mo old — obtain current credit check", "#F87171"))
                    elif _gate_needed and _gate_tier == "AMBER":
                        _ea_flags.append(("🟡", f"Financial data {int(_gate_age_months)}mo old — verify recency before award", "#FCD34D"))
                    elif "EDGAR" not in _ea_fin_src:
                        _ea_flags.append(("⚪", "No EDGAR verification — private company qualitative only", "#94A3B8"))

                    # Dimension coverage
                    _ea_default_count = len(DIMENSIONS) - _ea_dims_filled
                    if _ea_default_count >= 5:
                        _ea_flags.append(("🔴", f"{_ea_default_count} dimensions at midpoint — differentiation is price-driven only", "#F87171"))
                    elif _ea_default_count >= 2:
                        _ea_flags.append(("🟡", f"{_ea_default_count} dimensions at default midpoint — gather market data", "#FCD34D"))

                    # Competitive comparison
                    if _ea_n_sup == 1:
                        _ea_flags.append(("🟡", "Single supplier — no competitive benchmark; obtain at least one bid", "#FCD34D"))
                    elif _ea_n_sup == 2 and (_br_gap is not None and _br_gap < 3):
                        _ea_flags.append(("🟡", f"Score gap is only {round(_br_gap, 1)} pts — consider a third bid", "#FCD34D"))

                    # Blocker risk
                    if blocker is not None:
                        _ea_flags.append(("🟡", f"Active blocker: {blocker['Name']} ({blocker['Position']}) — align before presenting", "#FCD34D"))

                    if not _ea_flags:
                        _ea_flags.append(("🟢", "No critical data gaps detected for this evaluation", "#4ADE80"))

                    for _ef_icon, _ef_text, _ef_col in _ea_flags:
                        st.markdown(
                            f'<div style="display:flex;gap:0.5rem;align-items:flex-start;margin-bottom:0.5rem">'
                            f'<span style="flex-shrink:0">{_ef_icon}</span>'
                            f'<span style="font-size:0.78rem;color:{_ef_col};line-height:1.45">{html.escape(_ef_text)}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

            # ── EVALUATION HISTORY ────────────────────────────────
            st.markdown("---")
            with st.expander("📂 Prior Evaluations — History from This Session", expanded=False):
                st.markdown(
                    '<p style="color:#C4D3E8;font-size:0.92rem">All completed evaluations from this instance. '
                    'Use to compare past decisions, audit scoring changes, or benchmark prior awards.</p>',
                    unsafe_allow_html=True,
                )
                try:
                    _db = get_database()
                    _history = _db.get_evaluation_history(limit=20)
                    if _history:
                        for _ev in _history:
                            _ev_data = _ev.get("supplier_data", [])
                            _ev_scores = _ev.get("scores", {})
                            _ev_rec = _ev.get("recommendation", "—")
                            _ev_time = _ev.get("created_at", 0)
                            try:
                                from datetime import datetime as _dt
                                _ev_ts = _dt.fromtimestamp(_ev_time).strftime("%Y-%m-%d %H:%M") if _ev_time else "—"
                            except Exception:
                                _ev_ts = "—"
                            _n_sup = len(_ev_data)
                            _winner = next((s["Supplier"] for s in _ev_data if s.get("Supplier")), "—") if _ev_data else "—"
                            st.markdown(
                                f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);border-radius:8px;'
                                f'padding:0.6rem 1rem;margin-bottom:0.4rem;display:flex;align-items:center;gap:1.5rem;flex-wrap:wrap">'
                                f'<span style="font-family:monospace;font-size:0.82rem;color:#D0E0EF">{_ev_ts}</span>'
                                f'<strong style="color:#60A5FA;font-size:0.85rem">{html.escape(_ev["event_id"][:32])}</strong>'
                                f'<span style="font-size:0.82rem;color:#C4D3E8">{_n_sup} suppliers</span>'
                                f'<span style="font-size:0.82rem;color:#4ADE80">Recommended: <strong>{html.escape(_winner)}</strong></span>'
                                f'<span style="font-size:0.75rem;color:#D0E0EF;margin-left:auto">{html.escape(_ev_rec[:60])}{"…" if len(_ev_rec) > 60 else ""}</span>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )
                    else:
                        st.markdown(
                            '<div style="font-size:0.92rem;color:#A8BEDC;padding:0.5rem 0">'
                            'No prior evaluations saved yet. Complete an evaluation and the history will appear here.</div>',
                            unsafe_allow_html=True,
                        )
                except Exception as _hist_err:
                    st.markdown(f'<div style="font-size:0.8rem;color:#F87171">Could not load history: {html.escape(str(_hist_err))}</div>', unsafe_allow_html=True)

    with tab_stakeholders:
        c1, c2 = st.columns([1.0, 1.0])
        with c1:
            scatter = px.scatter(
                stake_df, x="Interest", y="Power", color="Position", text="Name",
                color_discrete_map=POSITION_COLORS, height=520,
            )
            scatter.update_traces(textposition="top center", marker=dict(size=14))
            scatter.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(13,21,38,0.8)",
                font=dict(color="#8EA3C3", family="Outfit"),
                xaxis=dict(
                    range=[0, 10.5], gridcolor="rgba(255,255,255,0.04)", color="#4A6080",
                    title=dict(text="INTEREST →", font=dict(size=10, color="#334155")),
                ),
                yaxis=dict(
                    range=[0, 10.5], gridcolor="rgba(255,255,255,0.04)", color="#4A6080",
                    title=dict(text="POWER →", font=dict(size=10, color="#334155")),
                ),
                margin=dict(l=40, r=20, t=30, b=40),
                legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#8EA3C3")),
                shapes=[
                    dict(type="line", x0=5.5, x1=5.5, y0=0, y1=10.5,
                         line=dict(color="rgba(255,255,255,0.1)", dash="dash", width=1)),
                    dict(type="line", x0=0, x1=10.5, y0=5.5, y1=5.5,
                         line=dict(color="rgba(255,255,255,0.1)", dash="dash", width=1)),
                ],
                annotations=[
                    dict(x=2.5, y=10.2, text="<b>KEEP SATISFIED</b>",
                         showarrow=False, font=dict(size=8, color="rgba(252,211,77,0.45)"),
                         xanchor="center"),
                    dict(x=8.5, y=10.2, text="<b>KEY PLAYER — MANAGE CLOSELY</b>",
                         showarrow=False, font=dict(size=8, color="rgba(248,113,113,0.55)"),
                         xanchor="center"),
                    dict(x=2.5, y=0.3, text="<b>MONITOR</b>",
                         showarrow=False, font=dict(size=8, color="rgba(100,116,139,0.5)"),
                         xanchor="center"),
                    dict(x=8.5, y=0.3, text="<b>KEEP INFORMED</b>",
                         showarrow=False, font=dict(size=8, color="rgba(74,222,128,0.45)"),
                         xanchor="center"),
                ],
            )
            st.plotly_chart(scatter, use_container_width=True)

        with c2:
            st.markdown("#### Talk Track by Stakeholder")
            for idx, (_, row) in enumerate(stake_df.iterrows()):
                color = POSITION_COLORS.get(row["Position"], "#94A3B8")
                mt = "0" if idx == 0 else "0.6rem"
                st.markdown(
                    f"""
                    <div style="
                        background:#0A1628;
                        border:1px solid rgba(148,163,184,0.14);
                        border-left:3px solid {color};
                        border-radius:8px;
                        padding:0.75rem 0.9rem;
                        margin-top:{mt};
                    ">
                        <div style="display:flex;align-items:baseline;gap:0.4rem;margin-bottom:0.25rem">
                            <strong style="color:#F1F5F9;font-size:0.88rem">{html.escape(sx(row['Name']))}</strong>
                            <span style="color:#94A3B8;font-size:0.82rem">— {html.escape(sx(row['Role']))}</span>
                        </div>
                        <div style="display:flex;flex-wrap:wrap;gap:0.35rem;margin-bottom:0.5rem">
                            <span style="background:rgba({','.join(str(int(color.lstrip('#')[i:i+2],16)) for i in (0,2,4))},0.15);color:{color};border:1px solid {color}40;border-radius:4px;padding:0.1rem 0.45rem;font-size:0.72rem;font-family:monospace">{html.escape(sx(row['Position']))}</span>
                            <span style="background:rgba(148,163,184,0.08);color:#94A3B8;border:1px solid rgba(148,163,184,0.15);border-radius:4px;padding:0.1rem 0.45rem;font-size:0.72rem;font-family:monospace">{html.escape(sx(row['Priority']))}</span>
                            <span style="background:rgba(148,163,184,0.08);color:#94A3B8;border:1px solid rgba(148,163,184,0.15);border-radius:4px;padding:0.1rem 0.45rem;font-size:0.72rem;font-family:monospace">Power {row['Power']}</span>
                            <span style="background:rgba(148,163,184,0.08);color:#94A3B8;border:1px solid rgba(148,163,184,0.15);border-radius:4px;padding:0.1rem 0.45rem;font-size:0.72rem;font-family:monospace">Interest {row['Interest']}</span>
                        </div>
                        <div style="font-size:0.82rem;color:#C4D3E8;line-height:1.55;padding:0.45rem 0.55rem;background:rgba(96,165,250,0.05);border-left:2px solid rgba(96,165,250,0.3);border-radius:0 4px 4px 0">
                            <span style="font-family:monospace;font-size:0.7rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.08em;display:block;margin-bottom:0.2rem">Talk Track</span>
                            {sx(row['Talk Track'])}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        st.markdown("#### Full Stakeholder Table")
        st.dataframe(stake_df[["Name", "Role", "Power", "Interest", "Position", "Priority", "Action"]],
                     use_container_width=True, hide_index=True)

    # ── MARKET TAB ─────────────────────────────────────────
    with tab_market:
        st.markdown("### Executive Snapshot")

        # ── 4 metric cards ──
        m1, m2, m3, m4 = st.columns(4)
        kq_c = {"Strategic": "#F87171", "Leverage": "#4ADE80", "Bottleneck": "#FCD34D", "Non-Critical": "#94A3B8"}.get(kraljic, "#60A5FA")
        _ws_color = _score_color(leader["Weighted Score"])
        _cf_color = _score_color(leader["Current Fit"])
        _ff_color = _score_color(leader["Future Fit"])
        for col, label, val, sub, accent in [
            (m1, "Recommended", leader["Supplier"], event_name, "#60A5FA"),
            (m2, "Overall Score", str(leader["Weighted Score"]), "weighted / 100", _ws_color),
            (m3, "Current Fit", str(leader["Current Fit"]), "execution strength", _cf_color),
            (m4, "Future Fit", str(leader["Future Fit"]), "strategic trajectory", _ff_color),
        ]:
            with col:
                st.markdown(
                    f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.22);border-radius:12px;'
                    f'padding:0.9rem;border-top:2px solid {accent}">'
                    f'<div style="font-size:0.78rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.35rem">{label}</div>'
                    f'<div style="font-size:{"1rem" if len(val)>6 else "1.7rem"};font-weight:700;color:{accent};font-family:monospace;line-height:1.1">{sx(val)}</div>'
                    f'<div style="font-size:0.82rem;color:#D0E0EF;margin-top:0.2rem">{sx(sub)}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("<div style='height:0.6rem'></div>", unsafe_allow_html=True)

        # ── 3 exec cards — compact ──
        e1, e2, e3 = st.columns(3)
        blocker_text_ov = block_risk_text(blocker, leader["Supplier"])
        cards_data = [
            (e1, "#0A1F4A", "#0D2D6B", "rgba(96,165,250,0.15)", "Recommendation",
             f"{sx(leader['Supplier'])} — {leader['Weighted Score']}/100",
             f"Best score under <strong style='color:#60A5FA'>{sx(kraljic)}</strong>. Current fit {leader['Current Fit']}, future fit {leader['Future Fit']}."),
            (e2, "#2A1500", "#3D2000", "rgba(252,211,77,0.15)", "Trade-off",
             "The choice is not just about price.",
             sx(make_tradeoff_text(leader, runner_up))),
            (e3, "#2A0808", "#3D1010", "rgba(248,113,113,0.15)", "Who Will Block This",
             "Know where the room may turn.",
             sx(blocker_text_ov)),
        ]
        for col, bg1, bg2, border_c, eyebrow, head, body in cards_data:
            with col:
                st.markdown(
                    f'<div style="background:linear-gradient(145deg,{bg1},{bg2});border:1px solid {border_c};'
                    f'border-radius:14px;padding:0.9rem 1rem;min-height:110px">'
                    f'<div style="font-size:0.6rem;color:rgba(226,232,240,0.45);text-transform:uppercase;letter-spacing:0.14em;margin-bottom:0.35rem">{eyebrow}</div>'
                    f'<div style="font-size:0.88rem;font-weight:700;color:#E2E8F0;margin-bottom:0.3rem">{head}</div>'
                    f'<div style="font-size:0.8rem;color:#C4D3E8;line-height:1.5">{body}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)

        # ── Main layout: radar left, market intel right ──
        ov_left, ov_right = st.columns([1.3, 0.9])

        with ov_left:
            # Recommendation logic
            rec_text = make_recommendation_text(leader, runner_up, leader_weakest_dim, kraljic)
            st.markdown(
                f'<div style="background:rgba(96,165,250,0.06);border-left:3px solid #60A5FA;'
                f'border-radius:0 10px 10px 0;padding:0.8rem 1rem;margin-bottom:0.8rem;'
                f'font-size:0.87rem;color:#E2E8F0;line-height:1.6;white-space:pre-line">{rec_text}</div>',
                unsafe_allow_html=True,
            )

            # Horizontal bar chart — top 3 suppliers, sorted by leader's score
            _bar_suppliers = ranked[:3]
            _bar_caption = f" · Top 3 of {len(ranked)}" if len(ranked) > 3 else ""
            st.markdown(
                f'<div style="font-size:0.66rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.4rem;font-weight:700">'
                f'Supplier Comparison<span style="font-weight:400;color:#64748B">{_bar_caption}</span></div>',
                unsafe_allow_html=True,
            )
            rc = ["#60A5FA", "#4ADE80", "#FCD34D", "#F87171"]
            _dim_short = {
                "Price / TCO": "Price / TCO", "SLA Strength": "SLA Strength",
                "Execution Risk": "Exec Risk", "Stakeholder Confidence": "Stakeholder",
                "Strategic Alignment": "Strategic Align", "Innovation Capacity": "Innovation",
                "Relationship Depth": "Relationship", "Commercial Flexibility": "Commercial",
                "ESG / Sustainability": "ESG", "Supplier Diversity": "Diversity",
            }
            # Sort dimensions by leader's score descending — strengths at top
            _dim_order = sorted(
                DIMENSIONS,
                key=lambda d: _bar_suppliers[0]["Scores"].get(d, 0),
                reverse=True,
            )
            _ylabels = [_dim_short.get(d, d) for d in _dim_order]
            fig = go.Figure()
            for idx, s in enumerate(_bar_suppliers):
                fig.add_trace(go.Bar(
                    x=[s["Scores"].get(d, 0) for d in _dim_order],
                    y=_ylabels,
                    name=s["Supplier"][:18],
                    orientation="h",
                    marker=dict(
                        color=rc[idx % len(rc)],
                        opacity=0.85,
                        line=dict(width=0),
                    ),
                    hovertemplate="%{y}: %{x}<extra>" + s["Supplier"] + "</extra>",
                ))
            fig.update_layout(
                barmode="group",
                xaxis=dict(
                    range=[0, 100],
                    tickvals=[0, 25, 50, 75, 100],
                    tickfont=dict(color="#94A3B8", size=10),
                    gridcolor="rgba(255,255,255,0.07)",
                    zeroline=False,
                ),
                yaxis=dict(
                    tickfont=dict(color="#CBD5E1", size=11),
                    autorange="reversed",
                ),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(13,21,38,0.4)",
                font=dict(color="#CBD5E1", family="Inter", size=11),
                height=420,
                margin=dict(l=10, r=10, t=44, b=10),
                legend=dict(
                    orientation="h", yanchor="bottom", y=1.02,
                    xanchor="center", x=0.5,
                    bgcolor="rgba(13,21,38,0.7)",
                    bordercolor="rgba(148,163,184,0.2)",
                    borderwidth=1,
                    font=dict(size=11, color="#E2E8F0"),
                ),
            )
            st.plotly_chart(fig, use_container_width=True)

            # Dimension breakdown table — replaces blank panel
            st.markdown('<div style="font-size:0.66rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.4rem;font-weight:700">Dimension Scores — All Suppliers</div>', unsafe_allow_html=True)
            header_html = '<div style="display:flex;gap:0.5rem;padding:0.25rem 0;border-bottom:1px solid rgba(148,163,184,0.15);margin-bottom:0.2rem"><span style="font-size:0.85rem;color:#D0E0EF;min-width:7rem">Dimension</span>'
            for s in ranked:
                header_html += f'<span style="font-size:0.85rem;color:{rc[ranked.index(s) % len(rc)]};font-weight:700;min-width:3rem;text-align:center">{sx(s["Supplier"][:8])}</span>'
            header_html += "</div>"
            st.markdown(header_html, unsafe_allow_html=True)
            for dim in DIMENSIONS:
                dim_scores = [s["Scores"][dim] for s in ranked]
                best_score = max(dim_scores)
                row_html = f'<div style="display:flex;gap:0.5rem;padding:0.2rem 0;border-bottom:1px solid rgba(148,163,184,0.05)"><span style="font-size:0.74rem;color:#C4D3E8;min-width:7rem">{dim.split("/")[0].strip()}</span>'
                for i, (s, score) in enumerate(zip(ranked, dim_scores)):
                    is_best = score == best_score
                    row_html += f'<span style="font-size:0.78rem;font-weight:{"700" if is_best else "400"};color:{"#E2E8F0" if is_best else "#64748B"};background:{"rgba(96,165,250,0.1)" if is_best else "transparent"};border-radius:4px;padding:0.08rem 0.3rem;min-width:3rem;text-align:center">{score}</span>'
                row_html += "</div>"
                st.markdown(row_html, unsafe_allow_html=True)

        with ov_right:
            # Category posture
            st.markdown(
                f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);border-radius:10px;padding:0.8rem;margin-bottom:0.7rem">'
                f'<div style="font-size:0.78rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.4rem;font-weight:700">Category Posture</div>'
                f'<div style="font-size:0.9rem;font-weight:700;color:{kq_c};margin-bottom:0.2rem">{sx(kraljic)}</div>'
                f'<div style="font-size:0.78rem;color:#C4D3E8;margin-bottom:0.4rem">{sx(kinfo["axis"])}</div>'
                f'<div style="font-size:0.77rem;color:#CBD5E1;line-height:1.5;margin-bottom:0.4rem">{sx(kinfo["desc"])}</div>'
                f'<div style="font-size:0.82rem;color:#D0E0EF">{sx(category_rule["tag"])} · {sx(category_rule["type"])}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # Active weights
            st.markdown('<div style="font-size:0.78rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.4rem;font-weight:700">Active Weights</div>', unsafe_allow_html=True)
            chips_w = "".join(
                f'<span style="display:inline-block;background:rgba(96,165,250,0.08);border:1px solid rgba(96,165,250,0.18);'
                f'color:#93C5FD;border-radius:999px;padding:0.1rem 0.4rem;font-size:0.66rem;font-family:monospace;font-weight:600;margin:0.12rem">'
                f'{d.split("/")[0].strip()}: {round(weights[d]*100,0):.0f}%</span>'
                for d in DIMENSIONS
            )
            st.markdown(f'<div style="margin-bottom:0.8rem">{chips_w}</div>', unsafe_allow_html=True)

            # ── LIVE MARKET INTELLIGENCE ──
            st.markdown(
                f'<div style="font-size:0.66rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.12em;'
                f'margin-bottom:0.4rem;font-weight:700">Market Leaders — {sx(selected_sub_name)}</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                '<div style="background:rgba(252,211,77,0.05);border-left:3px solid #FCD34D;border-radius:0 6px 6px 0;'
                'padding:0.4rem 0.6rem;font-size:0.82rem;color:#C4D3E8;margin-bottom:0.5rem">'
                '⚡ Live data via yfinance for public companies. Curated static data for private.</div>',
                unsafe_allow_html=True,
            )
            market_leaders_live = enrich_market_leaders_with_live_data(selected_sub_name)
            for idx, ldr in enumerate(market_leaders_live, 1):
                live = ldr.get("live")
                ticker_str = ldr.get("ticker", "")
                ticker_tag = ""
                if ticker_str and ticker_str not in ("Private", "—", ""):
                    ticker_tag = f'<span style="font-size:0.6rem;font-family:monospace;color:#60A5FA;background:rgba(96,165,250,0.1);border-radius:3px;padding:0.05rem 0.25rem;margin-left:0.3rem">{sx(ticker_str)}</span>'
                elif ticker_str == "Private":
                    ticker_tag = '<span style="font-size:0.6rem;color:#D0E0EF;margin-left:0.3rem">Private</span>'

                live_row_html = ""
                if live:
                    parts = []
                    if live["market_cap_fmt"] != "N/A":
                        parts.append(f'<span style="font-size:0.80rem;color:#4ADE80;font-family:monospace">{live["market_cap_fmt"]}</span>')
                    if live["revenue_growth"] != "N/A":
                        col_g = "#4ADE80" if "+" in live["revenue_growth"] else "#F87171"
                        parts.append(f'<span style="font-size:0.80rem;color:{col_g};font-family:monospace">Rev {live["revenue_growth"]}</span>')
                    if live["analyst_rating"] not in ("N/A", "None"):
                        parts.append(f'<span style="font-size:0.80rem;color:#FCD34D">{sx(live["analyst_rating"])}</span>')
                    if parts:
                        live_row_html = f'<div style="display:flex;gap:0.4rem;flex-wrap:wrap;margin-top:0.25rem">{"".join(parts)}</div>'

                st.markdown(
                    f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.12);border-radius:8px;'
                    f'padding:0.6rem 0.7rem;margin-bottom:0.35rem">'
                    f'<div style="display:flex;align-items:center;margin-bottom:0.2rem">'
                    f'<span style="font-size:0.80rem;color:#D0E0EF;font-family:monospace;margin-right:0.35rem">#{idx}</span>'
                    f'<strong style="font-size:0.8rem;color:#E2E8F0">{sx(ldr["name"])}</strong>'
                    f'{ticker_tag}</div>'
                    f'<div style="font-size:0.85rem;color:#C4D3E8;line-height:1.5"><strong style="color:#4ADE80">✓</strong> {sx(ldr["strength"])}</div>'
                    f'<div style="font-size:0.82rem;color:#F87171;line-height:1.5;margin-top:0.15rem">⚠ {sx(ldr["watch"])}</div>'
                    f'{live_row_html}</div>',
                    unsafe_allow_html=True,
                )

    # ── PPI / MACRO CONTEXT (bottom of Market tab) ──────────
    with tab_market:
        st.markdown("---")
        st.markdown("### Macro Market Context")
        st.markdown(
            '<p style="color:#C4D3E8;font-size:0.88rem">Producer Price Index and commodity signals relevant to this category. '
            'Source: U.S. Bureau of Labor Statistics public data (updated monthly). Use to anchor should-cost discussions and justify price movement claims in negotiation.</p>',
            unsafe_allow_html=True,
        )

        # BLS PPI series mapped by category keyword
        # Series IDs: PCU + 6-digit NAICS industry code + 6-digit product code
        _PPI_SERIES = {
            "technology":    {"id": "PCU511210511210", "label": "Software Publishing PPI", "unit": "Index"},
            "it":            {"id": "PCU511210511210", "label": "Software Publishing PPI", "unit": "Index"},
            "hr":            {"id": "PCU561310561310", "label": "Temporary Help Services PPI", "unit": "Index"},
            "staffing":      {"id": "PCU561310561310", "label": "Temporary Help Services PPI", "unit": "Index"},
            "logistics":     {"id": "PCU484000484000", "label": "Trucking & Freight PPI", "unit": "Index"},
            "manufacturing": {"id": "PCU331221331221", "label": "Steel Products PPI", "unit": "Index"},
            "packaging":     {"id": "PCU322100322100", "label": "Paper & Packaging PPI", "unit": "Index"},
            "marketing":     {"id": "PCU541810541810", "label": "Advertising Agencies PPI", "unit": "Index"},
            "finance":       {"id": "PCU522300522300", "label": "Financial Services PPI", "unit": "Index"},
            "services":      {"id": "PCU541110541110", "label": "Legal Services PPI", "unit": "Index"},
        }
        # Estimated annual inflation benchmarks (BLS/ISM/Gartner 2025) used as fallback
        _PPI_INFLATION_HINTS = {
            "technology": 3.5, "it": 3.5, "hr": 5.2, "staffing": 6.2,
            "logistics": 6.4, "manufacturing": 7.2, "packaging": 4.5,
            "marketing": 3.8, "finance": 4.1, "services": 5.8,
        }

        # Resolve PPI series from category key
        _cat_key_lower = (category_rule.get("tag", "") + " " + category).lower()
        _resolved_series = None
        for _kw, _series_data in _PPI_SERIES.items():
            if _kw in _cat_key_lower:
                _resolved_series = _series_data
                break
        if _resolved_series is None:
            _resolved_series = {"id": "PCEPI", "label": "PCE Price Index (General)", "unit": "Index"}

        @st.cache_data(ttl=86400, show_spinner=False)
        def _fetch_bls_ppi(series_id: str):
            try:
                _url = f"https://api.bls.gov/publicAPI/v1/timeseries/data/{series_id}"
                _resp = requests.get(_url, timeout=8)
                if _resp.status_code != 200:
                    return None
                _data = _resp.json()
                if _data.get("status") != "REQUEST_SUCCEEDED":
                    return None
                _series = _data.get("Results", {}).get("series", [])
                if not _series:
                    return None
                _points = _series[0].get("data", [])[:18]  # last 18 months
                _points.reverse()
                return [{"period": f'{p["year"]}-{p["period"].replace("M","").zfill(2)}', "value": float(p["value"])} for p in _points]
            except Exception:
                return None

        _ppi_data = _fetch_bls_ppi(_resolved_series["id"])

        ppi_l, ppi_r = st.columns([1.6, 1])
        with ppi_l:
            if _ppi_data and len(_ppi_data) >= 3:
                _ppi_df = pd.DataFrame(_ppi_data)
                _ppi_first = _ppi_df["value"].iloc[0]
                _ppi_last = _ppi_df["value"].iloc[-1]
                _ppi_chg = round(((_ppi_last - _ppi_first) / _ppi_first) * 100, 1) if _ppi_first > 0 else 0
                _ppi_color = "#F87171" if _ppi_chg > 0 else "#4ADE80"

                fig_ppi = go.Figure()
                fig_ppi.add_trace(go.Scatter(
                    x=_ppi_df["period"], y=_ppi_df["value"],
                    mode="lines+markers",
                    line=dict(color="#60A5FA", width=2),
                    marker=dict(size=4, color="#60A5FA"),
                    fill="tozeroy", fillcolor="rgba(96,165,250,0.08)",
                    name=_resolved_series["label"],
                ))
                fig_ppi.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,38,0.5)",
                    font=dict(color="#E2E8F0", size=10), height=220,
                    margin=dict(l=10, r=10, t=30, b=10),
                    xaxis=dict(color="#64748B", showgrid=False),
                    yaxis=dict(color="#64748B", gridcolor="rgba(255,255,255,0.05)"),
                    title=dict(text=f"{_resolved_series['label']} — Last 18 Months", font=dict(size=11, color="#94A3B8"), x=0),
                    showlegend=False,
                )
                st.plotly_chart(fig_ppi, use_container_width=True)
            else:
                _hint_key = next((k for k in _PPI_INFLATION_HINTS if k in _cat_key_lower), None)
                _hint_inf = _PPI_INFLATION_HINTS.get(_hint_key, 4.5) if _hint_key else 4.5
                _hint_color = "#F87171" if _hint_inf > 3 else "#4ADE80"
                st.markdown(
                    f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);border-radius:8px;padding:1rem;margin-bottom:0.5rem">'
                    f'<div style="font-size:0.6rem;color:#FCD34D;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.4rem">Category Inflation Benchmark (BLS/ISM/Gartner Estimate)</div>'
                    f'<div style="font-size:1.6rem;font-weight:700;color:{_hint_color};font-family:monospace">+{_hint_inf:.1f}% <span style="font-size:0.85rem;color:#94A3B8">per year</span></div>'
                    f'<div style="font-size:0.8rem;color:#C4D3E8;margin-top:0.4rem">{_resolved_series["label"]} · Live BLS data temporarily unavailable.</div>'
                    f'<div style="font-size:0.75rem;color:#64748B;margin-top:0.25rem">Use this benchmark to anchor supplier price increase negotiations. Refresh to retry live data.</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        with ppi_r:
            if _ppi_data and len(_ppi_data) >= 3:
                st.markdown(
                    f'<div style="background:#0D1526;border:1px solid rgba(96,165,250,0.18);border-radius:10px;padding:0.9rem 1rem;margin-bottom:0.6rem">'
                    f'<div style="font-size:0.6rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.4rem">18-Month Change</div>'
                    f'<div style="font-size:1.8rem;font-weight:700;color:{_ppi_color};font-family:monospace">{_ppi_chg:+.1f}%</div>'
                    f'<div style="font-size:0.82rem;color:#D0E0EF;margin-top:0.2rem">{_resolved_series["label"]}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f'<div style="background:#060D1A;border:1px solid rgba(252,211,77,0.12);border-radius:8px;padding:0.75rem 0.9rem">'
                    f'<div style="font-size:0.6rem;color:#FCD34D;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.4rem">Negotiation Implication</div>'
                    f'<div style="font-size:0.8rem;color:#CBD5E1;line-height:1.5">'
                    f'{"Prices in this category have risen " + str(abs(_ppi_chg)) + "% over 18 months. Supplier price increase requests are partially supported by market data — counter with a renewal cap, not flat rejection." if _ppi_chg > 3 else "Prices in this category are relatively stable. Any supplier price increase request above CPI requires category-specific justification. Use this as evidence." if _ppi_chg > 0 else "Prices in this category have declined " + str(abs(_ppi_chg)) + "% over 18 months. You have market data to push back on any price increase and argue for a reduction."}'
                    f'</div></div>',
                    unsafe_allow_html=True,
                )
            st.markdown(
                '<div style="font-size:0.80rem;color:#A8BEDC;margin-top:0.6rem">'
                '📊 BLS public data · No API key required · Updated monthly · '
                f'Series: {_resolved_series["id"]}</div>',
                unsafe_allow_html=True,
            )

    # ── MARKET TAB — SEC EDGAR Full-Text Supplier Filing Search ──
    with tab_market:
        st.markdown("---")
        st.markdown("#### SEC EDGAR Filing Intelligence")
        st.markdown(
            '<p class="muted">Search SEC full-text filings for any keyword — find how public suppliers '
            'are disclosing supply chain risks, price pressures, capacity constraints, or regulatory changes. '
            'This is free, real-time, and gives you an information edge over what suppliers tell you.</p>',
            unsafe_allow_html=True,
        )
        _edgar_kw_col, _edgar_btn_col = st.columns([3, 1])
        with _edgar_kw_col:
            _edgar_search_kw = st.text_input(
                "Search SEC filings for keyword",
                value=selected_sub_name.split("(")[0].strip()[:40],
                key="edgar_fts_kw",
                placeholder="e.g. 'supply chain disruption', 'price increase', 'capacity constraints'",
            )
        with _edgar_btn_col:
            _edgar_fetch = st.button("Search EDGAR", key="edgar_fts_fetch", type="primary")

        if _edgar_fetch and _edgar_search_kw.strip():
            with st.spinner("Searching SEC EDGAR full-text..."):
                try:
                    _fts_url    = "https://efts.sec.gov/LATEST/search-index"
                    _fts_params = {
                        "q": f'"{_edgar_search_kw.strip()}"',
                        "dateRange": "custom",
                        "startdt": "2023-01-01",
                        "enddt": "2025-12-31",
                        "forms": "10-K,10-Q",
                        "_source": "file_date,display_names,period_of_report,form_type",
                        "hits.hits.total.value": "true",
                        "hits.hits._source.period_of_report": "true",
                    }
                    _fts_resp = requests.get(
                        "https://efts.sec.gov/LATEST/search-index",
                        params={"q": f'"{_edgar_search_kw.strip()}"',
                                "dateRange": "custom",
                                "startdt": "2023-01-01",
                                "enddt": "2025-12-31",
                                "forms": "10-K,10-Q",
                                "_source": "file_date,display_names,period_of_report,form_type"},
                        headers={"User-Agent": "ProcureIQ research@procureiq.ai"},
                        timeout=12,
                    )
                    _fts_data = _fts_resp.json()
                    _hits_container = _fts_data.get("hits", {}) if isinstance(_fts_data, dict) else {}
                    if not isinstance(_hits_container, dict):
                        _hits_container = {}
                    _hits = _hits_container.get("hits", [])[:8]
                    _total_raw = _hits_container.get("total", 0)
                    _total = _total_raw.get("value", 0) if isinstance(_total_raw, dict) else int(_total_raw) if isinstance(_total_raw, (int, float)) else 0
                    if _hits:
                        st.markdown(
                            f'<div style="background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.15);'
                            f'border-radius:8px;padding:0.7rem 1rem;margin-bottom:1rem;font-size:0.88rem;color:#C4D3E8">'
                            f'<strong style="color:#60A5FA">{_total:,} total filings</strong> mention '
                            f'<em>"{html.escape(_edgar_search_kw.strip())}"</em> · Showing top {len(_hits)}</div>',
                            unsafe_allow_html=True,
                        )
                        for _hit in _hits:
                            _src  = _hit.get("_source", {})
                            _co   = ", ".join(d.get("name","") for d in _src.get("display_names", [])[:2])
                            _form = _src.get("form_type", "")
                            _dt   = _src.get("file_date", "")
                            _per  = _src.get("period_of_report", "")
                            st.markdown(
                                f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(96,165,250,0.10);'
                                f'border-radius:8px;padding:0.6rem 1rem;margin-bottom:0.35rem;'
                                f'display:flex;justify-content:space-between;align-items:center;gap:1rem">'
                                f'<div style="flex:1">'
                                f'<div style="font-weight:600;color:#E2E8F0;font-size:0.88rem">{html.escape(_co)}</div>'
                                f'<div style="font-size:0.78rem;color:#A8BEDC">Period: {html.escape(_per)}</div>'
                                f'</div>'
                                f'<div style="font-size:0.78rem;color:#60A5FA;white-space:nowrap">{html.escape(_form)} · {html.escape(_dt)}</div>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )
                    else:
                        st.info("No SEC filings matched that keyword. Try broader terms.")
                except Exception as _e:
                    st.warning(f"EDGAR search unavailable: {_e}")

        # ── Workflow nudge ─────────────────────────────────────────────────
        st.markdown(
            '<div style="background:rgba(29,78,216,0.08);border:1px solid rgba(96,165,250,0.25);'
            'border-radius:10px;padding:0.8rem 1.1rem;margin-top:1.5rem">'
            '<div style="font-size:0.78rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.09em;'
            'font-weight:700;margin-bottom:0.25rem">Next step in your workflow</div>'
            '<div style="font-size:0.92rem;color:#E2E8F0;font-weight:600">'
            '→ Open the <strong style="color:#60A5FA">Supplier Evaluation</strong> tab '
            'to score and compare your shortlisted suppliers.</div>'
            '<div style="font-size:0.82rem;color:#A8BEDC;margin-top:0.3rem;line-height:1.55">'
            'Use the price benchmarks and macro signals from this tab to calibrate '
            '<strong>Price / TCO</strong> scores. Use the Executive Snapshot above to '
            'confirm your current recommendation holds under market conditions.</div>'
            '</div>',
            unsafe_allow_html=True,
        )

    # ── STRATEGY TAB (Contracts + Sourcing Playbook) ───────
    with tab_strategy:
        render_supplier_intel_strip(leader)
        st.markdown("### RAQSCI Contract Guidance")
        st.markdown('<p class="muted">Procurement-specific contract guidance split by direct vs indirect. This is not legal redlining automation — it is structured thinking about what must be protected.</p>', unsafe_allow_html=True)

        st.markdown(
            f"""
            <div class="panel">
                <strong style="color:#F1F5F9">{category_rule['tag']}</strong>
                &nbsp; {type_badge(category_rule['type'])}<br>
                <span class="muted" style="font-size:0.92rem">{KRALJIC_INFO[kraljic]['axis']} · {KRALJIC_INFO[kraljic]['desc']}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        contract_set = category_raqsci(kraljic, category_rule)
        for section_name, content in contract_set.items():
            with st.expander(section_name, expanded=(section_name == "Requirements")):
                st.markdown(f'<div class="must-have"><strong>★ Must-have</strong><br><span style="font-size:0.88rem;color:#C4D3E8">{content["must"]}</span></div>', unsafe_allow_html=True)
                st.markdown(f'<div class="recommended"><strong>Strongly recommended</strong><br><span style="font-size:0.88rem;color:#C4D3E8">{content["recommended"]}</span></div>', unsafe_allow_html=True)
                st.markdown(f'<div class="nice-have"><strong>Nice to have</strong><br><span style="font-size:0.88rem;color:#C4D3E8">{content["nice"]}</span></div>', unsafe_allow_html=True)

        st.markdown('<div class="panel">', unsafe_allow_html=True)
        st.markdown('<div class="panel-eyebrow">What this still does not solve</div>', unsafe_allow_html=True)
        st.markdown(
            '<p class="muted">Detailed legal redlining · Should-cost modeling · Supplier capacity and plant-level constraints · Commodity index logic by region · Deep live private-supplier risk intelligence</p>',
            unsafe_allow_html=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("---")

    # ── SOURCING PLAYBOOK (merged into Strategy) ───────────
    with tab_strategy:
        st.markdown("### Sourcing Playbook")
        st.markdown(
            f'<p class="muted">Auto-generated for <strong style="color:#F1F5F9">{sx(selected_sub_name)}</strong> '
            f'({sx(category_rule["tag"])}) under a <strong style="color:#F1F5F9">{sx(kraljic)}</strong> posture. '
            f'Stakeholders, timeline, and risk flags all update dynamically from your Intake selections.</p>',
            unsafe_allow_html=True,
        )

        col_stake, col_timeline = st.columns([1, 1.2])

        with col_stake:
            st.markdown("#### Required Stakeholder Team")
            st.markdown(
                f'<p class="muted" style="font-size:0.78rem">Source: Subcategory — <strong style="color:#60A5FA">{sx(selected_sub_name)}</strong></p>',
                unsafe_allow_html=True,
            )

            # DYNAMIC: use subcategory-specific stakeholders, not generic category rules
            sub_stakes = selected_sub.get("stakeholders", [])
            rfp_stakes = category_rule.get("rfp_stakeholders", DEFAULT_RFP_STAKEHOLDERS)

            # Build must/recommended/nice from subcategory stakeholders
            # Subcategory gives a flat list — split into must (first half) vs recommended (rest)
            sub_must = sub_stakes[:max(3, len(sub_stakes)//2)]
            sub_rec  = sub_stakes[max(3, len(sub_stakes)//2):]
            sub_nice = rfp_stakes.get("nice", [])  # nice-to-have comes from category rule

            st.markdown(
                f"""
                <div class="rfp-stakeholder-group">
                    <div class="rfp-stakeholder-group-title" style="color:#F87171">
                        ★ Must Include — {sx(selected_sub_name)}
                    </div>
                    {"".join(f'<span class="rfp-stakeholder-pill">👤 {sx(s)}</span>' for s in sub_must)}
                </div>
                """,
                unsafe_allow_html=True,
            )

            if sub_rec:
                st.markdown(
                    f"""
                    <div class="rfp-stakeholder-group">
                        <div class="rfp-stakeholder-group-title" style="color:#FCD34D">
                            Strongly Recommended
                        </div>
                        {"".join(f'<span class="rfp-stakeholder-pill">👤 {sx(s)}</span>' for s in sub_rec)}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            st.markdown(
                f"""
                <div class="rfp-stakeholder-group">
                    <div class="rfp-stakeholder-group-title" style="color:#4ADE80">
                        Nice to Have
                    </div>
                    {"".join(f'<span class="rfp-stakeholder-pill">👤 {sx(s)}</span>' for s in sub_nice)}
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.markdown('<div class="panel" style="margin-top:0.5rem">', unsafe_allow_html=True)
            st.markdown('<div class="panel-eyebrow">Why Stakeholder Selection Matters</div>', unsafe_allow_html=True)
            st.markdown(
                f'<p style="font-size:0.85rem;color:#C4D3E8">In a <strong style="color:#F1F5F9">{sx(kraljic)}</strong> category like <strong style="color:#60A5FA">{sx(selected_sub_name)}</strong>, '
                f'sourcing does not fail because of suppliers — it fails because the wrong people are in the room too late. '
                f'The must-have list above is specific to this subcategory\'s risk profile: <em>{sx(selected_sub.get("key_risks",""))}</em>.</p>',
                unsafe_allow_html=True,
            )
            st.markdown("</div>", unsafe_allow_html=True)

        with col_timeline:
            st.markdown("#### Week-by-Week RFP Timeline")
            timeline = RFP_TIMELINE.get(kraljic, RFP_TIMELINE["Strategic"])

            # Build milestone tracker state
            if "milestone_status" not in st.session_state:
                st.session_state["milestone_status"] = {}

            _STATUS_OPTS   = ["Not Started", "In Progress", "Complete", "Blocked"]
            _STATUS_COLORS = {
                "Not Started": "#475569",
                "In Progress": "#60A5FA",
                "Complete":    "#4ADE80",
                "Blocked":     "#EF4444",
            }

            _milestone_rows = []
            for week_block in timeline:
                phase = week_block["phase"]
                color = PHASE_COLORS.get(phase, "#3B7BF8")
                st.markdown(
                    f'<div style="font-size:0.82rem;color:{color};font-family:monospace;'
                    f'letter-spacing:0.1em;text-transform:uppercase;margin-top:0.8rem;'
                    f'margin-bottom:0.2rem">{week_block["week"]} — {phase}</div>',
                    unsafe_allow_html=True,
                )
                for task in week_block["tasks"]:
                    _mk = f"ms_{week_block['week']}_{task[:24]}"
                    _mc1, _mc2, _mc3 = st.columns([3, 2, 2])
                    with _mc1:
                        st.markdown(
                            f'<div style="font-size:0.85rem;color:#E2E8F0;padding:0.3rem 0">{task}</div>',
                            unsafe_allow_html=True,
                        )
                    with _mc2:
                        _cur_status = st.session_state["milestone_status"].get(_mk, "Not Started")
                        _status = st.selectbox(
                            "Status",
                            _STATUS_OPTS,
                            index=_STATUS_OPTS.index(_cur_status),
                            key=f"sel_{_mk}",
                            label_visibility="collapsed",
                        )
                        st.session_state["milestone_status"][_mk] = _status
                    with _mc3:
                        _owner = st.text_input(
                            "Owner", value="", key=f"own_{_mk}",
                            placeholder="Owner name", label_visibility="collapsed",
                        )
                    _milestone_rows.append({
                        "Task": task,
                        "Week": week_block["week"],
                        "Phase": phase,
                        "Status": _status,
                        "Assigned To": _owner,
                        "Start Date": "",
                        "End Date": "",
                        "Duration": "",
                    })

            if st.button("Export Milestone Tracker (Smartsheets CSV)", key="export_milestones"):
                _ms_df  = pd.DataFrame(_milestone_rows)
                _ms_csv = _ms_df[["Task", "Start Date", "End Date", "Duration", "Assigned To", "Status"]].to_csv(index=False)
                st.download_button(
                    "⬇️ Download for Smartsheets",
                    data=_ms_csv,
                    file_name=f"RFP_Milestones_{selected_sub_name.replace(' ', '_')}.csv",
                    mime="text/csv",
                    key="dl_milestones",
                )

        # ── Capital Project Timeline Calculator ───────────────
        from datetime import date as _date, timedelta as _td
        st.markdown("---")
        st.markdown("#### Capital Project Timeline Calculator")
        st.markdown(
            '<p class="muted">Enter your equipment-on-site deadline. '
            'The system back-calculates when you must launch your RFP based on '
            "this category's manufacturing lead time and RFP duration.</p>",
            unsafe_allow_html=True,
        )
        _LEAD_TIME_DAYS = {"Strategic": 120, "Leverage": 60, "Bottleneck": 90, "Non-Critical": 30}
        _rfp_dur_days   = len(timeline) * 7
        _mfg_lead       = _LEAD_TIME_DAYS.get(kraljic, 90)

        _cpj_col1, _cpj_col2 = st.columns(2)
        with _cpj_col1:
            _need_by = st.date_input(
                "Equipment / Service Needed On-Site By",
                value=(_date.today() + _td(days=180)),
                key="capital_need_by",
                help="The hard project milestone date — drives back-calculation.",
            )
        with _cpj_col2:
            _install_days = st.number_input(
                "Installation / Commissioning Days",
                min_value=0, max_value=180, value=14, step=7,
                key="capital_install_days",
                help="Days between delivery and handover to operations.",
            )

        _delivery_deadline = _need_by - _td(days=int(_install_days))
        _rfp_launch_by     = _delivery_deadline - _td(days=_mfg_lead + _rfp_dur_days)
        _days_to_launch    = (_rfp_launch_by - _date.today()).days
        _lc = "#EF4444" if _days_to_launch < 0 else "#FCD34D" if _days_to_launch < 30 else "#4ADE80"

        st.markdown(
            f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:1rem;margin-top:1rem">'
            f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(96,165,250,0.15);border-radius:10px;padding:1rem;text-align:center">'
            f'<div style="font-size:0.78rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em">Delivery Deadline</div>'
            f'<div style="font-size:1.05rem;font-weight:700;color:#60A5FA">{_delivery_deadline.strftime("%b %d, %Y")}</div></div>'
            f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(96,165,250,0.15);border-radius:10px;padding:1rem;text-align:center">'
            f'<div style="font-size:0.78rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em">Launch RFP By</div>'
            f'<div style="font-size:1.05rem;font-weight:700;color:{_lc}">{_rfp_launch_by.strftime("%b %d, %Y")}</div></div>'
            f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(96,165,250,0.15);border-radius:10px;padding:1rem;text-align:center">'
            f'<div style="font-size:0.78rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em">Days Until Launch</div>'
            f'<div style="font-size:1.05rem;font-weight:700;color:{_lc}">{_days_to_launch:+d}</div></div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        if _days_to_launch < 0:
            st.error(
                f"You are {abs(_days_to_launch)} days past the required RFP launch date. "
                f"Escalate immediately or compress RFP scope to recover schedule."
            )
        elif _days_to_launch < 30:
            st.warning(
                f"Launch window closes in {_days_to_launch} days. "
                f"Initiate stakeholder kickoff this week."
            )
        else:
            st.success(
                f"On track — {_days_to_launch} days until required RFP launch. "
                f"Assumes {_mfg_lead}d manufacturing lead + {_rfp_dur_days}d RFP process for {kraljic} category."
            )

        # ── Category Maturity Assessment ──────────────────────
        st.markdown("---")
        st.markdown("#### Category Maturity Assessment")
        st.markdown(
            '<p class="muted">Score your current capability across five CIPS-aligned dimensions. '
            'The radar shows current state vs. target. Use this to anchor the conversation with leadership '
            'on what investment is required to reach the next maturity level.</p>',
            unsafe_allow_html=True,
        )
        _mat_dims = [
            "Data Quality",
            "Market Knowledge",
            "Supplier Relationships",
            "Contract Coverage",
            "Spend Under Management",
        ]
        _mat_cols = st.columns(5)
        _mat_scores = {}
        for _mi, _md in enumerate(_mat_dims):
            with _mat_cols[_mi]:
                _mat_scores[_md] = st.select_slider(
                    _md,
                    options=[1, 2, 3, 4, 5],
                    value=3,
                    key=f"mat_{_md.lower().replace(' ', '_')}",
                )

        _mat_avg   = sum(_mat_scores.values()) / len(_mat_scores)
        _mat_label = (
            "Initial — Reactive"        if _mat_avg < 2   else
            "Developing — Transactional" if _mat_avg < 3   else
            "Defined — Analytical"      if _mat_avg < 4   else
            "Managed — Strategic"       if _mat_avg < 4.5 else
            "Optimizing — Value-Creating"
        )
        _mat_next = {
            "Initial — Reactive":         "Build spend visibility: get a clean data extract and classify it. Set up at least one contract per category.",
            "Developing — Transactional": "Run a competitive RFP, document a one-page category strategy, map your supply market.",
            "Defined — Analytical":       "Establish preferred supplier relationships, implement SLA scorecards, publish a category plan to stakeholders.",
            "Managed — Strategic":        "Co-develop innovation roadmaps with top suppliers, link category performance to business outcomes.",
            "Optimizing — Value-Creating":"Benchmark against external peers, lead supplier-led innovation sessions, publish ESG supply chain data.",
        }

        import plotly.graph_objects as _go_mat
        _radar_fig = _go_mat.Figure()
        _radar_fig.add_trace(_go_mat.Scatterpolar(
            r=list(_mat_scores.values()) + [list(_mat_scores.values())[0]],
            theta=_mat_dims + [_mat_dims[0]],
            fill="toself",
            fillcolor="rgba(59,130,246,0.15)",
            line=dict(color="#3B82F6", width=2),
            name="Current",
        ))
        _radar_fig.add_trace(_go_mat.Scatterpolar(
            r=[5] * len(_mat_dims) + [5],
            theta=_mat_dims + [_mat_dims[0]],
            fill="toself",
            fillcolor="rgba(74,222,128,0.04)",
            line=dict(color="#4ADE80", width=1, dash="dot"),
            name="Target",
        ))
        _radar_fig.update_layout(
            polar=dict(
                radialaxis=dict(visible=True, range=[0, 5], color="#475569", tickfont=dict(size=9)),
                angularaxis=dict(color="#8EA3C3"),
            ),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#8EA3C3", family="Outfit"),
            showlegend=True,
            height=320,
            margin=dict(t=20, b=10, l=40, r=40),
        )
        st.plotly_chart(_radar_fig, use_container_width=True)
        st.markdown(
            f'<div style="background:rgba(96,165,250,0.05);border:1px solid rgba(96,165,250,0.12);'
            f'border-radius:10px;padding:0.9rem 1.2rem;margin-top:0.5rem">'
            f'<div style="font-size:0.78rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.3rem">'
            f'Maturity Level — {html.escape(selected_sub_name)}</div>'
            f'<div style="font-size:1rem;font-weight:700;color:#60A5FA">{_mat_label}</div>'
            f'<div style="font-size:0.82rem;color:#C4D3E8;margin-top:0.4rem">'
            f'Average: {_mat_avg:.1f}/5.0 · '
            f'<strong style="color:#F1F5F9">To move up:</strong> {_mat_next.get(_mat_label, "")}'
            f'</div></div>',
            unsafe_allow_html=True,
        )

        # ── BLS Producer Price Index + World Bank Commodity ──
        st.markdown("---")
        st.markdown("#### Market Price Intelligence — Live Indices")
        st.markdown(
            '<p class="muted">Real commodity and producer price data from the Bureau of Labor Statistics '
            'and World Bank. Use these to defend or challenge price escalation clauses in negotiations — '
            '"the BLS PPI for this category is up X% YoY" is a fact, not a negotiating position.</p>',
            unsafe_allow_html=True,
        )
        _bls_series = BLS_PPI_MAP.get(category)
        _wb_indicator = WB_COMMODITY_MAP.get(selected_sub_name)

        _ppi_col, _wb_col = st.columns(2)
        with _ppi_col:
            st.markdown(
                f'<div style="font-size:0.70rem;color:#60A5FA;text-transform:uppercase;'
                f'letter-spacing:0.1em;margin-bottom:0.4rem">BLS Producer Price Index — {html.escape(str(category))}</div>',
                unsafe_allow_html=True,
            )
            if _bls_series:
                _bls_data = fetch_bls_ppi([_bls_series])
                _series_data = _bls_data.get(_bls_series, [])
                if _series_data:
                    _vals = [d["value"] for d in _series_data]
                    _pds  = [d["period"] for d in _series_data]
                    _pct_change = ((_vals[0] - _vals[-1]) / _vals[-1] * 100) if len(_vals) > 1 else 0
                    _chg_color  = "#EF4444" if _pct_change > 3 else "#4ADE80" if _pct_change < -1 else "#FCD34D"
                    st.markdown(
                        f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(96,165,250,0.12);'
                        f'border-radius:10px;padding:0.8rem 1rem">'
                        f'<div style="font-size:1.4rem;font-weight:700;color:#F1F5F9">{_vals[0]:.1f}</div>'
                        f'<div style="font-size:0.80rem;color:{_chg_color};margin-top:0.1rem">'
                        f'{_pct_change:+.1f}% over last {len(_vals)} periods · '
                        f'<span style="color:#A8BEDC">Series {_bls_series}</span></div>'
                        f'<div style="font-size:0.75rem;color:#8BAAC4;margin-top:0.3rem">'
                        + " → ".join(f"{v:.1f}" for v in reversed(_vals)) +
                        f'</div></div>',
                        unsafe_allow_html=True,
                    )
                    if abs(_pct_change) > 3:
                        st.markdown(
                            f'<div style="background:rgba(252,211,77,0.06);border-left:3px solid #FCD34D;'
                            f'border-radius:0 8px 8px 0;padding:0.5rem 0.9rem;margin-top:0.5rem;font-size:0.85rem">'
                            f'<strong style="color:#FCD34D">Negotiation Signal:</strong> '
                            f'<span style="color:#F1F5F9">PPI moved {_pct_change:+.1f}% — '
                            f'{"challenge any escalation clause above this level" if _pct_change > 0 else "use as leverage to demand a price reduction or credit"}.'
                            f'</span></div>',
                            unsafe_allow_html=True,
                        )
                else:
                    st.markdown('<div style="color:#8BAAC4;font-size:0.85rem">BLS data unavailable — check internet connection or try again.</div>', unsafe_allow_html=True)
            else:
                st.markdown(
                    f'<div style="color:#8BAAC4;font-size:0.85rem">No BLS PPI series mapped for '
                    f'<em>{html.escape(category)}</em>. Contact data team to add.</div>',
                    unsafe_allow_html=True,
                )

        with _wb_col:
            st.markdown(
                f'<div style="font-size:0.70rem;color:#A78BFA;text-transform:uppercase;'
                f'letter-spacing:0.1em;margin-bottom:0.4rem">World Bank Commodity Index — {selected_sub_name}</div>',
                unsafe_allow_html=True,
            )
            if _wb_indicator:
                _wb_data = fetch_worldbank_commodity(_wb_indicator)
                if _wb_data:
                    _wb_vals = [d["value"] for d in _wb_data if d["value"] is not None]
                    _wb_dates = [d["date"] for d in _wb_data if d["value"] is not None]
                    _wb_chg   = ((_wb_vals[0] - _wb_vals[-1]) / _wb_vals[-1] * 100) if len(_wb_vals) > 1 else 0
                    _wb_color = "#EF4444" if _wb_chg > 5 else "#4ADE80" if _wb_chg < -3 else "#FCD34D"
                    st.markdown(
                        f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(167,139,250,0.15);'
                        f'border-radius:10px;padding:0.8rem 1rem">'
                        f'<div style="font-size:1.4rem;font-weight:700;color:#F1F5F9">{_wb_vals[0]:.1f}</div>'
                        f'<div style="font-size:0.80rem;color:{_wb_color};margin-top:0.1rem">'
                        f'{_wb_chg:+.1f}% · {_wb_dates[0]} vs {_wb_dates[-1]}</div>'
                        f'<div style="font-size:0.75rem;color:#8BAAC4;margin-top:0.3rem">Source: World Bank Open Data</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown('<div style="color:#8BAAC4;font-size:0.85rem">World Bank data unavailable for this subcategory.</div>', unsafe_allow_html=True)
            else:
                st.markdown(
                    f'<div style="color:#8BAAC4;font-size:0.85rem">No commodity price index mapped for '
                    f'<em>{html.escape(selected_sub_name)}</em>.</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("---")
        st.markdown("#### Risk Flags for This Evaluation")
        st.markdown(
            '<p class="muted">These flags are generated from your actual supplier scores, stakeholder map, and Kraljic position — not generic templates.</p>',
            unsafe_allow_html=True,
        )

        risk_flags = generate_rfp_risk_flags(leader, runner_up, blocker, kraljic, category_rule)

        for flag in risk_flags:
            tier = flag["tier"]
            tier_class = {"HIGH": "risk-flag-high", "MEDIUM": "risk-flag-medium", "HIDDEN": "risk-flag-hidden"}.get(tier, "risk-flag-medium")
            tier_color_class = {"HIGH": "risk-tier-high", "MEDIUM": "risk-tier-medium", "HIDDEN": "risk-tier-hidden"}.get(tier, "risk-tier-medium")

            st.markdown(
                f"""
                <div class="risk-flag {tier_class}">
                    <div class="risk-flag-icon">{flag['icon']}</div>
                    <div>
                        <div class="risk-flag-tier {tier_color_class}">{tier} RISK</div>
                        <strong style="color:#F1F5F9; font-size:0.9rem">{flag['title']}</strong>
                        <div class="risk-flag-body" style="margin-top:0.3rem">{flag['body']}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        # ── Auction Event Type Recommendation (belongs in Sourcing Playbook) ──
        st.markdown("---")
        st.markdown("#### Procurement Event Type Recommendation")
        st.markdown(
            '<p style="color:#C4D3E8;font-size:0.86rem">The event type determines how you structure competition — it is part of the sourcing strategy, not just an admin choice. Coupa and Ariba both support all types below.</p>',
            unsafe_allow_html=True,
        )
        auction_type_name, auction_rationale = recommend_auction_type(
            kraljic, num_suppliers,
            weights.get("Price / TCO", 0.15),
            switching_cost_answer,
            selected_sub.get("auction", "RFP"),
        )
        st.session_state["_last_auction_type"] = auction_type_name  # share with AI tab
        auction_info = AUCTION_TYPES.get(auction_type_name, AUCTION_TYPES["Rank Auction (Coupa / Ariba Standard)"])

        st.markdown(
            f'<div class="panel" style="border-left:4px solid {auction_info["color"]}">'
            f'<div class="panel-eyebrow">Recommended Event Type for {sx(selected_sub_name)}</div>'
            f'<div style="font-size:1.3rem;font-weight:800;color:{auction_info["color"]};margin-bottom:0.4rem">{sx(auction_type_name)}</div>'
            f'<div style="font-size:0.9rem;color:#F1F5F9;margin-bottom:0.55rem">{sx(auction_info["desc"])}</div>'
            f'<div class="soft-blue" style="font-size:0.92rem;margin-bottom:0.45rem"><strong>Why this fits:</strong> {sx(auction_rationale)}</div>'
            f'<div style="background:rgba(167,139,250,0.08);border-left:3px solid #A78BFA;border-radius:0 6px 6px 0;padding:0.55rem 0.8rem;font-size:0.76rem;color:#C4D3E8">'
            f'<strong style="color:#A78BFA;font-family:monospace">PLATFORM PATH</strong> &nbsp;'
            f'{sx(auction_info.get("coupa_ariba","Configurable in Coupa or Ariba Sourcing module"))}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── MEETING PREP (merged into Stakeholders) ────────────
    with tab_stakeholders:
        st.markdown("---")
        st.markdown("### Meeting Prep — Executive Summary Generator")
        st.markdown(
            '<p class="muted">One-click output you can copy directly into a Slack message, email to your CPO, or use as an opening statement in the room.</p>',
            unsafe_allow_html=True,
        )

        exec_summary = build_executive_summary(
            leader, runner_up, blocker, event_name, kraljic, category_rule, leader_weakest_dim
        )

        st.markdown('<div class="exec-summary-box">', unsafe_allow_html=True)
        st.markdown('<div class="panel-eyebrow" style="color:#60A5FA; margin-bottom:0.8rem">Executive Summary · Ready to Send</div>', unsafe_allow_html=True)

        # Format summary with highlights
        # Escape the full summary first, then selectively bold known-safe tokens
        _esc_summary = html.escape(exec_summary)
        _bold_tokens = [
            (html.escape(f"Recommendation: {leader['Supplier']}"),
             f"<strong>Recommendation: {html.escape(leader['Supplier'])}</strong>"),
            (html.escape(f"{leader['Weighted Score']}/100"),
             f"<strong>{leader['Weighted Score']}/100</strong>"),
            (html.escape(leader_weakest_dim),
             f"<strong>{html.escape(leader_weakest_dim)}</strong>"),
        ]
        formatted = _esc_summary
        for _plain, _bold in _bold_tokens:
            formatted = formatted.replace(_plain, _bold, 1)

        st.markdown(f'<div class="exec-summary-text">{formatted}</div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        st.text_area(
            "Copy-paste version (plain text)",
            exec_summary,
            height=160,
            key="exec_summary_copy",
        )

        st.markdown("---")
        st.markdown("#### What this tab is built for")
        st.markdown(
            """
            <div class="panel">
                <p class="muted" style="font-size:0.88rem">
                    Every sourcing recommendation has two moments that matter: the analysis that produces it,
                    and the meeting where it either survives or gets killed by politics. This tab bridges the gap.<br><br>
                    The summary above pulls from your live evaluation — supplier scores, stakeholder risk, Kraljic posture,
                    and contract exposure — and compresses it into the 5 sentences a CPO or executive sponsor actually needs.<br><br>
                    <strong style="color:#F1F5F9">Use it as:</strong> an email opener · a Slack summary · your first 30 seconds in the room · a pre-meeting alignment message to your champion.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown("#### Pre-Meeting Checklist")
        checklist_items = [
            f"Recommended supplier confirmed: {leader['Supplier']}",
            f"Weakest dimension identified and mitigation ready: {leader_weakest_dim}",
            f"Blocker engaged one-on-one before the meeting: {blocker['Name'] if blocker is not None else 'No critical blocker detected'}",
            "Champion briefed and ready to speak first",
            "Trade-off narrative rehearsed — know your answer when the room asks about price",
            "Contract negotiation targets aligned with recommendation logic",
        ]
        for item in checklist_items:
            st.markdown(
                f'<div class="soft-blue" style="margin-bottom:0.5rem; font-size:0.86rem">✓ {item}</div>',
                unsafe_allow_html=True,
            )

    # ── NEGOTIATE TAB (Negotiation + What-If) ──────────────
    with tab_negotiate:
        render_supplier_intel_strip(leader)
        st.markdown("### Negotiation Position")
        st.markdown(
            f'<p style="color:#C4D3E8;font-size:0.88rem">Everything here is calculated from your actual supplier data — '
            f'not static templates. Move supplier scores or prices and this tab updates.</p>',
            unsafe_allow_html=True,
        )

        # ── BATNA Bar — the core negotiation fact ────────────────────────────
        _batna_gap = round(leader["Weighted Score"] - runner_up["Weighted Score"], 1) if runner_up else None
        _price_delta_pct = round(((leader["Raw Price"] - runner_up["Raw Price"]) / runner_up["Raw Price"]) * 100, 1) if runner_up and runner_up["Raw Price"] > 0 else None

        st.markdown(
            f'<div style="background:linear-gradient(135deg,rgba(9,24,48,0.95),rgba(14,30,60,0.9));'
            f'border:1px solid rgba(96,165,250,0.25);border-radius:12px;padding:1.2rem 1.4rem;margin-bottom:1rem">'
            f'<div style="font-size:0.78rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.16em;margin-bottom:0.7rem;font-family:monospace">Your BATNA — Best Alternative to a Negotiated Agreement</div>'
            f'<div style="display:flex;gap:2.5rem;flex-wrap:wrap">'
            f'<div><div style="font-size:1.8rem;font-weight:700;color:#60A5FA;font-family:monospace;line-height:1">{html.escape(leader["Supplier"])}</div>'
            f'<div style="font-size:0.82rem;color:#D0E0EF;margin-top:0.2rem">Recommended · {leader["Weighted Score"]}/100</div></div>'
            + (f'<div style="display:flex;align-items:center;padding:0 1rem;border-left:1px solid rgba(96,165,250,0.12);border-right:1px solid rgba(96,165,250,0.12)">'
               f'<div style="text-align:center">'
               f'<div style="font-size:0.82rem;color:#D0E0EF;margin-bottom:0.2rem">Score gap</div>'
               f'<div style="font-size:1.4rem;font-weight:700;color:{"#4ADE80" if _batna_gap and _batna_gap >= 5 else "#FCD34D"};font-family:monospace">'
               f'+{_batna_gap} pts</div>'
               f'<div style="font-size:0.80rem;color:#D0E0EF">vs runner-up</div>'
               f'</div></div>'
               if runner_up else '') +
            (f'<div><div style="font-size:1.4rem;font-weight:700;color:{"#F87171" if _price_delta_pct and _price_delta_pct > 0 else "#4ADE80"};font-family:monospace;line-height:1">'
             f'{"+" if _price_delta_pct and _price_delta_pct > 0 else ""}{_price_delta_pct}%</div>'
             f'<div style="font-size:0.82rem;color:#D0E0EF;margin-top:0.2rem">price vs {html.escape(runner_up["Supplier"])}</div>'
             f'<div style="font-size:0.82rem;color:#A8BEDC">${leader["Raw Price"]:,.0f} vs ${runner_up["Raw Price"]:,.0f}</div>'
             f'</div>'
             if runner_up else '') +
            f'</div>'
            f'<div style="font-size:0.78rem;color:#C4D3E8;margin-top:0.8rem;padding-top:0.7rem;border-top:1px solid rgba(96,165,250,0.08)">'
            f'{"Your BATNA is strong — a " + str(_batna_gap) + "-pt lead means you can walk away if the supplier refuses key concessions." if _batna_gap and _batna_gap >= 5 else "Thin lead — be careful how hard you push. The runner-up is close enough that the room may question your recommendation."}'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Side-by-side supplier comparison ─────────────────────────────────
        st.markdown("#### Supplier-by-Supplier Comparison")
        st.markdown('<p style="color:#D0E0EF;font-size:0.8rem;margin-bottom:0.8rem">Every column is live data from the Suppliers tab. This changes as you update scores.</p>', unsafe_allow_html=True)

        # Header row
        header_cols = st.columns([1.4] + [1] * len(ranked))
        with header_cols[0]:
            st.markdown('<div style="font-size:0.78rem;color:#9EB8CE;text-transform:uppercase;letter-spacing:0.12em;padding:0.4rem 0">Dimension</div>', unsafe_allow_html=True)
        sup_colors = ["#60A5FA", "#4ADE80", "#FCD34D", "#F87171", "#A78BFA", "#FB923C", "#34D399", "#F472B6"]
        for idx, s in enumerate(ranked):
            with header_cols[idx + 1]:
                is_leader = idx == 0
                _cl = _confidence_label(s)
                _cc = _confidence_color(_cl)
                _sc = _score_color(s["Weighted Score"])
                st.markdown(
                    f'<div style="text-align:center;padding:0.4rem 0.2rem;border-radius:6px;'
                    f'background:{"rgba(29,78,216,0.1)" if is_leader else "transparent"};'
                    f'border:{"1px solid rgba(96,165,250,0.2)" if is_leader else "1px solid transparent"}">'
                    f'<div style="font-size:0.82rem;font-weight:700;color:{sup_colors[idx % len(sup_colors)]};font-family:monospace">{html.escape(sx(s["Supplier"][:12]))}</div>'
                    f'<div style="font-size:0.78rem;color:{_sc};font-weight:700">{s["Weighted Score"]}/100{"  ✓" if is_leader else ""}</div>'
                    f'<div style="font-size:0.62rem;color:{_cc};letter-spacing:0.06em">{_cl} CONFIDENCE</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        st.markdown('<div style="height:0.3rem"></div>', unsafe_allow_html=True)

        # Dimension rows
        dim_rows = DIMENSIONS + ["Weighted Score", "Price ($)", "Financial Health", "Current Fit", "Future Fit"]
        for dim in dim_rows:
            row_cols = st.columns([1.4] + [1] * len(ranked))
            # Get scores for this dimension across all suppliers
            if dim == "Weighted Score":
                vals = [s["Weighted Score"] for s in ranked]
                best_fn = max
                fmt_fn  = lambda v: f"{v}/100"
            elif dim == "Price ($)":
                vals = [s["Raw Price"] for s in ranked]
                best_fn = min  # lower price = better
                fmt_fn  = lambda v: f"${v:,.0f}"
            elif dim == "Financial Health":
                vals = [s["Financial Health"] for s in ranked]
                best_fn = max
                fmt_fn  = lambda v: f"{v}/100"
            elif dim == "Current Fit":
                vals = [s["Current Fit"] for s in ranked]
                best_fn = max
                fmt_fn  = lambda v: str(v)
            elif dim == "Future Fit":
                vals = [s["Future Fit"] for s in ranked]
                best_fn = max
                fmt_fn  = lambda v: str(v)
            else:
                vals = [s["Scores"][dim] for s in ranked]
                best_fn = max
                fmt_fn  = lambda v: str(v)

            best_val = best_fn(vals)
            with row_cols[0]:
                st.markdown(
                    f'<div style="font-size:0.75rem;color:#C4D3E8;padding:0.35rem 0;border-bottom:1px solid rgba(96,165,250,0.04)">'
                    f'{"─ " if dim in ["Price ($)","Financial Health","Current Fit","Future Fit"] else ""}{dim}</div>',
                    unsafe_allow_html=True,
                )
            for idx, (s, val) in enumerate(zip(ranked, vals)):
                is_best = val == best_val
                with row_cols[idx + 1]:
                    st.markdown(
                        f'<div style="text-align:center;padding:0.35rem 0.2rem;border-bottom:1px solid rgba(96,165,250,0.04);'
                        f'background:{"rgba(29,78,216,0.06)" if is_best else "transparent"}">'
                        f'<span style="font-family:monospace;font-size:0.78rem;font-weight:{"700" if is_best else "400"};'
                        f'color:{"#60A5FA" if is_best else "#64748B"}">{fmt_fn(val)}'
                        f'{"  ↑" if is_best and dim != "Price ($)" else ("  ↓" if is_best and dim == "Price ($)" else "")}'
                        f'</span></div>',
                        unsafe_allow_html=True,
                    )

        # ── Negotiation leverage points ───────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Negotiation Leverage Points")
        neg_l, neg_r = st.columns(2)

        with neg_l:
            st.markdown('<div style="font-size:0.80rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.14em;margin-bottom:0.6rem;font-family:monospace">Where you have leverage</div>', unsafe_allow_html=True)
            for point in default_negotiation_points(kraljic, category_rule, leader_weakest_dim):
                st.markdown(
                    f'<div style="background:#060D1A;border-left:2px solid #3B82F6;padding:0.55rem 0.8rem;'
                    f'margin-bottom:0.4rem;border-radius:0 6px 6px 0;font-size:0.82rem;color:#CBD5E1">'
                    f'→ {point}</div>',
                    unsafe_allow_html=True,
                )

        with neg_r:
            st.markdown('<div style="font-size:0.80rem;color:#F87171;text-transform:uppercase;letter-spacing:0.14em;margin-bottom:0.6rem;font-family:monospace">What the supplier will push back on</div>', unsafe_allow_html=True)
            # Dynamic pushback points based on actual dimension scores
            pushbacks = []
            if leader["Scores"].get("Price / TCO", 100) > 70:
                pushbacks.append(f"Price — {leader['Supplier']} scores well on price already. They know it.")
            if leader["Scores"].get("Execution Risk", 0) < 60:
                pushbacks.append(f"SLA penalty language — their execution score is {leader['Scores'].get('Execution Risk', 0)}/100. They'll resist penalties.")
            if leader["Scores"].get("Commercial Flexibility", 0) < 60:
                pushbacks.append(f"Renewal cap language — commercial flexibility scores low ({leader['Scores'].get('Commercial Flexibility',0)}/100).")
            if leader["Financial Health"] < 65:
                pushbacks.append(f"Payment terms — financial health at {leader['Financial Health']}/100 means cash flow matters to them.")
            if not pushbacks:
                pushbacks = ["Strong overall supplier — be prepared for minimal concessions.", "Focus on contract structure, not price reductions."]
            for pb in pushbacks:
                st.markdown(
                    f'<div style="background:rgba(248,113,113,0.05);border-left:2px solid #F87171;'
                    f'padding:0.55rem 0.8rem;margin-bottom:0.4rem;border-radius:0 6px 6px 0;font-size:0.82rem;color:#CBD5E1">'
                    f'⚠ {pb}</div>',
                    unsafe_allow_html=True,
                )

        # ── SEC filing context for negotiation ───────────────────────────────
        if leader.get("SEC Context") and leader["SEC Context"].get("found"):
            sec_ctx = leader["SEC Context"]
            recent_filings = sec_ctx.get("recent_filings", [])
            if recent_filings:
                st.markdown("---")
                st.markdown(f'<div style="font-size:0.80rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.14em;margin-bottom:0.5rem;font-family:monospace">SEC Filing Intel — {html.escape(sx(leader["Supplier"]))}</div>', unsafe_allow_html=True)
                st.markdown(
                    '<p style="color:#D0E0EF;font-size:0.78rem">Recent filings signal financial health and corporate activity. Use to inform negotiation posture.</p>',
                    unsafe_allow_html=True,
                )
                filings_html = " &nbsp;·&nbsp; ".join(f'<span style="font-family:monospace;font-size:0.85rem;color:#60A5FA;background:rgba(96,165,250,0.08);border-radius:4px;padding:0.1rem 0.4rem">{f}</span>' for f in recent_filings[:4])
                st.markdown(filings_html, unsafe_allow_html=True)

    # ── WHAT-IF (merged into Negotiate) ──────────────────────
    with tab_negotiate:
        st.markdown("---")
        st.markdown("### What-If Sensitivity Analysis")
        # ── CFO/CPO Framing — the question this tab answers ──────────────────
        st.markdown(
            f'<div style="background:rgba(252,211,77,0.06);border-left:3px solid #FCD34D;'
            f'border-radius:0 10px 10px 0;padding:0.9rem 1.1rem;margin-bottom:1rem">'
            f'<div style="font-size:0.78rem;color:#FCD34D;text-transform:uppercase;letter-spacing:0.14em;'
            f'margin-bottom:0.35rem;font-family:monospace">The Question the Room Will Ask</div>'
            f'<div style="font-size:0.92rem;color:#E2E8F0;font-weight:600;margin-bottom:0.3rem">'
            f'"What happens to your recommendation if {sx(leader["Supplier"])}\'s price goes up 15%?"</div>'
            f'<div style="font-size:0.82rem;color:#C4D3E8;line-height:1.6">'
            f'This tab answers that — live. Adjust any price, SLA, or weight below and the scoring '
            f'recalculates instantly. If the recommendation flips, that\'s what you need to defend. '
            f'If it holds, that\'s your closing argument. '
            f'<strong style="color:#CBD5E1">Use this before every presentation.</strong></div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Base state summary ───────────────────────────────
        st.markdown(
            f'<div style="background:#0D1526;border:1px solid rgba(96,165,250,0.2);border-radius:10px;'
            f'padding:0.8rem 1rem;margin-bottom:1rem;display:flex;gap:2rem;flex-wrap:wrap">'
            f'<span style="font-size:0.85rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.1em">Base State</span>'
            f'<span style="font-size:0.85rem;color:#E2E8F0">Winner: <strong style="color:#60A5FA">{sx(leader["Supplier"])}</strong></span>'
            f'<span style="font-size:0.85rem;color:#E2E8F0">Score: <strong>{leader["Weighted Score"]}/100</strong></span>'
            f'<span style="font-size:0.85rem;color:#E2E8F0">Gap to runner-up: <strong style="color:#4ADE80">'
            f'{round(leader["Weighted Score"] - runner_up["Weighted Score"], 1) if runner_up else "N/A"} pts</strong></span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        wa_col, wb_col = st.columns([1, 1])

        with wa_col:
            st.markdown("#### Price Shock Simulator")
            st.markdown(
                '<p style="font-size:0.82rem;color:#C4D3E8;margin-bottom:0.8rem">'
                'Apply a price change to any supplier. The tool re-scores all suppliers and shows if the recommendation flips.</p>',
                unsafe_allow_html=True,
            )
            price_shocks = {}
            for _shock_i, s in enumerate(ranked):
                shock_pct = st.slider(
                    f"{s['Supplier']} price change",
                    min_value=-30, max_value=50, value=0, step=5,
                    format="%d%%",
                    key=f"shock_{_shock_i}_{s['Supplier'][:30]}",
                    help=f"Current: ${s['Raw Price']:,.0f}. Shift to see score impact.",
                )
                if shock_pct != 0:
                    price_shocks[s["Supplier"]] = shock_pct / 100.0
                    new_price = s["Raw Price"] * (1 + shock_pct / 100.0)
                    st.markdown(
                        f'<div style="font-size:0.76rem;color:#FCD34D;margin-top:-0.4rem;margin-bottom:0.4rem;">'
                        f'→ ${new_price:,.0f} ({shock_pct:+d}%)</div>',
                        unsafe_allow_html=True,
                    )

            st.markdown("#### SLA Override")
            sla_overrides = {}
            for _sla_i, s in enumerate(ranked):
                new_sla = st.selectbox(
                    f"{s['Supplier']} SLA",
                    ["(no change)", "Strong", "Moderate", "Weak"],
                    key=f"sla_shock_{_sla_i}_{s['Supplier'][:30]}",
                )
                if new_sla != "(no change)":
                    sla_overrides[s["Supplier"]] = new_sla

        with wb_col:
            st.markdown("#### Weight Stress Test")
            st.markdown(
                '<p style="font-size:0.82rem;color:#C4D3E8;margin-bottom:0.8rem">'
                'Shift the evaluation weight of any dimension to see how sensitive the recommendation is.</p>',
                unsafe_allow_html=True,
            )
            weight_overrides_raw = {}
            for dim in DIMENSIONS:
                current_pct = round(weights[dim] * 100)
                new_w = st.slider(
                    dim,
                    min_value=0, max_value=25, value=current_pct, step=1,
                    format="%d%%",
                    key=f"w_shock_{dim}",
                )
                weight_overrides_raw[dim] = new_w

            # Normalize to 0-1
            total_w = sum(weight_overrides_raw.values())
            weight_overrides = {k: v / total_w for k, v in weight_overrides_raw.items()} if total_w > 0 else weights

        # ── Run sensitivity analysis ─────────────────────────
        st.markdown("---")

        fin_scores_map = {s["Supplier"]: s["Financial Health"] for s in scored_suppliers}
        suppliers_for_whatif = [
            {
                "Supplier": s["Supplier"], "Raw Price": s["Raw Price"],
                "SLA Strength": s["SLA Strength"], "Execution Risk": s["Execution Risk"],
                "Stakeholder Confidence": s["Stakeholder Confidence"],
                "Strategic Alignment": s["Strategic Alignment"],
                "Innovation Capacity": s["Innovation Capacity"],
                "Relationship Depth": s["Relationship Depth"],
                "Commercial Flexibility": s["Commercial Flexibility"],
            }
            for s in suppliers
        ]

        shocked_results = run_sensitivity_analysis(
            suppliers_for_whatif, weights, fin_scores_map,
            price_shocks, weight_overrides, sla_overrides, {}
        )

        shocked_winner = shocked_results[0]["Supplier"] if shocked_results else leader["Supplier"]
        recommendation_flipped = shocked_winner != leader["Supplier"]

        # ── Result banner ────────────────────────────────────
        if any(price_shocks) or sla_overrides or weight_overrides != weights:
            if recommendation_flipped:
                st.markdown(
                    f'<div style="background:rgba(248,113,113,0.1);border:1px solid rgba(248,113,113,0.3);'
                    f'border-radius:12px;padding:1rem 1.2rem;margin-bottom:0.8rem">'
                    f'<div style="font-size:0.82rem;color:#F87171;text-transform:uppercase;letter-spacing:0.12em;'
                    f'margin-bottom:0.3rem;font-weight:700">⚠ Recommendation Flipped</div>'
                    f'<div style="font-size:1rem;color:#E2E8F0;font-weight:700">'
                    f'{html.escape(sx(shocked_winner))} now leads under these conditions</div>'
                    f'<div style="font-size:0.92rem;color:#C4D3E8;margin-top:0.3rem">'
                    f'Your base recommendation of {html.escape(sx(leader["Supplier"]))} is sensitive to the changes above. '
                    f'This is the conversation to prepare for in the room.</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f'<div style="background:rgba(74,222,128,0.08);border:1px solid rgba(74,222,128,0.25);'
                    f'border-radius:12px;padding:1rem 1.2rem;margin-bottom:0.8rem">'
                    f'<div style="font-size:0.82rem;color:#4ADE80;text-transform:uppercase;letter-spacing:0.12em;'
                    f'margin-bottom:0.3rem;font-weight:700">✓ Recommendation Holds</div>'
                    f'<div style="font-size:1rem;color:#E2E8F0;font-weight:700">'
                    f'{html.escape(sx(leader["Supplier"]))} still leads under these conditions</div>'
                    f'<div style="font-size:0.92rem;color:#C4D3E8;margin-top:0.3rem">'
                    f'The recommendation is robust to these shocks. Use this as evidence in the room.</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Score comparison chart ────────────────────────────
        if shocked_results:
            st.markdown("#### Score Comparison — Base vs. Shocked")
            chart_data = []
            base_map = {s["Supplier"]: s["Weighted Score"] for s in scored_suppliers}
            for sr in shocked_results:
                chart_data.append({
                    "Supplier": sr["Supplier"],
                    "Base Score": base_map.get(sr["Supplier"], 0),
                    "Shocked Score": round(sr["Shocked Score"], 1),
                    "Delta": round(sr["Shocked Score"] - base_map.get(sr["Supplier"], 0), 1),
                })
            chart_df = pd.DataFrame(chart_data)

            # Bar chart
            fig_w = go.Figure()
            colors_w = ["#60A5FA", "#4ADE80", "#FCD34D", "#F87171"]
            for i, (_, row) in enumerate(chart_df.iterrows()):
                fig_w.add_trace(go.Bar(
                    name=row["Supplier"],
                    x=[f"{row['Supplier']} Base", f"{row['Supplier']} Shocked"],
                    y=[row["Base Score"], row["Shocked Score"]],
                    marker_color=[colors_w[i % len(colors_w)], colors_w[i % len(colors_w)]],
                    marker_opacity=[0.5, 1.0],
                    text=[f"{row['Base Score']}", f"{row['Shocked Score']} ({row['Delta']:+.1f})"],
                    textposition="outside",
                ))
            fig_w.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(13,21,38,0.5)",
                font=dict(color="#E2E8F0", size=11),
                height=320,
                margin=dict(l=20, r=20, t=20, b=20),
                showlegend=False,
                yaxis=dict(range=[0, 110], gridcolor="rgba(255,255,255,0.06)", color="#64748B"),
                xaxis=dict(color="#64748B"),
                bargap=0.25,
            )
            st.plotly_chart(fig_w, use_container_width=True)

            # Delta table
            st.markdown("#### Detailed Impact Table")
            for row in chart_data:
                delta_color = "#4ADE80" if row["Delta"] >= 0 else "#F87171"
                delta_icon = "↑" if row["Delta"] > 0 else ("↓" if row["Delta"] < 0 else "→")
                is_winner = row["Supplier"] == shocked_winner
                st.markdown(
                    f'<div style="background:{"rgba(96,165,250,0.08)" if is_winner else "#0D1526"};'
                    f'border:1px solid {"rgba(96,165,250,0.3)" if is_winner else "rgba(148,163,184,0.15)"};'
                    f'border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.35rem;'
                    f'display:flex;align-items:center;gap:1.5rem">'
                    f'<strong style="color:#E2E8F0;min-width:8rem">{sx(row["Supplier"])}</strong>'
                    f'<span style="font-size:0.82rem;color:#D0E0EF">Base: <strong style="color:#C4D3E8">{row["Base Score"]}</strong></span>'
                    f'<span style="font-size:0.82rem;color:#D0E0EF">Shocked: <strong style="color:#E2E8F0">{row["Shocked Score"]}</strong></span>'
                    f'<span style="font-family:monospace;font-size:0.85rem;color:{delta_color};font-weight:700">'
                    f'{delta_icon} {row["Delta"]:+.1f} pts</span>'
                    f'{"<span style='font-size:0.82rem;color:#60A5FA;font-weight:700;margin-left:auto'>LEADS UNDER SHOCK</span>" if is_winner else ""}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Breakeven analysis ───────────────────────────────
        st.markdown("---")
        st.markdown("#### Breakeven Analysis")
        if runner_up:
            gap = round(leader["Weighted Score"] - runner_up["Weighted Score"], 1)
            price_weight_pct = round(weights.get("Price / TCO", 0.15) * 100, 1)
            # How much price increase to leader closes the gap?
            # Price score change per % increase depends on price spread
            price_spread = max(s["Raw Price"] for s in scored_suppliers) - min(s["Raw Price"] for s in scored_suppliers)
            if price_spread > 0 and weights.get("Price / TCO", 0) > 0:
                leader_price_now = leader["Raw Price"]
                # Approximate: each 10% price increase costs ~(price_weight * score_range) points
                approx_pts_per_10pct = weights.get("Price / TCO", 0.15) * 40  # rough estimate
                pct_to_flip = (gap / approx_pts_per_10pct) * 10 if approx_pts_per_10pct > 0 else 999
                st.markdown(
                    f'<div style="background:rgba(167,139,250,0.06);border-left:3px solid #A78BFA;'
                    f'border-radius:0 8px 8px 0;padding:0.8rem 1rem;font-size:0.88rem;color:#E2E8F0">'
                    f'<strong style="color:#A78BFA">Breakeven Estimate:</strong> '
                    f'{sx(leader["Supplier"])} can absorb approximately a '
                    f'<strong>+{round(pct_to_flip, 0):.0f}%</strong> price increase before '
                    f'{sx(runner_up["Supplier"])} becomes the better-scoring option '
                    f'(assuming all other dimensions stay constant). '
                    f'<br><span style="font-size:0.78rem;color:#D0E0EF;margin-top:0.3rem;display:block">'
                    f'Price weight is {price_weight_pct}% of the total evaluation. '
                    f'Current score gap is {gap} pts. Use the sliders above to test specific scenarios.</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.markdown('<p style="color:#D0E0EF;font-size:0.86rem">Add a second supplier to see breakeven analysis.</p>', unsafe_allow_html=True)

        # ── BATNA WORKSHEET ───────────────────────────────────
        st.markdown("---")
        st.markdown("### BATNA Worksheet")
        st.markdown(
            '<p style="color:#C4D3E8;font-size:0.88rem">Based on Fisher &amp; Ury\'s <em>Getting to Yes</em> framework. '
            'Complete before any supplier negotiation. The goal is to know your walk-away before the conversation starts.</p>',
            unsafe_allow_html=True,
        )

        bw_col1, bw_col2 = st.columns(2)

        with bw_col1:
            st.markdown(
                '<div style="font-size:0.80rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.14em;'
                'margin-bottom:0.6rem;font-family:monospace">1. Define Your BATNA</div>',
                unsafe_allow_html=True,
            )
            batna_alt = st.text_area(
                "What is your best alternative if this negotiation fails?",
                value=f"Continue with incumbent / extend current contract" if leader.get("Supplier") else "Extend status quo",
                key="batna_alternative",
                height=70,
                help="Your BATNA is what you'll actually do if you walk away. It must be real and executable.",
            )
            batna_alt_cost = st.number_input(
                "Estimated cost of BATNA alternative ($)",
                min_value=0.0, value=float(runner_up["Raw Price"]) if runner_up else 0.0,
                step=10000.0, key="batna_alt_cost",
                help="What does the alternative actually cost? This is your walk-away anchor.",
            )
            st.markdown(
                '<div style="font-size:0.80rem;color:#F59E0B;text-transform:uppercase;letter-spacing:0.14em;'
                'margin-bottom:0.6rem;margin-top:1rem;font-family:monospace">2. Define the ZOPA</div>',
                unsafe_allow_html=True,
            )
            zopa_floor = st.number_input(
                "Your walk-away price (maximum you'll pay) $",
                min_value=0.0,
                value=float(round(leader["Raw Price"] * 1.05)) if leader else 0.0,
                step=10000.0, key="batna_zopa_floor",
                help="If the supplier won't meet this price, you walk. Base it on your BATNA cost above.",
            )
            zopa_target = st.number_input(
                "Your target price (what you're trying to achieve) $",
                min_value=0.0,
                value=float(round(leader["Raw Price"] * 0.90)) if leader else 0.0,
                step=10000.0, key="batna_zopa_target",
                help="Realistic target, not best-case. Supported by market data or benchmark.",
            )
            zopa_open = st.number_input(
                "Your opening anchor (first offer or counter) $",
                min_value=0.0,
                value=float(round(leader["Raw Price"] * 0.82)) if leader else 0.0,
                step=10000.0, key="batna_zopa_open",
                help="Anchoring high (or low) influences the negotiation midpoint. Should be ambitious but defensible.",
            )

        with bw_col2:
            st.markdown(
                '<div style="font-size:0.80rem;color:#4ADE80;text-transform:uppercase;letter-spacing:0.14em;'
                'margin-bottom:0.6rem;font-family:monospace">3. Concession Strategy</div>',
                unsafe_allow_html=True,
            )
            concession_1 = st.text_input(
                "Concession 1 — what you'll give up first (low value to you)",
                value="Payment terms — offer Net 30 instead of Net 60",
                key="batna_conc1",
            )
            concession_2 = st.text_input(
                "Concession 2 — what you'll give up second",
                value="Contract term — extend from 2 to 3 years in exchange for price lock",
                key="batna_conc2",
            )
            concession_3 = st.text_input(
                "Concession 3 — last resort only",
                value="Volume commitment — guarantee 80% of spend in exchange for rate cap",
                key="batna_conc3",
            )
            st.markdown(
                '<div style="font-size:0.80rem;color:#F87171;text-transform:uppercase;letter-spacing:0.14em;'
                'margin-bottom:0.6rem;margin-top:1rem;font-family:monospace">4. Non-Negotiables (Never Give)</div>',
                unsafe_allow_html=True,
            )
            non_neg = st.text_area(
                "What will you NOT concede regardless of pressure?",
                value=f"SLA service credit schedule · Annual renewal cap · Data portability on exit · {leader_weakest_dim} improvement commitment",
                key="batna_non_neg",
                height=70,
                help="Know your non-negotiables before the meeting. Conceding them under pressure signals weakness.",
            )

        # ── BATNA Summary Card ────────────────────────────────
        _price_gap_pct = round(((zopa_floor - zopa_target) / zopa_floor) * 100, 1) if zopa_floor > 0 else 0
        _anchor_gap_pct = round(((leader["Raw Price"] - zopa_open) / leader["Raw Price"]) * 100, 1) if leader["Raw Price"] > 0 else 0
        st.markdown(
            f'<div style="background:linear-gradient(135deg,rgba(9,24,48,0.95),rgba(14,30,60,0.9));'
            f'border:1px solid rgba(96,165,250,0.2);border-radius:12px;padding:1.2rem 1.4rem;margin-top:0.8rem">'
            f'<div style="font-size:0.78rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.16em;margin-bottom:0.8rem;font-family:monospace">BATNA Summary — Pre-Negotiation Position</div>'
            f'<div style="display:flex;gap:2rem;flex-wrap:wrap">'
            f'<div><div style="font-size:0.80rem;color:#D0E0EF;margin-bottom:0.15rem">Walk-Away Anchor</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:#F87171;font-family:monospace">${zopa_floor:,.0f}</div></div>'
            f'<div><div style="font-size:0.80rem;color:#D0E0EF;margin-bottom:0.15rem">Target Price</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:#FCD34D;font-family:monospace">${zopa_target:,.0f}</div></div>'
            f'<div><div style="font-size:0.80rem;color:#D0E0EF;margin-bottom:0.15rem">Opening Anchor</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:#4ADE80;font-family:monospace">${zopa_open:,.0f}</div></div>'
            f'<div><div style="font-size:0.80rem;color:#D0E0EF;margin-bottom:0.15rem">Negotiation Range</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:#60A5FA;font-family:monospace">${abs(zopa_floor - zopa_open):,.0f}</div></div>'
            f'<div><div style="font-size:0.80rem;color:#D0E0EF;margin-bottom:0.15rem">Anchor vs. Ask</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:#A78BFA;font-family:monospace">{_anchor_gap_pct:.1f}% below quoted</div></div>'
            f'</div>'
            f'<div style="font-size:0.78rem;color:#C4D3E8;margin-top:0.8rem;padding-top:0.6rem;border-top:1px solid rgba(96,165,250,0.08)">'
            f'BATNA alternative: {html.escape(batna_alt[:80])} · Est. cost: ${batna_alt_cost:,.0f}'
            f'</div></div>',
            unsafe_allow_html=True,
        )

    # ── COMMS & AI TAB (Excel + Emails + AI Assist) ────────
    with tab_comms:
        _comms_section = st.radio(
            "Section",
            ["📧 Emails", "🤖 AI Assist", "📊 Excel I/O", "📄 Contract Reader", "🌐 Supplier Portal", "📝 Contract Generator"],
            horizontal=True,
            label_visibility="collapsed",
            key="comms_section_nav",
        )
        st.markdown("---")

    with tab_comms:
        if _comms_section == "📊 Excel I/O":
            st.markdown("### Excel Import / Export")
            imp_col, exp_col = st.columns(2)
            with imp_col:
                st.markdown("#### Import Supplier Data")
                st.markdown('<p class="muted">Upload a filled supplier template (.xlsx) and ProcureIQ will populate the evaluation automatically.</p>', unsafe_allow_html=True)
                uploaded_xl = st.file_uploader("Upload Supplier Template (.xlsx)", type=["xlsx", "xls"], key="excel_upload")
                if uploaded_xl:
                    parsed = parse_supplier_excel(uploaded_xl)
                    if parsed:
                        st.success(f"✅ Loaded {len(parsed)} suppliers from Excel. Switch to the Suppliers tab — data is pre-populated.")
                        st.session_state["imported_suppliers"] = parsed
                    else:
                        st.error("❌ Could not parse the file. Make sure it uses the Import Template format (available in the Export section).")
                st.markdown('<div class="soft-blue" style="font-size:0.92rem; margin-top:0.8rem">', unsafe_allow_html=True)
                st.markdown("**Required columns:** Supplier Name, Quoted Price ($). All other columns are optional but improve accuracy.", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
            with exp_col:
                st.markdown("#### Export Full Evaluation")
                st.markdown('<p class="muted">Download a formatted .xlsx with 5 sheets: Executive Summary, Supplier Ranking, Stakeholder Map, Risk Flags, and Import Template.</p>', unsafe_allow_html=True)
                if st.button("Generate Export (.xlsx)", type="primary", use_container_width=True):
                    risk_flags_for_export = generate_rfp_risk_flags(leader, runner_up, blocker, kraljic, category_rule)
                    exec_sum_for_export = build_executive_summary(leader, runner_up, blocker, event_name, kraljic, category_rule, leader_weakest_dim)
                    auction_type_for_export, _ = recommend_auction_type(
                        kraljic, num_suppliers,
                        weights.get("Price / TCO", 0.15),
                        switching_cost_answer,
                        selected_sub.get("auction", "RFP"),
                    )
                    xl_bytes = build_export_excel(
                        ranked, stake_df,
                        event_name=event_name,
                        category=category,
                        kraljic=kraljic,
                        auction_type=auction_type_for_export,
                        exec_summary=exec_sum_for_export,
                        risk_flags=risk_flags_for_export,
                    )
                    st.download_button(
                        label="⬇️ Download ProcureIQ Export",
                        data=xl_bytes,
                        file_name=f"ProcureIQ_{event_name.replace(' ','_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                st.markdown("---")
                st.markdown("#### Executive One-Pager")
                st.markdown('<p class="muted">Single-page decision brief for CPOs and executive sponsors. Download as PDF (ready to send) or HTML (editable).</p>', unsafe_allow_html=True)
                _op_col1, _op_col2 = st.columns(2)
                with _op_col1:
                    if st.button("Generate PDF Brief", use_container_width=True, key="gen_pdf", type="primary"):
                        _op_risk_flags = generate_rfp_risk_flags(leader, runner_up, blocker, kraljic, category_rule)
                        _op_action = build_90day_action_plan(
                            leader, runner_up, blocker, kraljic, category_rule,
                            leader_weakest_dim, selected_sub, intake_answers,
                        )
                        _op_html = build_executive_onepager_html(
                            event_name, category, selected_sub_name, kraljic,
                            leader, runner_up, ranked, _op_risk_flags, _op_action,
                            category_rule, leader_weakest_dim,
                        )
                        try:
                            import weasyprint as _wp
                            _pdf_bytes = _wp.HTML(string=_op_html).write_pdf()
                            st.download_button(
                                label="⬇️ Download PDF",
                                data=_pdf_bytes,
                                file_name=f"ProcureIQ_Brief_{event_name.replace(' ','_')}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                                key="dl_pdf",
                            )
                        except Exception as _pdf_err:
                            st.error(f"PDF generation failed: {_pdf_err}")
                with _op_col2:
                    if st.button("Generate HTML Brief", use_container_width=True, key="gen_onepager"):
                        _op_risk_flags = generate_rfp_risk_flags(leader, runner_up, blocker, kraljic, category_rule)
                        _op_action = build_90day_action_plan(
                            leader, runner_up, blocker, kraljic, category_rule,
                            leader_weakest_dim, selected_sub, intake_answers,
                        )
                        _op_html = build_executive_onepager_html(
                            event_name, category, selected_sub_name, kraljic,
                            leader, runner_up, ranked, _op_risk_flags, _op_action,
                            category_rule, leader_weakest_dim,
                        )
                        st.download_button(
                            label="⬇️ Download HTML",
                            data=_op_html.encode("utf-8"),
                            file_name=f"ProcureIQ_Brief_{event_name.replace(' ','_')}.html",
                            mime="text/html",
                            use_container_width=True,
                            key="dl_onepager",
                        )
            st.markdown('<p style="font-size:0.92rem;color:#C4D3E8;margin-top:0.8rem">The Import Template sheet in the export file shows the exact column format required for upload. Fill it in Excel and upload above.</p>', unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("#### Category Strategy Brief")
            st.markdown(
                '<p class="muted">One-page CIPS-format category strategy document — '
                'market sizing, supply market analysis, strategic objectives, recommended sourcing approach. '
                'Ready to present to CPO or category review board.</p>',
                unsafe_allow_html=True,
            )
            if st.button("Generate Category Strategy PDF", type="secondary", key="gen_cat_strategy", use_container_width=True):
                _sourcing_approach = (
                    f"Full competitive RFP with weighted scoring across RAQSCI dimensions. "
                    f"Given the {kraljic} posture, prioritise supplier relationship management "
                    f"and risk mitigation alongside cost."
                    if kraljic in ("Strategic", "Bottleneck") else
                    f"Reverse auction or streamlined RFP to maximise price competition. "
                    f"Award based on total cost of ownership, not unit price alone."
                )
                _stakes_str  = ", ".join(selected_sub.get("stakeholders", []))
                _saved_scores = {
                    k.replace("mat_", "").replace("_", " ").title(): v
                    for k, v in st.session_state.items()
                    if k.startswith("mat_")
                }
                _mat_section = "".join(
                    f"<tr><td>{dim}</td><td style='text-align:center'>{score}/5</td>"
                    f"<td>{'&#9608;' * score + '&#9617;' * (5 - score)}</td></tr>"
                    for dim, score in _saved_scores.items()
                ) if _saved_scores else ""
                _cs_html = (
                    "<!DOCTYPE html><html><head><meta charset='utf-8'>"
                    "<style>"
                    "body{font-family:Arial,sans-serif;color:#1a1a2e;margin:48px;font-size:11.5px;line-height:1.55}"
                    "h1{font-size:17px;color:#1e3a8a;border-bottom:2px solid #1e3a8a;padding-bottom:6px;margin-bottom:4px}"
                    "h2{font-size:12px;color:#1e3a8a;margin-top:18px;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.06em}"
                    ".meta{font-size:9.5px;color:#B0C4DC;margin-bottom:16px}"
                    ".kv{display:flex;gap:12px;margin:10px 0;flex-wrap:wrap}"
                    ".kv-item{flex:1;min-width:120px;background:#f8fafc;border-left:3px solid #3b82f6;padding:6px 10px}"
                    ".kv-label{font-size:8.5px;text-transform:uppercase;letter-spacing:0.08em;color:#64748b}"
                    ".kv-val{font-size:13px;font-weight:700;color:#1e3a8a}"
                    "table{width:100%;border-collapse:collapse;margin-top:8px;font-size:10.5px}"
                    "th{background:#1e3a8a;color:white;padding:5px 8px;text-align:left;font-size:10px}"
                    "td{border-bottom:1px solid #e2e8f0;padding:5px 8px}"
                    ".risk{background:#fef2f2;border-left:3px solid #ef4444;padding:4px 8px;margin:3px 0;font-size:10.5px}"
                    ".footer{margin-top:24px;font-size:8.5px;color:#94a3b8;border-top:1px solid #e2e8f0;padding-top:8px}"
                    "</style></head><body>"
                    f"<h1>Category Strategy &#8212; {html.escape(selected_sub_name)}</h1>"
                    f"<div class='meta'>Generated {_date.today().strftime('%B %d, %Y')} &nbsp;&middot;&nbsp; ProcureIQ &nbsp;&middot;&nbsp; Event: {html.escape(event_name)}</div>"
                    "<div class='kv'>"
                    f"<div class='kv-item'><div class='kv-label'>Category</div><div class='kv-val'>{html.escape(selected_sub_name)}</div></div>"
                    f"<div class='kv-item'><div class='kv-label'>Parent</div><div class='kv-val'>{html.escape(category)}</div></div>"
                    f"<div class='kv-item'><div class='kv-label'>Kraljic Posture</div><div class='kv-val'>{html.escape(kraljic)}</div></div>"
                    f"<div class='kv-item'><div class='kv-label'>Contract Type</div><div class='kv-val'>{html.escape(selected_sub.get('contract_type',''))}</div></div>"
                    f"<div class='kv-item'><div class='kv-label'>Switching Cost</div><div class='kv-val'>{html.escape(selected_sub.get('switching_cost',''))}</div></div>"
                    "</div>"
                    "<h2>Supply Market Analysis</h2>"
                    f"<p>{html.escape(selected_sub.get('notes',''))}</p>"
                    "<h2>Key Risks</h2>"
                    f"<div class='risk'>{html.escape(selected_sub.get('key_risks',''))}</div>"
                    "<h2>Required Stakeholder Team</h2>"
                    f"<p>{html.escape(_stakes_str)}</p>"
                    "<h2>Strategic Objectives &amp; KPIs</h2>"
                    "<table><tr><th>Objective</th><th>KPI</th><th>Target</th></tr>"
                    "<tr><td>Cost Reduction</td><td>Savings vs. Baseline</td><td>5&#8211;15%</td></tr>"
                    "<tr><td>Supply Assurance</td><td>On-Time Delivery Rate</td><td>&gt;98%</td></tr>"
                    "<tr><td>Risk Mitigation</td><td>Single-Source Exposure</td><td>&lt;30%</td></tr>"
                    "<tr><td>Contract Coverage</td><td>% Spend Under Contract</td><td>&gt;90%</td></tr>"
                    "<tr><td>Supplier Performance</td><td>Quarterly Scorecard Score</td><td>&gt;80/100</td></tr>"
                    "</table>"
                    "<h2>Recommended Sourcing Approach</h2>"
                    f"<p>Method: <strong>{html.escape(selected_sub.get('auction','RFP'))}</strong>. {html.escape(_sourcing_approach)}</p>"
                    + (
                        "<h2>Category Maturity</h2>"
                        "<table><tr><th>Dimension</th><th>Score</th><th>Visual</th></tr>"
                        + _mat_section +
                        "</table>"
                        if _mat_section else ""
                    ) +
                    "<div class='footer'>ProcureIQ &middot; Confidential &middot; For internal procurement use only</div>"
                    "</body></html>"
                )
                try:
                    import weasyprint as _wp2
                    _cat_pdf = _wp2.HTML(string=_cs_html).write_pdf()
                    st.download_button(
                        "⬇️ Download Category Strategy PDF",
                        data=_cat_pdf,
                        file_name=f"Category_Strategy_{selected_sub_name.replace(' ', '_')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="dl_cat_strategy",
                    )
                except Exception as _cs_err:
                    st.error(f"PDF generation failed: {_cs_err}")
                    st.download_button(
                        "⬇️ Download Category Strategy HTML",
                        data=_cs_html.encode("utf-8"),
                        file_name=f"Category_Strategy_{selected_sub_name.replace(' ', '_')}.html",
                        mime="text/html",
                        use_container_width=True,
                        key="dl_cat_strategy_html",
                    )

        # ── AUDIT LOG UI ────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 🔍 Evaluation Audit Log")
        st.markdown(
            '<p class="muted">Every scored evaluation and override is logged here. '
            'This record is required under SOX and FCPA procurement controls. '
            'Export to attach to your sourcing file.</p>',
            unsafe_allow_html=True,
        )

        # Log current evaluation score as an audit entry
        if st.button("📝 Log This Evaluation to Audit Trail", key="log_audit_btn",
                     help="Creates an immutable audit record of the current scoring state."):
            try:
                _adb = get_database()
                _adb.log_audit_event(
                    user_id=st.session_state.get("authenticated_user", "anonymous"),
                    action="supplier_evaluation_scored",
                    resource=event_name,
                    details={
                        "event_name": event_name,
                        "category": category,
                        "subcategory": selected_sub_name,
                        "kraljic": kraljic,
                        "recommendation": leader["Supplier"],
                        "score": leader["Weighted Score"],
                        "score_range": f"{max(0,round(leader['Weighted Score']-8))}–{min(100,round(leader['Weighted Score']+8))}",
                        "score_gap": round(leader["Weighted Score"] - runner_up["Weighted Score"], 1) if runner_up else None,
                        "coi_disclosures": {
                            st.session_state.get(f"name_{j}", f"Supplier {j+1}"): st.session_state.get(f"coi_description_{j}")
                            for j in range(num_suppliers)
                            if st.session_state.get(f"coi_flag_{j}")
                        },
                        "num_suppliers": num_suppliers,
                        "suppliers_evaluated": [s["Supplier"] for s in ranked],
                    }
                )
                st.success("Audit record logged.")
            except Exception as _ae:
                st.error(f"Audit log failed: {_ae}")

        # Read and display audit log
        try:
            _adb2 = get_database()
            _audit_entries = _adb2.get_audit_log(limit=20)
            if _audit_entries:
                _audit_rows = []
                for _entry in _audit_entries:
                    from datetime import datetime as _dt
                    _ts = _dt.fromtimestamp(_entry["timestamp"]).strftime("%Y-%m-%d %H:%M") if _entry["timestamp"] else "—"
                    _details = _entry.get("details") or {}
                    _audit_rows.append({
                        "Timestamp": _ts,
                        "User": _entry.get("user_id", "—"),
                        "Action": _entry.get("action", "—"),
                        "Event": _entry.get("resource", "—"),
                        "Recommendation": _details.get("recommendation", "—") if isinstance(_details, dict) else "—",
                        "Score": _details.get("score", "—") if isinstance(_details, dict) else "—",
                    })
                _audit_df = pd.DataFrame(_audit_rows)
                st.dataframe(_audit_df, use_container_width=True, hide_index=True)
                # Export
                _audit_csv = _audit_df.to_csv(index=False)
                st.download_button(
                    label="📥 Export Audit Log (CSV)",
                    data=_audit_csv,
                    file_name=f"ProcureIQ_AuditLog_{event_name.replace(' ','_')}.csv",
                    mime="text/csv",
                    key="audit_log_export",
                )
            else:
                st.info("No audit records yet. Click 'Log This Evaluation' above to create the first record.")
        except Exception as _ae2:
            st.caption(f"Audit log unavailable: {_ae2}")

    # ── EMAILS (merged into Comms) ─────────────────────────
    with tab_comms:
      if _comms_section == "📧 Emails":
        st.markdown("### Stakeholder Email Generator")
        st.markdown('<p class="muted">Pre-written emails for every key moment in the sourcing process. Edit before sending — these are starting points, not final drafts.</p>', unsafe_allow_html=True)

        emails = generate_stakeholder_emails(
            leader, runner_up, stake_df, event_name, leader_weakest_dim, kraljic, blocker
        )

        for email_type, email_data in emails.items():
            with st.expander(f"📧 {email_type}", expanded=(email_type == "Executive Sponsor Summary")):
                st.markdown(
                    f'<div class="soft-blue" style="font-size:0.82rem; margin-bottom:0.6rem"><strong>Subject:</strong> {html.escape(sx(email_data["subject"]))}</div>',
                    unsafe_allow_html=True,
                )
                edited = st.text_area(
                    "Email Body (edit before sending)",
                    email_data["body"],
                    height=220,
                    key=f"email_{email_type}",
                )

        st.markdown('<div class="panel" style="margin-top:1rem">', unsafe_allow_html=True)
        st.markdown('<div class="panel-eyebrow">Email Integration Options</div>', unsafe_allow_html=True)
        st.markdown(
            """
            <p class="muted" style="font-size:0.86rem">
                <strong style="color:#F1F5F9">Zero-infrastructure option:</strong>
                Each email above can be copied and pasted directly into Outlook, Gmail, or Teams.<br><br>
                <strong style="color:#F1F5F9">For enterprise deployment:</strong>
                Add a SendGrid API key to Streamlit secrets (<code>SENDGRID_API_KEY</code>) and emails
                can be sent directly from this interface with one click — 100 emails/day free tier.<br><br>
                <strong style="color:#F1F5F9">For Outlook integration:</strong>
                The mailto: link format can be added to each email so clicking opens it pre-populated
                in your default email client.
            </p>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

    # ── AI ASSIST (merged into Comms) ──────────────────────
    with tab_comms:
      if _comms_section == "🤖 AI Assist":
        st.markdown("### AI Command Center")
        st.markdown(
            f'<p style="color:#C4D3E8;font-size:0.92rem">Context pre-loaded: <strong style="color:#60A5FA">{html.escape(sx(selected_sub_name))}</strong> · '
            f'<strong style="color:#F1F5F9">{html.escape(sx(leader["Supplier"]))}</strong> recommended · '
            f'Weakest: <strong style="color:#F87171">{html.escape(sx(leader_weakest_dim))}</strong></p>',
            unsafe_allow_html=True,
        )

        # ── Section 1: Mode Selector ──
        st.markdown("#### Choose Your AI Mission")
        mode_cols = st.columns(len(AI_PROMPT_MODES))
        selected_mode = st.session_state.get("ai_mode", list(AI_PROMPT_MODES.keys())[0])

        for i, (mode_name, mode_data) in enumerate(AI_PROMPT_MODES.items()):
            with mode_cols[i]:
                is_active = mode_name == selected_mode
                is_active = mode_name == selected_mode
                border_color = "#3B82F6" if is_active else "rgba(96,165,250,0.15)"
                bg_color = "rgba(29,78,216,0.12)" if is_active else "rgba(6,13,26,0.8)"
                title_color = "#60A5FA" if is_active else "#CBD5E1"
                btn_label = "✓ Active" if is_active else "Select"
                st.markdown(
                    f'<div style="background:{bg_color};border:1.5px solid {border_color};'
                    f'border-radius:10px;padding:0.85rem 0.7rem;text-align:center;margin-bottom:0.4rem;'
                    f'transition:all 0.15s ease">'
                    f'<div style="font-size:1.4rem;margin-bottom:0.35rem;line-height:1">{mode_name.split()[0]}</div>'
                    f'<div style="font-size:0.76rem;font-weight:700;color:{title_color};letter-spacing:0.01em">'
                    f'{" ".join(mode_name.split()[1:])}</div>'
                    f'<div style="font-size:0.82rem;color:#D0E0EF;margin-top:0.3rem;line-height:1.45">'
                    f'{mode_data["desc"]}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                if st.button(btn_label, key=f"mode_btn_{i}", use_container_width=True,
                             type="primary" if is_active else "secondary"):
                    st.session_state["ai_mode"] = mode_name
                    st.rerun()

        selected_mode = st.session_state.get("ai_mode", list(AI_PROMPT_MODES.keys())[0])
        mode_data = AI_PROMPT_MODES[selected_mode]

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

        # ── Section 2: Guardrails Display ──
        with st.expander("⛔ Active Guardrails — What This Prompt Prevents", expanded=False):
            st.markdown(
                '<p style="font-size:0.82rem;color:#C4D3E8">These guardrails are embedded in your prompt to prevent AI hallucination, generic advice, and false validation. They force the AI to be specific and critical.</p>',
                unsafe_allow_html=True,
            )
            for g in mode_data["guardrails"]:
                st.markdown(
                    f'<div style="background:rgba(248,113,113,0.06);border-left:3px solid #F87171;padding:0.5rem 0.85rem;margin-bottom:0.35rem;border-radius:0 6px 6px 0;font-size:0.92rem;color:#C4D3E8">⛔ {sx(g)}</div>',
                    unsafe_allow_html=True,
                )

        # ── Section 3: Build and display the prompt ──
        _auction_for_ai = st.session_state.get("_last_auction_type", "Rank Auction (Coupa / Ariba Standard)")
        ai_prompt_v2 = build_ai_prompt_v2(
            event_name, category, selected_sub_name, kraljic,
            leader, runner_up, leader_weakest_dim, blocker,
            stake_df, _auction_for_ai, intake_answers,
            selected_sub, selected_mode,
        )

        st.markdown("#### Generated Prompt — Ready to Use")
        st.text_area(
            f"Prompt for: {selected_mode}",
            ai_prompt_v2,
            height=360,
            key="ai_prompt_v2_box",
        )

        # ── Section 4: In-App Claude Analysis ──────────────────
        st.markdown("---")
        st.markdown(
            '<div style="display:flex;align-items:center;gap:0.6rem;margin-bottom:0.4rem">'
            '<span style="font-size:1rem;font-weight:700;color:#F1F5F9">Analyze with Claude</span>'
            '<span style="font-family:monospace;font-size:0.78rem;color:#D97706;border:1px solid rgba(217,119,6,0.3);'
            'border-radius:4px;padding:0.06rem 0.35rem;letter-spacing:0.08em">IN-APP · ANTHROPIC</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        _has_api_key = bool(_get_api_key())
        if _has_api_key and _ANTHROPIC_AVAILABLE:
            st.markdown(
                '<p style="font-size:0.82rem;color:#D0E0EF">Run the analysis directly in ProcureIQ. '
                'The full evaluation context is pre-loaded — no copy-paste required.</p>',
                unsafe_allow_html=True,
            )
            _claude_btn_col, _ = st.columns([1, 2])
            with _claude_btn_col:
                if st.button("Run Analysis in Claude", type="primary", use_container_width=True, key="run_claude_btn"):
                    st.session_state.pop("claude_response", None)
                    st.session_state["claude_mode"] = selected_mode
                    st.session_state["claude_streaming"] = True
                    st.rerun()

            if st.session_state.get("claude_streaming"):
                _resp_mode = st.session_state.get("claude_mode", "")
                st.markdown(
                    f'<div style="font-family:monospace;font-size:0.6rem;color:#D97706;text-transform:uppercase;'
                    f'letter-spacing:0.14em;margin-bottom:0.5rem;margin-top:0.8rem">Claude · {sx(_resp_mode)} · Streaming</div>',
                    unsafe_allow_html=True,
                )
                with st.container():
                    _streamed = st.write_stream(stream_claude_api(ai_prompt_v2))
                st.session_state["claude_response"] = _streamed
                st.session_state["claude_streaming"] = False

            if "claude_response" in st.session_state and not st.session_state.get("claude_streaming"):
                _resp_mode = st.session_state.get("claude_mode", "")
                st.markdown(
                    f'<div style="font-family:monospace;font-size:0.6rem;color:#D97706;text-transform:uppercase;'
                    f'letter-spacing:0.14em;margin-bottom:0.4rem;margin-top:0.6rem">Claude · {sx(_resp_mode)} · Complete</div>',
                    unsafe_allow_html=True,
                )
                display_ai_governance_banner()
                st.write(st.session_state["claude_response"])
                if st.button("Clear Response", key="clear_claude", type="secondary"):
                    del st.session_state["claude_response"]
                    st.rerun()
        else:
            st.markdown(
                '<div style="background:rgba(29,78,216,0.06);border:1px solid rgba(96,165,250,0.15);'
                'border-radius:10px;padding:1rem 1.2rem">'
                '<div style="font-family:monospace;font-size:0.6rem;color:#60A5FA;text-transform:uppercase;'
                'letter-spacing:0.12em;margin-bottom:0.5rem">Enable In-App AI</div>'
                '<div style="font-size:0.92rem;color:#C4D3E8;line-height:1.6">'
                'Set <code style="color:#93C5FD;background:rgba(96,165,250,0.1);padding:0.1rem 0.3rem;border-radius:3px">'
                'ANTHROPIC_API_KEY</code> in your environment or Streamlit secrets to analyze directly in ProcureIQ. '
                'The prompt above is ready to paste into any external AI tool in the meantime.</div>'
                '</div>',
                unsafe_allow_html=True,
            )

        # ── Section 5: External AI tools ─────────────────────
        st.markdown("---")
        st.markdown(
            '<div style="font-size:0.78rem;color:#D0E0EF;text-transform:uppercase;letter-spacing:0.14em;'
            'margin-bottom:0.5rem;font-family:monospace">Or open in external tool</div>',
            unsafe_allow_html=True,
        )
        tool_cols = st.columns(len(AI_TOOLS))
        for i, tool in enumerate(AI_TOOLS):
            with tool_cols[i]:
                st.markdown(
                    f'<a href="{tool["url"]}" target="_blank" style="text-decoration:none">'
                    f'<div style="background:#0A1628;border:1px solid rgba(96,165,250,0.15);border-radius:10px;'
                    f'padding:0.7rem 0.5rem;text-align:center;transition:all 0.15s ease">'
                    f'<div style="font-size:1.3rem">{tool["icon"]}</div>'
                    f'<div style="font-weight:700;color:#F1F5F9;font-size:0.78rem;margin-top:0.15rem">{tool["name"]}</div>'
                    f'<div style="font-size:0.80rem;color:{tool["color"]};margin-top:0.1rem">↗ Open</div>'
                    f'</div></a>',
                    unsafe_allow_html=True,
                )

        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

        # ── Section 5: RFP Question Generator ──
        st.markdown("---")
        st.markdown("#### RFP Question Bank")
        st.markdown(
            f'<p style="color:#C4D3E8;font-size:0.88rem">Capability questions specific to <strong style="color:#60A5FA">{sx(selected_sub_name)}</strong> — for supplier discovery, early comparison, and RFP response evaluation.</p>',
            unsafe_allow_html=True,
        )
        rfp_questions = get_rfp_questions(selected_sub_name)
        for i, q in enumerate(rfp_questions, 1):
            st.markdown(
                f'<div style="background:#0A1628;border:1px solid rgba(96,165,250,0.2);border-radius:6px;padding:0.7rem 1rem;margin-bottom:0.4rem;display:flex;gap:0.8rem;align-items:flex-start">'
                f'<span style="font-family:monospace;font-size:0.82rem;color:#60A5FA;font-weight:700;min-width:1.4rem;margin-top:0.1rem">Q{i}</span>'
                f'<span style="font-size:0.87rem;color:#F1F5F9;line-height:1.5">{sx(q)}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        rfp_q_text = "\n".join(f"Q{i}. {q}" for i, q in enumerate(rfp_questions, 1))
        st.text_area("Copy all questions", rfp_q_text, height=120, key="rfp_q_copy")

        # ── Market Intelligence now lives in Overview tab ──
        st.markdown("---")
        st.markdown(
            '<div style="background:rgba(96,165,250,0.06);border-left:3px solid #60A5FA;border-radius:0 8px 8px 0;'
            'padding:0.7rem 1rem;font-size:0.87rem;color:#E2E8F0">'
            '📊 <strong style="color:#60A5FA">Live Market Intelligence</strong> has moved to the '
            '<strong>Overview tab</strong> — right column, below the active weights. '
            'It now shows alongside the supplier ranking for faster comparison.</div>',
            unsafe_allow_html=True,
        )


        st.markdown("</div>", unsafe_allow_html=True)

    # ── SPEND & RISK TAB ──────────────────────────────────
    with tab_spend:
        # ── PORTFOLIO DASHBOARD ──────────────────────────────────────────────
        _port_db = get_database()
        _port_events_raw = _port_db.get_portfolio_events(limit=20)
        _port_events = [r.get("value", {}) for r in _port_events_raw if r.get("value")]

        st.markdown(
            '<div style="font-size:0.7rem;color:#60A5FA;text-transform:uppercase;'
            'letter-spacing:0.14em;font-weight:700;margin-bottom:0.4rem">'
            'Portfolio Dashboard</div>',
            unsafe_allow_html=True,
        )

        if not _port_events:
            st.markdown(
                '<div style="background:rgba(96,165,250,0.05);border:1px dashed rgba(96,165,250,0.2);'
                'border-radius:10px;padding:1.2rem 1.5rem;margin-bottom:1.2rem">'
                '<span style="font-size:0.85rem;color:#94A3B8">'
                '📂 No events saved yet — complete an evaluation in the <strong>Supplier Evaluation</strong> '
                'tab and click <strong>Save to Portfolio</strong> to populate this dashboard.'
                '</span></div>',
                unsafe_allow_html=True,
            )
        else:
            # ── Summary metrics ──────────────────────────────────────────────
            _p_total   = len(_port_events)
            _p_cats    = len({e.get("category", "") for e in _port_events if e.get("category")})
            _p_scores  = [e.get("score", 0) for e in _port_events if e.get("score")]
            _p_avg_score = round(sum(_p_scores) / len(_p_scores)) if _p_scores else 0
            from collections import Counter as _Counter
            _cat_counts = _Counter(e.get("category", "Unknown") for e in _port_events)
            _p_top_cat  = _cat_counts.most_common(1)[0][0] if _cat_counts else "—"

            _pcols = st.columns(4)
            _metric_blocks = [
                ("Total Events", str(_p_total), "#60A5FA"),
                ("Categories Covered", str(_p_cats), "#34D399"),
                ("Avg Recommended Score", f"{_p_avg_score}", "#A78BFA"),
                ("Most Active Category", _p_top_cat[:22] + ("…" if len(_p_top_cat) > 22 else ""), "#F59E0B"),
            ]
            for _pc, (_pm_label, _pm_val, _pm_color) in zip(_pcols, _metric_blocks):
                with _pc:
                    st.markdown(
                        f'<div style="background:rgba(15,23,42,0.6);border:1px solid rgba(255,255,255,0.07);'
                        f'border-radius:10px;padding:0.9rem 1rem;text-align:center">'
                        f'<div style="font-size:1.45rem;font-weight:700;color:{_pm_color}">{_pm_val}</div>'
                        f'<div style="font-size:0.7rem;color:#94A3B8;margin-top:0.2rem">{_pm_label}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            st.markdown('<div style="height:1rem"></div>', unsafe_allow_html=True)

            # ── Kraljic distribution ─────────────────────────────────────────
            _kral_counts = _Counter(e.get("kraljic", "Unknown") for e in _port_events)
            _kral_palette = {
                "Strategic":    ("#7C3AED", "rgba(124,58,237,0.15)"),
                "Leverage":     ("#2563EB", "rgba(37,99,235,0.15)"),
                "Bottleneck":   ("#D97706", "rgba(217,119,6,0.15)"),
                "Non-Critical": ("#059669", "rgba(5,150,105,0.15)"),
            }
            _kral_col, _cat_col = st.columns(2)

            with _kral_col:
                st.markdown(
                    '<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.1em;font-weight:600;margin-bottom:0.6rem">Kraljic Distribution</div>',
                    unsafe_allow_html=True,
                )
                for _kq in ["Strategic", "Leverage", "Bottleneck", "Non-Critical"]:
                    _kc = _kral_counts.get(_kq, 0)
                    _kpct = int(_kc / _p_total * 100) if _p_total else 0
                    _kcolor, _kbg = _kral_palette.get(_kq, ("#94A3B8", "rgba(148,163,184,0.12)"))
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.4rem">'
                        f'<div style="width:90px;font-size:0.72rem;color:#CBD5E1">{_kq}</div>'
                        f'<div style="flex:1;background:rgba(255,255,255,0.05);border-radius:4px;height:14px;overflow:hidden">'
                        f'<div style="width:{_kpct}%;background:{_kcolor};height:100%;border-radius:4px;'
                        f'transition:width 0.3s"></div></div>'
                        f'<div style="width:28px;text-align:right;font-size:0.72rem;color:{_kcolor};font-weight:600">{_kc}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            with _cat_col:
                st.markdown(
                    '<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase;'
                    'letter-spacing:0.1em;font-weight:600;margin-bottom:0.6rem">Top Categories</div>',
                    unsafe_allow_html=True,
                )
                for _cname, _ccount in _cat_counts.most_common(6):
                    _cpct = int(_ccount / _p_total * 100) if _p_total else 0
                    _cshort = _cname[:24] + ("…" if len(_cname) > 24 else "")
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.4rem">'
                        f'<div style="width:120px;font-size:0.72rem;color:#CBD5E1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{_cshort}</div>'
                        f'<div style="flex:1;background:rgba(255,255,255,0.05);border-radius:4px;height:14px;overflow:hidden">'
                        f'<div style="width:{_cpct}%;background:#3B82F6;height:100%;border-radius:4px"></div></div>'
                        f'<div style="width:28px;text-align:right;font-size:0.72rem;color:#60A5FA;font-weight:600">{_ccount}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            st.markdown('<div style="height:0.8rem"></div>', unsafe_allow_html=True)

            # ── Risk Heat Map ─────────────────────────────────────────────────
            st.markdown(
                '<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase;'
                'letter-spacing:0.1em;font-weight:600;margin-bottom:0.5rem">Data Provenance &amp; Risk Signals</div>',
                unsafe_allow_html=True,
            )

            def _freshness_cell(period_end_str, fin_source):
                """Return (icon, label, css-color) for EDGAR freshness."""
                if not fin_source or fin_source == "User Assessment":
                    return "⚪", "Manual", "#94A3B8"
                if not period_end_str:
                    return "—", "Unknown", "#64748B"
                try:
                    from datetime import date as _date
                    _pe_d = _date.fromisoformat(period_end_str[:10])
                    _age_m = (_date.today() - _pe_d).days / 30.44
                    if _age_m <= 12:
                        return "🟢", "Fresh", "#4ADE80"
                    elif _age_m <= 18:
                        return "🟡", "Verify", "#FCD34D"
                    else:
                        return "🔴", "Stale", "#F87171"
                except Exception:
                    return "—", "Unknown", "#64748B"

            _hm_header = (
                '<table style="width:100%;border-collapse:collapse;font-size:0.73rem">'
                '<thead><tr style="border-bottom:1px solid rgba(255,255,255,0.1)">'
                '<th style="text-align:left;padding:0.3rem 0.5rem;color:#64748B;font-weight:600;white-space:nowrap">Event</th>'
                '<th style="text-align:left;padding:0.3rem 0.5rem;color:#64748B;font-weight:600;white-space:nowrap">Posture</th>'
                '<th style="text-align:left;padding:0.3rem 0.5rem;color:#64748B;font-weight:600;white-space:nowrap">Recommendation</th>'
                '<th style="text-align:left;padding:0.3rem 0.5rem;color:#64748B;font-weight:600;white-space:nowrap">Fin. Source</th>'
                '<th style="text-align:left;padding:0.3rem 0.5rem;color:#64748B;font-weight:600;white-space:nowrap">EDGAR Period</th>'
                '<th style="text-align:center;padding:0.3rem 0.5rem;color:#64748B;font-weight:600;white-space:nowrap">Freshness</th>'
                '<th style="text-align:center;padding:0.3rem 0.5rem;color:#64748B;font-weight:600;white-space:nowrap">High-Risk Sup.</th>'
                '</tr></thead><tbody>'
            )

            _hm_rows = []
            for _hm_ev in _port_events:
                _hm_name    = html.escape(str(_hm_ev.get("event_name", "—"))[:32])
                _hm_kral    = _hm_ev.get("kraljic", "—")
                _hm_rec     = html.escape(str(_hm_ev.get("recommendation", "—"))[:22])
                _hm_score   = _hm_ev.get("score", "")
                _hm_rec_str = f"{_hm_rec} ({_hm_score})" if _hm_score else _hm_rec

                _hm_fin_src  = _hm_ev.get("fin_source", "")
                _hm_period   = _hm_ev.get("edgar_period_end", "")
                _hm_fin_risk = _hm_ev.get("fin_risk", "")
                _hm_hr_count = _hm_ev.get("high_risk_count", None)
                _hm_def      = _hm_ev.get("defensibility", "")

                _hm_icon, _hm_fresh_lbl, _hm_fresh_col = _freshness_cell(_hm_period, _hm_fin_src)

                # Source badge
                if not _hm_fin_src:
                    _src_display = '<span style="color:#64748B">Not captured</span>'
                elif "EDGAR" in _hm_fin_src:
                    _src_display = f'<span style="color:#60A5FA">{html.escape(_hm_fin_src)}</span>'
                else:
                    _src_display = f'<span style="color:#94A3B8">{html.escape(_hm_fin_src)}</span>'

                # Period display
                _period_display = html.escape(_hm_period[:10]) if _hm_period else '<span style="color:#64748B">—</span>'

                # Freshness cell
                _fresh_display = (
                    f'<span style="color:{_hm_fresh_col};font-weight:600">{_hm_icon} {_hm_fresh_lbl}</span>'
                )

                # High-risk count cell
                if _hm_hr_count is None:
                    _hr_display = '<span style="color:#64748B">—</span>'
                elif _hm_hr_count == 0:
                    _hr_display = '<span style="color:#4ADE80;font-weight:600">0</span>'
                else:
                    _hr_display = f'<span style="color:#F87171;font-weight:700">{_hm_hr_count}</span>'

                # Kraljic badge color
                _kq_color = {"Strategic":"#A78BFA","Leverage":"#60A5FA","Bottleneck":"#FCD34D","Non-Critical":"#34D399"}.get(_hm_kral,"#94A3B8")

                _hm_rows.append(
                    f'<tr style="border-bottom:1px solid rgba(255,255,255,0.04)">'
                    f'<td style="padding:0.35rem 0.5rem;color:#CBD5E1;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_hm_name}</td>'
                    f'<td style="padding:0.35rem 0.5rem;white-space:nowrap"><span style="color:{_kq_color};font-weight:600">{html.escape(_hm_kral)}</span></td>'
                    f'<td style="padding:0.35rem 0.5rem;color:#CBD5E1;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_hm_rec_str}</td>'
                    f'<td style="padding:0.35rem 0.5rem;white-space:nowrap">{_src_display}</td>'
                    f'<td style="padding:0.35rem 0.5rem;color:#CBD5E1;white-space:nowrap">{_period_display}</td>'
                    f'<td style="padding:0.35rem 0.5rem;text-align:center;white-space:nowrap">{_fresh_display}</td>'
                    f'<td style="padding:0.35rem 0.5rem;text-align:center">{_hr_display}</td>'
                    f'</tr>'
                )

            st.markdown(
                '<div style="background:rgba(15,23,42,0.5);border:1px solid rgba(255,255,255,0.07);'
                'border-radius:10px;padding:0.5rem 0.8rem;margin-bottom:0.6rem;overflow-x:auto">'
                + _hm_header + "".join(_hm_rows) + "</tbody></table></div>",
                unsafe_allow_html=True,
            )

            st.markdown('<div style="height:0.6rem"></div>', unsafe_allow_html=True)

            # ── Events table ─────────────────────────────────────────────────
            st.markdown(
                '<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase;'
                'letter-spacing:0.1em;font-weight:600;margin-bottom:0.6rem">Saved Events</div>',
                unsafe_allow_html=True,
            )

            _kral_badge_style = {
                "Strategic":    "background:rgba(124,58,237,0.18);color:#A78BFA;border:1px solid rgba(124,58,237,0.35)",
                "Leverage":     "background:rgba(37,99,235,0.18);color:#60A5FA;border:1px solid rgba(37,99,235,0.35)",
                "Bottleneck":   "background:rgba(217,119,6,0.18);color:#FCD34D;border:1px solid rgba(217,119,6,0.35)",
                "Non-Critical": "background:rgba(5,150,105,0.18);color:#34D399;border:1px solid rgba(5,150,105,0.35)",
            }

            for _pe in _port_events:
                _pe_id    = _pe.get("event_id", "")
                _pe_name  = _pe.get("event_name", "Unnamed Event")
                _pe_cat   = _pe.get("category", "—")
                _pe_sub   = _pe.get("subcategory", "—")
                _pe_kral  = _pe.get("kraljic", "—")
                _pe_rec   = _pe.get("recommendation", "—")
                _pe_score = _pe.get("score", "—")
                _pe_gap   = _pe.get("score_gap", None)
                _pe_saved = _pe.get("saved_at", "")[:10] if _pe.get("saved_at") else "—"
                _kbadge   = _kral_badge_style.get(_pe_kral, "background:rgba(148,163,184,0.12);color:#94A3B8;border:1px solid rgba(148,163,184,0.2)")
                _gap_str  = f" (+{_pe_gap:.0f}pt lead)" if isinstance(_pe_gap, (int, float)) and _pe_gap else ""

                _row_left, _row_right = st.columns([6, 1])
                with _row_left:
                    st.markdown(
                        f'<div style="background:rgba(15,23,42,0.5);border:1px solid rgba(255,255,255,0.06);'
                        f'border-radius:8px;padding:0.65rem 1rem;margin-bottom:0.35rem">'
                        f'<div style="display:flex;align-items:center;gap:0.6rem;flex-wrap:wrap">'
                        f'<span style="font-size:0.82rem;font-weight:600;color:#E2E8F0">{_pe_name}</span>'
                        f'<span style="font-size:0.68rem;padding:0.1rem 0.5rem;border-radius:4px;{_kbadge}">{_pe_kral}</span>'
                        f'</div>'
                        f'<div style="font-size:0.72rem;color:#94A3B8;margin-top:0.25rem">'
                        f'{_pe_cat} › {_pe_sub} &nbsp;·&nbsp; '
                        f'<span style="color:#CBD5E1">Rec: {_pe_rec}</span> &nbsp;·&nbsp; '
                        f'Score: <span style="color:#60A5FA;font-weight:600">{_pe_score}</span>{_gap_str}'
                        f' &nbsp;·&nbsp; <span style="color:#64748B">{_pe_saved}</span>'
                        f'</div></div>',
                        unsafe_allow_html=True,
                    )
                with _row_right:
                    st.markdown('<div style="height:0.4rem"></div>', unsafe_allow_html=True)
                    if _pe_id and st.button("Load", key=f"load_portfolio_{_pe_id}", use_container_width=True):
                        st.session_state["_piq_restore_session_id"] = _pe_id
                        st.rerun()

        st.markdown(
            '<hr style="border:none;border-top:1px solid rgba(255,255,255,0.07);margin:1.4rem 0"/>',
            unsafe_allow_html=True,
        )
        # ── END PORTFOLIO DASHBOARD ──────────────────────────────────────────

        _has_named_suppliers = any(
            st.session_state.get(f"name_{k}", "").strip() and
            st.session_state.get(f"name_{k}", "").strip() != f"Supplier {k+1}"
            for k in range(num_suppliers)
        )
        if not _has_named_suppliers:
            st.markdown(
                '<div style="background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.18);'
                'border-radius:12px;padding:1.4rem 1.6rem;margin:1rem 0">'
                '<div style="font-size:0.75rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.12em;'
                'font-weight:700;margin-bottom:0.5rem">No Spend Data Yet</div>'
                '<div style="font-size:0.88rem;color:#CBD5E1;line-height:1.6">'
                'Enter supplier names in the <strong>Supplier Evaluation</strong> tab and set contract '
                'start dates to populate the renewal calendar, spend classification, and anomaly detection here.'
                '</div>'
                '</div>',
                unsafe_allow_html=True,
            )

        st.markdown("### Spend Intelligence — Auto-Classification")
        
        # ── CONTRACT RENEWAL CALENDAR ────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📅 Contract Renewal Calendar")
        st.markdown(
            '<p class="muted">Upcoming contract expirations and renewal actions. Set contract terms in the Supplier tab to see renewal dates here.</p>',
            unsafe_allow_html=True,
        )
        
        # Build renewal calendar from real contract dates
        from datetime import datetime as _dt, date as _date, timedelta as _td
        _notice_days_map = {"30 days": 30, "60 days": 60, "90 days": 90, "120 days": 120, "180 days": 180, "1 year": 365}
        renewal_events = []
        today = _date.today()
        for i in range(num_suppliers):
            sp_name    = st.session_state.get(f"name_{i}", f"Supplier {i+1}")
            tco_term   = st.session_state.get(f"tco_term_{i}", 3)
            start_raw   = st.session_state.get(f"contract_start_{i}", str(today))
            notice_str  = st.session_state.get(f"notice_period_{i}", "60 days")
            auto_renews = st.session_state.get(f"auto_renews_{i}", False)
            unbid_count = int(st.session_state.get(f"unbid_cycles_{i}", 0))
            try:
                start_date  = _date.fromisoformat(str(start_raw))
            except Exception:
                start_date  = today
            notice_days  = _notice_days_map.get(notice_str, 60)
            expiry_date  = start_date + _td(days=365 * tco_term)
            action_date  = expiry_date - _td(days=notice_days)
            days_to_act  = (action_date - today).days
            days_to_exp  = (expiry_date - today).days

            if days_to_act <= 0 and auto_renews:
                urgency_label = "⛔ AUTO-RENEWED — Window Missed"
                urgency_color = "rgba(239,68,68,0.15)"
                border_color  = "#EF4444"
            elif days_to_act <= 0:
                urgency_label = "🔴 OVERDUE — Act Now"
                urgency_color = "rgba(239,68,68,0.12)"
                border_color  = "#F87171"
            elif days_to_act <= 30:
                urgency_label = "🔴 < 30d to action deadline"
                urgency_color = "rgba(239,68,68,0.08)"
                border_color  = "#F87171"
            elif days_to_act <= 60:
                urgency_label = "🟠 < 60d to action deadline"
                urgency_color = "rgba(249,115,22,0.08)"
                border_color  = "#FB923C"
            elif days_to_act <= 90:
                urgency_label = "🟡 < 90d to action deadline"
                urgency_color = "rgba(234,179,8,0.08)"
                border_color  = "#FCD34D"
            else:
                urgency_label = "🟢 > 90d — Monitor"
                urgency_color = "rgba(74,222,128,0.04)"
                border_color  = "#4ADE80"

            _zombie_tag = f" · ⚠ Zombie: {unbid_count}x unbid" if unbid_count >= 2 else ""
            _auto_tag   = " · AUTO-RENEWS" if auto_renews else ""

            renewal_events.append({
                "Supplier":         sp_name,
                "Start Date":       start_date.strftime("%Y-%m-%d"),
                "Expiry Date":      expiry_date.strftime("%Y-%m-%d"),
                "Notice Period":    notice_str,
                "Action Deadline":  action_date.strftime("%Y-%m-%d"),
                "Days to Deadline": days_to_act,
                "Days to Expiry":   days_to_exp,
                "Term (years)":     tco_term,
                "Auto-Renews":      "Yes" if auto_renews else "No",
                "Unbid Cycles":     unbid_count,
                "Status":           urgency_label,
                "_zombie":          _zombie_tag,
                "_auto":            _auto_tag,
                "_bg":              urgency_color,
                "_border":          border_color,
            })

        if renewal_events:
            renewal_df = pd.DataFrame(renewal_events).sort_values("Days to Deadline")

            # Summary strip
            overdue    = sum(1 for e in renewal_events if e["Days to Deadline"] <= 0)
            urgent     = sum(1 for e in renewal_events if 0 < e["Days to Deadline"] <= 60)
            auto_count = sum(1 for e in renewal_events if e["Auto-Renews"] == "Yes")
            zombie_ct  = sum(1 for e in renewal_events if e["Unbid Cycles"] >= 2)
            st.markdown(
                f'<div style="display:flex;gap:1rem;margin-bottom:1rem;flex-wrap:wrap">'
                f'<div style="background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.25);border-radius:8px;padding:0.5rem 1rem">'
                f'<div style="font-size:0.6rem;color:#F87171;text-transform:uppercase;letter-spacing:0.1em">Overdue</div>'
                f'<div style="font-size:1.4rem;font-weight:700;color:#F87171;font-family:var(--mono)">{overdue}</div></div>'
                f'<div style="background:rgba(249,115,22,0.08);border:1px solid rgba(249,115,22,0.25);border-radius:8px;padding:0.5rem 1rem">'
                f'<div style="font-size:0.6rem;color:#FB923C;text-transform:uppercase;letter-spacing:0.1em">Act within 60d</div>'
                f'<div style="font-size:1.4rem;font-weight:700;color:#FB923C;font-family:var(--mono)">{urgent}</div></div>'
                f'<div style="background:rgba(74,222,128,0.04);border:1px solid rgba(74,222,128,0.15);border-radius:8px;padding:0.5rem 1rem">'
                f'<div style="font-size:0.6rem;color:#4ADE80;text-transform:uppercase;letter-spacing:0.1em">Monitored</div>'
                f'<div style="font-size:1.4rem;font-weight:700;color:#4ADE80;font-family:var(--mono)">{len(renewal_events)-overdue-urgent}</div></div>'
                f'<div style="background:rgba(239,68,68,0.06);border:1px solid rgba(239,68,68,0.2);border-radius:8px;padding:0.5rem 1rem">'
                f'<div style="font-size:0.6rem;color:#EF4444;text-transform:uppercase;letter-spacing:0.1em">Auto-Renew Risk</div>'
                f'<div style="font-size:1.4rem;font-weight:700;color:#EF4444;font-family:var(--mono)">{auto_count}</div></div>'
                f'<div style="background:rgba(245,158,11,0.06);border:1px solid rgba(245,158,11,0.2);border-radius:8px;padding:0.5rem 1rem">'
                f'<div style="font-size:0.6rem;color:#F59E0B;text-transform:uppercase;letter-spacing:0.1em">Zombie Contracts</div>'
                f'<div style="font-size:1.4rem;font-weight:700;color:#F59E0B;font-family:var(--mono)">{zombie_ct}</div></div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            for _, row in renewal_df.iterrows():
                _z = str(row.get("_zombie", ""))
                _a = str(row.get("_auto", ""))
                _zombie_html = (
                    '<span style="color:#F59E0B;font-size:0.75rem;margin-left:0.5rem">' + _z + "</span>"
                    if _z else ""
                )
                _auto_html = (
                    '<span style="color:#EF4444;font-size:0.85rem;margin-left:0.5rem">' + _a + "</span>"
                    if _a else ""
                )
                st.markdown(
                    f'<div style="background:{row["_bg"]};border-left:3px solid {row["_border"]};'
                    f'border-radius:0 8px 8px 0;padding:0.7rem 1rem;margin-bottom:0.5rem;'
                    f'display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:0.5rem">'
                    f'<div>'
                    f'<div style="font-weight:700;color:#E2E8F0;font-size:0.9rem">'
                    f'{html.escape(str(row["Supplier"]))}{_zombie_html}{_auto_html}'
                    f'</div>'
                    f'<div style="font-size:0.85rem;color:#C4D3E8;margin-top:0.15rem">'
                    f'Started {row["Start Date"]} &nbsp;·&nbsp; Expires {row["Expiry Date"]} &nbsp;·&nbsp; Notice: {row["Notice Period"]}</div>'
                    f'</div>'
                    f'<div style="text-align:right">'
                    f'<div style="font-family:var(--mono);font-size:0.85rem;color:{row["_border"]}">{row["Status"]}</div>'
                    f'<div style="font-family:var(--mono);font-size:0.82rem;color:#D0E0EF;margin-top:0.1rem">'
                    f'Act by {row["Action Deadline"]} ({row["Days to Deadline"]}d)</div>'
                    f'</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # Export renewal calendar
            _export_df = renewal_df.drop(columns=["_bg", "_border", "_zombie", "_auto"])
            st.download_button(
                label="📥 Export Renewal Calendar (.csv)",
                data=_export_df.to_csv(index=False),
                file_name=f"contract_renewals_{_dt.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                key="renewal_download",
            )
        else:
            st.info("No suppliers entered yet. Add suppliers in the Suppliers tab to see renewal calendar.")
        
        st.markdown("---")
        st.markdown("#### 💰 Savings Tracking")
        st.markdown(
            '<p class="muted">Define your savings baseline and track negotiated vs. realized savings against your target.</p>',
            unsafe_allow_html=True,
        )
        
        savings_col1, savings_col2, savings_col3 = st.columns(3)
        with savings_col1:
            baseline_method = st.radio(
                "Savings Baseline Methodology",
                ["Year-over-Year", "Should-Cost", "Market Benchmark"],
                help="Select how savings are calculated:\n"
                     "• Year-over-Year: Current year vs. prior year\n"
                     "• Should-Cost: Industry cost model\n"
                     "• Market Benchmark: Peer company reference prices",
                key="savings_baseline_method",
                horizontal=True
            )
        
        with savings_col2:
            if baseline_method == "Year-over-Year":
                baseline_value = st.number_input(
                    "Prior Year Annual Spend ($)",
                    min_value=0.0,
                    value=float(1000000),
                    step=10000.0,
                    key="yoy_baseline",
                    help="Last year's total annual spending on this category"
                )
                st.session_state["savings_baseline_value"] = baseline_value
            elif baseline_method == "Should-Cost":
                baseline_value = st.number_input(
                    "Should-Cost Target ($)",
                    min_value=0.0,
                    value=float(900000),
                    step=10000.0,
                    key="shouldcost_baseline",
                    help="Industry benchmark or internal target cost"
                )
                st.session_state["savings_baseline_value"] = baseline_value
            else:  # Market Benchmark
                baseline_value = st.number_input(
                    "Market Benchmark ($)",
                    min_value=0.0,
                    value=float(950000),
                    step=10000.0,
                    key="market_baseline",
                    help="Comparable company pricing from recent benchmarks"
                )
                st.session_state["savings_baseline_value"] = baseline_value
        
        with savings_col3:
            # Calculate projected savings from winner
            winner_price = st.session_state.get(f"raw_price_0", 1000000)  # First supplier is usually winner
            if baseline_value and winner_price:
                negotiated_savings = baseline_value - winner_price
                savings_pct = (negotiated_savings / baseline_value) * 100 if baseline_value > 0 else 0
                
                color = "#4ADE80" if savings_pct > 0 else "#EF4444"
                _sav_label = "PROJECTED SAVINGS" if savings_pct > 0 else "PROJECTED OVERSPEND"
                st.markdown(
                    f'<div style="background:rgba(74,222,128,0.08);border:1px solid rgba(74,222,128,0.25);'
                    f'border-radius:12px;padding:1.2rem 1.4rem;margin-top:0.5rem;text-align:center">'
                    f'<div style="font-size:0.78rem;color:{color};text-transform:uppercase;'
                    f'letter-spacing:0.14em;font-weight:700;margin-bottom:0.4rem">{_sav_label}</div>'
                    f'<div style="font-size:2.8rem;font-weight:800;color:{color};'
                    f'font-family:monospace;letter-spacing:-0.02em;line-height:1">'
                    f'${abs(negotiated_savings):,.0f}</div>'
                    f'<div style="font-size:1rem;font-weight:600;color:{color};margin-top:0.2rem">'
                    f'{savings_pct:+.1f}% vs. baseline</div>'
                    f'<div style="font-size:0.78rem;color:#A8BEDC;margin-top:0.4rem">'
                    f'Baseline ${baseline_value:,.0f} → Winner ${winner_price:,.0f}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        
        st.markdown("---")
        st.markdown(
            '<p class="muted">Upload a PO or invoice export to automatically classify spend, identify fragmentation, and surface sourcing opportunities.</p>',
            unsafe_allow_html=True,
        )

        uploaded_spend = st.file_uploader(
            "Upload PO / Invoice Export (CSV or Excel)",
            type=["csv", "xlsx", "xls"],
            key="spend_upload",
        )

        if uploaded_spend is not None:
            try:
                # ── AGENT #20: ERP Connector — auto-detect schema ──────────
                _erp_result = run_erp_connector(
                    file_bytes=uploaded_spend.read(),
                    file_name=uploaded_spend.name,
                )
                uploaded_spend.seek(0)  # Reset for downstream read
                if not _erp_result.get("error") and _erp_result.get("erp_detected", "Generic") != "Generic":
                    _erp_name = _erp_result["erp_detected"]
                    _erp_conf = int(_erp_result["confidence"] * 100)
                    st.markdown(
                        f'<div style="background:rgba(96,165,250,0.08);border:1px solid rgba(96,165,250,0.2);'
                        f'border-radius:8px;padding:0.5rem 0.9rem;margin-bottom:0.6rem;font-size:0.85rem">'
                        f'📄 <strong style="color:#60A5FA">File Format Detected:</strong> {html.escape(_erp_name)} '
                        f'<span style="color:#94A3B8">({_erp_conf}% confidence)</span> · '
                        f'{_erp_result.get("row_count",0):,} rows · '
                        f'Total: <strong style="color:#4ADE80">${_erp_result.get("total_spend",0):,.0f}</strong> · '
                        f'CapEx: ${_erp_result.get("capex_spend",0):,.0f} · '
                        f'OpEx: ${_erp_result.get("opex_spend",0):,.0f}'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    # Store normalized records for anomaly agent
                    if _erp_result.get("normalized_records"):
                        st.session_state["_erp_normalized_records"] = _erp_result["normalized_records"]
                    if _erp_result.get("quality_issues"):
                        for _qi in _erp_result["quality_issues"]:
                            st.warning(f"Data quality: {_qi}")

                if uploaded_spend.name.lower().endswith(".csv"):
                    spend_df = pd.read_csv(uploaded_spend)
                else:
                    spend_df = pd.read_excel(uploaded_spend)

                if spend_df.empty:
                    st.warning("Uploaded file is empty.")
                else:
                    col_list = spend_df.columns.tolist()
                    st.session_state["_spend_df_cols"] = col_list  # Used by ERP column mapper below

                    # Auto-detect columns — SAP standard field names + generic aliases
                    _SAP_DESC  = ["short text", "material description", "desc", "item", "product", "service", "line", "detail", "narr", "text"]
                    _SAP_AMT   = ["net amount", "amount", "total", "spend", "cost", "price", "value", "net", "dmbtr", "wrbtr"]
                    _SAP_SUP   = ["vendor name", "vendor", "supplier", "payee", "merchant", "lifnr"]
                    _SAP_PLANT = ["plant", "facility", "site", "location", "werk"]
                    _SAP_GL    = ["gl account", "gl acct", "g/l account", "account", "hkont", "saknr"]
                    _SAP_DTYPE = ["spend type", "doc type", "document type", "capex", "opex", "asset class", "blart"]

                    auto_desc  = next((c for c in col_list if any(k in c.lower() for k in _SAP_DESC)), col_list[0])
                    auto_amt   = next((c for c in col_list if any(k in c.lower() for k in _SAP_AMT)), col_list[min(1, len(col_list)-1)])
                    auto_sup   = next((c for c in col_list if any(k in c.lower() for k in _SAP_SUP)), None)
                    auto_plant = next((c for c in col_list if any(k in c.lower() for k in _SAP_PLANT)), None)
                    auto_gl    = next((c for c in col_list if any(k in c.lower() for k in _SAP_GL)), None)
                    auto_dtype = next((c for c in col_list if any(k in c.lower() for k in _SAP_DTYPE)), None)

                    sc1, sc2, sc3 = st.columns(3)
                    with sc1:
                        desc_col = st.selectbox("Description Column", col_list, index=col_list.index(auto_desc), key="spend_desc_col")
                    with sc2:
                        amt_col = st.selectbox("Amount Column", col_list, index=col_list.index(auto_amt), key="spend_amt_col")
                    with sc3:
                        sup_opts = ["(none)"] + col_list
                        default_sup_idx = sup_opts.index(auto_sup) if auto_sup and auto_sup in sup_opts else 0
                        supplier_col = st.selectbox("Supplier Column (optional)", sup_opts, index=default_sup_idx, key="spend_supplier_col")

                    sc4, sc5, sc6 = st.columns(3)
                    with sc4:
                        plant_opts = ["(none)"] + col_list
                        plant_col = st.selectbox(
                            "Facility / Plant Column", plant_opts,
                            index=plant_opts.index(auto_plant) if auto_plant and auto_plant in plant_opts else 0,
                            key="spend_plant_col",
                            help="Enables cross-facility fragmentation view (SAP: Plant / Werk).",
                        )
                    with sc5:
                        gl_opts = ["(none)"] + col_list
                        gl_col = st.selectbox(
                            "GL Account Column", gl_opts,
                            index=gl_opts.index(auto_gl) if auto_gl and auto_gl in gl_opts else 0,
                            key="spend_gl_col",
                            help="GL account enables CapEx vs OpEx split (SAP: HKont / Saknr).",
                        )
                    with sc6:
                        dtype_opts = ["(none)"] + col_list
                        dtype_col = st.selectbox(
                            "Spend Type Column (CapEx / OpEx)", dtype_opts,
                            index=dtype_opts.index(auto_dtype) if auto_dtype and auto_dtype in dtype_opts else 0,
                            key="spend_dtype_col",
                            help="Labels each line as Capital or Operating spend.",
                        )

                    # ── Classify with TF-IDF ──
                    if _SKLEARN_AVAILABLE and TfidfVectorizer is not None and cosine_similarity is not None:
                        corpus, subcat_labels_list, parent_labels_list = [], [], []
                        for pk, pd_data in CATEGORY_TAXONOMY.items():
                            for sub in pd_data["subcategories"]:
                                corpus.append(sub["name"] + " " + sub.get("notes", ""))
                                subcat_labels_list.append(sub["name"])
                                parent_labels_list.append(pd_data["label"])

                        _vec = TfidfVectorizer(ngram_range=(1, 2), max_features=8000)
                        _tfidf = _vec.fit_transform(corpus)
                        _descs = spend_df[desc_col].fillna("").astype(str).tolist()
                        _desc_vec = _vec.transform(_descs)
                        _sims = cosine_similarity(_desc_vec, _tfidf)
                        _best_idx = _sims.argmax(axis=1)
                        _best_scores = _sims.max(axis=1)

                        spend_df = spend_df.copy()
                        spend_df["Auto Category"] = [parent_labels_list[i] for i in _best_idx]
                        spend_df["Auto Subcategory"] = [subcat_labels_list[i] for i in _best_idx]
                        spend_df["Match Confidence"] = [round(float(s), 2) for s in _best_scores]
                    else:
                        spend_df = spend_df.copy()
                        spend_df["Auto Category"] = "Unclassified"
                        spend_df["Auto Subcategory"] = "Unclassified"
                        spend_df["Match Confidence"] = 0.0
                        st.info("scikit-learn not available — install it to enable auto-classification.")

                    # ── Show classified table ──
                    st.markdown("#### Classified Line Items (first 100 rows)")
                    display_cols = [desc_col, "Auto Category", "Auto Subcategory", "Match Confidence"]
                    if supplier_col != "(none)":
                        display_cols = [desc_col, supplier_col, "Auto Category", "Auto Subcategory", "Match Confidence"]
                    st.dataframe(spend_df[display_cols].head(100), use_container_width=True)

                    # ── Spend by Category chart ──
                    spend_df[amt_col] = pd.to_numeric(spend_df[amt_col], errors="coerce").fillna(0)
                    cat_spend = spend_df.groupby("Auto Category")[amt_col].sum().reset_index()
                    cat_spend.columns = ["Category", "Total Spend"]
                    cat_spend = cat_spend.sort_values("Total Spend", ascending=False)

                    fig_spend = px.bar(
                        cat_spend, x="Category", y="Total Spend",
                        title="Spend by Auto-Classified Category",
                        color="Total Spend",
                        color_continuous_scale=[[0, "#1E3A5F"], [0.5, "#3B82F6"], [1, "#60A5FA"]],
                    )
                    fig_spend.update_layout(
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,38,0.7)",
                        font=dict(color="#8EA3C3", family="Outfit"),
                        xaxis=dict(tickangle=-30, gridcolor="rgba(255,255,255,0.04)"),
                        yaxis=dict(gridcolor="rgba(255,255,255,0.04)"),
                        coloraxis_showscale=False,
                        height=340, margin=dict(t=40, b=10, l=10, r=10),
                    )
                    st.plotly_chart(fig_spend, use_container_width=True)

                    # ── Fragmentation alerts ──
                    if supplier_col != "(none)":
                        st.markdown("#### Fragmentation Alerts")
                        frag = spend_df.groupby("Auto Category")[supplier_col].nunique().reset_index()
                        frag.columns = ["Category", "Unique Suppliers"]
                        frag = frag[frag["Unique Suppliers"] >= 3].sort_values("Unique Suppliers", ascending=False)
                        if not frag.empty:
                            for _, row in frag.iterrows():
                                row_category = row["Category"]
                                series = cast(pd.Series, cat_spend.loc[cat_spend["Category"] == row_category, "Total Spend"])
                                cat_total = float(pd.to_numeric(series, errors="coerce").sum())
                                st.markdown(
                                    f'<div style="background:rgba(252,211,77,0.06);border-left:3px solid #FCD34D;'
                                    f'border-radius:0 8px 8px 0;padding:0.6rem 1rem;margin-bottom:0.4rem;font-size:0.86rem">'
                                    f'⚠️ <strong style="color:#FCD34D">{row["Category"]}</strong> — '
                                    f'{row["Unique Suppliers"]} unique suppliers · '
                                    f'${cat_total:,.0f} total spend · Consider consolidation</div>',
                                    unsafe_allow_html=True,
                                )
                        else:
                            st.success("No significant fragmentation detected across categories.")

                        # Top supplier per category
                        st.markdown("#### Top Supplier per Category")
                        top_sup = (
                            spend_df.groupby(["Auto Category", supplier_col])[amt_col]
                            .sum()
                            .reset_index()
                        )
                        top_sup.columns = ["Category", "Supplier", "Spend"]
                        top_sup = top_sup.loc[top_sup.groupby("Category")["Spend"].idxmax()]
                        top_sup = top_sup.sort_values("Spend", ascending=False).reset_index(drop=True)
                        st.dataframe(top_sup, use_container_width=True)

                        # ── Multi-Facility Cross-Reference ──────────────────
                        if plant_col != "(none)" and plant_col in spend_df.columns:
                            st.markdown("#### Cross-Facility Supplier Fragmentation")
                            st.markdown(
                                '<p class="muted">Suppliers appearing at multiple facilities as separate POs — '
                                'prime targets for a national contract that consolidates site-by-site purchasing.</p>',
                                unsafe_allow_html=True,
                            )
                            _mf = (
                                spend_df.groupby(supplier_col)
                                .agg(
                                    Facilities=(plant_col, "nunique"),
                                    Total_Spend=(amt_col, "sum"),
                                    PO_Count=(amt_col, "count"),
                                )
                                .reset_index()
                            )
                            _mf = _mf[_mf["Facilities"] >= 2].sort_values("Total_Spend", ascending=False).head(20)
                            _mf.columns = ["Supplier", "# Facilities", "Total Spend", "PO Count"]
                            if not _mf.empty:
                                for _, _mfr in _mf.iterrows():
                                    _consl_opp = _mfr["Total Spend"] * 0.08
                                    st.markdown(
                                        f'<div style="background:rgba(74,222,128,0.05);border-left:3px solid #4ADE80;'
                                        f'border-radius:0 8px 8px 0;padding:0.6rem 1rem;margin-bottom:0.4rem;font-size:0.85rem">'
                                        f'<strong style="color:#4ADE80">{_mfr["Supplier"]}</strong> — '
                                        f'{_mfr["# Facilities"]} facilities · {int(_mfr["PO Count"])} POs · '
                                        f'${_mfr["Total Spend"]:,.0f} total · '
                                        f'<span style="color:#FCD34D">Est. consolidation opportunity: ${_consl_opp:,.0f}</span></div>',
                                        unsafe_allow_html=True,
                                    )
                            else:
                                st.info("No multi-facility supplier fragmentation detected.")

                    # ── Pareto / 80-20 concentration analysis ──────────────
                    st.markdown("---")
                    st.markdown("#### Pareto Concentration — Where Your Money Goes")
                    _total_spend = float(spend_df[amt_col].sum())
                    if _total_spend > 0 and not cat_spend.empty:
                        cat_spend_sorted = cat_spend.sort_values("Total Spend", ascending=False).reset_index(drop=True)
                        cat_spend_sorted["Cumulative %"] = (cat_spend_sorted["Total Spend"].cumsum() / _total_spend * 100).round(1)
                        cat_spend_sorted["Spend %"] = (cat_spend_sorted["Total Spend"] / _total_spend * 100).round(1)
                        _80pct_cats = (cat_spend_sorted["Cumulative %"] <= 80).sum() + 1
                        st.markdown(
                            f'<div style="background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.15);'
                            f'border-radius:10px;padding:0.8rem 1.2rem;margin-bottom:0.8rem">'
                            f'<div style="font-family:var(--mono);font-size:0.78rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:0.3rem">80/20 Rule</div>'
                            f'<div style="font-size:1rem;color:#F1F5F9">'
                            f'<strong style="color:#60A5FA">{_80pct_cats}</strong> of {len(cat_spend_sorted)} categories '
                            f'represent 80% of your ${_total_spend:,.0f} total spend. '
                            f'Start sourcing events here — these are where savings compound fastest.</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                        _fig_pareto = px.bar(
                            cat_spend_sorted, x="Category", y="Spend %",
                            title="Spend Concentration by Category (%)",
                            color="Cumulative %",
                            color_continuous_scale=[[0,"#1D4ED8"],[0.5,"#3B82F6"],[1,"#60A5FA"]],
                            text="Spend %",
                        )
                        _fig_pareto.add_scatter(
                            x=cat_spend_sorted["Category"], y=cat_spend_sorted["Cumulative %"],
                            mode="lines+markers", name="Cumulative %", yaxis="y2",
                            line=dict(color="#4ADE80", width=2), marker=dict(size=5),
                        )
                        _fig_pareto.update_layout(
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,38,0.6)",
                            font=dict(color="#8EA3C3"), xaxis=dict(tickangle=-30),
                            yaxis=dict(title="Spend %", gridcolor="rgba(255,255,255,0.04)"),
                            yaxis2=dict(title="Cumulative %", overlaying="y", side="right", range=[0, 110], showgrid=False),
                            coloraxis_showscale=False, showlegend=False,
                            height=340, margin=dict(t=40,b=10,l=10,r=10),
                        )
                        _fig_pareto.update_traces(texttemplate="%{text:.0f}%", textposition="outside", selector=dict(type="bar"))
                        st.plotly_chart(_fig_pareto, use_container_width=True)

                    # ── Sourcing opportunity scoring ────────────────────────
                    st.markdown("---")
                    st.markdown("#### Sourcing Opportunity Ranking")
                    st.markdown(
                        '<p class="muted">Each category scored on spend size, supplier count, and last-bid recency. '
                        'P1 = launch now. P2 = this quarter. P3 = monitor.</p>',
                        unsafe_allow_html=True,
                    )
                    _opp_rows = []
                    for _, _cr in cat_spend.iterrows():
                        _cat_name = _cr["Category"]
                        _cat_amt  = float(_cr["Total Spend"])
                        _n_sups   = int(spend_df[spend_df["Auto Category"] == _cat_name][supplier_col].nunique()) if supplier_col != "(none)" else 1
                        # Spend score
                        _s_score = 30 if _cat_amt >= 1_000_000 else 15 if _cat_amt >= 250_000 else 8 if _cat_amt >= 50_000 else 2
                        # Fragmentation score
                        _f_score = 20 if _n_sups >= 5 else 12 if _n_sups >= 3 else 5 if _n_sups >= 2 else 0
                        _total_opp = _s_score + _f_score
                        _priority = "P1 — Launch Now" if _total_opp >= 35 else "P2 — This Quarter" if _total_opp >= 18 else "P3 — Monitor"
                        _p_color  = "#F87171" if "P1" in _priority else "#FCD34D" if "P2" in _priority else "#94A3B8"
                        _opp_rows.append({
                            "Category": _cat_name, "Total Spend": _cat_amt,
                            "Suppliers": _n_sups, "Opp Score": _total_opp,
                            "Priority": _priority, "_color": _p_color,
                        })
                    _opp_rows.sort(key=lambda x: x["Opp Score"], reverse=True)
                    for _opp in _opp_rows:
                        st.markdown(
                            f'<div style="display:flex;justify-content:space-between;align-items:center;'
                            f'background:rgba(4,9,15,0.5);border:1px solid rgba(96,165,250,0.07);'
                            f'border-radius:8px;padding:0.6rem 1rem;margin-bottom:0.4rem;flex-wrap:wrap;gap:0.4rem">'
                            f'<div><div style="font-weight:700;color:#E2E8F0;font-size:0.9rem">{html.escape(str(_opp["Category"]))}</div>'
                            f'<div style="font-size:0.85rem;color:#D0E0EF">${_opp["Total Spend"]:,.0f} spend &nbsp;·&nbsp; {_opp["Suppliers"]} supplier{"s" if _opp["Suppliers"] != 1 else ""}</div></div>'
                            f'<div style="display:flex;align-items:center;gap:0.8rem">'
                            f'<span style="font-family:var(--mono);font-size:0.75rem;color:{_opp["_color"]};'
                            f'border:1px solid {_opp["_color"]}44;background:{_opp["_color"]}11;'
                            f'border-radius:4px;padding:0.15rem 0.5rem">{html.escape(_opp["Priority"])}</span>'
                            f'<span style="font-family:var(--mono);font-size:0.82rem;color:#A8BEDC">Score: {_opp["Opp Score"]}</span>'
                            f'</div></div>',
                            unsafe_allow_html=True,
                        )

                    # ── Export ──
                    csv_classified = spend_df.to_csv(index=False)
                    st.download_button(
                        "⬇ Export Classified Data (CSV)",
                        csv_classified,
                        "spend_classified.csv",
                        "text/csv",
                        key="spend_export",
                    )
            except Exception as _e:
                st.error(f"Error reading file: {_e}")

        else:
            # No file — show instructions + sample template
            st.markdown(
                '<div style="background:#0A1628;border:1px solid rgba(96,165,250,0.15);border-radius:12px;'
                'padding:1.4rem 1.6rem;margin:0.8rem 0">'
                '<div style="font-size:0.82rem;text-transform:uppercase;letter-spacing:0.12em;color:#60A5FA;font-weight:700;margin-bottom:0.8rem">HOW IT WORKS</div>'
                '<div style="display:flex;flex-direction:column;gap:0.5rem;font-size:0.88rem;color:#CBD5E1">'
                '<div>1. Export your PO, invoice, or spend data to CSV or Excel</div>'
                '<div>2. Upload it above — the tool auto-detects description and amount columns</div>'
                '<div>3. Each line item is classified against the ProcureIQ category taxonomy using TF-IDF text matching</div>'
                '<div>4. You get: spend by category chart · fragmentation alerts · top supplier per category · exportable output</div>'
                '</div></div>',
                unsafe_allow_html=True,
            )

            sample_df = pd.DataFrame({
                "Description": [
                    "Cloud hosting and compute services", "IT help desk managed support",
                    "Background screening services", "Payroll processing Q1",
                    "Temp labor - warehouse operations", "ERP license renewal",
                    "Corporate travel management", "Employee wellness platform",
                ],
                "Amount": [150000, 80000, 15000, 120000, 95000, 280000, 45000, 22000],
                "Supplier": ["AWS", "TechCo MSP", "HireRight", "ADP", "Staffing Inc", "SAP", "Concur", "WellnessApp"],
            })
            csv_sample = sample_df.to_csv(index=False)
            st.download_button(
                "⬇ Download Sample Template",
                csv_sample,
                "spend_template.csv",
                "text/csv",
                key="spend_sample_dl",
            )

        # ── USASpending.gov Federal Benchmark ─────────────────
        st.markdown("---")
        st.markdown("#### Federal Contract Benchmark — USASpending.gov")
        st.markdown(
            '<p class="muted">Real federal government contract awards for this category. '
            'Use these as market-rate benchmarks and supplier discovery signals. '
            'Data sourced live from USASpending.gov — the official US federal spend database.</p>',
            unsafe_allow_html=True,
        )
        _usa_keyword = USASPENDING_KEYWORD_MAP.get(
            selected_sub_name,
            selected_sub_name.split("(")[0].strip()[:40]
        )
        _usa_col1, _usa_col2 = st.columns([3, 1])
        with _usa_col2:
            _usa_custom = st.text_input(
                "Search keyword (optional)",
                value=_usa_keyword,
                key="usa_keyword_override",
                help="Override the auto-detected keyword for this category.",
            )
        _kw_to_use = _usa_custom.strip() if _usa_custom.strip() else _usa_keyword
        with _usa_col1:
            st.markdown(
                f'<div style="font-size:0.85rem;color:#C4D3E8;padding:0.6rem 0">'
                f'Searching: <strong style="color:#60A5FA">"{html.escape(_kw_to_use)}"</strong> · '
                f'Federal awards 2022–2025</div>',
                unsafe_allow_html=True,
            )

        if st.button("Pull Federal Benchmarks", key="usa_fetch", type="primary"):
            with st.spinner("Querying USASpending.gov..."):
                _usa_awards = fetch_usaspending_awards(_kw_to_use, limit=8)
            if _usa_awards:
                _total_usa = sum(float(a.get("Award Amount") or 0) for a in _usa_awards)
                st.markdown(
                    f'<div style="background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.15);'
                    f'border-radius:10px;padding:0.8rem 1.2rem;margin-bottom:1rem">'
                    f'<div style="font-size:0.80rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.1em">Top {len(_usa_awards)} Federal Awards Found</div>'
                    f'<div style="font-size:1.3rem;font-weight:700;color:#F1F5F9">'
                    f'${_total_usa:,.0f} combined · avg ${_total_usa/len(_usa_awards):,.0f}</div>'
                    f'<div style="font-size:0.82rem;color:#A8BEDC">Use these award amounts as your market rate upper bound. '
                    f'Federal contracts often carry 10-20% premium over commercial pricing.</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                for _aw in _usa_awards:
                    _aw_amt = float(_aw.get("Award Amount") or 0)
                    _aw_rec = str(_aw.get("Recipient Name") or "Unknown")
                    _aw_desc = str(_aw.get("Description") or "")[:120]
                    _aw_agency = str(_aw.get("Awarding Agency Name") or "")
                    _aw_date = str(_aw.get("Action Date") or "")
                    st.markdown(
                        f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(96,165,250,0.10);'
                        f'border-radius:8px;padding:0.65rem 1rem;margin-bottom:0.4rem;'
                        f'display:flex;justify-content:space-between;align-items:flex-start;gap:1rem">'
                        f'<div style="flex:1;min-width:0">'
                        f'<div style="font-weight:600;color:#E2E8F0;font-size:0.88rem">{html.escape(_aw_rec)}</div>'
                        f'<div style="font-size:0.80rem;color:#A8BEDC;margin-top:0.1rem">{html.escape(_aw_desc)}</div>'
                        f'<div style="font-size:0.75rem;color:#8BAAC4;margin-top:0.1rem">'
                        f'{html.escape(_aw_agency)} · {html.escape(_aw_date)}</div>'
                        f'</div>'
                        f'<div style="font-weight:700;color:#4ADE80;font-family:monospace;'
                        f'font-size:0.92rem;white-space:nowrap">${_aw_amt:,.0f}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.info(
                    "No federal awards found for this keyword. "
                    "Try a broader search term — e.g. 'maintenance services' instead of a specific subcategory name."
                )

    # end with main_col

    # ── ERP FIELD MAPPING ─────────────────────────────────────────────
    with tab_spend:
        st.markdown("---")
        st.markdown("#### 🔌 ERP / P2P Field Mapping")
        st.markdown(
            '<p class="muted">Map your ERP or P2P export columns to ProcureIQ fields. '
            'Supports SAP S/4HANA, Oracle Fusion, Coupa, and Ariba export formats. '
            'GL account ranges auto-classify spend as CapEx or OpEx.</p>',
            unsafe_allow_html=True,
        )

        _ERP_SYSTEMS = {
            "SAP S/4HANA": {
                "fields": {
                    "Supplier": ["NAME1", "LIFNR", "Vendor Name", "Vendor"],
                    "Amount":   ["DMBTR", "WRBTR", "Amount in Local Currency", "Net Value"],
                    "Description": ["TXZ01", "MAKTX", "Short Text", "Material Description"],
                    "GL Account": ["SAKNR", "HKONT", "G/L Account", "GL Account"],
                    "Plant": ["WERKS", "Plant", "Werk"],
                    "Cost Center": ["KOSTL", "Cost Center"],
                    "PO Number": ["EBELN", "Purchasing Document"],
                    "Invoice Date": ["BLDAT", "BUDAT", "Document Date"],
                },
                "capex_gl_ranges": ["0100-0999", "1000-1999 (Assets)"],
                "opex_gl_ranges":  ["4000-5999 (Operating)", "6000-6999 (G&A)"],
                "note": "SAP MIRO / ME2M / FBL1N exports. Use transaction MB51 for material movements.",
            },
            "Oracle Fusion": {
                "fields": {
                    "Supplier": ["VENDOR_NAME", "SUPPLIER_NAME", "Party Name"],
                    "Amount":   ["AMOUNT", "ENTERED_AMOUNT", "ACCOUNTED_AMOUNT"],
                    "Description": ["DESCRIPTION", "ITEM_DESCRIPTION", "LINE_DESCRIPTION"],
                    "GL Account": ["CODE_COMBINATION_ID", "ACCOUNT", "GL_CODE"],
                    "Cost Center": ["COST_CENTER", "DEPARTMENT"],
                    "PO Number": ["PO_NUMBER", "ORDER_NUMBER"],
                    "Invoice Date": ["INVOICE_DATE", "ACCOUNTING_DATE"],
                },
                "capex_gl_ranges": ["1500-1999 (Fixed Assets)", "1200-1499 (Intangibles)"],
                "opex_gl_ranges":  ["5000-5999 (Expenses)", "6000-6999 (Admin)"],
                "note": "Oracle FBDI / OTBI export. Use Payables Invoice Report or PO Accrual Reports.",
            },
            "Coupa": {
                "fields": {
                    "Supplier": ["Supplier Name", "supplier_name", "Vendor"],
                    "Amount":   ["Amount", "Total Price", "amount_in_usd"],
                    "Description": ["Description", "Line Description", "Commodity"],
                    "GL Account": ["Account", "Charge Account", "account_code"],
                    "Cost Center": ["Cost Center", "Department"],
                    "PO Number": ["PO Number", "Purchase Order"],
                    "Invoice Date": ["Invoice Date", "Created At"],
                },
                "capex_gl_ranges": ["15XX (Capital)", "17XX (Construction in Progress)"],
                "opex_gl_ranges":  ["50XX–69XX (Operating)"],
                "note": "Coupa Spend360 / Transaction Export. Commodity codes map directly to ProcureIQ subcategories.",
            },
            "SAP Ariba": {
                "fields": {
                    "Supplier": ["SupplierName", "VendorName", "supplier_name"],
                    "Amount":   ["InvoiceAmount", "TotalAmount", "NetAmount"],
                    "Description": ["LineDescription", "Description", "CommodityName"],
                    "GL Account": ["GLAccount", "AccountCode"],
                    "Cost Center": ["CostCenter", "BusinessUnit"],
                    "PO Number": ["PurchaseOrderNumber", "OrderId"],
                    "Invoice Date": ["InvoiceDate", "PostingDate"],
                },
                "capex_gl_ranges": ["1XXXXX (Assets)"],
                "opex_gl_ranges":  ["5XXXXX / 6XXXXX (Operations)"],
                "note": "Ariba Spend Analysis / Network reports. UNSPSC commodity codes auto-map to subcategories.",
            },
        }

        _GL_CAPEX_KEYWORDS = ["capital", "asset", "equipment", "construction", "capex", "infrastructure", "facility", "1xxx", "15xx", "16xx"]
        _GL_OPEX_KEYWORDS  = ["operating", "opex", "expense", "service", "maintenance", "software", "4xxx", "5xxx", "6xxx", "7xxx"]

        _erp_sys = st.selectbox("ERP / P2P System", list(_ERP_SYSTEMS.keys()), key="erp_system_select")
        _erp_cfg = _ERP_SYSTEMS[_erp_sys]

        _erp_c1, _erp_c2 = st.columns(2)
        with _erp_c1:
            st.markdown(
                f'<div style="background:#060D1A;border:1px solid rgba(96,165,250,0.15);border-radius:10px;padding:1rem;margin-bottom:0.8rem">'
                f'<div style="font-size:0.78rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:0.6rem">Standard Field Aliases</div>',
                unsafe_allow_html=True,
            )
            for _piq_field, _aliases in _erp_cfg["fields"].items():
                _alias_str = " · ".join(f"<code style='background:rgba(96,165,250,0.1);border-radius:3px;padding:0.1rem 0.3rem;font-size:0.78rem;color:#93C5FD'>{a}</code>" for a in _aliases[:3])
                st.markdown(
                    f'<div style="display:flex;gap:0.5rem;margin-bottom:0.4rem;align-items:baseline">'
                    f'<span style="font-size:0.85rem;font-weight:600;color:#E2E8F0;min-width:110px">{_piq_field}</span>'
                    f'<span style="font-size:0.82rem;color:#94A3B8">→</span> {_alias_str}'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            st.markdown('</div>', unsafe_allow_html=True)
            st.caption(_erp_cfg["note"])

        with _erp_c2:
            st.markdown(
                f'<div style="background:#060D1A;border:1px solid rgba(96,165,250,0.15);border-radius:10px;padding:1rem;margin-bottom:0.8rem">'
                f'<div style="font-size:0.78rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:0.6rem">CapEx vs OpEx GL Classification</div>'
                f'<div style="font-size:0.88rem;color:#4ADE80;font-weight:600;margin-bottom:0.3rem">CapEx GL Ranges</div>'
                + "".join(
                    f'<div style="font-size:0.84rem;color:#C4D3E8;margin-bottom:0.2rem">· {r}</div>'
                    for r in _erp_cfg["capex_gl_ranges"]
                )
                + f'<div style="font-size:0.88rem;color:#60A5FA;font-weight:600;margin-bottom:0.3rem;margin-top:0.6rem">OpEx GL Ranges</div>'
                + "".join(
                    f'<div style="font-size:0.84rem;color:#C4D3E8;margin-bottom:0.2rem">· {r}</div>'
                    for r in _erp_cfg["opex_gl_ranges"]
                )
                + '</div>',
                unsafe_allow_html=True,
            )

            # GL Account classifier
            st.markdown("##### Quick GL Classifier")
            _gl_input = st.text_input(
                "Enter GL Account / Code",
                placeholder="e.g. 0100-0001 or 5010",
                key="gl_classify_input",
            )
            if _gl_input.strip():
                _gl_low = _gl_input.strip().lower()
                _is_capex = any(kw in _gl_low for kw in _GL_CAPEX_KEYWORDS) or (
                    _gl_low[:2] in ("01", "10", "11", "12", "13", "14", "15", "16", "17")
                )
                _is_opex = any(kw in _gl_low for kw in _GL_OPEX_KEYWORDS) or (
                    _gl_low[:2] in ("40", "41", "42", "43", "44", "45", "50", "51", "52", "53",
                                    "54", "55", "60", "61", "62", "63", "64", "65", "70", "71")
                )
                if _is_capex and not _is_opex:
                    st.markdown(
                        f'<div style="background:rgba(74,222,128,0.08);border:1px solid rgba(74,222,128,0.25);'
                        f'border-radius:8px;padding:0.6rem 0.9rem;font-size:0.9rem;color:#4ADE80;font-weight:600">'
                        f'CAPEX — Capital expenditure. Budget against CapEx plan; likely requires Finance approval.</div>',
                        unsafe_allow_html=True,
                    )
                elif _is_opex and not _is_capex:
                    st.markdown(
                        f'<div style="background:rgba(96,165,250,0.08);border:1px solid rgba(96,165,250,0.25);'
                        f'border-radius:8px;padding:0.6rem 0.9rem;font-size:0.9rem;color:#60A5FA;font-weight:600">'
                        f'OPEX — Operating expenditure. Run-rate cost, typically within procurement authority.</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        f'<div style="background:rgba(251,191,36,0.08);border:1px solid rgba(251,191,36,0.25);'
                        f'border-radius:8px;padding:0.6rem 0.9rem;font-size:0.9rem;color:#FCD34D">'
                        f'AMBIGUOUS — Review account description. Could be mixed-use (e.g. IT infrastructure that straddles CapEx/OpEx).</div>',
                        unsafe_allow_html=True,
                    )

        # Column mapper for uploaded spend files
        if st.session_state.get("_spend_df_cols"):
            st.markdown("##### Auto-Map Your Upload Columns")
            _upload_cols = st.session_state.get("_spend_df_cols", [])
            _mapped = {}
            _aliases_flat = {a.lower(): piq_f for piq_f, aliases in _erp_cfg["fields"].items() for a in aliases}
            for _uc in _upload_cols:
                _match = _aliases_flat.get(_uc.lower()) or _aliases_flat.get(_uc.replace(" ", "_").lower())
                if _match:
                    _mapped[_uc] = _match
            if _mapped:
                st.markdown(
                    '<div style="font-size:0.82rem;color:#4ADE80;margin-bottom:0.4rem">Auto-detected column mappings:</div>',
                    unsafe_allow_html=True,
                )
                for _uc, _pf in _mapped.items():
                    st.markdown(
                        f'<div style="font-size:0.84rem;color:#E2E8F0;margin-bottom:0.2rem">'
                        f'<code style="color:#93C5FD">{html.escape(_uc)}</code> → <strong>{_pf}</strong></div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.caption("No automatic column matches found. Check that your export uses standard field names for this ERP system.")

    # ── CONTRACT READER ──────────────────────────────────────
    with tab_comms:
      if _comms_section == "📄 Contract Reader":
        st.markdown("### Contract Reader & Risk Highlighter")
        st.markdown(
            '<p class="muted">Paste or upload any contract. '
            'The reader scans for 12 high-risk clause categories — auto-renewal traps, '
            'liability caps, indemnification asymmetry, governing law, termination rights, '
            'IP ownership, and more. Works for any industry and any contract type.</p>',
            unsafe_allow_html=True,
        )

        _cr_col1, _cr_col2 = st.columns([1, 1])
        with _cr_col1:
            _contract_name = st.text_input(
                "Contract Name / Reference",
                placeholder="e.g. MSA with Grainger — 2024 Renewal",
                key="cr_name",
            )
            _counterparty = st.text_input(
                "Counterparty (Supplier / Vendor)",
                placeholder="e.g. Grainger Industrial Supply",
                key="cr_counterparty",
            )
        with _cr_col2:
            _cr_jurisdiction = st.text_input(
                "Governing Law / Jurisdiction",
                placeholder="e.g. State of Georgia, USA",
                key="cr_jurisdiction",
                help="Where this contract will be executed and disputes resolved. Different states/countries have very different default rules on limitation of liability, implied warranties, and indemnification.",
            )
            _contract_type_cr = st.selectbox(
                "Contract Type",
                ["Master Service Agreement", "Supply Agreement", "SaaS / License", "Statement of Work",
                 "Lease Agreement", "Professional Services", "Distribution Agreement", "Other"],
                key="cr_contract_type",
            )

        _contract_text = st.text_area(
            "Paste Contract Text",
            height=220,
            placeholder="Paste the full contract or specific clauses here. The reader will scan for risk flags across all 12 categories.",
            key="cr_text",
        )

        _uploaded_contract = st.file_uploader(
            "Or upload contract (.txt or .pdf text extract)",
            type=["txt"],
            key="cr_upload",
        )
        if _uploaded_contract is not None:
            _contract_text = _uploaded_contract.read().decode("utf-8", errors="ignore")
            st.success(f"Loaded {len(_contract_text):,} characters from {_uploaded_contract.name}")

        if st.button("Analyze Contract", type="primary", key="cr_analyze", use_container_width=True):
            if not _contract_text.strip():
                st.warning("Paste contract text or upload a file first.")
            else:
                import re as _re

                # ── Risk pattern library — 12 categories ────────────
                _RISK_PATTERNS = [
                    {
                        "category": "Auto-Renewal Trap",
                        "tier": "HIGH",
                        "patterns": [
                            r"auto.?renew", r"automatically renew", r"unless.*notice.*given",
                            r"shall renew.*unless", r"evergreen", r"successive.*term",
                            r"notice of non.?renewal", r"termination notice.*prior",
                        ],
                        "why": "Contracts that auto-renew without action lock you into another full term. The risk is missing the notice window — often 60-180 days before expiry — which is your true action deadline, not the contract end date.",
                        "what_to_do": "Negotiate: (1) change notice period to 30 days, (2) add a mutual opt-out right, (3) require vendor to send written renewal reminder 90 days before deadline.",
                    },
                    {
                        "category": "Limitation of Liability Cap",
                        "tier": "HIGH",
                        "patterns": [
                            r"limitation of liability", r"limit.*liability.*not exceed",
                            r"aggregate liability.*shall not", r"liability.*capped at",
                            r"in no event.*liable", r"maximum liability",
                        ],
                        "why": "Liability caps that are set to 1x or 3x annual contract value are standard vendor positions — and almost always inadequate for data breaches, supply failures, or critical service outages where damages far exceed fees paid.",
                        "what_to_do": "Push for: (1) carve-outs from the cap for data breaches, fraud, and gross negligence; (2) cap set at minimum 12 months fees; (3) separate cap for IP indemnification.",
                    },
                    {
                        "category": "Indemnification Asymmetry",
                        "tier": "HIGH",
                        "patterns": [
                            r"indemnif", r"hold harmless", r"defend.*indemnify",
                            r"indemnity.*obligation", r"third.?party claim",
                        ],
                        "why": "One-sided indemnification — where you indemnify the vendor for broad categories but get narrow protection in return — is the most common expensive mistake in commercial contracts.",
                        "what_to_do": "Require mutual indemnification. Your indemnification obligations should be limited to your own acts/omissions. Vendor should indemnify you for IP infringement, data breaches, and their negligence.",
                    },
                    {
                        "category": "Governing Law & Jurisdiction",
                        "tier": "MEDIUM",
                        "patterns": [
                            r"govern(?:ed|ing) by.*law", r"laws of the state of",
                            r"jurisdiction.*courts", r"exclusive jurisdiction",
                            r"choice of law", r"venue.*shall be",
                            r"dispute.*arbitrat", r"binding arbitration",
                        ],
                        "why": "The governing state determines default rules on implied warranties, damage limitations, and enforcement. Mandatory arbitration clauses can block class actions and limit discovery rights.",
                        "what_to_do": "Negotiate for your home jurisdiction. If vendor insists on arbitration: require (1) the right to seek injunctive relief in court, (2) seat of arbitration in your city, (3) no waiver of class action for claims under $X.",
                    },
                    {
                        "category": "Termination for Convenience",
                        "tier": "HIGH",
                        "patterns": [
                            r"terminat.*convenience", r"terminat.*without cause",
                            r"may terminat.*upon.*notice", r"right to terminat",
                            r"terminat.*for any reason",
                        ],
                        "why": "If only the vendor has termination for convenience rights, you are locked in for the full term with no exit. This is extremely common in vendor-drafted contracts.",
                        "what_to_do": "Require mutual termination for convenience with 30-90 day notice. If vendor refuses, negotiate a buyout formula rather than full remaining fees as the exit cost.",
                    },
                    {
                        "category": "IP Ownership Risk",
                        "tier": "HIGH",
                        "patterns": [
                            r"intellectual property.*vendor", r"vendor.*retain.*ip",
                            r"work.*for.*hire", r"all.*right.*title.*interest",
                            r"assign.*all.*ip", r"derivative work",
                            r"proprietary.*remains", r"supplier.*owns",
                        ],
                        "why": "Custom deliverables, data models, and work product default to the creator unless explicitly assigned. Without a work-for-hire clause, the vendor legally owns what you paid to build.",
                        "what_to_do": "Require explicit work-for-hire language: 'All work product created under this agreement is a work made for hire and is the sole property of [Company].' Add an assignment as backup.",
                    },
                    {
                        "category": "Warranty Disclaimer",
                        "tier": "MEDIUM",
                        "patterns": [
                            r"disclaim.*warrant", r"as.is", r"as is",
                            r"no warrant", r"without warrant",
                            r"implied warrant.*disclaim", r"merchantability",
                        ],
                        "why": "Blanket warranty disclaimers eliminate all implied warranties — including the implied warranty of merchantability and fitness for purpose. This shifts all performance risk to you.",
                        "what_to_do": "Require express warranties: uptime SLAs, conformance to specifications, compliance with applicable law, and non-infringement. Put financial consequences on SLA failures.",
                    },
                    {
                        "category": "Price Escalation Clause",
                        "tier": "MEDIUM",
                        "patterns": [
                            r"price.*increas", r"rate.*adjustm", r"CPI.*adjustm",
                            r"annual.*increas", r"escalat", r"inflationary adjustm",
                            r"adjust.*CPI", r"cost of living",
                        ],
                        "why": "Uncapped price escalation is how a 3-year contract becomes 40% more expensive by year three without any renegotiation right.",
                        "what_to_do": "Cap annual escalation at the lesser of CPI or X% (typically 3-5%). Require 90-day notice before any price change. Add a right to terminate if you decline the increase.",
                    },
                    {
                        "category": "Data Ownership & Privacy",
                        "tier": "HIGH",
                        "patterns": [
                            r"data.*vendor.*use", r"anonymiz.*data", r"aggregate.*data",
                            r"personal data", r"personally identifiable",
                            r"data.*resell", r"data.*third part",
                            r"GDPR", r"CCPA", r"data process",
                        ],
                        "why": "Many SaaS vendors include the right to use your anonymized data for benchmarking or AI model training. This means your spend patterns, supplier names, and pricing data can flow to competitors.",
                        "what_to_do": "Require: (1) you own all data you input; (2) vendor cannot use your data to train models or benchmark without explicit consent; (3) data deletion within 30 days of termination.",
                    },
                    {
                        "category": "Force Majeure Scope",
                        "tier": "MEDIUM",
                        "patterns": [
                            r"force majeure", r"act of god", r"beyond.*control",
                            r"pandemic", r"epidemic", r"government.*order",
                            r"supply.*shortage", r"labor.*dispute",
                        ],
                        "why": "Overly broad force majeure clauses let vendors pause performance for supply shortages, labor disputes, or 'any event beyond reasonable control' — without financial consequence or your right to terminate.",
                        "what_to_do": "Limit force majeure to true acts of God and government orders. Require: (1) vendor notification within 48 hours; (2) vendor must use commercially reasonable efforts to mitigate; (3) your right to terminate after 30 days.",
                    },
                    {
                        "category": "Assignment Without Consent",
                        "tier": "MEDIUM",
                        "patterns": [
                            r"assign.*without.*consent", r"may assign", r"successor.*assign",
                            r"change of control", r"merger.*acquisition",
                            r"assign.*right.*obligation",
                        ],
                        "why": "If the vendor can assign the contract without your consent, you could find yourself bound to a competitor or an unknown third party after an M&A transaction.",
                        "what_to_do": "Require your written consent for any assignment. Add: in the event of a change of control, you have the right to terminate without penalty within 90 days.",
                    },
                    {
                        "category": "Acceptance & Payment Terms",
                        "tier": "MEDIUM",
                        "patterns": [
                            r"net 30", r"net 60", r"net 90",
                            r"payment.*due.*upon", r"invoice.*due",
                            r"late.*fee", r"interest.*overdue",
                            r"deemed.*accept", r"accept.*unless.*object",
                        ],
                        "why": "Deemed acceptance clauses — where deliverables are accepted unless you object within X days — shift quality risk to you. Aggressive payment terms and late interest rates are hidden cost escalators.",
                        "what_to_do": "Negotiate: (1) explicit written acceptance required; (2) Net 45+ payment terms; (3) late interest capped at prime rate; (4) right to withhold payment on disputed invoices without penalty.",
                    },
                ]

                _txt_lower = _contract_text.lower()
                _findings  = []

                for _risk in _RISK_PATTERNS:
                    _matches = []
                    for _pat in _risk["patterns"]:
                        for _m in _re.finditer(_pat, _txt_lower):
                            start = max(0, _m.start() - 80)
                            end   = min(len(_contract_text), _m.end() + 120)
                            _excerpt = _contract_text[start:end].strip().replace("\n", " ")
                            _matches.append(_excerpt)
                    if _matches:
                        _findings.append({
                            "category": _risk["category"],
                            "tier":     _risk["tier"],
                            "count":    len(_matches),
                            "excerpts": _matches[:3],
                            "why":      _risk["why"],
                            "what_to_do": _risk["what_to_do"],
                        })

                # ── Jurisdiction warning ──────────────────────────
                _juris_note = ""
                _juris_input = (_cr_jurisdiction or "").lower()
                if any(s in _juris_input for s in ["delaware", "de"]):
                    _juris_note = "Delaware: Highly vendor-favorable. Courts enforce limitation of liability caps strictly. Negotiate liability carve-outs explicitly."
                elif any(s in _juris_input for s in ["california", "ca"]):
                    _juris_note = "California: Strong implied warranties and consumer protections. Non-compete clauses are unenforceable. CCPA data obligations apply."
                elif any(s in _juris_input for s in ["new york", "ny"]):
                    _juris_note = "New York: UCC governs goods contracts. Courts enforce contracts as written — ambiguous terms will not be read in your favor."
                elif any(s in _juris_input for s in ["texas", "tx"]):
                    _juris_note = "Texas: Business-friendly jurisdiction. Consequential damage waivers are broadly enforced. Force majeure clauses interpreted narrowly."
                elif any(s in _juris_input for s in ["georgia", "ga"]):
                    _juris_note = "Georgia: UCC-based. Non-compete enforceability reformed in 2011 — now enforceable with reasonable limits. Indemnification caps strictly enforced."
                elif any(s in _juris_input for s in ["uk", "england", "wales", "united kingdom"]):
                    _juris_note = "English Law: Strong implied terms under Sale of Goods Act. GDPR (UK) applies to all personal data. Consequential damage waivers generally enforceable."
                elif any(s in _juris_input for s in ["canada", "ontario", "british columbia", "alberta"]):
                    _juris_note = "Canadian jurisdiction: PIPEDA / provincial privacy law applies. Consumer contracts have implied fitness warranties. Arbitration clauses in B2B contracts are generally enforceable."
                elif _juris_input:
                    _juris_note = f"Jurisdiction '{_cr_jurisdiction}' detected. Verify local laws on: limitation of liability enforceability, implied warranty rules, arbitration requirements, and data privacy obligations."

                # ── Display results ───────────────────────────────
                _high   = [f for f in _findings if f["tier"] == "HIGH"]
                _medium = [f for f in _findings if f["tier"] == "MEDIUM"]
                _clean  = [r for r in _RISK_PATTERNS if not any(f["category"] == r["category"] for f in _findings)]

                st.markdown("---")
                _r1, _r2, _r3, _r4 = st.columns(4)
                with _r1:
                    st.markdown(
                        f'<div style="background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.25);border-radius:8px;padding:0.6rem 1rem">'
                        f'<div style="font-size:0.80rem;color:#F87171;text-transform:uppercase;letter-spacing:0.1em">High Risk</div>'
                        f'<div style="font-size:1.6rem;font-weight:700;color:#F87171;font-family:monospace">{len(_high)}</div></div>',
                        unsafe_allow_html=True,
                    )
                with _r2:
                    st.markdown(
                        f'<div style="background:rgba(252,211,77,0.06);border:1px solid rgba(252,211,77,0.2);border-radius:8px;padding:0.6rem 1rem">'
                        f'<div style="font-size:0.80rem;color:#FCD34D;text-transform:uppercase;letter-spacing:0.1em">Medium Risk</div>'
                        f'<div style="font-size:1.6rem;font-weight:700;color:#FCD34D;font-family:monospace">{len(_medium)}</div></div>',
                        unsafe_allow_html=True,
                    )
                with _r3:
                    st.markdown(
                        f'<div style="background:rgba(74,222,128,0.04);border:1px solid rgba(74,222,128,0.15);border-radius:8px;padding:0.6rem 1rem">'
                        f'<div style="font-size:0.80rem;color:#4ADE80;text-transform:uppercase;letter-spacing:0.1em">Clean</div>'
                        f'<div style="font-size:1.6rem;font-weight:700;color:#4ADE80;font-family:monospace">{len(_clean)}</div></div>',
                        unsafe_allow_html=True,
                    )
                with _r4:
                    _risk_score = min(100, len(_high) * 14 + len(_medium) * 6)
                    _rs_color   = "#EF4444" if _risk_score >= 50 else "#FCD34D" if _risk_score >= 25 else "#4ADE80"
                    st.markdown(
                        f'<div style="background:rgba(96,165,250,0.05);border:1px solid rgba(96,165,250,0.15);border-radius:8px;padding:0.6rem 1rem">'
                        f'<div style="font-size:0.80rem;color:#60A5FA;text-transform:uppercase;letter-spacing:0.1em">Risk Score</div>'
                        f'<div style="font-size:1.6rem;font-weight:700;color:{_rs_color};font-family:monospace">{_risk_score}/100</div></div>',
                        unsafe_allow_html=True,
                    )

                if _juris_note:
                    st.markdown(
                        f'<div style="background:rgba(167,139,250,0.06);border-left:3px solid #A78BFA;'
                        f'border-radius:0 8px 8px 0;padding:0.7rem 1rem;margin:1rem 0;font-size:0.90rem">'
                        f'<strong style="color:#A78BFA">Jurisdiction Note — {html.escape(_cr_jurisdiction)}</strong><br/>'
                        f'<span style="color:#E2E8F0">{html.escape(_juris_note)}</span></div>',
                        unsafe_allow_html=True,
                    )

                if not _findings:
                    st.success("No risk patterns detected. Either the contract is clean or the text didn't parse — verify the paste captured full clause text.")
                else:
                    for _f in _findings:
                        _tier_color  = "#EF4444" if _f["tier"] == "HIGH" else "#FCD34D"
                        _tier_bg     = "rgba(239,68,68,0.06)" if _f["tier"] == "HIGH" else "rgba(252,211,77,0.05)"
                        _tier_border = "rgba(239,68,68,0.3)" if _f["tier"] == "HIGH" else "rgba(252,211,77,0.25)"
                        with st.expander(
                            f"{_f['tier']} — {_f['category']} ({_f['count']} clause match{'es' if _f['count'] > 1 else ''})",
                            expanded=(_f["tier"] == "HIGH"),
                        ):
                            st.markdown(
                                f'<div style="background:{_tier_bg};border:1px solid {_tier_border};'
                                f'border-radius:8px;padding:0.8rem 1rem;margin-bottom:0.8rem">'
                                f'<div style="font-size:0.80rem;color:{_tier_color};text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.3rem">Why This Is Risky</div>'
                                f'<div style="font-size:0.90rem;color:#F1F5F9">{html.escape(_f["why"])}</div></div>',
                                unsafe_allow_html=True,
                            )
                            st.markdown(
                                f'<div style="background:rgba(74,222,128,0.04);border:1px solid rgba(74,222,128,0.15);'
                                f'border-radius:8px;padding:0.8rem 1rem;margin-bottom:0.8rem">'
                                f'<div style="font-size:0.80rem;color:#4ADE80;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.3rem">What To Negotiate</div>'
                                f'<div style="font-size:0.90rem;color:#F1F5F9">{html.escape(_f["what_to_do"])}</div></div>',
                                unsafe_allow_html=True,
                            )
                            if _f["excerpts"]:
                                st.markdown(
                                    f'<div style="font-size:0.80rem;color:#A8BEDC;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.3rem">Contract Excerpts Found</div>',
                                    unsafe_allow_html=True,
                                )
                                for _ex in _f["excerpts"]:
                                    st.markdown(
                                        f'<div style="background:rgba(13,21,38,0.8);border-left:3px solid {_tier_color};'
                                        f'border-radius:0 6px 6px 0;padding:0.5rem 0.8rem;margin-bottom:0.4rem;'
                                        f'font-family:monospace;font-size:0.82rem;color:#C4D3E8">'
                                        f'...{html.escape(_ex)}...</div>',
                                        unsafe_allow_html=True,
                                    )

                if _clean:
                    with st.expander(f"Clean Clauses — {len(_clean)} categories with no flags"):
                        for _c in _clean:
                            st.markdown(
                                f'<div style="display:flex;align-items:center;gap:0.5rem;padding:0.3rem 0;'
                                f'font-size:0.88rem;color:#4ADE80">✓ {_c["category"]}</div>',
                                unsafe_allow_html=True,
                            )

                # Export report
                _report_lines = [
                    f"CONTRACT RISK REPORT",
                    f"Contract: {_contract_name or 'Unnamed'}",
                    f"Counterparty: {_counterparty or 'N/A'}",
                    f"Jurisdiction: {_cr_jurisdiction or 'Not specified'}",
                    f"Type: {_contract_type_cr}",
                    f"Risk Score: {_risk_score}/100",
                    f"High Risk Issues: {len(_high)}  |  Medium Risk: {len(_medium)}  |  Clean: {len(_clean)}",
                    f"",
                ]
                if _juris_note:
                    _report_lines += [f"JURISDICTION NOTE: {_juris_note}", ""]
                for _f in _findings:
                    _report_lines += [
                        f"--- {_f['tier']}: {_f['category']} ({_f['count']} match) ---",
                        f"Why risky: {_f['why']}",
                        f"What to negotiate: {_f['what_to_do']}",
                        "",
                    ]
                st.download_button(
                    "⬇️ Export Risk Report (.txt)",
                    data="\n".join(_report_lines),
                    file_name=f"Contract_Risk_{(_contract_name or 'report').replace(' ','_')}.txt",
                    mime="text/plain",
                    key="cr_export",
                )

    # ── SUPPLIER PORTAL MANAGEMENT ─────────────────────────────────────
    with tab_comms:
      if _comms_section == "🌐 Supplier Portal":
        st.markdown("### Supplier Portal")
        st.markdown(
            '<p class="muted">Share a branded portal link with your suppliers. They submit their '
            'response without needing a ProcureIQ account. All submissions land here for your review.</p>',
            unsafe_allow_html=True,
        )

        # ── Link Generator ──
        st.markdown("#### Generate Portal Link")
        _portal_event_ref = st.text_input(
            "Event Reference (sent to suppliers so you know which RFP this is for)",
            value=event_name or generate_id("EVT-", 6),
            key="portal_event_ref",
        )
        _base_url = st.text_input(
            "Your ProcureIQ URL (the address where this app is running)",
            placeholder="https://your-app.streamlit.app",
            key="portal_base_url",
            help="Copy the URL from your browser address bar, without a trailing slash.",
        )
        if _base_url.strip():
            import hmac as _hmac_mod, hashlib as _hl_mod
            _portal_secret = os.getenv("SECRET_KEY", "")
            if not _portal_secret:
                st.warning(
                    "Supplier portal is disabled: SECRET_KEY environment variable is not set. "
                    "Add SECRET_KEY to your .env file to generate portal links."
                )
            else:
                _portal_event_id = _portal_event_ref.strip()
                _portal_token = _hmac_mod.new(
                    _portal_secret.encode(), _portal_event_id.encode(), _hl_mod.sha256
                ).hexdigest()[:24]
                _portal_link = (
                    f"{_base_url.strip()}?mode=portal"
                    f"&event={_portal_event_id}"
                    f"&token={_portal_token}"
                )
                st.markdown(
                    f'<div style="background:#060D1A;border:1px solid rgba(96,165,250,0.3);border-radius:8px;'
                    f'padding:0.75rem 1rem;font-family:monospace;font-size:0.88rem;color:#60A5FA;'
                    f'word-break:break-all;margin-bottom:0.5rem">{html.escape(_portal_link)}</div>',
                    unsafe_allow_html=True,
                )
                st.caption("Share this link with suppliers. The token in the URL prevents unauthorized access.")

        st.markdown("---")
        st.markdown("#### Submitted Responses")

        # ── View Submissions ──
        try:
            import datetime as _spdt
            _filter_ref = _portal_event_ref.strip() if _portal_event_ref.strip() else ""
            _sp_raw = get_database().get_portal_submissions(event_ref=_filter_ref, limit=30)
            _sp_rows = [(_r["key"], _r["value"]) for _r in _sp_raw]

            if _sp_rows:
                for _spr in _sp_rows:
                    _sprec = _spr[1] if isinstance(_spr[1], dict) else {}
                    _sp_company = html.escape(str(_sprec.get("company", "—")))
                    _sp_contact = html.escape(str(_sprec.get("contact", "—")))
                    _sp_email   = html.escape(str(_sprec.get("email", "—")))
                    _sp_price   = _sprec.get("price", 0)
                    _sp_lead    = _sprec.get("lead_days", "—")
                    _sp_div     = html.escape(str(_sprec.get("diversity", "—")))
                    _sp_ts      = _sprec.get("submitted_at", 0)
                    _sp_date    = _spdt.datetime.fromtimestamp(_sp_ts).strftime("%b %d, %Y %H:%M") if _sp_ts else "—"
                    _sp_ref     = html.escape(str(_sprec.get("event_ref", "—")))
                    _sp_certs   = html.escape(str(_sprec.get("certs", "")))
                    _sp_notes   = html.escape(str(_sprec.get("notes", "")))
                    _sp_loc     = html.escape(str(_sprec.get("location", "—")))
                    _sp_coi     = "✅ Confirmed" if _sprec.get("coi_clear") else "⚠️ Not confirmed"

                    with st.expander(f"{_sp_company}  ·  ${_sp_price:,.0f}  ·  {_sp_date}"):
                        _sc1, _sc2, _sc3 = st.columns(3)
                        with _sc1:
                            st.markdown(
                                f'<div style="font-size:0.82rem;color:#94A3B8">Contact</div>'
                                f'<div style="font-size:0.92rem;color:#E2E8F0;font-weight:600">{_sp_contact}</div>'
                                f'<div style="font-size:0.82rem;color:#60A5FA">{_sp_email}</div>'
                                f'<div style="font-size:0.82rem;color:#94A3B8;margin-top:0.4rem">Location: {_sp_loc}</div>',
                                unsafe_allow_html=True,
                            )
                        with _sc2:
                            st.markdown(
                                f'<div style="font-size:0.82rem;color:#94A3B8">Quote</div>'
                                f'<div style="font-size:1.1rem;color:#4ADE80;font-weight:700">${_sp_price:,.2f}</div>'
                                f'<div style="font-size:0.82rem;color:#94A3B8">Lead time: {_sp_lead} days</div>'
                                f'<div style="font-size:0.82rem;color:#94A3B8">Diversity: {_sp_div}</div>',
                                unsafe_allow_html=True,
                            )
                        with _sc3:
                            st.markdown(
                                f'<div style="font-size:0.82rem;color:#94A3B8">COI</div>'
                                f'<div style="font-size:0.88rem;color:#E2E8F0">{_sp_coi}</div>'
                                f'<div style="font-size:0.82rem;color:#94A3B8;margin-top:0.4rem">Certifications</div>'
                                f'<div style="font-size:0.85rem;color:#CBD5E1">{_sp_certs or "None listed"}</div>',
                                unsafe_allow_html=True,
                            )
                        if _sp_notes:
                            st.markdown(
                                f'<div style="background:#060D1A;border-left:3px solid rgba(96,165,250,0.3);'
                                f'border-radius:0 6px 6px 0;padding:0.5rem 0.8rem;margin-top:0.5rem;'
                                f'font-size:0.85rem;color:#C4D3E8"><strong>Notes:</strong> {_sp_notes}</div>',
                                unsafe_allow_html=True,
                            )
                        # Pre-fill supplier tab shortcut
                        if st.button(f"Load into Supplier #{min(num_suppliers, 10)} slot", key=f"portal_load_{_spr[0]}",
                                     help="Pre-populates the next available supplier slot in the Suppliers tab."):
                            # Find first empty supplier slot
                            _empty_slot = 0
                            for _si in range(num_suppliers):
                                if not st.session_state.get(f"supplier_name_{_si}", "").strip():
                                    _empty_slot = _si
                                    break
                            st.session_state[f"supplier_name_{_empty_slot}"]     = _sprec.get("company", "")
                            st.session_state[f"supplier_contact_{_empty_slot}"]  = _sprec.get("contact", "")
                            st.session_state[f"supplier_location_{_empty_slot}"] = _sprec.get("location", "")
                            st.session_state[f"supplier_price_{_empty_slot}"]    = float(_sprec.get("price", 0))
                            st.success(f"Pre-filled supplier slot #{_empty_slot + 1}. Switch to the Suppliers tab.")
            else:
                st.info("No portal submissions yet for this event reference. Share the portal link with suppliers to collect responses.")
        except Exception as _spe:
            st.error(f"Could not load submissions: {_spe}")

    # ── AGENT #2: Contract Generation ─────────────────────────────────
    # Hidden by default — set PROCUREIQ_EXPERIMENTAL=true to enable.
    # Rationale: LLM-generated contract language creates a credibility risk in professional demos.
    # The disclaimer is present but insufficient for any legal or compliance audience.
    with tab_comms:
      if _comms_section == "📝 Contract Generator":
        if os.getenv("PROCUREIQ_EXPERIMENTAL", "").lower() != "true":
            st.info(
                "**Contract Drafting Assistant is not enabled in this build.**  \n"
                "This feature is experimental. LLM-generated contract language is an AI starting point only "
                "and must not be executed without qualified legal review.  \n"
                "To enable: set `PROCUREIQ_EXPERIMENTAL=true` in your environment.",
                icon="🔬",
            )
        else:
          st.markdown("### Contract Drafting Assistant")
          st.markdown(
              '<p class="muted">AI generates a structured contract draft — MSA, SOW, Supply Agreement, '
              'or SaaS license — calibrated to your Kraljic posture, RAQSCI requirements, and jurisdiction.</p>',
              unsafe_allow_html=True,
          )
          st.warning(
              "**Legal disclaimer:** Output from this tool is an AI-generated starting point for internal "
              "discussion only. It is **not** a legally reviewed contract and **must not** be executed, "
              "signed, or sent to suppliers without review and approval by qualified legal counsel. "
              "Nothing here constitutes legal advice."
          )

          _cg_c1, _cg_c2 = st.columns(2)
          with _cg_c1:
              _cg_type = st.selectbox(
                  "Contract Type",
                  ["Master Service Agreement", "Supply Agreement", "Statement of Work",
                   "SaaS / Software License", "Professional Services Agreement",
                   "Distribution Agreement", "Non-Disclosure Agreement"],
                  key="cg_contract_type",
              )
              _cg_buyer    = st.text_input("Buyer Organization", value="[Your Organization]", key="cg_buyer")
              _cg_supplier = st.text_input(
                  "Supplier Name",
                  value=leader["Supplier"] if ranked else "",
                  key="cg_supplier",
              )
              _cg_value    = st.number_input(
                  "Annual Contract Value ($)",
                  min_value=0.0,
                  value=float(st.session_state.get("ctrl_annual_value", 500000) or 500000),
                  step=50000.0,
                  format="%.0f",
                  key="cg_value",
              )
              _cg_term     = st.number_input("Contract Term (years)", min_value=1, max_value=10, value=3, key="cg_term")

          with _cg_c2:
              _cg_jurisdiction = st.text_input(
                  "Governing Law / Jurisdiction",
                  value="State of Delaware, USA",
                  key="cg_jurisdiction",
              )
              _cg_kraljic = st.selectbox(
                  "Kraljic Posture",
                  ["Strategic", "Leverage", "Bottleneck", "Non-Critical"],
                  index=["Strategic","Leverage","Bottleneck","Non-Critical"].index(
                      st.session_state.get("ctrl_kraljic", "Strategic") or "Strategic"
                  ),
                  key="cg_kraljic",
              )
              _cg_payment  = st.selectbox("Payment Terms", ["Net 30", "Net 45", "Net 60", "Net 15", "2/10 Net 30"], key="cg_payment")
              _cg_notice   = st.number_input("Termination Notice (days)", min_value=30, max_value=365, value=90, key="cg_notice")
              _cg_sla      = st.text_area("SLA Targets", placeholder="e.g. 99.9% uptime, 4hr response time, <2% defect rate", key="cg_sla", height=80)

          st.markdown("##### RAQSCI Requirements to Include")
          _raqsci_cols = st.columns(6)
          _raqsci_labels = ["Regulatory", "Assurance", "Quality", "Service", "Cost", "Innovation"]
          _raqsci_keys   = ["R", "A", "Q", "S", "C", "I"]
          _raqsci_flags  = {}
          for _ri, (_rl, _rk) in enumerate(zip(_raqsci_labels, _raqsci_keys)):
              with _raqsci_cols[_ri]:
                  _raqsci_flags[_rl] = st.checkbox(_rl, value=True, key=f"cg_raqsci_{_rk}")

          if st.button("⚡ Generate Contract", key="cg_generate_btn", type="primary"):
              with st.spinner("AI drafting contract … (30-60 seconds for a full document)"):
                  _cg_result = run_contract_generation_agent(
                      contract_type=_cg_type,
                      buyer_name=_cg_buyer,
                      supplier_name=_cg_supplier,
                      category=safe_category or category,
                      subcategory=selected_sub_name,
                      annual_value=_cg_value,
                      contract_term_years=int(_cg_term),
                      jurisdiction=_cg_jurisdiction,
                      kraljic=_cg_kraljic,
                      raqsci_flags=_raqsci_flags,
                      sla_targets=_cg_sla,
                      payment_terms=_cg_payment,
                      notice_period_days=int(_cg_notice),
                      api_key=_get_api_key(),
                      provider=_get_provider(),
                  )
              st.session_state["_cg_result"] = _cg_result

          _cgr = st.session_state.get("_cg_result")
          if _cgr:
              if _cgr.get("error"):
                  st.error(f"Generation failed: {_cgr['error']}")
              elif _cgr.get("html"):
                  st.success(
                      f"Contract generated — {_cgr.get('word_count', 0):,} words · "
                      f"{_cgr.get('contract_type')} · {_cgr.get('kraljic')} posture"
                  )
                  # PDF download
                  try:
                      import weasyprint as _wp_cg
                      _cg_pdf = _wp_cg.HTML(string=_cgr["html"]).write_pdf()
                      _cg_fname = f"{(_cg_supplier or 'Contract').replace(' ','_')}_{_cg_type.replace(' ','_')}.pdf"
                      st.download_button(
                          "⬇️ Download Contract PDF",
                          data=_cg_pdf,
                          file_name=_cg_fname,
                          mime="application/pdf",
                          key="cg_pdf_dl",
                      )
                  except ImportError:
                      st.info("Install weasyprint for PDF export. Showing HTML preview:")

                  # HTML preview (truncated)
                  with st.expander("Preview Contract (first 3,000 characters)", expanded=False):
                      st.markdown(
                          f'<div style="background:#060D1A;border:1px solid rgba(148,163,184,0.15);'
                          f'border-radius:8px;padding:1rem;font-family:Georgia,serif;'
                          f'font-size:0.85rem;color:#E2E8F0;line-height:1.7;max-height:500px;overflow-y:auto">'
                          f'{_cgr["html"][:3000]}…</div>',
                          unsafe_allow_html=True,
                      )

    # ── AGENT #34: Tenant Provisioning — Admin Panel ───────────────────
    with tab_comms:
      if _comms_section == "🌐 Supplier Portal":
        st.markdown("---")
        if os.getenv("PROCUREIQ_EXPERIMENTAL", "").lower() != "true":
            st.info(
                "**Multi-Tenant Organization Provisioning** is not enabled in this deployment. "
                "Set `PROCUREIQ_EXPERIMENTAL=true` to activate.",
                icon="🔬",
            )
        else:
            _admin_expander = st.expander("⚙️ Admin: Provision New Organization", expanded=False)
            with _admin_expander:
                st.markdown(
                    '<p class="muted">Create a new organization workspace with isolated data, '
                    'custom branding, and an admin user. Each org gets its own org_id scope.</p>',
                    unsafe_allow_html=True,
                )
                _prov_c1, _prov_c2 = st.columns(2)
                with _prov_c1:
                    _prov_name  = st.text_input("Organization Name", placeholder="Acme Corp", key="prov_name")
                    _prov_email = st.text_input("Admin Email", placeholder="admin@acme.com", key="prov_email")
                    _prov_pw    = st.text_input("Admin Password", type="password", key="prov_pw")
                with _prov_c2:
                    _prov_plan  = st.selectbox("Plan", ["starter", "professional", "enterprise"], key="prov_plan")
                    _prov_color = st.color_picker("Brand Color", value="#3B82F6", key="prov_color")
                    _prov_logo  = st.text_input("Logo URL (optional)", key="prov_logo")

                if st.button("🚀 Provision Organization", key="prov_org_btn", type="primary"):
                    if _prov_name and _prov_email and _prov_pw:
                        with st.spinner("Provisioning …"):
                            _prov_result = provision_organization(
                                org_name=_prov_name,
                                admin_email=_prov_email,
                                admin_password=_prov_pw,
                                plan=_prov_plan,
                                logo_url=_prov_logo,
                                primary_color=_prov_color,
                            )
                        if _prov_result.get("success"):
                            st.success(
                                f"✅ Organization **{_prov_name}** provisioned! "
                                f"org_id: `{_prov_result['org_id']}` · "
                                f"Admin: {_prov_email}"
                            )
                            for _step in _prov_result.get("setup_steps", []):
                                _step_icon = "✓" if _step["status"] == "OK" else "⚠"
                                st.caption(f"{_step_icon} {_step['step']}: {_step['status']}")
                        else:
                            st.error(f"Provisioning failed: {_prov_result.get('error')}")
                    else:
                        st.warning("Organization name, admin email, and password are required.")

                # List existing orgs
                _orgs = list_organizations()
                if _orgs:
                    st.markdown("##### Provisioned Organizations")
                    for _org in _orgs:
                        st.markdown(
                            f'<div style="background:#060D1A;border:1px solid rgba(148,163,184,0.1);'
                            f'border-radius:8px;padding:0.5rem 0.8rem;margin-bottom:0.3rem;'
                            f'font-size:0.85rem;color:#E2E8F0">'
                            f'<strong>{html.escape(_org["name"])}</strong> '
                            f'<span style="color:#60A5FA">{_org["org_id"]}</span> · '
                            f'<span style="color:#94A3B8">{_org["plan"]} · {_org["user_count"]} users</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

    # ── AGENT #13: Spend Anomaly — wire into Spend Intel ──────────────
    with tab_spend:
        _normalized_recs = st.session_state.get("_erp_normalized_records")
        if _normalized_recs:
            st.markdown("---")
            st.markdown("#### 🔍 Spend Anomaly Detection Agent")
            st.markdown(
                '<p class="muted">Scans your spend data for duplicate invoices, spend spikes, '
                'split transactions near approval thresholds, and single-supplier concentration risk.</p>',
                unsafe_allow_html=True,
            )
            _anom_threshold = st.number_input(
                "Approval threshold ($) for split-transaction detection",
                min_value=1000.0, max_value=500000.0, value=10000.0, step=1000.0,
                key="anom_threshold",
            )
            if st.button("🔍 Detect Anomalies", key="run_anomaly_btn", type="primary"):
                with st.spinner("Scanning for anomalies …"):
                    _anom_result = run_spend_anomaly_agent(
                        records=_normalized_recs,
                        amount_col="amount_num",
                        supplier_col="supplier",
                        category_col="Auto Category",
                        approval_threshold=_anom_threshold,
                        api_key=_get_api_key(),
                        provider=_get_provider(),
                    )
                st.session_state["_anom_result"] = _anom_result

            _ar = st.session_state.get("_anom_result")
            if _ar and not _ar.get("error"):
                _stats = _ar.get("summary_stats", {})
                _syn   = _ar.get("synthesis", {})
                _a1, _a2, _a3, _a4 = st.columns(4)
                for _col, _lbl, _val, _col_hex in [
                    (_a1, "Anomalies Found",    _stats.get("total_anomalies", 0),     "#F87171"),
                    (_a2, "High Severity",      _stats.get("high_severity", 0),        "#F87171"),
                    (_a3, "Amount Flagged",     f"${_stats.get('total_flagged_amount',0):,.0f}", "#FCD34D"),
                    (_a4, "Risk Rating",        _syn.get("risk_rating", "—"),          "#60A5FA"),
                ]:
                    _col.markdown(
                        f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);'
                        f'border-radius:10px;padding:0.8rem;text-align:center">'
                        f'<div style="font-size:0.75rem;color:#94A3B8;text-transform:uppercase;letter-spacing:0.08em">{_lbl}</div>'
                        f'<div style="font-size:1.3rem;font-weight:700;color:{_col_hex}">{_val}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                if _syn.get("executive_summary"):
                    st.markdown(
                        f'<div style="background:rgba(248,113,113,0.06);border:1px solid rgba(248,113,113,0.2);'
                        f'border-radius:10px;padding:0.9rem 1.1rem;margin:0.8rem 0;color:#E2E8F0;font-size:0.9rem">'
                        f'{html.escape(_syn["executive_summary"])}</div>',
                        unsafe_allow_html=True,
                    )

                for _anom in _ar.get("anomalies", [])[:20]:
                    _sev = _anom.get("severity", "MEDIUM")
                    _sev_color = {"HIGH": "#F87171", "MEDIUM": "#FCD34D"}.get(_sev, "#94A3B8")
                    _anom_type = _anom.get("type", "").replace("_", " ")
                    st.markdown(
                        f'<div style="background:#060D1A;border-left:3px solid {_sev_color};'
                        f'border-radius:0 8px 8px 0;padding:0.5rem 0.8rem;margin-bottom:0.3rem">'
                        f'<span style="font-size:0.78rem;color:{_sev_color};font-weight:700;text-transform:uppercase">{_sev} · {_anom_type}</span> '
                        f'<span style="font-size:0.82rem;color:#C4D3E8">— {html.escape(str(_anom.get("supplier", "")))}</span><br/>'
                        f'<span style="font-size:0.84rem;color:#E2E8F0">{html.escape(str(_anom.get("reason", "")))}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

    # ── AGENT: Spend Forecasting ────────────────────────────────────────
    with tab_spend:
        st.markdown("---")
        st.markdown("#### 📈 Spend Forecasting Agent")
        st.markdown(
            '<p class="muted">Paste monthly spend history to project the next 12 months with '
            'inflation-adjusted regression, 90% confidence intervals, and — with an AI key — '
            'CFO-ready savings recommendations.</p>',
            unsafe_allow_html=True,
        )
        _sf_c1, _sf_c2 = st.columns([2, 1])
        with _sf_c1:
            _sf_raw = st.text_area(
                "Monthly spend history (comma or newline separated, oldest first)",
                placeholder="e.g.  85000, 91000, 88000, 95000, 102000, 98000",
                height=100,
                key="sf_spend_input",
            )
        with _sf_c2:
            _sf_cat  = st.selectbox("Category", list(CATEGORY_TAXONOMY.keys()), key="sf_cat",
                                    index=list(CATEGORY_TAXONOMY.keys()).index(_wt_parent)
                                    if "_wt_parent" in dir() and "_wt_parent" in CATEGORY_TAXONOMY else 0)
            _sf_sub  = st.text_input("Subcategory (optional)", value=selected_sub_name if "selected_sub_name" in dir() else "",
                                     key="sf_sub", placeholder="e.g. IT Managed Services")
            _sf_seas = st.checkbox("Apply seasonal index", value=True, key="sf_seasonality")

        if st.button("📈 Run Forecast", key="sf_run_btn", type="primary"):
            _sf_vals = _sf_parse_csv(_sf_raw)
            if len(_sf_vals) < 3:
                st.warning("Please enter at least 3 months of spend data (comma or newline separated).")
            else:
                with st.spinner("Projecting spend …"):
                    _sf_result = run_spend_forecast(
                        monthly_spend=_sf_vals,
                        category=_sf_cat,
                        subcategory=_sf_sub.strip(),
                        apply_seasonality=_sf_seas,
                        api_key=_get_api_key(),
                        provider=_get_provider(),
                    )
                st.session_state["_sf_result"] = _sf_result

        _sfr = st.session_state.get("_sf_result")
        if _sfr and not _sfr.get("error") and _sfr.get("forecast"):
            _fc = _sfr["forecast"]
            _sm = _fc.get("summary", {})
            _fc_months = _fc.get("forecast_months", [])

            # KPI strip
            _sf_k1, _sf_k2, _sf_k3, _sf_k4 = st.columns(4)
            _sf_delta_color = "#F87171" if (_sm.get("forecast_vs_run_rate_pct", 0) or 0) > 5 else "#4ADE80"
            for _sfcol, _lbl, _val, _chex in [
                (_sf_k1, "Run-Rate Annual",   f"${_sm.get('run_rate_annual', 0):,.0f}",                     "#60A5FA"),
                (_sf_k2, "12-Mo Forecast",    f"${_sm.get('total_12mo_forecast', 0):,.0f}",                 "#FCD34D"),
                (_sf_k3, "vs Run-Rate",       f"{_sm.get('forecast_vs_run_rate_pct', 0):+.1f}%",            _sf_delta_color),
                (_sf_k4, "Inflation Factor",  f"{_sfr.get('inflation_rate_pct', 0):.1f}% p.a.",             "#A78BFA"),
            ]:
                _sfcol.markdown(
                    f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);'
                    f'border-radius:10px;padding:0.8rem;text-align:center">'
                    f'<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase;letter-spacing:0.08em">{_lbl}</div>'
                    f'<div style="font-size:1.25rem;font-weight:700;color:{_chex}">{_val}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # Forecast table
            if _fc_months:
                import datetime as _dt_mod
                _today = _dt_mod.date.today()
                _month_labels = []
                for _mi in range(len(_fc_months)):
                    _d = _today.replace(day=1)
                    _mo = (_d.month + _mi) % 12 + 1
                    _yr = _d.year + (_d.month + _mi) // 12
                    _month_labels.append(f"{_dt_mod.date(1900, _mo, 1).strftime('%b')} {_yr}")

                st.markdown("##### Monthly Projection")
                _fc_rows = ""
                for _mi, (_mlbl, _mdata) in enumerate(zip(_month_labels, _fc_months)):
                    _bar_pct = int(min(100, _mdata["projected"] / max(1, _sm.get("run_rate_annual", 1) / 12) * 100))
                    _fc_rows += (
                        f'<tr>'
                        f'<td style="padding:0.35rem 0.6rem;color:#94A3B8;font-size:0.82rem">{_mlbl}</td>'
                        f'<td style="padding:0.35rem 0.6rem;color:#E2E8F0;font-weight:600;font-size:0.88rem">'
                        f'${_mdata["projected"]:,.0f}</td>'
                        f'<td style="padding:0.35rem 0.6rem;color:#64748B;font-size:0.78rem">'
                        f'${_mdata["ci_low"]:,.0f} – ${_mdata["ci_high"]:,.0f}</td>'
                        f'<td style="padding:0.35rem 0.6rem;width:120px">'
                        f'<div style="background:rgba(96,165,250,0.15);border-radius:4px;height:8px">'
                        f'<div style="background:#60A5FA;width:{_bar_pct}%;height:8px;border-radius:4px"></div>'
                        f'</div></td>'
                        f'</tr>'
                    )
                st.markdown(
                    f'<table style="width:100%;border-collapse:collapse;font-family:inherit">'
                    f'<thead><tr>'
                    f'<th style="padding:0.4rem 0.6rem;color:#64748B;font-size:0.75rem;text-align:left;font-weight:500">Month</th>'
                    f'<th style="padding:0.4rem 0.6rem;color:#64748B;font-size:0.75rem;text-align:left;font-weight:500">Projected</th>'
                    f'<th style="padding:0.4rem 0.6rem;color:#64748B;font-size:0.75rem;text-align:left;font-weight:500">90% CI</th>'
                    f'<th style="padding:0.4rem 0.6rem;color:#64748B;font-size:0.75rem;text-align:left;font-weight:500">vs. avg</th>'
                    f'</tr></thead><tbody>{_fc_rows}</tbody></table>',
                    unsafe_allow_html=True,
                )

            # CFO narrative (LLM)
            _sfn = _sfr.get("narrative", {})
            if _sfn and not _sfn.get("error"):
                st.markdown("##### CFO Briefing")
                if _sfn.get("headline"):
                    st.markdown(
                        f'<div style="background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.18);'
                        f'border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.8rem;color:#C4D3E8;font-size:0.9rem;font-style:italic">'
                        f'{html.escape(_sfn["headline"])}</div>',
                        unsafe_allow_html=True,
                    )
                _levers = _sfn.get("savings_levers", [])
                if _levers:
                    st.markdown("**Savings Levers**")
                    for _lv in _levers:
                        _eff_col = {"Low": "#4ADE80", "Medium": "#FCD34D", "High": "#F87171"}.get(_lv.get("effort", ""), "#94A3B8")
                        st.markdown(
                            f'<div style="background:#060D1A;border-left:3px solid #60A5FA;'
                            f'border-radius:0 8px 8px 0;padding:0.5rem 0.8rem;margin-bottom:0.3rem">'
                            f'<strong style="color:#E2E8F0">{html.escape(_lv.get("lever",""))}</strong> '
                            f'<span style="color:#4ADE80;font-size:0.85rem">−{_lv.get("estimated_savings_pct",0):.0f}%</span> · '
                            f'<span style="color:{_eff_col};font-size:0.78rem">{_lv.get("effort","")} effort</span> · '
                            f'<span style="color:#94A3B8;font-size:0.78rem">{_lv.get("timeline","")}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                if _sfn.get("narrative"):
                    st.markdown(
                        f'<div style="background:rgba(13,21,38,0.7);border:1px solid rgba(148,163,184,0.12);'
                        f'border-radius:8px;padding:0.8rem 1rem;margin-top:0.6rem;color:#C4D3E8;font-size:0.88rem;line-height:1.6">'
                        f'{html.escape(_sfn["narrative"])}</div>',
                        unsafe_allow_html=True,
                    )

    # ── AGENT: Contract Obligation Tracker ─────────────────────────────
    with tab_comms:
        st.markdown("---")
        st.markdown("#### 📋 Contract Obligation Tracker")
        st.markdown(
            '<p class="muted">Paste any contract text (MSA, SOW, SLA, NDA) to extract renewal dates, '
            'notice periods, price escalation triggers, performance milestones, and hidden risk traps. '
            'Works with or without an AI key — regex fallback catches the most common clauses.</p>',
            unsafe_allow_html=True,
        )
        _ct_label = st.text_input("Contract label (for reference)", value="Vendor MSA", key="ct_label",
                                   placeholder="e.g. Microsoft EA 2025, AWS MSA")
        _ct_text  = st.text_area(
            "Paste contract text here",
            height=200,
            key="ct_contract_text",
            placeholder="Paste the full contract text or key sections. The AI will extract all obligations, dates, and risk flags.",
        )
        if st.button("📋 Extract Obligations", key="ct_run_btn", type="primary"):
            if not _ct_text.strip():
                st.warning("Please paste contract text to analyze.")
            else:
                with st.spinner("Extracting obligations …"):
                    _ct_result = run_contract_tracker(
                        contract_text=_ct_text,
                        contract_label=_ct_label.strip() or "Contract",
                        api_key=_get_api_key(),
                        provider=_get_provider(),
                    )
                st.session_state["_ct_result"] = _ct_result

        _ctr = st.session_state.get("_ct_result")
        if _ctr and not _ctr.get("error"):
            _ct_counts = _ctr.get("obligation_counts", {})
            _ct_mode   = _ctr.get("extraction_mode", "regex_fallback")
            _ct_summ   = _ctr.get("contract_summary", {})

            # Mode badge + summary
            _mode_color = "#4ADE80" if _ct_mode == "llm" else "#FCD34D"
            _mode_label = "AI Extraction" if _ct_mode == "llm" else "Regex Fallback"
            st.markdown(
                f'<div style="display:flex;gap:0.5rem;margin-bottom:0.8rem;flex-wrap:wrap;align-items:center">'
                f'<span style="background:rgba(96,165,250,0.1);border:1px solid rgba(96,165,250,0.25);'
                f'border-radius:6px;padding:0.2rem 0.6rem;font-size:0.75rem;color:#60A5FA">'
                f'{html.escape(_ctr.get("label",""))}</span>'
                f'<span style="background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.1);'
                f'border-radius:6px;padding:0.2rem 0.6rem;font-size:0.75rem;color:{_mode_color}">'
                f'{_mode_label}</span>'
                + (f'<span style="font-size:0.78rem;color:#94A3B8">Supplier: {html.escape(_ct_summ.get("supplier_name") or "—")}</span>' if _ct_summ.get("supplier_name") else "")
                + (f'<span style="font-size:0.78rem;color:#94A3B8">Expires: {html.escape(_ct_summ.get("expiry_date") or "—")}</span>' if _ct_summ.get("expiry_date") else "")
                + f'</div>',
                unsafe_allow_html=True,
            )

            # Count pills
            _ct_p1, _ct_p2, _ct_p3, _ct_p4 = st.columns(4)
            for _ctcol, _sev, _chex in [
                (_ct_p1, "CRITICAL", "#F87171"),
                (_ct_p2, "HIGH",     "#FB923C"),
                (_ct_p3, "MEDIUM",   "#FCD34D"),
                (_ct_p4, "LOW",      "#4ADE80"),
            ]:
                _cnt = _ct_counts.get(_sev, 0)
                _ctcol.markdown(
                    f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);'
                    f'border-radius:10px;padding:0.7rem;text-align:center">'
                    f'<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase">{_sev}</div>'
                    f'<div style="font-size:1.3rem;font-weight:700;color:{_chex}">{_cnt}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # Obligation list
            _obs = _ctr.get("obligations", [])
            if _obs:
                st.markdown("##### Obligation Calendar")
                for _ob in _obs[:25]:
                    _osev  = _ob.get("severity", "MEDIUM")
                    _ocolor = {"CRITICAL": "#F87171", "HIGH": "#FB923C", "MEDIUM": "#FCD34D", "LOW": "#4ADE80"}.get(_osev, "#94A3B8")
                    _days  = _ob.get("days_until")
                    _days_str = f"Due in {_days}d" if _days is not None and _days >= 0 else ("PAST DUE" if _days is not None else "No date")
                    _urg_col = _ob.get("urgency_color", "#60A5FA")
                    st.markdown(
                        f'<div style="background:#060D1A;border-left:3px solid {_ocolor};'
                        f'border-radius:0 8px 8px 0;padding:0.5rem 0.9rem;margin-bottom:0.3rem;'
                        f'display:flex;justify-content:space-between;align-items:flex-start;gap:1rem">'
                        f'<div style="flex:1">'
                        f'<div style="font-size:0.78rem;color:{_ocolor};font-weight:700;text-transform:uppercase;margin-bottom:0.15rem">'
                        f'{_osev} · {html.escape(_ob.get("type",""))}</div>'
                        f'<div style="color:#E2E8F0;font-size:0.86rem">{html.escape(_ob.get("label","")[:100])}</div>'
                        f'<div style="color:#94A3B8;font-size:0.78rem;margin-top:0.2rem">'
                        f'{html.escape((_ob.get("consequence") or "")[:100])}</div>'
                        f'</div>'
                        f'<div style="text-align:right;white-space:nowrap">'
                        f'<div style="color:{_urg_col};font-size:0.82rem;font-weight:600">{_days_str}</div>'
                        f'<div style="color:#64748B;font-size:0.72rem">{html.escape(_ob.get("due_date") or "")}</div>'
                        f'</div></div>',
                        unsafe_allow_html=True,
                    )

            # Risk flags
            _rfs = _ctr.get("risk_flags", [])
            if _rfs:
                st.markdown("##### Risk Flags")
                for _rf in _rfs:
                    _rf_sev = _rf.get("severity", "HIGH")
                    _rf_col = {"CRITICAL": "#F87171", "HIGH": "#FB923C", "MEDIUM": "#FCD34D"}.get(_rf_sev, "#94A3B8")
                    st.markdown(
                        f'<div style="background:rgba(248,113,113,0.06);border:1px solid rgba(248,113,113,0.18);'
                        f'border-radius:8px;padding:0.5rem 0.8rem;margin-bottom:0.3rem">'
                        f'<strong style="color:{_rf_col}">⚠ {html.escape(_rf.get("flag",""))}</strong>'
                        f'<span style="color:#C4D3E8;font-size:0.84rem"> — {html.escape(_rf.get("detail","")[:180])}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            # Recommended actions
            _recs = _ctr.get("recommended_actions", [])
            if _recs:
                st.markdown("##### Recommended Actions")
                for _rec in _recs:
                    _urgency = _rec.get("urgency", "90 days")
                    _urg_col = {"Immediate": "#F87171", "30 days": "#FB923C", "60 days": "#FCD34D"}.get(_urgency, "#4ADE80")
                    st.markdown(
                        f'<div style="background:#060D1A;border:1px solid rgba(148,163,184,0.1);'
                        f'border-radius:8px;padding:0.5rem 0.8rem;margin-bottom:0.3rem;'
                        f'display:flex;justify-content:space-between;align-items:center;gap:1rem">'
                        f'<div style="color:#E2E8F0;font-size:0.86rem">{html.escape(_rec.get("action","")[:160])}</div>'
                        f'<div style="background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.1);'
                        f'border-radius:6px;padding:0.2rem 0.6rem;font-size:0.75rem;color:{_urg_col};white-space:nowrap">'
                        f'{html.escape(_urgency)}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

    # ── AGENT: Supplier Risk Monitor ───────────────────────────────────
    with tab_market:
        st.markdown("---")
        st.markdown("#### 🛡 Supplier Risk Monitor")
        st.markdown(
            '<p class="muted">Scans SEC EDGAR 8-K filings, XBRL financial ratios, and live news '
            'for each supplier to compute a composite 0–100 risk score. No API key required for '
            'the data layer — AI key unlocks portfolio-level narrative synthesis.</p>',
            unsafe_allow_html=True,
        )

        # Build supplier list from session state slots
        _srm_suppliers = []
        for _si in range(10):
            _sn = st.session_state.get(f"name_{_si}", "").strip()
            _tk = st.session_state.get(f"ticker_{_si}", "").strip().upper()
            if _sn and _sn not in (f"Supplier {_si+1}", ""):
                _srm_suppliers.append({"name": _sn, "ticker": _tk})

        if _srm_suppliers:
            st.markdown(
                f'<div style="background:rgba(96,165,250,0.06);border:1px solid rgba(96,165,250,0.15);'
                f'border-radius:8px;padding:0.6rem 1rem;margin-bottom:0.8rem;font-size:0.85rem;color:#A8BEDC">'
                f'<strong style="color:#60A5FA">{len(_srm_suppliers)} suppliers</strong> loaded from Suppliers tab. '
                f'Suppliers without a ticker use company name for SEC lookup.</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info("Add suppliers in the Suppliers tab first, then return here to scan.")

        _srm_col1, _srm_col2 = st.columns([2, 1])
        with _srm_col1:
            _srm_extra = st.text_area(
                "Add suppliers not in your list (one per line: Name, TICKER)",
                placeholder="Amazon, AMZN\nMicrosoft, MSFT\nAcme Logistics, (leave blank if private)",
                height=90,
                key="srm_extra_suppliers",
            )
        with _srm_col2:
            _srm_days = st.slider("8-K lookback (days)", 30, 365, 90, step=30, key="srm_lookback")
            _srm_btn  = st.button("Run Supplier Risk Check (SEC + Financial Health)", key="srm_run_btn", type="primary", use_container_width=True)

        # Parse extra suppliers
        _all_srm_suppliers = list(_srm_suppliers)
        if _srm_extra.strip():
            for _line in _srm_extra.strip().split("\n"):
                _parts = [p.strip() for p in _line.split(",")]
                if _parts and _parts[0]:
                    _all_srm_suppliers.append({
                        "name": _parts[0],
                        "ticker": _parts[1] if len(_parts) > 1 else "",
                    })

        if _srm_btn:
            if not _all_srm_suppliers:
                st.warning("No suppliers to scan. Add suppliers in the Suppliers tab or use the input above.")
            else:
                with st.spinner(f"Scanning {len(_all_srm_suppliers)} suppliers via SEC EDGAR & news …"):
                    _srm_result = run_supplier_risk_monitor(
                        suppliers=_all_srm_suppliers,
                        api_key=_get_api_key(),
                        provider=_get_provider(),
                        days_lookback=_srm_days,
                    )
                st.session_state["_srm_result"] = _srm_result

        _srmr = st.session_state.get("_srm_result")
        if _srmr and not _srmr.get("error"):
            _srm_scores = _srmr.get("supplier_scores", [])
            _srm_high   = _srmr.get("high_risk_count", 0)
            _srm_med    = _srmr.get("medium_risk_count", 0)
            _srm_total  = _srmr.get("total_suppliers", 0)

            # KPI strip
            _sk1, _sk2, _sk3 = st.columns(3)
            for _scol, _lbl, _val, _hex in [
                (_sk1, "Suppliers Scanned", _srm_total,  "#60A5FA"),
                (_sk2, "High Risk",          _srm_high,  "#F87171"),
                (_sk3, "Medium Risk",         _srm_med,  "#FCD34D"),
            ]:
                _scol.markdown(
                    f'<div style="background:#0D1526;border:1px solid rgba(148,163,184,0.15);'
                    f'border-radius:10px;padding:0.8rem;text-align:center">'
                    f'<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase;letter-spacing:0.08em">{_lbl}</div>'
                    f'<div style="font-size:1.3rem;font-weight:700;color:{_hex}">{_val}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # Portfolio narrative (LLM)
            _srm_syn = _srmr.get("portfolio_summary", {})
            if _srm_syn and not _srm_syn.get("error"):
                if _srm_syn.get("headline"):
                    st.markdown(
                        f'<div style="background:rgba(248,113,113,0.06);border:1px solid rgba(248,113,113,0.2);'
                        f'border-radius:10px;padding:0.8rem 1.1rem;margin:0.8rem 0;color:#E2E8F0;font-size:0.9rem">'
                        f'<strong style="color:#F87171">Portfolio Risk: {html.escape(_srm_syn.get("portfolio_risk_rating","—"))}</strong> — '
                        f'{html.escape(_srm_syn.get("headline",""))}</div>',
                        unsafe_allow_html=True,
                    )
                _imm = _srm_syn.get("immediate_actions", [])
                if _imm:
                    st.markdown("**Immediate Actions**")
                    for _ia in _imm:
                        _ia_urg = _ia.get("urgency", "90 days")
                        _ia_col = {"Immediate": "#F87171", "30 days": "#FB923C", "90 days": "#4ADE80"}.get(_ia_urg, "#94A3B8")
                        st.markdown(
                            f'<div style="background:#060D1A;border-left:3px solid {_ia_col};'
                            f'border-radius:0 8px 8px 0;padding:0.4rem 0.8rem;margin-bottom:0.25rem;font-size:0.85rem">'
                            f'<strong style="color:#E2E8F0">{html.escape(_ia.get("supplier",""))}</strong> '
                            f'<span style="color:#C4D3E8">— {html.escape(_ia.get("action","")[:120])}</span> '
                            f'<span style="color:{_ia_col};font-size:0.75rem">({_ia_urg})</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

            # Supplier score cards
            st.markdown("##### Supplier Risk Scores")
            for _ss in _srm_scores:
                _ss_score = _ss.get("risk_score", 0)
                _ss_color = _ss.get("risk_color", "#94A3B8")
                _ss_tier  = _ss.get("risk_tier", "LOW")
                _ss_name  = _ss.get("company", "")
                _ss_tk    = _ss.get("ticker", "")
                _ss_8k    = _ss.get("recent_8k_count", 0)
                _ss_news  = len(_ss.get("news_signals", []))
                _ss_flags = _ss.get("flags", [])

                with st.expander(
                    f"{_ss_name}  ·  Score {_ss_score}/100  ·  {_ss_tier} RISK"
                    + (f"  [{_ss_tk}]" if _ss_tk else ""),
                    expanded=(_ss_tier == "HIGH"),
                ):
                    _sb_col1, _sb_col2 = st.columns([1, 3])
                    with _sb_col1:
                        st.markdown(
                            f'<div style="background:{_ss_color}22;border:2px solid {_ss_color};'
                            f'border-radius:12px;padding:1rem;text-align:center">'
                            f'<div style="font-size:2.2rem;font-weight:800;color:{_ss_color}">{_ss_score}</div>'
                            f'<div style="font-size:0.72rem;color:#94A3B8;text-transform:uppercase">{_ss_tier} RISK</div>'
                            f'<div style="font-size:0.75rem;color:#64748B;margin-top:0.4rem">'
                            f'{_ss_8k} 8-K filings · {_ss_news} news signals</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    with _sb_col2:
                        _ratios = _ss.get("ratios", {})
                        if _ratios:
                            _rlines = []
                            if _ratios.get("revenue_growth_pct") is not None:
                                _rlines.append(f"Revenue growth: {_ratios['revenue_growth_pct']:+.1f}%")
                            if _ratios.get("debt_to_assets_current") is not None:
                                _rlines.append(f"Debt/assets: {_ratios['debt_to_assets_current']:.2f}")
                            if _ratios.get("profit_margin_current") is not None:
                                _rlines.append(f"Profit margin: {_ratios['profit_margin_current']:.1%}")
                            if _ratios.get("cash") is not None:
                                _cash_b = _ratios["cash"] / 1e9
                                _rlines.append(f"Cash: ${_cash_b:.1f}B")
                            if _rlines:
                                st.markdown(
                                    '<div style="background:#060D1A;border:1px solid rgba(148,163,184,0.1);'
                                    'border-radius:8px;padding:0.5rem 0.8rem;margin-bottom:0.5rem;'
                                    'display:flex;gap:1.5rem;flex-wrap:wrap">'
                                    + "".join(
                                        f'<span style="font-size:0.82rem;color:#C4D3E8">{html.escape(r)}</span>'
                                        for r in _rlines
                                    )
                                    + '</div>',
                                    unsafe_allow_html=True,
                                )
                        for _fl in _ss_flags[:8]:
                            _fl_sev   = _fl.get("severity", "MEDIUM")
                            _fl_color = {"HIGH": "#F87171", "MEDIUM": "#FCD34D"}.get(_fl_sev, "#94A3B8")
                            st.markdown(
                                f'<div style="background:#060D1A;border-left:3px solid {_fl_color};'
                                f'border-radius:0 8px 8px 0;padding:0.3rem 0.7rem;margin-bottom:0.25rem;font-size:0.82rem">'
                                f'<span style="color:{_fl_color};font-weight:700">{_fl_sev}</span> · '
                                f'<span style="color:#94A3B8">{html.escape(_fl.get("source",""))}</span> — '
                                f'<span style="color:#C4D3E8">{html.escape(_fl.get("detail","")[:120])}</span>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )


# =========================================================
# RENDER
# =========================================================

# ── SUPPLIER PORTAL MODE ─────────────────────────────────────────────
# Accessed via: ?mode=portal&event=EVT-XXXXXX
# Suppliers land on a simplified form to submit their own data.
# No authentication required for portal — submissions stored in SQLite.
def render_supplier_portal():
    """Self-service supplier submission form, activated via ?mode=portal."""
    import time as _ptime
    _qp  = st.query_params
    _event_ref = _qp.get("event", "OPEN")

    st.markdown(
        '<div style="max-width:720px;margin:0 auto;padding:2rem 1rem">'
        '<div style="display:flex;align-items:center;gap:1rem;margin-bottom:2rem">'
        '<div style="font-size:2rem;font-weight:800;color:#60A5FA;letter-spacing:-0.04em">ProcureIQ</div>'
        '<div style="font-size:0.85rem;color:#94A3B8;background:#060D1A;border:1px solid rgba(96,165,250,0.2);'
        'border-radius:6px;padding:0.2rem 0.6rem">Supplier Portal</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        f'<h2 style="color:#F1F5F9;font-size:1.4rem;margin-bottom:0.3rem">Supplier Response Form</h2>'
        f'<p style="color:#94A3B8;font-size:0.88rem;margin-bottom:1.5rem">Event reference: <strong style="color:#60A5FA">{html.escape(_event_ref)}</strong> — '
        f'Complete all fields and submit. Your response is stored securely and reviewed by the procurement team.</p>',
        unsafe_allow_html=True,
    )

    with st.form("portal_submission_form"):
        st.markdown("##### Company Information")
        _p_company   = st.text_input("Company / Legal Entity Name *", placeholder="Acme Corp LLC")
        _p_contact   = st.text_input("Primary Contact Name *", placeholder="Jane Smith")
        _p_email     = st.text_input("Contact Email *", placeholder="jane@acme.com")
        _p_location  = st.text_input("HQ Location (City, State/Country)", placeholder="Houston, TX")
        _p_ein       = st.text_input("EIN / VAT / Business Registration Number", placeholder="12-3456789")

        st.markdown("##### Commercial Proposal")
        _p_price     = st.number_input("Total Quoted Price (USD)", min_value=0.0, step=1000.0, format="%.2f")
        _p_unit      = st.text_input("Unit / Pricing Basis", placeholder="Per unit, per year, lump sum…")
        _p_lead_days = st.number_input("Lead Time (days)", min_value=0, max_value=730, step=1)
        _p_warranty  = st.text_input("Warranty / SLA Commitment", placeholder="12 months parts & labour…")

        st.markdown("##### Capability & Compliance")
        _p_certs     = st.text_area("Certifications (ISO, SOC 2, GDPR, etc.)", placeholder="ISO 9001:2015, SOC 2 Type II…", height=80)
        _p_refs      = st.text_area("Reference Customers (optional)", placeholder="3 customers in similar industry…", height=80)
        _p_diversity = st.selectbox("Diversity Classification", ["None / Not applicable", "Minority-Owned (MBE)", "Women-Owned (WBE)", "Veteran-Owned (VOSB)", "Small Business (SB)", "HUBZone", "Other"])
        _p_coi       = st.checkbox("I confirm no conflict of interest with the procuring organisation")

        st.markdown("##### Additional Comments")
        _p_notes     = st.text_area("Notes / Exceptions / Value-adds", height=100)

        _submitted = st.form_submit_button("Submit Response", type="primary")

    if _submitted:
        if not _p_company.strip() or not _p_contact.strip() or not _p_email.strip():
            st.error("Company name, contact name, and email are required.")
        else:
            try:
                _portal_rec = {
                    "event_ref":   _event_ref,
                    "company":     _p_company.strip(),
                    "contact":     _p_contact.strip(),
                    "email":       _p_email.strip(),
                    "location":    _p_location.strip(),
                    "ein":         _p_ein.strip(),
                    "price":       _p_price,
                    "unit":        _p_unit.strip(),
                    "lead_days":   int(_p_lead_days),
                    "warranty":    _p_warranty.strip(),
                    "certs":       _p_certs.strip(),
                    "refs":        _p_refs.strip(),
                    "diversity":   _p_diversity,
                    "coi_clear":   _p_coi,
                    "notes":       _p_notes.strip(),
                    "submitted_at": _ptime.time(),
                }
                get_database().save_portal_submission(_event_ref, generate_id("SUB-", 6), _portal_rec)
                st.success(
                    f"Thank you, **{html.escape(_p_company.strip())}**! Your response has been submitted. "
                    "The procurement team will be in touch within 2 business days."
                )
                st.balloons()
            except Exception as _pe:
                st.error(f"Submission failed: {_pe}")

    st.markdown("</div>", unsafe_allow_html=True)


# ── AI PROVIDER SIDEBAR ──────────────────────────────────────────────
_render_api_key_sidebar()

# ── ROUTING ───────────────────────────────────────────────────────────
# Check for portal mode BEFORE authentication so suppliers can submit without an account.
# Portal URLs must include a valid HMAC token generated by the link generator.
_qp_check = st.query_params
if _qp_check.get("mode") == "portal":
    import hmac as _portal_hmac, hashlib as _portal_hl
    _p_event = _qp_check.get("event", "")
    _p_token = _qp_check.get("token", "")
    _p_secret = os.getenv("SECRET_KEY", "")
    # Fail closed: portal is disabled if SECRET_KEY is not configured.
    # Do not fall back to a default — a known fallback secret in a public repo
    # allows anyone to forge valid portal tokens for any event ID.
    if not _p_secret:
        st.error(
            "Supplier portal is disabled: SECRET_KEY environment variable is not configured. "
            "Set SECRET_KEY in your .env file or deployment environment to enable the portal."
        )
        st.stop()
    _p_expected = _portal_hmac.new(
        _p_secret.encode(), _p_event.encode(), _portal_hl.sha256
    ).hexdigest()[:24]
    _p_valid = _portal_hmac.compare_digest(_p_token, _p_expected) if _p_token else False
    if not _p_valid:
        st.error(
            "Invalid or missing portal access token. "
            "Please use the link generated from the Stakeholder Comms tab."
        )
        st.stop()
    render_supplier_portal()
    st.stop()

# Require authentication
username = require_authentication()

if st.session_state.entered_express:
    render_express()
elif st.session_state.entered_scan:
    render_quickscan()
elif st.session_state.entered_dashboard:
    render_dashboard()
else:
    render_cover()
